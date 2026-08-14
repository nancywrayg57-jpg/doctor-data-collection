from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
import time
import urllib.error
import urllib.request
from collections import Counter
from dataclasses import dataclass, replace
from datetime import date
from difflib import SequenceMatcher
from html import unescape
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "缺少依赖：需要 requests、beautifulsoup4、openpyxl。请使用已安装这些库的本机 Python 运行。"
    ) from exc


ROOT = Path(r"D:\workspace\信息收集整理")
VAULT = ROOT / "医生画像仓库"
SOURCE_DIR = VAULT / "99_资料来源"
WORK_DIR = ROOT / "work"
LEDGER_PATH = SOURCE_DIR / "珠三角三甲医院官网入口台账.xlsx"
WORKBOOK_BUILDER = WORK_DIR / "build_doctor_workbook.mjs"
MASTER_BASENAME = "珠三角三甲医院_医生画像自动采集总底表"
MASTER_JSON_PATH = WORK_DIR / f"{MASTER_BASENAME}_payload.json"
MASTER_CSV_PATH = SOURCE_DIR / f"{MASTER_BASENAME}.csv"
MASTER_XLSX_PATH = SOURCE_DIR / f"{MASTER_BASENAME}.xlsx"
MASTER_PREVIEW_PATH = WORK_DIR / f"{MASTER_BASENAME}_preview.png"
MASTER_REPORT_PATH = SOURCE_DIR / f"{MASTER_BASENAME}_更新报告.md"
BUNDLED_NODE = Path(
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "node"
    / "bin"
    / "node.exe"
)

BASE_HEADERS = [
    "序号",
    "医院",
    "姓名",
    "科室_分类页",
    "科室_列表卡片",
    "职称_关键词",
    "职称身份原文",
    "重点优先级",
    "重点关注范围",
    "重点疾病标签",
    "擅长诊疗方向摘录",
    "亮眼经历线索",
    "列表简介",
    "详情正文摘录",
    "来源类型",
    "来源链接",
    "照片链接",
    "照片文件",
    "采集入口",
    "采集方式",
    "采集日期",
    "详情页状态",
    "已建画像",
    "异常提示",
    "复核状态",
]

TITLE_TERMS = [
    "主任中医师",
    "副主任中医师",
    "主治中医师",
    "主任医师",
    "副主任医师",
    "主治医师",
    "住院医师",
    "医师",
    "主任护师",
    "副主任护师",
    "主管护师",
    "研究员",
    "副研究员",
    "教授",
    "副教授",
    "博士生导师",
    "硕士生导师",
    "医学博士",
    "医学硕士",
    "博士",
    "硕士",
    "科主任",
    "副主任",
    "首席专家",
    "院长",
    "副院长",
]

GROUP_KEYWORDS = {
    "慢性病": [
        "高血压",
        "糖尿病",
        "冠心病",
        "心力衰竭",
        "心衰",
        "高脂血症",
        "动脉粥样硬化",
        "慢性",
        "代谢",
    ],
    "肿瘤": [
        "肿瘤",
        "癌",
        "瘤",
        "淋巴瘤",
        "白血病",
        "放疗",
        "化疗",
        "靶向",
        "鼻咽癌",
        "肺癌",
        "肝癌",
        "胃癌",
        "肠癌",
        "乳腺癌",
    ],
    "生殖疾病": [
        "生殖",
        "不孕",
        "卵巢",
        "辅助生殖",
        "试管",
        "胚胎",
        "妇科内分泌",
        "多囊",
        "月经",
    ],
    "免疫/风湿/感染": [
        "风湿",
        "免疫",
        "感染",
        "免疫功能",
        "红斑狼疮",
        "类风湿",
        "强直性脊柱炎",
    ],
    "术后恢复/康复": [
        "术后",
        "康复",
        "恢复",
        "营养",
        "姑息",
        "疼痛",
        "创面",
        "烧伤",
        "伤口",
        "功能障碍",
    ],
    "疑难重症": [
        "疑难",
        "危重",
        "重症",
        "急危重",
        "难治",
        "复杂",
        "多学科",
        "MDT",
        "高危",
        "罕见",
        "转化治疗",
    ],
}

PRIORITY_DEPARTMENTS = [
    "呼吸",
    "消化",
    "内分泌",
    "血液",
    "风湿",
    "心血管",
    "胃肠",
    "肝胆",
    "胸部肿瘤",
    "腹盆部肿瘤",
    "中医肿瘤",
    "肿瘤介入",
    "鼻咽癌",
    "乳腺肿瘤",
    "放射肿瘤",
    "妇科",
    "生殖",
    "康复",
    "营养",
]

HIGHLIGHT_TERMS = [
    "科主任",
    "首席专家",
    "主任医师",
    "副主任医师",
    "博士生导师",
    "硕士生导师",
    "政府特殊津贴",
    "突出贡献",
    "高层次人才",
    "领军人才",
    "学科带头人",
    "国家重点研发",
    "重点研发",
    "访问",
    "研修",
    "进修",
    "美国",
    "英国",
    "德国",
    "国家自然科学基金",
    "SCI",
    "发表",
    "专利",
    "指南",
    "名医",
    "好医生",
    "人才",
    "奖",
    "委员会",
    "协会",
    "学会",
    "建设",
    "率先",
    "多学科",
    "MDT",
    "疑难",
    "危重",
    "精准",
]

GENERIC_ADAPTER_ID = "generic_official_template"
GDSKIN_ADAPTER_ID = "gdskin_aspnet_expert"
NY5Y_ADAPTER_ID = "ny5y_official_expert"
GDZY5413_ADAPTER_ID = "gdzy5413_official_specialist"
GYKQYY_ADAPTER_ID = "gykqyy_public_doctor_api"
GYFYYY_ADAPTER_ID = "gyfyyy_static_department_tree"
GY3Y_ADAPTER_ID = "gy3y_static_team_directory"
GZBRAIN_ADAPTER_ID = "gzbrain_static_expert_directory"
GZSZYY_ADAPTER_ID = "gzszyy_department_expert_directory"
GZSYS_ADAPTER_ID = "gzsys_drupal_doctor_cards"
FAHSYSU_ADAPTER_ID = "fahsysu_drupal_expert_directory"
GDGH_ADAPTER_ID = "gdghospital_static_department_expert"
GDGH_EXPECTED_GROUP_COUNT = 26
GDGH_EXPECTED_DEPARTMENT_COUNT = 83
GDGH_EXPECTED_RELATION_COUNT = 1343
GDGH_EXPECTED_NURSING_COUNT = 9
GDMCH_ADAPTER_ID = "gdmch_paginated_expert_photo"
GDMCH_EXPECTED_PAGE_COUNT = 111
GDMCH_EXPECTED_RELATION_COUNT = 884
GDMCH_EXPECTED_NON_DOCTOR_COUNT = 51
GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT = 658
GDMCH_EXPECTED_DEFAULT_PHOTO_COUNT = 225
GDMCH_EXPECTED_SAME_NAME_GROUPS = {
    "郭庆禄": ["34640", "34931"],
    "周真": ["32647", "32821"],
    "刘颖": ["32499", "33007"],
    "何裕": ["32750", "33134"],
}
GDMCH_SAME_IDENTITY_DETAIL_GROUPS = {frozenset({"34640", "34931"})}
GDMCH_EXPECTED_FINAL_IDENTITY_COUNT = (
    GDMCH_EXPECTED_RELATION_COUNT - GDMCH_EXPECTED_NON_DOCTOR_COUNT - 1
)
FAHSYSU_EXPECTED_SAME_NAME_GROUPS = {
    "庄锦涛": ["29148", "31480"],
    "涂响安": ["735", "31481"],
    "匡铭": ["650", "5582"],
    "梁力建": ["653", "21325"],
    "王伟": ["5592", "25409"],
    "刘敏": ["5684", "25838"],
    "陈宇": ["5708", "5784"],
    "何潇芳": ["38113", "38613"],
}
FAHSYSU_CAMPUS_MARKERS = (
    "院本部",
    "本部",
    "东院区",
    "东院",
    "南沙",
    "南院区",
    "黄埔",
    "院区",
)
GZSZYY_CARE_SITE_PATHS = {
    "/district1_zzlyq/": "珠玑院区",
    "/district1_thxyq/": "天河新院区",
    "/district1_tdfy/": "同德院区",
    "/district1_wymzb/": "五羊门诊部",
    "/district1_tdmzb/": "同德门诊部",
}
GZSZYY_SAME_IDENTITY_DETAIL_GROUPS = {
    frozenset({"ELe31Mb6", "JxboyNeg"}),  # 林少贞
    frozenset({"4QbYVOdz", "X7ax9byv"}),  # 唐瑾秋
    frozenset({"LDdwkmd1", "QBeXY8ay"}),  # 高三德
}
GZSZYY_DISTINCT_SAME_NAME_DETAIL_GROUPS = {
    frozenset({"3YaOggax", "WZdP6yaK"}),  # 王健：检验病理与外科两种身份
}
GZSZYY_CAMPUS_LABELS = {
    "珠玑路院区",
    "同德围分院",
    "同德综合门诊部",
    "五羊门诊部",
}
GYKQYY_DIRECTORY_API = "https://www.gykqyy.com/api/article/getZhuanjiaList"
GYKQYY_DETAIL_API = "https://www.gykqyy.com/api/article/getArticleDetail"
GENERIC_MAX_PAGES_DEFAULT = 60
GDSKIN_ENTRY_METADATA = {
    "901": {
        "category_name": "首席专家",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "902": {
        "category_name": "知名专家",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "906": {
        "category_name": "皮肤内科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "910": {
        "category_name": "外阴皮肤病/性病科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "913": {
        "category_name": "整形美容外科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "915": {
        "category_name": "中医皮肤科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "917": {
        "category_name": "激光美肤中心",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "921": {
        "category_name": "皮肤外科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "922": {
        "category_name": "儿童皮肤科",
        "affiliation": "南方医科大学皮肤病医院（官网专家团队栏目）",
    },
    "924": {
        "category_name": "珠江新城医学美容中心",
        "affiliation": "南方医科大学皮肤病医院·珠江新城医学美容中心",
    },
}
GDSKIN_GENERAL_CATEGORIES = {"首席专家", "知名专家"}
GDSKIN_EXPECTED_ENTRY_COUNTS = {
    "901": 1,
    "902": 3,
    "906": 29,
    "910": 7,
    "913": 8,
    "915": 14,
    "917": 6,
    "921": 4,
    "922": 5,
    "924": 0,
}
GDSKIN_EXPECTED_PAGE_COUNTS = {entry_id: (2 if entry_id == "906" else 1) for entry_id in GDSKIN_ENTRY_METADATA}
GDSKIN_DOCTOR_ROLE_TERMS = [
    "主任医师",
    "副主任医师",
    "主治医师",
    "住院医师",
    "主管医师",
    "医师",
    "医生",
    "首席专家",
]
NY5Y_ENTRY_METADATA = {
    "100": {
        "path": "/zhuanjia_mingyi.php",
        "category_name": "专家风采",
        "affiliation": "南方医科大学第五附属医院（官网专家风采栏目）",
    },
    "162": {
        "path": "/zhuanjia_lingnan.php",
        "category_name": "岭南名医",
        "affiliation": "南方医科大学第五附属医院（官网岭南名医荣誉栏目）",
    },
}
NY5Y_GENERAL_CATEGORIES = {"专家风采", "岭南名医"}
NY5Y_EXPECTED_ENTRY_COUNTS = {"100": 133, "162": 80}
NY5Y_EXPECTED_UNIQUE_COUNT = 134
NY5Y_EXPECTED_CROSS_ENTRY_DUPLICATES = 79
GDZY5413_ENTRY_METADATA = {
    "851": {
        "category_name": "名医名家",
        "affiliation": "广东省第二中医院（官网名医名家栏目）",
    },
    "852": {
        "category_name": "各科专家",
        "affiliation": "广东省第二中医院（官网各科专家栏目）",
    },
}
GENERIC_DETAIL_PATH_HINTS = [
    "doctor",
    "expert",
    "zhuanjia",
    "specialist",
    "physician",
    "staff",
    "team",
    "doctorinfo",
    "expertinfo",
    "node",
]
GENERIC_PATH_BLOCK_HINTS = [
    "news",
    "notice",
    "article",
    "video",
    "cggg",
    "cgjggg",
    "cgzb",
    "dzcggg",
    "scdygg",
    "qtgsxx",
    "xwzx",
    "mtbd",
    "zxtg",
    "zlkp",
    "boshihou",
    "bgpt",
    "tender",
    "zhaobiao",
    "caigou",
    "research",
    "keyan",
    "home",
]
GENERIC_STRONG_TITLE_TERMS = [
    "主任医师",
    "副主任医师",
    "主治医师",
    "住院医师",
    "医师",
    "主任护师",
    "副主任护师",
    "主管护师",
    "研究员",
    "副研究员",
    "教授",
    "副教授",
    "博士生导师",
    "硕士生导师",
    "科主任",
    "首席专家",
]
GENERIC_DETAIL_TEXT_HINTS = [
    "姓名",
    "医生",
    "专家",
    "科室",
    "职称",
    "擅长",
    "专长",
    "简介",
    "出诊",
    "门诊",
]
GENERIC_IGNORED_LINK_TEXTS = [
    "首页",
    "更多",
    "更多>>",
    "查看",
    "查看详情",
    "详情",
    "上一页",
    "下一页",
    "尾页",
    "末页",
    "返回",
    "预约挂号",
    "在线预约",
    "登录",
]
GENERIC_NAME_BLOCK_TERMS = [
    "医院",
    "中心",
    "科室",
    "专科",
    "专家",
    "医生",
    "主任",
    "医师",
    "教授",
    "导师",
    "博士",
    "硕士",
    "门诊",
    "简介",
    "擅长",
    "详情",
    "预约",
    "全部",
    "面包屑",
    "院士",
    "风采",
    "临床",
    "护理",
    "名医",
    "名家",
    "列表",
    "团队",
    "学科",
    "护理",
    "公告",
    "新闻",
    "患者",
    "就诊须知",
    "住院须知",
    "联系我们",
    "门诊时间",
]

PROFILE_NOISE_TAGS = {
    "aside",
    "button",
    "footer",
    "form",
    "nav",
    "noscript",
    "script",
    "style",
    "svg",
    "template",
}
PROFILE_NOISE_ATTRIBUTE_TOKENS = {
    "action-tools",
    "actions",
    "breadcrumb",
    "breadcrumbs",
    "crumb",
    "crumbs",
    "footer",
    "header",
    "hidden",
    "menu",
    "nav",
    "navigation",
    "pager",
    "pagination",
    "print",
    "share",
    "sidebar",
    "side-nav",
    "sidenav",
    "site-footer",
    "site-header",
    "social",
    "sr-only",
    "toolbar",
    "visually-hidden",
}
PROFILE_INLINE_NOISE_PHRASES = [
    "加载更多",
    "打印本页",
    "关闭窗口",
    "返回顶部",
]


@dataclass
class HospitalTarget:
    city: str
    hospital: str
    homepage: str
    entry_url: str
    difficulty: str
    review: str
    adapter_id: str
    entry_urls: tuple[str, ...] = ()
    ledger_entry_url: str = ""


def effective_entry_urls(target: HospitalTarget) -> list[str]:
    values = target.entry_urls or (target.entry_url,)
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = clean_text(value)
        key = canonical_url(cleaned)
        if cleaned and key not in seen:
            seen.add(key)
            unique.append(cleaned)
    return unique


def clean_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", unescape(value or "")).strip()


def clip(value: str | None, max_len: int) -> str:
    text = clean_text(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def first_nonempty(*values: str | None) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def safe_file_part(value: str) -> str:
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", value).strip("_")
    return text or "unnamed"


def canonical_url(value: str) -> str:
    parsed = urlparse(value)
    netloc = parsed.netloc.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return parsed._replace(netloc=netloc, fragment="").geturl().rstrip("/")


def fetch(session: requests.Session, url: str, retries: int = 3) -> tuple[int | None, str, str]:
    last_error = ""
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, timeout=35)
            if response.status_code == 200:
                response.encoding = response.apparent_encoding or response.encoding or "utf-8"
                return response.status_code, response.text, ""
            last_error = f"HTTP {response.status_code}"
        except Exception as exc:  # noqa: BLE001 - keep collection failure visible
            last_error = str(exc)
        time.sleep(0.8 * attempt)
    return None, "", last_error


def fetch_standard_public_get(url: str, retries: int = 3) -> tuple[int | None, str, str]:
    """Use a plain non-browser GET for sites that reject browser-like clients.

    No cookies, browser fingerprint headers, proxy, challenge solver, or private endpoint is used.
    """

    last_error = ""
    for attempt in range(1, retries + 1):
        try:
            request = urllib.request.Request(url, method="GET")
            with urllib.request.urlopen(request, timeout=35) as response:
                status = int(response.status)
                body = response.read()
                if status == 200:
                    charset = response.headers.get_content_charset() or "utf-8"
                    return status, body.decode(charset, errors="replace"), ""
                last_error = f"HTTP {status}"
        except urllib.error.HTTPError as exc:
            last_error = f"HTTP {exc.code}"
        except Exception as exc:  # noqa: BLE001 - keep collection failure visible
            last_error = str(exc)
        time.sleep(0.8 * attempt)
    return None, "", last_error


def create_official_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36; "
                "public official-site collection"
            )
        }
    )
    return session


def extract_terms(text: str, terms: list[str]) -> list[str]:
    found: list[str] = []
    used_spans: list[tuple[int, int]] = []
    for term in sorted(terms, key=len, reverse=True):
        matched = False
        for match in re.finditer(re.escape(term), text):
            span = match.span()
            if any(max(span[0], used[0]) < min(span[1], used[1]) for used in used_spans):
                continue
            used_spans.append(span)
            matched = True
        if matched and term not in found:
            found.append(term)
    return [term for term in terms if term in found]


def extract_sentences(text: str, terms: list[str], limit: int = 4, max_len: int = 420) -> str:
    parts = [clean_text(part) for part in re.split(r"(?<=[。！？；;])|\n", text) if clean_text(part)]
    hits: list[str] = []
    for part in parts:
        if any(term in part for term in terms):
            hits.append(part)
        if len(hits) >= limit:
            break
    return clip(" ".join(hits), max_len)


def group_tags(text: str) -> tuple[list[str], list[str]]:
    groups: list[str] = []
    tags: list[str] = []
    for group, keywords in GROUP_KEYWORDS.items():
        matches = extract_terms(text, keywords)
        if matches:
            groups.append(group)
            tags.extend(matches)
    return groups, list(dict.fromkeys(tags))


def strip_article_tail(text: str) -> str:
    cleaned = clean_text(text)
    markers = [
        "上一篇：",
        "下一篇：",
        "相关文章：",
        "上一篇:",
        "下一篇:",
        "相关文章:",
        "最新文章",
        "南部战区空军医院 南部战区空军医院",
    ]
    positions = [cleaned.find(marker) for marker in markers if cleaned.find(marker) >= 0]
    if positions:
        cleaned = cleaned[: min(positions)]
    return clean_text(cleaned)


def infer_department(text: str) -> str:
    source = clean_text(text)
    if not source:
        return ""
    pattern = (
        r"([\u4e00-\u9fff、]{1,24}?"
        r"(?:内科|外科|妇科|儿科|骨科|皮肤科|肿瘤科|口腔科|营养科|保健办|科|中心|病房|门诊))"
        r"(?=主任|副主任|主治|医师|专家|博士|硕士|，|,|；|;|\s|$)"
    )
    candidates = [clean_text(match) for match in re.findall(pattern, source)]
    candidates = [
        candidate
        for candidate in candidates
        if not any(term in candidate for term in ["本中心", "聘任", "攻读", "获得", "入职"])
    ]
    candidates = [candidate for candidate in candidates if candidate]
    return candidates[-1] if candidates else ""


def clean_generic_department(value: str | None) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"^(?:科室|所在科室|出诊科室|专业|专科)\s*[:：]\s*", "", text)
    stop_pattern = (
        r"\s*(?:介绍|简介|个人简介|职称|职务|专业擅长|擅长|医疗专长|诊疗专长|"
        r"技术专长|业务专长|专长|研究方向)\s*[:：]"
        r"|\s*(?=(?:19|20)\d{2}年)"
        r"|[。！？；;]"
    )
    text = clean_text(re.split(stop_pattern, text, maxsplit=1)[0])
    match = re.match(
        r"^[\u4e00-\u9fff、]{1,28}?(?:科|中心|病房|门诊)(?:[一二三四五六七八九十\d]+室)?",
        text,
    )
    if match:
        return clean_text(match.group(0))
    return text if len(text) <= 30 else ""


def extract_gdzy5413_department(title_text: str | None, profile_text: str | None) -> str:
    source = clean_text(" ".join(value for value in [title_text, profile_text] if value))
    patterns = [
        r"广东省第二中医院([\u4e00-\u9fff\d]{1,20}(?:科|中心)(?:[一二三四五六七八九十\d]+区)?)\s*(?:主任|区长|负责人)",
        r"([\u4e00-\u9fff\d]{1,20}(?:科|中心)(?:[一二三四五六七八九十\d]+区)?)\s*(?:主任|区长|负责人)",
        r"([\u4e00-\u9fff\d]{1,20}科)主任(?:中|西)医师",
    ]
    blocked = {"医务科"}
    for pattern in patterns:
        match = re.search(pattern, source)
        if not match:
            continue
        candidate = clean_text(match.group(1))
        if candidate and candidate not in blocked:
            return candidate
    return ""


def has_department_text_pollution(raw_value: str | None, cleaned_value: str | None) -> bool:
    raw = clean_text(raw_value)
    cleaned = clean_text(cleaned_value)
    return bool(raw and cleaned and raw != cleaned)


def read_ledger(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["入口台账"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        item = {headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))}
        if item.get("医院名称"):
            rows.append(item)
    return rows


def normalize_bottom_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for header in BASE_HEADERS:
        value = row.get(header, "")
        if value is None:
            normalized[header] = ""
        elif header == "序号":
            normalized[header] = value
        else:
            normalized[header] = clean_text(str(value))
    return normalized


def doctor_row_key(row: dict[str, Any]) -> tuple[str, ...]:
    hospital = clean_text(str(row.get("医院") or ""))
    name = clean_text(str(row.get("姓名") or ""))
    department = first_nonempty(
        str(row.get("科室_分类页") or ""),
        str(row.get("科室_列表卡片") or ""),
    )
    source = clean_text(str(row.get("来源链接") or ""))
    warnings = clean_text(str(row.get("异常提示") or ""))
    if "同名待甄别" in warnings and hospital and source:
        return ("source", hospital, canonical_url(source))
    if hospital and name and department:
        return ("doctor", hospital, name, department)
    if hospital and source:
        return ("source", hospital, canonical_url(source))
    return ("row", hospital, name, source)


def row_quality_score(row: dict[str, Any]) -> int:
    quality_fields = [
        "职称_关键词",
        "职称身份原文",
        "重点关注范围",
        "重点疾病标签",
        "擅长诊疗方向摘录",
        "亮眼经历线索",
        "列表简介",
        "详情正文摘录",
        "来源链接",
    ]
    nonempty = sum(1 for field in quality_fields if clean_text(str(row.get(field) or "")))
    length = sum(len(clean_text(str(row.get(field) or ""))) for field in quality_fields)
    return nonempty * 1000 + length


def read_bottom_table_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "自动采集底表" not in workbook.sheetnames:
        return []
    sheet = workbook["自动采集底表"]
    raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [clean_text(str(header or "")) for header in raw_headers]
    index_by_header = {header: index for index, header in enumerate(headers) if header}
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values or not any(values):
            continue
        item = {}
        for header in BASE_HEADERS:
            column_index = index_by_header.get(header)
            item[header] = values[column_index] if column_index is not None and column_index < len(values) else ""
        normalized = normalize_bottom_row(item)
        if not normalized["医院"] and not normalized["姓名"] and not normalized["来源链接"]:
            continue
        rows.append(normalized)
    return rows


def load_existing_rows_for_master() -> tuple[list[dict[str, Any]], str, bool]:
    if MASTER_XLSX_PATH.exists():
        return read_bottom_table_rows(MASTER_XLSX_PATH), str(MASTER_XLSX_PATH), True

    legacy_paths = [
        path
        for path in sorted(SOURCE_DIR.glob("*_全院医生自动采集底表.xlsx"))
        if path.name != MASTER_XLSX_PATH.name
    ]
    rows: list[dict[str, Any]] = []
    for path in legacy_paths:
        rows.extend(read_bottom_table_rows(path))
    source_label = "；".join(str(path) for path in legacy_paths) if legacy_paths else "无既有底表"
    return rows, source_label, False


def merge_rows_for_master(
    existing_rows: list[dict[str, Any]],
    incoming_rows: list[dict[str, Any]],
    preserve_existing: bool,
    refresh_incoming: bool = False,
) -> tuple[list[dict[str, Any]], int, int, int, int]:
    merged: list[dict[str, Any]] = []
    index_by_key: dict[tuple[str, ...], int] = {}
    existing_duplicates = 0
    incoming_added = 0
    incoming_skipped = 0
    incoming_refreshed = 0

    for row in existing_rows:
        normalized = normalize_bottom_row(row)
        key = doctor_row_key(normalized)
        if key in index_by_key:
            existing_duplicates += 1
            if not preserve_existing and row_quality_score(normalized) > row_quality_score(merged[index_by_key[key]]):
                merged[index_by_key[key]] = normalized
            continue
        index_by_key[key] = len(merged)
        merged.append(normalized)

    for row in incoming_rows:
        normalized = normalize_bottom_row(row)
        key = doctor_row_key(normalized)
        if key in index_by_key:
            if refresh_incoming:
                merged[index_by_key[key]] = normalized
                incoming_refreshed += 1
            else:
                incoming_skipped += 1
            continue
        index_by_key[key] = len(merged)
        merged.append(normalized)
        incoming_added += 1

    for index, row in enumerate(merged, start=1):
        row["序号"] = index
    return merged, incoming_added, incoming_skipped, incoming_refreshed, existing_duplicates


def build_hospital_batches(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        hospital = clean_text(str(row.get("医院") or "未识别医院"))
        grouped.setdefault(hospital, []).append(row)

    batches: list[dict[str, Any]] = []
    for hospital, hospital_rows in sorted(grouped.items()):
        dates = sorted({clean_text(str(row.get("采集日期") or "")) for row in hospital_rows if row.get("采集日期")})
        review_count = sum(1 for row in hospital_rows if clean_text(str(row.get("复核状态") or "")) != "已复核")
        entry_urls = list(
            dict.fromkeys(
                clean_text(str(row.get("采集入口") or ""))
                for row in hospital_rows
                if clean_text(str(row.get("采集入口") or ""))
            )
        )
        batches.append(
            {
                "医院": hospital,
                "医生数": len(hospital_rows),
                "采集日期": "、".join(dates),
                "待复核数": review_count,
                "已建画像数": sum(1 for row in hospital_rows if clean_text(str(row.get("已建画像") or "")) == "是"),
                "采集入口": "、".join(entry_urls),
            }
        )
    return batches


def sync_profile_flags(rows: list[dict[str, Any]], profile_links: set[str]) -> None:
    normalized_profile_links = {canonical_url(link) for link in profile_links if canonical_url(link)}
    for row in rows:
        source = canonical_url(str(row.get("来源链接") or ""))
        row["已建画像"] = "是" if source and source in normalized_profile_links else "否"


def covered_department_names(rows: list[dict[str, Any]]) -> list[str]:
    return sorted(
        {
            clean_text(department)
            for row in rows
            if "非医生页面或姓名异常" not in str(row.get("异常提示") or "")
            for department in clean_text(str(row.get("科室_分类页") or "")).split("、")
            if clean_text(department)
        }
    )


def build_master_payload(
    today: str,
    incoming_payload: dict[str, Any] | None = None,
    refresh_incoming: bool = False,
    replace_incoming_hospital: bool = False,
    batch_meta_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    existing_rows, source_label, preserve_existing = load_existing_rows_for_master()
    incoming_rows = incoming_payload["rows"] if incoming_payload else []
    incoming_hospital = clean_text(
        str((incoming_payload or {}).get("meta", {}).get("hospital") or "")
    )
    if replace_incoming_hospital and incoming_hospital:
        existing_rows = [
            row
            for row in existing_rows
            if clean_text(str(row.get("医院") or "")) != incoming_hospital
        ]
    rows, added, skipped, refreshed, existing_duplicates = merge_rows_for_master(
        existing_rows,
        incoming_rows,
        preserve_existing=preserve_existing,
        refresh_incoming=refresh_incoming,
    )
    sync_profile_flags(rows, collect_existing_profile_links())

    category_counter = Counter()
    priority_counter = Counter()
    group_counter = Counter()
    warning_counter = Counter()
    for row in rows:
        department = first_nonempty(str(row.get("科室_分类页") or ""), str(row.get("科室_列表卡片") or ""))
        if department:
            category_counter[department] += 1
        priority = clean_text(str(row.get("重点优先级") or "普通"))
        if priority:
            priority_counter[priority] += 1
        for group in clean_text(str(row.get("重点关注范围") or "")).split("、"):
            if group:
                group_counter[group] += 1
        for warning in clean_text(str(row.get("异常提示") or "")).split("；"):
            if warning:
                warning_counter[warning] += 1

    incoming_meta = incoming_payload.get("meta", {}) if incoming_payload else {}
    batch_meta = incoming_meta if incoming_payload else (batch_meta_override or {})
    hospital_batches = build_hospital_batches(rows)
    return {
        "meta": {
            "city": "珠三角",
            "hospital": "珠三角三甲医院医生画像总表",
            "homepage": "",
            "entry_url": "详见各行采集入口",
            "adapter_id": "multi_hospital_official_pipeline",
            "collected_at": today,
            "category_count": len({row.get("采集入口") for row in rows if row.get("采集入口")}),
            "raw_card_rows": batch_meta.get("raw_card_rows", 0),
            "unique_doctor_count": len(rows),
            "category_error_count": batch_meta.get("category_error_count", 0),
            "detail_error_count": batch_meta.get("detail_error_count", 0),
            "existing_profile_count": sum(1 for row in rows if clean_text(str(row.get("已建画像") or "")) == "是"),
            "ledger_review": "多院汇总，详见官网入口台账",
            "ledger_difficulty": "多院汇总",
            "source_seed": source_label,
            "current_batch_hospital": (
                batch_meta.get("hospital", "")
                if incoming_payload
                else batch_meta.get("current_batch_hospital", "")
            ),
            "current_batch_rows": (
                len(incoming_rows)
                if incoming_payload
                else int(batch_meta.get("current_batch_rows") or 0)
            ),
            "new_rows_added": (
                added if incoming_payload else int(batch_meta.get("new_rows_added") or 0)
            ),
            "duplicate_rows_skipped": (
                skipped
                if incoming_payload
                else int(batch_meta.get("duplicate_rows_skipped") or 0)
            ),
            "existing_rows_refreshed": (
                refreshed
                if incoming_payload
                else int(batch_meta.get("existing_rows_refreshed") or 0)
            ),
            "existing_duplicate_rows": (
                existing_duplicates
                if incoming_payload
                else int(batch_meta.get("existing_duplicate_rows") or 0)
            ),
            "hospital_count": len(hospital_batches),
        },
        "categories": [],
        "category_errors": incoming_payload.get("category_errors", []) if incoming_payload else [],
        "detail_errors": incoming_payload.get("detail_errors", []) if incoming_payload else [],
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "hospital_batches": hospital_batches,
        "rows": rows,
    }


def dedicated_adapter_for(entry_url: str) -> str:
    parsed = urlparse(entry_url or "")
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    if host.removeprefix("www.") == "gdskin.com" and path == "/showclass.aspx":
        entry_ids = [
            value
            for name, value in parse_qsl(parsed.query, keep_blank_values=True)
            if name.lower() == "id"
        ]
        if entry_ids and entry_ids[0] in GDSKIN_ENTRY_METADATA:
            return GDSKIN_ADAPTER_ID
    if ny5y_entry_kind(entry_url):
        return NY5Y_ADAPTER_ID
    if gdzy5413_entry_kind(entry_url):
        return GDZY5413_ADAPTER_ID
    if host.removeprefix("www.") == "gykqyy.com" and path == "/list.html":
        category_ids = [
            value
            for name, value in parse_qsl(parsed.query, keep_blank_values=True)
            if name.lower() == "category"
        ]
        if category_ids == ["55"]:
            return GYKQYY_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gyfyyy.cn"
        and path == "/cn/ks/"
        and not parsed.query
        and not parsed.fragment
    ):
        return GYFYYY_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gy3y.cn"
        and path == "/ks/team.html"
        and not parsed.query
        and not parsed.fragment
    ):
        return GY3Y_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gzbrain.cn"
        and path == "/myzj/list.html"
        and not parsed.query
        and not parsed.fragment
    ):
        return GZBRAIN_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gzszyy.com"
        and path == "/expert/"
        and not parsed.query
        and not parsed.fragment
    ):
        return GZSZYY_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gzsys.org.cn"
        and path.rstrip("/") == "/doctor/592/search"
        and not parsed.query
        and not parsed.fragment
    ):
        return GZSYS_ADAPTER_ID
    if (
        host.removeprefix("www.") == "fahsysu.org.cn"
        and path.rstrip("/") == "/page/6945"
        and not parsed.query
        and not parsed.fragment
    ):
        return FAHSYSU_ADAPTER_ID
    if (
        host.removeprefix("www.") == "gdghospital.org.cn"
        and path.rstrip("/") == "/departmentsearch/list.html"
        and not parsed.query
        and not parsed.fragment
    ):
        return GDGH_ADAPTER_ID
    if (
        host.removeprefix("www.") == "e3861.com"
        and path.rstrip("/") == "/keshizhuanjia/zhuanjiajieshao"
        and not parsed.query
        and not parsed.fragment
    ):
        return GDMCH_ADAPTER_ID
    if "gzzoc.org.cn" in host and "/expert-introduction" in path:
        return "gzzoc_drupal_doctor"
    if "nbkjyy.mil.cn" in host and "/expert" in path:
        return "nbkjyy_static_expert"
    return ""


def adapter_for(entry_url: str, include_generic: bool = False) -> str:
    adapter_id = dedicated_adapter_for(entry_url)
    if adapter_id:
        return adapter_id
    parsed = urlparse(entry_url or "")
    if include_generic and parsed.scheme in {"http", "https"} and parsed.netloc:
        return GENERIC_ADAPTER_ID
    return ""


def confirmed_a_targets(rows: list[dict[str, Any]], include_generic: bool = False) -> list[HospitalTarget]:
    """Return every human-confirmed target; A is priority metadata, not authorization."""
    targets: list[HospitalTarget] = []
    for row in rows:
        review = clean_text(str(row.get("人工复核结果") or ""))
        difficulty = clean_text(str(row.get("采集难度_初判") or ""))
        entry_url = clean_text(str(row.get("医生目录入口_候选") or ""))
        adapter_id = adapter_for(entry_url, include_generic=include_generic)
        if review != "确认可采集":
            continue
        if not adapter_id:
            continue
        targets.append(
            HospitalTarget(
                city=clean_text(str(row.get("城市") or "")),
                hospital=clean_text(str(row.get("医院名称") or "")),
                homepage=clean_text(str(row.get("官网首页_候选") or "")),
                entry_url=entry_url,
                difficulty=difficulty,
                review=review,
                adapter_id=adapter_id,
            )
        )
    return targets


def collect_existing_profile_links() -> set[str]:
    links: set[str] = set()
    if not VAULT.exists():
        return links
    for path in VAULT.rglob("*.md"):
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        for match in re.finditer(r"来源链接[:：]\s*(https?://\S+)", text):
                links.add(canonical_url(match.group(1)))
    return links


def with_query_param(url: str, key: str, value: str | int) -> str:
    parsed = urlparse(url)
    pairs = [(name, item) for name, item in parse_qsl(parsed.query, keep_blank_values=True) if name != key]
    pairs.append((key, str(value)))
    return urlunparse(parsed._replace(query=urlencode(pairs)))


def comparable_host(url: str) -> str:
    host = urlparse(url).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def gdskin_entry_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdskin.com" or parsed.path.lower() != "/showclass.aspx":
        return ""
    for name, value in parse_qsl(parsed.query, keep_blank_values=True):
        if name.lower() == "id" and value in GDSKIN_ENTRY_METADATA:
            return value
    return ""


def gdskin_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdskin.com" or parsed.path.lower() != "/shownews.aspx":
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    if len(pairs) != 1 or pairs[0][0].lower() != "id" or not pairs[0][1].isdigit():
        return ""
    return pairs[0][1]


def ny5y_entry_kind(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "ny5y.cn":
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    if len(pairs) != 1 or pairs[0][0].lower() != "id":
        return ""
    entry_id = pairs[0][1]
    metadata = NY5Y_ENTRY_METADATA.get(entry_id)
    if not metadata or parsed.path.lower() != metadata["path"]:
        return ""
    return entry_id


def ny5y_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "ny5y.cn" or parsed.path.lower() != "/yisheng_xq.php":
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    if len(pairs) != 1 or pairs[0][0].lower() != "id" or not pairs[0][1].isdigit():
        return ""
    return pairs[0][1]


def gdzy5413_entry_kind(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdzy5413.com" or parsed.path.lower() != "/main/famousdoctorinfo.aspx":
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    if len(pairs) != 3:
        return ""
    params = {name.lower(): value for name, value in pairs}
    if len(params) != 3 or params.get("fid") != "81" or params.get("pid") != "850":
        return ""
    cid = params.get("cid", "")
    return cid if cid in GDZY5413_ENTRY_METADATA else ""


def gdzy5413_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdzy5413.com" or parsed.path.lower() != "/main/doctor/specialist.aspx":
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    if len(pairs) != 1 or pairs[0][0].lower() != "typeid" or not pairs[0][1].isdigit():
        return ""
    return pairs[0][1]


def gdzy5413_ksdoctor_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if (
        comparable_host(parsed.geturl()) != "gdzy5413.com"
        or parsed.path.lower() != "/main/ks/templet2/ksdoctorinfo.aspx"
    ):
        return ""
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    allowed = {"bid", "typeid", "cid", "ksid", "id"}
    if len(pairs) != 5 or {name.lower() for name, _value in pairs} != allowed:
        return ""
    params = {name.lower(): value for name, value in pairs}
    if any(not params[name].isdigit() for name in allowed):
        return ""
    if not (params["bid"] == params["cid"] and params["typeid"] == params["ksid"]):
        return ""
    return params["id"]


def generic_detail_identity(url: str) -> str:
    detail_id = gdskin_detail_id(url)
    if detail_id:
        return f"gdskin:{detail_id}"
    detail_id = ny5y_detail_id(url)
    if detail_id:
        return f"ny5y:{detail_id}"
    detail_id = gdzy5413_detail_id(url)
    if detail_id:
        return f"gdzy5413:{detail_id}"
    detail_id = gdzy5413_ksdoctor_detail_id(url)
    return f"gdzy5413:ksdoctor:{detail_id}" if detail_id else canonical_url(url)


def is_same_official_host(base_url: str, candidate_url: str) -> bool:
    return comparable_host(base_url) == comparable_host(candidate_url)


def is_collectable_url(base_url: str, candidate_url: str) -> bool:
    parsed = urlparse(candidate_url)
    if parsed.scheme not in {"http", "https"}:
        return False
    if not is_same_official_host(base_url, candidate_url):
        return False
    path = parsed.path.lower()
    if any(path.endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".gif", ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".zip"]):
        return False
    if any(block in path for block in ["/login", "/register", "/search", "/user", "/admin"]):
        return False
    return True


def compact_visible_text(element: Any, max_len: int = 1200) -> str:
    return clip(clean_text(element.get_text(" ", strip=True)), max_len)


def strip_profile_navigation_text(value: str | None) -> str:
    text = clean_text(value)
    if not text:
        return ""

    separator = r"(?:/|>|＞|»|›|→)"
    label = r"(?:姓名|科室|所在科室|职称|职务|专业擅长|擅长|专长|简介|个人简介|一、|二、)"
    breadcrumb_patterns = [
        (
            r"^.*?(?:你)?当前所在的位置\s*[:：]?\s*.*?"
            r"(?:专家|医师)信息\s*[^。！？；;\n]{0,160}?"
            r"科室\s*[:：]\s*[\u4e00-\u9fff、（）()]{1,30}?"
            r"(?:科(?:一室|二室)?|中心)\s*"
        ),
        rf"(?:临床专家\s*)?面包屑\s*首页\s*(?:{separator}\s*[^/＞>»›→]{{1,60}}?\s*){{1,10}}(?={label}|$)",
        rf"(?:你)?当前所在的位置\s*[:：]?\s*[^/＞>»›→]{{0,60}}\s*(?:{separator}\s*[^/＞>»›→]{{1,60}}?\s*){{1,10}}(?={label}|$)",
        rf"当前位置\s*[:：]?\s*首页\s*(?:{separator}\s*[^/＞>»›→]{{1,60}}?\s*){{1,10}}(?={label}|$)",
    ]
    for pattern in breadcrumb_patterns:
        text = re.sub(pattern, " ", text, count=1)
    for phrase in PROFILE_INLINE_NOISE_PHRASES:
        text = text.replace(phrase, " ")
    return clean_text(text)


def contains_navigation_text(value: str | None) -> bool:
    text = clean_text(value)
    if not text:
        return False
    return any(
        marker in text
        for marker in ["你当前所在的位置", "当前所在的位置", "当前位置", "面包屑", "首页 >", "首页＞"]
    )


def extract_clean_highlights(value: str | None) -> str:
    cleaned_source = strip_profile_navigation_text(value)
    highlights = extract_sentences(cleaned_source, HIGHLIGHT_TERMS)
    if contains_navigation_text(highlights):
        return ""
    return highlights


def element_attribute_tokens(element: Any) -> set[str]:
    values: list[str] = []
    element_id = element.get("id")
    if element_id:
        values.append(str(element_id))
    values.extend(str(value) for value in (element.get("class") or []))
    return {
        token
        for value in values
        for token in re.split(r"\s+", value.lower())
        if token
    }


def is_profile_noise_element(element: Any) -> bool:
    name = str(getattr(element, "name", "") or "").lower()
    if name in PROFILE_NOISE_TAGS:
        return True
    if clean_text(element.get("role")).lower() == "navigation":
        return True
    if element.has_attr("hidden") or clean_text(element.get("aria-hidden")).lower() == "true":
        return True

    tokens = element_attribute_tokens(element)
    if tokens & PROFILE_NOISE_ATTRIBUTE_TOKENS:
        return True
    return any("breadcrumb" in token for token in tokens)


def extract_navigation_context(soup: BeautifulSoup) -> str:
    candidates: list[tuple[int, str]] = []
    seen: set[str] = set()
    for element in soup.find_all(True):
        tokens = element_attribute_tokens(element)
        role = clean_text(element.get("role")).lower()
        has_breadcrumb_token = (
            bool(tokens & {"breadcrumb", "breadcrumbs", "crumb", "crumbs"})
            or any("breadcrumb" in token for token in tokens)
        )
        is_navigation = role == "navigation" or has_breadcrumb_token
        if not is_navigation:
            continue
        text = compact_visible_text(element, 1200)
        if not text or text in seen:
            continue
        seen.add(text)
        score = 100 if has_breadcrumb_token else 0
        if "面包屑" in text or "当前位置" in text:
            score += 50
        if "首页" in text and any(separator in text for separator in ["/", ">", "＞", "»", "›"]):
            score += 20
        candidates.append((score, text))
    contexts = [text for _, text in sorted(candidates, key=lambda item: item[0], reverse=True)]
    return clip(" ".join(contexts), 2400)


def remove_profile_noise_elements(soup: BeautifulSoup) -> None:
    for element in reversed(soup.find_all(True)):
        if is_profile_noise_element(element):
            element.decompose()


def nearest_card_text(anchor: Any) -> str:
    parts = [clean_text(anchor.get_text(" ", strip=True))]
    image = anchor.find("img")
    if image:
        parts.append(clean_text(image.get("alt")))
        parts.append(clean_text(image.get("title")))

    seed_text = clean_text(" ".join(part for part in parts if part))
    parent = anchor.parent
    for _ in range(4):
        if not parent:
            break
        text = compact_visible_text(parent)
        link_count = len(parent.find_all("a"))
        if 4 <= len(text) <= 700:
            strong_title_hits = extract_terms(text, GENERIC_STRONG_TITLE_TERMS)
            if link_count <= 4 and (strong_title_hits or "擅长" in text or "专长" in text or looks_like_person_name(seed_text)):
                return strip_profile_navigation_text(" ".join([seed_text, text]))
        parent = parent.parent
    return strip_profile_navigation_text(seed_text)


def looks_like_person_name(value: str) -> bool:
    candidate = clean_text(value).strip("：:，,、；;（）()[]【】<>《》")
    if "·" in candidate:
        if not re.fullmatch(r"[\u4e00-\u9fff·]{3,8}", candidate):
            return False
    elif not re.fullmatch(r"[\u4e00-\u9fff]{2,4}", candidate):
        return False
    if any(term in candidate for term in GENERIC_NAME_BLOCK_TERMS):
        return False
    return True


def matches_generic_directory_detail_url(entry_url: str, candidate_url: str) -> bool:
    if gdskin_entry_id(entry_url):
        return bool(gdskin_detail_id(candidate_url))
    if ny5y_entry_kind(entry_url):
        return bool(ny5y_detail_id(candidate_url))
    if (gdzy5413_id := gdzy5413_entry_kind(entry_url)):
        return bool(
            gdzy5413_detail_id(candidate_url)
            or (gdzy5413_id == "852" and gdzy5413_ksdoctor_detail_id(candidate_url))
        )
    entry_match = re.fullmatch(r"/section/(\d+)/?", urlparse(entry_url).path)
    if comparable_host(entry_url) != "smukqyy.cn" or not entry_match:
        return True
    directory_id = re.escape(entry_match.group(1))
    return bool(re.fullmatch(rf"/prods/{directory_id}/\d+/?", urlparse(candidate_url).path))


def generic_record_quality(
    name: str,
    source_link: str,
    entry_url: str,
    detail: dict[str, str],
    item: dict[str, str],
) -> tuple[bool, list[str]]:
    valid = looks_like_person_name(name) and matches_generic_directory_detail_url(entry_url, source_link)
    warnings: list[str] = []
    if gdskin_entry_id(entry_url):
        role_text = " ".join(
            [
                clean_text(detail.get("title_field")),
                clean_text(item.get("list_title")),
            ]
        )
        if not any(term in role_text for term in GDSKIN_DOCTOR_ROLE_TERMS):
            valid = False
            warnings.append("专家团队列表身份不属于医生角色")
    if not valid:
        warnings.append("非医生页面或姓名异常")
    if detail.get("department_polluted") == "yes" or has_department_text_pollution(
        item.get("department"), clean_generic_department(item.get("department"))
    ):
        warnings.append("科室原文含正文，已清洗")
    if detail.get("specialty_navigation_polluted") == "yes":
        warnings.append("擅长原文含导航文本，已清空")
    return valid, warnings


def classify_generic_record(
    valid_doctor_record: bool,
    combined_text: str,
    title_hits: list[str],
) -> tuple[str, list[str], list[str]]:
    if not valid_doctor_record:
        return "", [], []
    groups, tags = group_tags(combined_text)
    priority = "普通"
    if any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups:
        priority = "高"
    elif any(term != "医师" for term in title_hits):
        priority = "中"
    return priority, groups, tags


def extract_person_name(*texts: str | None) -> str:
    for raw in texts:
        text = clean_text(raw)
        if not text:
            continue
        text = re.sub(r"(姓名|专家姓名|医生姓名)\s*[:：]", " ", text)
        title_positions = [text.find(term) for term in TITLE_TERMS if text.find(term) > 0]
        fragments = [text]
        if title_positions:
            fragments.insert(0, text[: min(title_positions)])
        for fragment in fragments:
            for match in re.finditer(r"[\u4e00-\u9fff·]{2,8}", fragment):
                candidate = match.group(0)
                if looks_like_person_name(candidate):
                    return candidate
    return ""


def generic_link_score(href: str, context: str, entry_url: str = "") -> int:
    parsed = urlparse(href)
    path = parsed.path.lower()
    if gdskin_entry_id(entry_url) and gdskin_detail_id(href):
        has_name = bool(extract_person_name(context))
        has_doctor_role = any(term in context for term in GDSKIN_DOCTOR_ROLE_TERMS)
        return 12 if has_name and has_doctor_role else -10
    if ny5y_entry_kind(entry_url) and ny5y_detail_id(href):
        return 12
    if gdzy5413_entry_kind(entry_url) and (
        gdzy5413_detail_id(href) or gdzy5413_ksdoctor_detail_id(href)
    ):
        return 12
    if any(hint in path for hint in GENERIC_PATH_BLOCK_HINTS):
        return -10

    path_hint = any(hint in path for hint in GENERIC_DETAIL_PATH_HINTS)
    has_numeric_detail = bool(re.search(r"[/_-]\d{2,}(\.html?)?$", path) or re.search(r"/node/\d+$", path))
    strong_title_hits = extract_terms(context, GENERIC_STRONG_TITLE_TERMS)
    has_name = bool(extract_person_name(context))
    detail_text_signal = any(term in context for term in ["姓名", "科室", "职称", "擅长", "专长"])
    if path_hint and not (has_name or strong_title_hits or detail_text_signal):
        return -10
    if path_hint and not has_numeric_detail and not has_name and not (strong_title_hits and detail_text_signal):
        return -10
    if not path_hint and not (has_name and (strong_title_hits or detail_text_signal)):
        return -10

    score = 0
    if path_hint:
        score += 4
    if has_numeric_detail:
        score += 2
    if strong_title_hits:
        score += 4
    if has_name:
        score += 3
    if any(term in context for term in GENERIC_DETAIL_TEXT_HINTS):
        score += 2
    if "擅长" in context or "专长" in context:
        score += 2
    short_text = clean_text(context)
    if short_text in GENERIC_IGNORED_LINK_TEXTS:
        score -= 4
    return score


def discover_generic_detail_links(html: str, page_url: str, entry_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    gdskin_id = gdskin_entry_id(entry_url)
    ny5y_id = ny5y_entry_kind(entry_url)
    gdzy5413_id = gdzy5413_entry_kind(entry_url)
    strict_smukq_directory = comparable_host(entry_url) == "smukqyy.cn" and bool(
        re.fullmatch(r"/section/\d+/?", urlparse(entry_url).path)
    )
    anchors = (
        soup.select("table.masterTitleH a[href]")
        if gdskin_id
        else soup.select("li.xinyutitle1 a[href]")
        if gdzy5413_id == "851"
        else soup.select("div.contentinfo div.pudocname a[href]")
        if gdzy5413_id == "852"
        else soup.find_all("a", href=True)
    )
    for anchor in anchors:
        href = urljoin(page_url, anchor["href"])
        if canonical_url(href) == canonical_url(page_url):
            continue
        if not is_collectable_url(entry_url, href):
            continue
        if not matches_generic_directory_detail_url(entry_url, href):
            continue
        gdzy5413_card = anchor.find_parent("li", class_="xinyutitle1") if gdzy5413_id == "851" else None
        gdzy5413_department_block = (
            anchor.find_parent("div", class_="contentinfo") if gdzy5413_id == "852" else None
        )
        context = (
            compact_visible_text(gdzy5413_card, 700)
            if gdzy5413_card
            else
            strip_profile_navigation_text(compact_visible_text(anchor, 700))
            if gdskin_id or ny5y_id
            else nearest_card_text(anchor)
        )
        score = generic_link_score(href, context, entry_url)
        if score < 6 and not strict_smukq_directory:
            continue
        key = generic_detail_identity(href)
        if key in seen:
            continue
        seen.add(key)
        name_element = gdzy5413_card.select_one(".docnameall") if gdzy5413_card else None
        honor_element = gdzy5413_card.select_one(".docjich") if gdzy5413_card else None
        name = (
            re.sub(r"\s+", "", clean_text(name_element.get_text(" ", strip=True)))
            if name_element
            else re.sub(r"\s+", "", clean_text(anchor.get_text(" ", strip=True)))
            if gdzy5413_department_block
            else extract_person_name(context)
        )
        list_title = clean_text(honor_element.get_text(" ", strip=True)) if honor_element else context
        gdzy5413_department = ""
        if gdzy5413_department_block:
            department_element = gdzy5413_department_block.select_one(".ks_title")
            gdzy5413_department = clean_generic_department(
                compact_visible_text(department_element, 200) if department_element else ""
            )
        rows.append(
            {
                "source_link": href,
                "name": name,
                "list_title": clip(list_title, 500),
                "department": (
                    GDSKIN_ENTRY_METADATA[gdskin_id]["category_name"]
                    if gdskin_id
                    else NY5Y_ENTRY_METADATA[ny5y_id]["category_name"]
                    if ny5y_id
                    else gdzy5413_department
                    if gdzy5413_id == "852"
                    else extract_gdzy5413_department(list_title, "")
                    if gdzy5413_id == "851"
                    else infer_department(context)
                ),
                "description": clip(list_title if gdzy5413_id == "851" else context, 700),
                "list_page": page_url,
                "score": str(score),
            }
        )
    return rows


def discover_gdskin_excluded_links(html: str, page_url: str, entry_url: str) -> list[dict[str, str]]:
    if not gdskin_entry_id(entry_url):
        return []
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select("table.masterTitleH a[href]"):
        href = urljoin(page_url, anchor["href"])
        detail_id = gdskin_detail_id(href)
        if not detail_id or detail_id in seen:
            continue
        seen.add(detail_id)
        context = strip_profile_navigation_text(compact_visible_text(anchor, 700))
        if any(term in context for term in GDSKIN_DOCTOR_ROLE_TERMS):
            continue
        rows.append(
            {
                "entry_url": entry_url,
                "source_link": href,
                "list_title": context,
                "reason": "专家团队列表身份不属于医生角色，已排除",
            }
        )
    return rows


def gdzy5413_raw_detail_relation_count(html: str, page_url: str, entry_url: str) -> int:
    if not gdzy5413_entry_kind(entry_url):
        return 0
    soup = BeautifulSoup(html, "html.parser")
    return sum(
        1
        for anchor in soup.find_all("a", href=True)
        if gdzy5413_detail_id(urljoin(page_url, anchor["href"]))
        or (
            gdzy5413_entry_kind(entry_url) == "852"
            and gdzy5413_ksdoctor_detail_id(urljoin(page_url, anchor["href"]))
        )
    )


def discover_generic_list_pages(entry_url: str, html: str, max_pages: int) -> list[str]:
    if ny5y_entry_kind(entry_url) or gdzy5413_entry_kind(entry_url):
        # Owner 指定入口均按现场证据视为单页目录；其他栏目或模板不得自行扩围。
        return [entry_url]
    soup = BeautifulSoup(html, "html.parser")
    pages = [entry_url]
    seen = {canonical_url(entry_url)}
    entry_path = urlparse(entry_url).path.rstrip("/")
    for anchor in soup.find_all("a", href=True):
        text = clean_text(anchor.get_text(" ", strip=True))
        href = urljoin(entry_url, anchor["href"])
        if not is_collectable_url(entry_url, href):
            continue
        parsed = urlparse(href)
        path = parsed.path.rstrip("/")
        query = parsed.query.lower()
        page_like = False
        if re.fullmatch(r"\d{1,3}", text):
            page_like = True
        if any(word in text for word in ["下一页", "尾页", "末页"]):
            page_like = True
        if re.search(r"(?:page|p|pageindex|current|start)=\d+", query):
            page_like = True
        if re.search(r"(?:list_|index_)?\d{1,3}\.html?$", path.lower()):
            page_like = True
        if not page_like:
            continue
        if entry_path and path and not path.startswith(entry_path.rsplit("/", 1)[0]):
            continue
        key = canonical_url(href)
        if key in seen:
            continue
        seen.add(key)
        pages.append(href)
        if len(pages) >= max_pages:
            break
    return pages


def discover_gdskin_postback_documents(
    session: requests.Session,
    entry_url: str,
    entry_html: str,
    max_pages: int,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    documents = [{"page": "1", "url": entry_url, "html": entry_html}]
    errors: list[dict[str, str]] = []
    soup = BeautifulSoup(entry_html, "html.parser")
    postback_pages: dict[int, str] = {}
    for anchor in soup.find_all("a", href=True):
        match = re.search(r"__doPostBack\('([^']+)','Page\$(\d+)'\)", anchor["href"])
        if not match:
            continue
        page_number = int(match.group(2))
        if 1 < page_number <= max_pages:
            postback_pages[page_number] = match.group(1)

    base_payload: dict[str, str] = {}
    for field in soup.select("form input[name]"):
        if clean_text(field.get("type")).lower() in {"submit", "button", "image", "file"}:
            continue
        base_payload[str(field["name"])] = str(field.get("value") or "")

    for page_number, event_target in sorted(postback_pages.items()):
        payload = dict(base_payload)
        payload["__EVENTTARGET"] = event_target
        payload["__EVENTARGUMENT"] = f"Page${page_number}"
        page_error = ""
        for attempt in range(1, 4):
            try:
                response = session.post(entry_url, data=payload, timeout=35)
                if response.status_code == 200:
                    response.encoding = response.apparent_encoding or response.encoding or "utf-8"
                    documents.append(
                        {
                            "page": str(page_number),
                            "url": f"{entry_url}#postback-page-{page_number}",
                            "html": response.text,
                        }
                    )
                    page_error = ""
                    break
                page_error = f"HTTP {response.status_code}"
            except Exception as exc:  # noqa: BLE001 - keep collection failure visible
                page_error = str(exc)
            time.sleep(0.8 * attempt)
        if page_error:
            errors.append(
                {
                    "page": str(page_number),
                    "entry_url": entry_url,
                    "url": f"{entry_url}#postback-page-{page_number}",
                    "error": page_error,
                }
            )
    return documents, errors


def best_detail_scope_text(soup: BeautifulSoup) -> str:
    selectors = [
        "article",
        "main",
        ".main",
        ".content",
        ".cont",
        ".detail",
        ".article",
        ".doctor-detail",
        ".expert-detail",
        ".field-body",
        ".views-field-body",
    ]
    candidates: list[tuple[int, str]] = []
    for selector in selectors:
        for element in soup.select(selector):
            text = strip_profile_navigation_text(compact_visible_text(element, 6000))
            if len(text) < 20:
                continue
            label_score = sum(1 for term in ["姓名", "科室", "职称", "擅长", "简介", "专长"] if term in text)
            title_score = len(extract_terms(text, TITLE_TERMS))
            candidates.append((len(text) + label_score * 500 + title_score * 400, text))
    if candidates:
        return strip_article_tail(max(candidates, key=lambda item: item[0])[1])
    return strip_article_tail(strip_profile_navigation_text(compact_visible_text(soup, 6000)))


def first_visible_heading_text(soup: BeautifulSoup) -> str:
    for heading in soup.find_all(["h1", "h2", "h3"]):
        classes = " ".join(heading.get("class") or []).lower()
        if "hidden" in classes or "breadcrumb" in classes:
            continue
        text = clean_text(heading.get_text(" ", strip=True))
        if text and not any(term in text for term in ["面包屑", "当前位置"]):
            return text
    return ""


def extract_breadcrumb_department(text: str) -> str:
    source = clean_text(text)
    patterns = [
        r"临床科室\s*/\s*[^/]{1,30}\s*/\s*([^/]{1,24}?)\s*/\s*临床专家",
        r"科室导航\s*/\s*[^/]{1,30}\s*/\s*([^/]{1,24}?)\s*/",
        r"首页\s*/\s*临床科室\s*/\s*[^/]{1,30}\s*/\s*([^/]{1,24}?)\s*/",
    ]
    for pattern in patterns:
        match = re.search(pattern, source)
        if not match:
            continue
        candidate = clean_text(match.group(1)).strip("/ ")
        if candidate and not any(term in candidate for term in ["系列", "临床专家", "首页"]):
            return candidate
    return ""


def extract_labeled_value_any(text: str, labels: list[str], stop_labels: list[str]) -> str:
    value = extract_labeled_value(text, labels, stop_labels)
    if value:
        return value
    lines = [clean_text(line) for line in re.split(r"[\n\r]+", text) if clean_text(line)]
    for label in labels:
        value = extract_label_from_lines(lines, label, stop_labels)
        if value:
            return value
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = "|".join(re.escape(label) for label in stop_labels + ["一、", "二、", "三、", "四、", "五、", "更新时间", "专科门诊时间", "门诊时间"])
    pattern = rf"(?:^|[\s，,；;。])(?:{label_pattern})\s+(.+?)(?=(?:{stop_pattern})\s*[:：]?|$)"
    match = re.search(pattern, clean_text(text), flags=re.S)
    return clean_text(match.group(1)) if match else ""


def extract_explicit_labeled_value(text: str, labels: list[str], stop_labels: list[str]) -> str:
    value = extract_labeled_value(text, labels, stop_labels)
    if value:
        return value
    lines = [clean_text(line) for line in re.split(r"[\n\r]+", text) if clean_text(line)]
    for label in labels:
        value = extract_label_from_lines(lines, label, stop_labels)
        if value:
            return value
    return ""


def parse_generic_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
    h1 = first_visible_heading_text(soup)
    navigation_context = extract_navigation_context(soup)
    remove_profile_noise_elements(soup)
    raw_scope_text = compact_visible_text(soup, 6000)
    raw_highlights = extract_sentences(raw_scope_text, HIGHLIGHT_TERMS)
    highlight_navigation_polluted = contains_navigation_text(raw_highlights)
    main_text = best_detail_scope_text(soup)
    compact = clean_text("\n".join([h1, title, main_text]))
    name = first_nonempty(
        extract_labeled_value_any(
            compact,
            ["姓名", "专家姓名", "医生姓名", "名称"],
            ["科室", "所在科室", "职称", "专业擅长", "擅长", "专长", "简介"],
        ),
        extract_person_name(h1, title, fallback.get("name", ""), fallback.get("list_title", "")),
    )
    department_raw = first_nonempty(
        extract_labeled_value_any(
            compact,
            ["科室", "所在科室", "出诊科室", "专业", "专科"],
            ["职称", "职务", "专业擅长", "擅长", "专长", "简介", "个人简介"],
        ),
        extract_breadcrumb_department(navigation_context),
        fallback.get("department", ""),
        infer_department(compact[:500]),
    )
    department = clean_generic_department(department_raw)
    title_field = first_nonempty(
        extract_labeled_value_any(
            compact,
            ["职称", "职务", "专家职称", "技术职称"],
            ["科室", "所在科室", "专业擅长", "擅长", "专长", "简介", "个人简介"],
        ),
        "、".join(extract_terms(compact[:1000], TITLE_TERMS)),
        fallback.get("list_title", ""),
    )
    specialty_labels = [
        "专业擅长",
        "擅长",
        "医疗专长",
        "诊疗专长",
        "技术专长",
        "业务专长",
        "专长",
        "研究方向",
    ]
    specialty_stop_labels = [
        "个人简介",
        "简介",
        "介绍",
        "专家简介",
        "医生简介",
        "出诊",
        "门诊",
        "社会任职",
        "学术任职",
    ]
    specialty_raw = strip_article_tail(
        extract_explicit_labeled_value(
            compact,
            specialty_labels,
            specialty_stop_labels,
        )
    )
    specialty_preclean = strip_article_tail(
        extract_explicit_labeled_value(raw_scope_text, specialty_labels, specialty_stop_labels)
    )
    specialty_navigation_polluted = contains_navigation_text(specialty_raw) or contains_navigation_text(
        specialty_preclean
    )
    specialty = "" if specialty_navigation_polluted else specialty_raw
    profile_text = strip_article_tail(
        first_nonempty(
            extract_labeled_value_any(
                compact,
                ["个人简介", "专家简介", "医生简介", "简介", "详细介绍"],
                ["专业擅长", "擅长", "专长", "出诊", "门诊", "上一篇", "下一篇"],
            ),
            main_text,
        )
    )
    return {
        "name": clean_text(name),
        "department": clean_text(department),
        "department_raw": clean_text(department_raw),
        "department_polluted": "yes" if has_department_text_pollution(department_raw, department) else "no",
        "title_field": clean_text(title_field),
        "specialty": clean_text(specialty),
        "specialty_raw": clean_text(first_nonempty(specialty_raw, specialty_preclean)),
        "specialty_navigation_polluted": "yes" if specialty_navigation_polluted else "no",
        "highlight_navigation_polluted": "yes" if highlight_navigation_polluted else "no",
        "profile_text": clean_text(profile_text),
        "title_text": first_nonempty(h1, title),
    }


def parse_ny5y_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")

    def selected_text(selector: str, max_len: int) -> str:
        element = soup.select_one(selector)
        return clip(element.get_text(" ", strip=True), max_len) if element else ""

    name_element = soup.select_one(".yuanzhang")
    direct_name = clean_text(
        " ".join(str(value) for value in name_element.find_all(string=True, recursive=False))
    ) if name_element else ""
    name = first_nonempty(direct_name, fallback.get("name", ""))
    department_raw = selected_text(".suoshulei", 200)
    department = clean_generic_department(re.sub(r"^进入\s*", "", department_raw))
    if department in NY5Y_GENERAL_CATEGORIES:
        department = ""
    title_field = first_nonempty(selected_text(".xq_zhicheng", 500), fallback.get("list_title", ""))
    specialty_raw = selected_text(".xq_content", 1200)
    specialty = clean_text(re.sub(r"^\s*擅长\s*[:：]\s*", "", specialty_raw))
    profile_text = selected_text(".xq_xiangxi_jieshao_xq", 6000)
    raw_highlights = extract_sentences(profile_text, HIGHLIGHT_TERMS)
    return {
        "name": clean_text(name),
        "department": clean_text(department),
        "department_raw": clean_text(department_raw),
        # “进入”是详情页科室链接的固定 UI 前缀，不属于正文污染。
        "department_polluted": "no",
        "title_field": clean_text(title_field),
        "specialty": clean_text(specialty),
        "specialty_raw": clean_text(specialty_raw),
        "specialty_navigation_polluted": "yes" if contains_navigation_text(specialty) else "no",
        "highlight_navigation_polluted": "yes" if contains_navigation_text(raw_highlights) else "no",
        "profile_text": clean_text(profile_text),
        "title_text": clean_text(name),
    }


def parse_gdzy5413_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    profile_element = (
        soup.select_one("#news_info_plAll .news_info_s")
        or soup.select_one("#news_info_plAll")
        or soup.select_one(".news_info_s")
    )
    profile_text = compact_visible_text(profile_element, 6000) if profile_element else ""
    profile_text = re.sub(r"\s*了解详情[>＞]*\s*$", "", profile_text)
    name = re.sub(r"\s+", "", clean_text(fallback.get("name", "")))
    title_field = first_nonempty(
        fallback.get("list_title", ""),
        re.split(r"[。；;]", profile_text, maxsplit=1)[0] if profile_text else "",
    )
    department = extract_gdzy5413_department(title_field, profile_text)
    specialty_raw = ""
    specialty_match = re.search(
        r"(?:主攻方向为|主攻|专长于|专长|尤其擅长|擅长治疗|擅长诊治|擅长|擅治|长于|善于|善用|专科)\s*[:：]?\s*(.+)$",
        profile_text,
    )
    if specialty_match:
        specialty_raw = clean_text(specialty_match.group(1))
    specialty_raw = re.split(r"(?<=。)\s*(?:主持|承担|发表|获|出版|现任)", specialty_raw, maxsplit=1)[0]
    return {
        "name": name,
        "department": department,
        "department_raw": department,
        "department_polluted": "no",
        "title_field": title_field,
        "specialty": clip(specialty_raw, 1200),
        "specialty_raw": clip(specialty_raw, 1200),
        "specialty_navigation_polluted": "no",
        "highlight_navigation_polluted": "no",
        "profile_text": profile_text,
        "title_text": name,
    }


def parse_gdzy5413_ksdoctor_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    content = soup.select_one(".newslistbg_m_c") or soup.select_one("table.newslist")
    content_text = compact_visible_text(content, 12000) if content else ""
    breadcrumb_element = content.select_one(".typeall_right") if content else None
    breadcrumb = compact_visible_text(breadcrumb_element, 500) if breadcrumb_element else ""
    department_match = re.search(r">\s*([^>]{1,40}?)\s*>\s*专家介绍", breadcrumb)
    department = clean_generic_department(department_match.group(1) if department_match else "")
    if not department:
        department = clean_generic_department(fallback.get("department", ""))

    basic_match = re.search(r"【基本资料】\s*(.*?)(?=【医生简介】|【出诊安排】|$)", content_text)
    profile_match = re.search(r"【医生简介】\s*(.*?)(?=【出诊安排】|$)", content_text)
    basic_text = clean_text(basic_match.group(1)) if basic_match else ""
    profile_text = clean_text(profile_match.group(1)) if profile_match else ""
    name_match = re.search(r"姓名\s*[:：]\s*([\u4e00-\u9fff·\s]{2,10}?)(?=\s*职称\s*[:：]|$)", basic_text)
    title_match = re.search(r"职称\s*[:：]\s*(.*?)(?=\s*擅长\s*[:：]|$)", basic_text)
    specialty_match = re.search(r"擅长\s*[:：]\s*(.*)$", basic_text)
    name = re.sub(
        r"\s+",
        "",
        clean_text(name_match.group(1) if name_match else fallback.get("name", "")),
    )
    title_field = clean_text(title_match.group(1)) if title_match else ""
    specialty = clean_text(specialty_match.group(1)) if specialty_match else ""
    return {
        "name": name,
        "department": department,
        "department_raw": department,
        "department_polluted": "no",
        "title_field": title_field,
        "specialty": clip(specialty, 1200),
        "specialty_raw": clip(specialty, 1200),
        "specialty_navigation_polluted": "no",
        "highlight_navigation_polluted": "no",
        "profile_text": clip(profile_text, 6000),
        "title_text": name,
    }


def parse_gdskin_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    content_element = soup.select_one(".labelContent")
    if not content_element:
        return parse_generic_detail(html, fallback)

    title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
    content = strip_article_tail(compact_visible_text(content_element, 6000))
    content = re.sub(r"^预约挂号\s*", "", content)
    content_paragraphs: list[str] = []
    for paragraph in content_element.find_all("p"):
        paragraph_text = strip_article_tail(compact_visible_text(paragraph, 2400))
        paragraph_text = re.sub(r"^预约挂号\s*", "", paragraph_text)
        if paragraph_text and paragraph_text not in content_paragraphs:
            content_paragraphs.append(paragraph_text)
    name = first_nonempty(
        fallback.get("name"),
        extract_person_name(title, content[:160], fallback.get("list_title", "")),
    )

    specialty_raw = ""
    profile_text = ""
    for paragraph_index, paragraph_text in enumerate(content_paragraphs):
        specialty_match = re.match(r"^(?:专长|擅长)\s*[:：]\s*(.+)$", paragraph_text, flags=re.S)
        if not specialty_match:
            continue
        specialty_parts = re.split(r"简介\s*[:：]", specialty_match.group(1), maxsplit=1)
        specialty_raw = clean_text(specialty_parts[0])
        if len(specialty_parts) > 1:
            profile_text = clean_text(specialty_parts[1])
        else:
            following_paragraphs = [
                clean_text(re.sub(r"^简介\s*[:：]\s*", "", value))
                for value in content_paragraphs[paragraph_index + 1 :]
            ]
            profile_text = clean_text(" ".join(value for value in following_paragraphs if value))
        break
    if not specialty_raw:
        specialty_match = re.search(
            r"(?:专长|擅长)\s*[:：]\s*(.+?)(?=简介\s*[:：]|$)",
            content,
            flags=re.S,
        )
        specialty_raw = clean_text(specialty_match.group(1)) if specialty_match else ""
    specialty_navigation_polluted = contains_navigation_text(specialty_raw)
    specialty = "" if specialty_navigation_polluted else specialty_raw

    if not profile_text:
        profile_match = re.search(r"简介\s*[:：]\s*(.+)$", content, flags=re.S)
        profile_text = clean_text(profile_match.group(1)) if profile_match else ""

    first_label_positions = [
        position
        for marker in ["专长：", "专长:", "擅长：", "擅长:", "简介：", "简介:"]
        if (position := content.find(marker)) >= 0
    ]
    title_field = content[: min(first_label_positions)] if first_label_positions else content
    if name:
        title_field = re.sub(rf"^\s*{re.escape(name)}\s*", "", title_field, count=1)
    title_field = clean_text(title_field).strip("、，,；;_- ")
    if not first_label_positions and not profile_text:
        biography_match = re.search(
            r"(?=(?:(?:博士|硕士|本科)毕业于|毕业于|主要从事|从事|曾赴|曾任|主持|参与|以第一作者|发表))",
            title_field,
        )
        if biography_match and biography_match.start() > 0:
            candidate_title = clean_text(title_field[: biography_match.start()]).strip("、，,；;_- ")
            if extract_terms(candidate_title, TITLE_TERMS):
                profile_text = clean_text(title_field[biography_match.start() :])
                title_field = candidate_title

    department_raw = clean_text(fallback.get("department"))
    department = clean_generic_department(department_raw)
    raw_highlights = extract_sentences(content, HIGHLIGHT_TERMS)
    return {
        "name": clean_text(name),
        "department": department,
        "department_raw": department_raw,
        "department_polluted": "yes" if has_department_text_pollution(department_raw, department) else "no",
        "title_field": title_field,
        "specialty": specialty,
        "specialty_raw": specialty_raw,
        "specialty_navigation_polluted": "yes" if specialty_navigation_polluted else "no",
        "highlight_navigation_polluted": "yes" if contains_navigation_text(raw_highlights) else "no",
        "profile_text": profile_text,
        "title_text": first_nonempty(name, title),
    }


def gzzoc_list_pages(entry_url: str, html: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    max_page = 0
    for anchor in soup.select(".pager a[href]"):
        parsed = urlparse(urljoin(entry_url, anchor["href"]))
        for name, value in parse_qsl(parsed.query, keep_blank_values=True):
            if name == "page" and value.isdigit():
                max_page = max(max_page, int(value))
    return [with_query_param(entry_url, "page", page) for page in range(max_page + 1)]


def parse_gzzoc_list_page(html: str, page_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    for item in soup.select("li.list-item"):
        anchor = item.select_one(".item-name a[href]") or item.select_one('a[href*="/node/"]')
        if not anchor:
            continue
        href = urljoin(page_url, anchor["href"])
        if not re.search(r"/node/\d+$", urlparse(href).path):
            continue
        name = clean_text((item.select_one(".item-name") or anchor).get_text(" ", strip=True))
        list_title = clean_text((item.select_one(".item-post") or item).get_text(" ", strip=True))
        department = clean_text((item.select_one(".item-dept") or item).get_text(" ", strip=True))
        description = clean_text((item.select_one(".item-desc") or item).get_text(" ", strip=True))
        rows.append(
            {
                "source_link": href,
                "name": name,
                "list_title": list_title,
                "department": department,
                "description": description,
                "list_page": page_url,
            }
        )
    return rows


def extract_label_from_lines(lines: list[str], label: str, stop_labels: list[str]) -> str:
    stop_set = {clean_text(stop.rstrip(":：")) for stop in stop_labels}
    label_clean = clean_text(label.rstrip(":："))
    for index, line in enumerate(lines):
        normalized = clean_text(line.rstrip(":："))
        value = ""
        if normalized == label_clean:
            values: list[str] = []
            for next_line in lines[index + 1 :]:
                next_normalized = clean_text(next_line.rstrip(":："))
                if next_normalized in stop_set:
                    break
                if next_line:
                    values.append(next_line)
            return clean_text(" ".join(values))
        for marker in (f"{label_clean}:", f"{label_clean}："):
            if line.startswith(marker):
                value = clean_text(line[len(marker) :])
                break
        if value:
            for stop in stop_labels:
                for marker in (f"{stop}:", f"{stop}：", stop):
                    position = value.find(marker)
                    if position > 0:
                        value = value[:position]
            return clean_text(value)
    return ""


def fetch_json(
    session: requests.Session,
    url: str,
    params: dict[str, Any] | None = None,
    retries: int = 3,
) -> tuple[int | None, dict[str, Any], str]:
    last_status: int | None = None
    last_error = ""
    retryable_statuses = {429, 500, 502, 503, 504}
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, params=params, timeout=35)
        except Exception as exc:  # noqa: BLE001 - keep collection failure visible
            last_error = str(exc)
            if attempt < retries:
                time.sleep(0.8 * attempt)
                continue
            return None, {}, last_error
        last_status = response.status_code
        content_type = clean_text(response.headers.get("Content-Type"))
        if response.status_code != 200:
            last_error = f"HTTP {response.status_code}"
            if response.status_code in retryable_statuses and attempt < retries:
                time.sleep(0.8 * attempt)
                continue
            return response.status_code, {}, last_error
        if "json" not in content_type.lower():
            return response.status_code, {}, f"非 JSON 响应：{content_type or '未声明 Content-Type'}"
        try:
            payload = response.json()
        except ValueError as exc:
            return response.status_code, {}, f"JSON 解析失败：{exc}"
        if not isinstance(payload, dict):
            return response.status_code, {}, "JSON 顶层不是对象"
        return response.status_code, payload, ""
    return last_status, {}, last_error or "JSON 请求失败"


def flatten_gykqyy_directory_groups(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    relations: list[dict[str, Any]] = []
    for group in groups:
        group_name = clean_text(str(group.get("name") or ""))
        for department in group.get("child") or []:
            if not isinstance(department, dict):
                continue
            department_name = clean_text(str(department.get("name") or ""))
            for doctor in department.get("child") or []:
                if not isinstance(doctor, dict):
                    continue
                relations.append(
                    {
                        "group": group_name,
                        "department": department_name,
                        "doctor": doctor,
                    }
                )
    return relations


def select_gykqyy_trial_doctors(
    doctors: list[dict[str, Any]],
    max_doctors: int | None,
) -> list[dict[str, Any]]:
    if not max_doctors or len(doctors) <= max_doctors:
        return doctors[:]
    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()
    covered_departments: set[str] = set()
    for doctor in doctors:
        departments = [
            clean_text(str(value))
            for value in doctor.get("departments", [])
            if clean_text(str(value))
        ]
        if departments and any(value not in covered_departments for value in departments):
            selected.append(doctor)
            selected_ids.add(str(doctor["id"]))
            covered_departments.update(departments)
        if len(selected) >= max_doctors:
            return selected
    for doctor in doctors:
        if str(doctor["id"]) in selected_ids:
            continue
        selected.append(doctor)
        selected_ids.add(str(doctor["id"]))
        if len(selected) >= max_doctors:
            break
    return selected


def gykqyy_profile_text(detail: dict[str, Any]) -> str:
    html = str(detail.get("content") or "")
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    return strip_profile_navigation_text(soup.get_text(" ", strip=True))


def collect_gykqyy(target: HospitalTarget, today: str, max_doctors: int | None = None) -> dict[str, Any]:
    session = create_official_session()
    session.headers.update({"Referer": target.entry_url})
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页读取失败：{entry_error}")
    if GYKQYY_DIRECTORY_API not in entry_html or GYKQYY_DETAIL_API not in entry_html:
        raise RuntimeError("入口页未同时公开医生目录与详情接口，拒绝调用页面外接口。")

    directory_status, directory_payload, directory_error = fetch_json(session, GYKQYY_DIRECTORY_API)
    data = directory_payload.get("data") if isinstance(directory_payload, dict) else None
    if directory_status != 200 or not isinstance(data, dict):
        raise RuntimeError(f"官网医生目录接口读取失败：{directory_error or '响应结构异常'}")
    if directory_payload.get("code") != 1:
        raise RuntimeError(f"官网医生目录接口返回失败：{directory_payload.get('msg')}")

    groups = [item for item in data.get("list") or [] if isinstance(item, dict)]
    banner = [item for item in data.get("banner") or [] if isinstance(item, dict)]
    banner_by_id = {
        clean_text(str(item.get("id") or "")): item
        for item in banner
        if clean_text(str(item.get("id") or ""))
    }
    relations = flatten_gykqyy_directory_groups(groups)
    departments = [
        {
            "group": clean_text(str(group.get("name") or "")),
            "id": str(department.get("id") or ""),
            "name": clean_text(str(department.get("name") or "")),
            "doctor_relation_count": len(department.get("child") or []),
        }
        for group in groups
        for department in (group.get("child") or [])
        if isinstance(department, dict)
    ]

    by_id: dict[str, dict[str, Any]] = {}
    for relation in relations:
        doctor = relation["doctor"]
        doctor_id = clean_text(str(doctor.get("id") or ""))
        if not doctor_id:
            continue
        item = by_id.setdefault(
            doctor_id,
            {
                **doctor,
                "id": doctor_id,
                "departments": [],
                "groups": [],
            },
        )
        department = relation["department"]
        if department and department not in item["departments"]:
            item["departments"].append(department)
        group_name = relation["group"]
        if group_name and group_name not in item["groups"]:
            item["groups"].append(group_name)

    all_doctors = sorted(
        (
            {
                **item,
                **{
                    key: value
                    for key, value in banner_by_id.get(doctor_id, {}).items()
                    if value not in {None, ""}
                },
                "id": doctor_id,
            }
            for doctor_id, item in by_id.items()
        ),
        key=lambda item: (-int(item.get("weigh") or 0), int(str(item["id"]))),
    )
    ids_by_name: dict[str, list[str]] = {}
    for doctor in all_doctors:
        doctor_name = clean_text(str(doctor.get("title") or ""))
        if doctor_name:
            ids_by_name.setdefault(doctor_name, []).append(str(doctor["id"]))
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in ids_by_name.items()
        if len(set(ids)) > 1
    }
    selected_doctors = select_gykqyy_trial_doctors(all_doctors, max_doctors)
    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    excluded_detail_candidates: list[dict[str, str]] = []
    identity_reconciliation: list[dict[str, Any]] = []
    for item in selected_doctors:
        doctor_id = str(item["id"])
        link = f"{target.entry_url}&id={doctor_id}"
        detail_status, detail_payload, detail_error = fetch_json(
            session,
            GYKQYY_DETAIL_API,
            params={"category_id": 55, "article_id": doctor_id},
        )
        detail_data = detail_payload.get("data") if isinstance(detail_payload, dict) else None
        detail = detail_data.get("detail") if isinstance(detail_data, dict) else None
        if detail_status != 200 or detail_payload.get("code") != 1 or not isinstance(detail, dict):
            detail_errors.append({"source_link": link, "error": detail_error or "详情接口响应结构异常"})
            detail = {}

        name = first_nonempty(str(detail.get("title") or ""), str(item.get("title") or ""))
        departments_for_doctor = [clean_text(str(value)) for value in item.get("departments", [])]
        department = "、".join(value for value in departments_for_doctor if value)
        if not name:
            excluded_detail_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": "",
                    "source_link": link,
                    "reason": "目录与详情均无姓名，核心追溯字段缺失，排除正式行与画像",
                }
            )
            continue
        list_department = clean_text(str(item.get("keshi") or ""))
        title_field = clean_text(str(item.get("zhicheng") or ""))
        profile_text = gykqyy_profile_text(detail)
        list_intro = clean_text(str(item.get("intro") or ""))
        detail_intro = clean_text(str(detail.get("intro") or ""))
        specialty = clip(first_nonempty(detail_intro, list_intro), 520)
        combined_text = "\n".join(
            [target.hospital, department, list_department, title_field, specialty, profile_text]
        )
        title_hits = extract_terms(title_field, TITLE_TERMS) or extract_terms(profile_text, TITLE_TERMS)
        title_identity = first_nonempty(title_field, "、".join(title_hits))
        groups_found, tags = group_tags(combined_text)
        highlights = extract_clean_highlights(profile_text)
        warnings: list[str] = []
        if not detail:
            warnings.append("详情接口读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if name in same_name_groups:
            warnings.append("同名待甄别")
        priority = "普通"
        if any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found:
            priority = "高"
        elif any(term != "医师" for term in title_hits):
            priority = "中"

        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": list_department,
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": highlights,
                "列表简介": clip(list_intro, 700),
                "详情正文摘录": clip(profile_text, 1800),
                "来源类型": "医院官网",
                "来源链接": link,
                "采集入口": target.entry_url,
                "采集方式": "官网页面公开同域目录接口+官网页面公开同域详情接口",
                "采集日期": today,
                "详情页状态": "200" if detail else "失败",
                "已建画像": "是" if canonical_url(link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        identity_reconciliation.append(
            {
                "detail_id": doctor_id,
                "name": name,
                "resolution": "同名不同 ID 分行保留" if name in same_name_groups else "唯一 ID 保留",
                "departments": departments_for_doctor,
                "source_link": link,
                "reason": "同名待甄别" if name in same_name_groups else "官网科室树唯一详情 ID",
            }
        )

    category_counter = Counter(row["科室_分类页"] for row in rows if row["科室_分类页"])
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group
        for row in rows
        for group in str(row["重点关注范围"]).split("、")
        if group
    )
    warning_counter = Counter(
        warning
        for row in rows
        for warning in str(row["异常提示"]).split("；")
        if warning
    )
    directory_names = [clean_text(str(item.get("title") or "")) for item in all_doctors]
    nonblank_names = {name for name in directory_names if name}
    named_detail_count = sum(1 for name in directory_names if name)
    blank_name_detail_count = len(directory_names) - named_detail_count
    tree_ids = {str(item["id"]) for item in all_doctors}
    banner_only = [
        item
        for item in banner
        if clean_text(str(item.get("id") or "")) not in tree_ids
    ]

    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #22（与官网入口台账一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": 1,
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(all_doctors),
            "sample_entry_coverage_count": 1,
            "sample_entry_categories": ["医生团队（category=55）"],
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(all_doctors),
            "census_named_detail_count": named_detail_count,
            "census_blank_name_detail_count": blank_name_detail_count,
            "census_unique_nonblank_name_count": len(nonblank_names),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "census_department_count": len(departments),
            "census_group_count": len(groups),
            "census_banner_count": len(banner),
            "excluded_non_doctor_count": len(banner_only),
            "pagination_count": 1,
            "pagination_method": "医生专区由单次 getZhuanjiaList 请求一次性返回，无 page/pageNo 参数",
            "directory_api": GYKQYY_DIRECTORY_API,
            "detail_api": GYKQYY_DETAIL_API,
            "api_source_evidence": "医生目录 HTML 内联 Vue 脚本的 axios.get 明确声明两个同域公开接口",
            "category_error_count": 0,
            "detail_error_count": len(detail_errors),
            "gykqyy_final_row_count": len(rows),
            "gykqyy_same_name_separate_row_count": sum(
                1 for row in rows if clean_text(str(row.get("姓名") or "")) in same_name_groups
            ),
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": departments,
        "entry_reconnaissance": [
            {
                "category_name": "医生团队（category=55）",
                "entry_url": target.entry_url,
                "page_nature": "动态 Vue 医生目录；单次同域公开接口载入",
                "list_page_count": 1,
                "raw_detail_relation_count": len(relations),
                "unique_detail_count": len(all_doctors),
                "out_of_scope_detail_count": len(banner_only),
                "affiliation": target.hospital,
                "independent_entity_check": "官网同域、无鉴权、无内部参数",
            }
        ],
        "excluded_candidates": [
            {
                "entry_url": target.entry_url,
                "list_title": clean_text(str(item.get("title") or "")),
                "source_link": f"{target.entry_url}&id={item.get('id')}",
                "reason": "焦点推荐记录未出现在科室医生树中且姓名为空，排除",
            }
            for item in banner_only
        ]
        + excluded_detail_candidates,
        "gykqyy_identity_reconciliation": identity_reconciliation,
        "category_errors": [],
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def gyfyyy_department_path(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gyfyyy.cn" or parsed.query or parsed.fragment:
        return ""
    path = parsed.path
    if not re.fullmatch(r"/cn/ks/(?:[^/]+/){1,2}", path, flags=re.IGNORECASE):
        return ""
    return path


def gyfyyy_detail_id(url: str | None, department_url: str | None = None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gyfyyy.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"(/cn/ks/(?:[^/]+/){1,2})doctor_(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""
    if department_url and match.group(1).lower() != gyfyyy_department_path(department_url).lower():
        return ""
    return match.group(2)


def discover_gyfyyy_departments(html: str, entry_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    departments: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select("a[href]"):
        department_url = urljoin(entry_url, str(anchor.get("href") or ""))
        if not gyfyyy_department_path(department_url):
            continue
        key = canonical_url(department_url)
        if key in seen:
            continue
        department = clean_text(anchor.get_text(" ", strip=True))
        if not department:
            continue
        seen.add(key)
        departments.append(
            {
                "name": department,
                "department_url": department_url,
                "team_url": urljoin(department_url, "doctorList.html"),
            }
        )
    return departments


GYFYYY_NURSING_IDENTITY_TERMS = ("主任护师", "副主任护师", "主管护师", "护师", "护士")
GYFYYY_MEDICAL_IDENTITY_TERMS = ("医师", "医士", "研究员", "教授")


def gyfyyy_nursing_only_identity(text: str | None) -> bool:
    value = clean_text(text)
    return bool(
        value
        and any(term in value for term in GYFYYY_NURSING_IDENTITY_TERMS)
        and not any(term in value for term in GYFYYY_MEDICAL_IDENTITY_TERMS)
    )


def discover_gyfyyy_doctor_relations(
    html: str,
    department: dict[str, str],
) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    relations: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for anchor in soup.select("section.doctors.team a[href]"):
        source_link = urljoin(department["team_url"], str(anchor.get("href") or ""))
        doctor_id = gyfyyy_detail_id(source_link, department["department_url"])
        if not doctor_id or doctor_id in seen_ids:
            continue
        seen_ids.add(doctor_id)
        card = anchor.find_parent("li") or anchor
        relations.append(
            {
                "id": doctor_id,
                "department": department["name"],
                "department_url": department["department_url"],
                "team_url": department["team_url"],
                "source_link": source_link,
                "list_title": clean_text(card.get_text(" ", strip=True)),
            }
        )
    return relations


def select_gyfyyy_trial_doctors(
    doctors: list[dict[str, Any]],
    max_doctors: int | None,
) -> list[dict[str, Any]]:
    return select_gykqyy_trial_doctors(doctors, max_doctors)


GYFYYY_SCHEDULE_LABEL_PATTERN = re.compile(
    r"(?:专家门诊时间|门诊出诊时间|出诊时间|开诊时间|开诊院区)\s*[:：]?\s*.*?"
    r"(?=(?:擅长|专长|简介|教育经历|职称|专业方向|专业特长|现任|本人|从事|研究方向|"
    r"社会任职|学术任职|主持|参与|发表|中华|中国|广东|广州市|医学博士|主任医师|"
    r"副主任医师|主治医师|教授|博士研究生导师|泌尿外科主任|科室副主任)\s*[:：]?|$)",
    flags=re.IGNORECASE,
)
GYFYYY_WEEKLY_CLINIC_PATTERN = re.compile(
    r"(?:专家门诊|特需门诊|专科门诊)[^。；]*?(?:每?周[一二三四五六日天]|周末)[^。；]*(?:[。；]|$)",
    flags=re.IGNORECASE,
)
GYFYYY_BARE_SCHEDULE_PATTERN = re.compile(
    r"(?:[^。；]{0,40}(?:门诊|出诊)\s*[（(][^）)]*[）)]\s*[:：]?\s*)?"
    r"(?:每?周[一二三四五六日天](?:[、,，及和至到\-]\s*(?:周)?[一二三四五六日天])*)"
    r"(?:上午|下午|晚上|夜诊|全天|早上)?(?:\s*[:：]?\s*\d{1,2}[:：]\d{2}\s*(?:--?|至)\s*\d{1,2}[:：]\d{2})?"
    r"(?:出诊)?(?:[。；]|$)",
    flags=re.IGNORECASE,
)


def strip_gyfyyy_schedule_text(value: str | None) -> str:
    """Remove public appointment schedules, which are outside profile fields."""

    text = clean_text(value)
    previous = None
    while text != previous:
        previous = text
        text = GYFYYY_SCHEDULE_LABEL_PATTERN.sub(" ", text)
    text = GYFYYY_WEEKLY_CLINIC_PATTERN.sub(" ", text)
    text = GYFYYY_BARE_SCHEDULE_PATTERN.sub(" ", text)
    return clean_text(text)


def parse_gyfyyy_detail(html: str, fallback: dict[str, Any]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    card = soup.select_one("section.doctorcard")
    intro = soup.select_one("section.doctorintro")
    name = clean_text(card.select_one("strong").get_text(" ", strip=True)) if card and card.select_one("strong") else ""
    title = "、".join(
        value
        for value in (
            clean_text(node.get_text(" ", strip=True)) for node in (card.select("b") if card else [])
        )
        if value
    )
    specialty = " ".join(
        value
        for value in (
            clean_text(node.get_text(" ", strip=True)) for node in (card.select("p") if card else [])
        )
        if value
    )
    specialty = clean_text(re.sub(r"^(?:擅长\s*[:：]?\s*)+", "", specialty))
    profile_text = strip_gyfyyy_schedule_text(" ".join(
        value
        for value in (
            clean_text(node.get_text(" ", strip=True)) for node in (intro.select("p") if intro else [])
        )
        if value
    ))
    return {
        "name": first_nonempty(name, str(fallback.get("name") or "")),
        "title": first_nonempty(title, str(fallback.get("list_title") or "")),
        "specialty": clip(specialty, 520),
        "profile_text": clip(profile_text, 1800),
    }


def gy3y_detail_id(url: str | None, department_url: str | None = None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gy3y.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"(/ks/(?:hp/)?[^/]+/[^/]+/)doctor_(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""
    if department_url:
        expected = urlparse(clean_text(department_url))
        if (
            comparable_host(expected.geturl()) != "gy3y.cn"
            or expected.query
            or expected.fragment
            or expected.path.lower() != match.group(1).lower()
        ):
            return ""
    return match.group(2)


def gzbrain_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gzbrain.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(r"/myzj/info_itemid_(\d+)\.html", parsed.path, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def gzbrain_list_page_number(url: str | None) -> int | None:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gzbrain.cn" or parsed.query or parsed.fragment:
        return None
    if parsed.path.lower() == "/myzj/list.html":
        return 1
    match = re.fullmatch(r"/myzj/list_page_(\d+)\.html", parsed.path, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def discover_gzbrain_list_pages(html: str, entry_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    page_numbers = {1}
    for anchor in soup.select(".p_page a[href]"):
        page_url = urljoin(entry_url, str(anchor.get("href") or ""))
        page_number = gzbrain_list_page_number(page_url)
        if page_number:
            page_numbers.add(page_number)
    last_page = max(page_numbers)
    return [
        entry_url if page_number == 1 else urljoin(entry_url, f"list_page_{page_number}.html")
        for page_number in range(1, last_page + 1)
    ]


def split_gzbrain_title_department(value: str | None) -> tuple[str, str]:
    text = clean_text(value)
    title_hits = extract_terms(text, TITLE_TERMS)
    remainder = text
    for term in sorted(title_hits, key=len, reverse=True):
        remainder = remainder.replace(term, " ")
    remainder = clean_text(re.sub(r"^[、,，;；\s]+|[、,，;；\s]+$", "", remainder))
    departments = re.findall(
        r"[\u4e00-\u9fff（）()]{1,20}(?:科|中心|门诊)"
        r"(?:、[\u4e00-\u9fff（）()]{1,20}(?:科|中心|门诊))*",
        remainder,
    )
    department = clean_generic_department(departments[-1] if departments else remainder)
    return "、".join(title_hits), department


def parse_gzbrain_list_page(html: str, page_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for anchor in soup.select("div.expert_list ul.ul > li > a[href]"):
        source_link = urljoin(page_url, str(anchor.get("href") or ""))
        detail_id = gzbrain_detail_id(source_link)
        if not detail_id or detail_id in seen_ids:
            continue
        seen_ids.add(detail_id)
        name_node = anchor.select_one("div.txt h2")
        title_node = anchor.select_one("div.txt h3")
        specialty_node = anchor.select_one("div.txt h4")
        name = clean_text(name_node.get_text(" ", strip=True) if name_node else "")
        title_department = clean_text(title_node.get_text(" ", strip=True) if title_node else "")
        title, department = split_gzbrain_title_department(title_department)
        specialty = clean_text(specialty_node.get_text(" ", strip=True) if specialty_node else "")
        specialty = clean_text(re.sub(r"^专长\s*[:：]\s*", "", specialty))
        schedule_node = anchor.select_one("div.txt p")
        schedule = clean_text(schedule_node.get_text(" ", strip=True) if schedule_node else "")
        rows.append(
            {
                "id": detail_id,
                "name": name,
                "title": title,
                "title_department": title_department,
                "department": department,
                "specialty": specialty,
                "schedule": schedule,
                "source_link": source_link,
                "list_page": page_url,
            }
        )
    return rows


def filter_gzbrain_profile_text(value: str | None) -> tuple[str, int]:
    text = strip_gyfyyy_schedule_text(value)
    if not text:
        return "", 0
    kept: list[str] = []
    excluded = 0
    for sentence in re.split(r"(?<=[。！？；;])\s*", text):
        sentence = clean_text(sentence)
        if not sentence:
            continue
        if contains_gzbrain_patient_case_text(sentence):
            excluded += 1
            continue
        kept.append(sentence)
    return clean_text(" ".join(kept)), excluded


def contains_gzbrain_patient_case_text(value: str | None) -> bool:
    text = clean_text(value)
    return bool(
        any(term in text for term in ["患者案例", "病例详情", "患者评价"])
        or re.search(r"(?:病例|个案)\s*[:：]", text)
        or re.search(
            r"患者[^。！？；;]{0,30}(?:\d{1,3}\s*岁|男性|女性|男士|女士|某某|[\u4e00-\u9fff]某)",
            text,
        )
        or re.search(r"\d{1,3}\s*岁[^。！？；;]{0,30}患者", text)
    )


def parse_gzbrain_detail(html: str, fallback: dict[str, str]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    container = soup.select_one("div.single_con")
    if not container:
        return {
            "name": clean_text(fallback.get("name")),
            "title": clean_text(fallback.get("title")),
            "specialty": clean_text(fallback.get("specialty")),
            "profile_text": "",
            "patient_case_exclusion_count": 0,
        }
    name_node = container.select_one("div.single-header h2")
    title_node = container.select_one("div.single-header h3")
    specialty = ""
    for paragraph in container.select("div.single_tex p"):
        text = clean_text(paragraph.get_text(" ", strip=True))
        if re.match(r"^专长\s*[:：]", text):
            specialty = clean_text(
                re.sub(r"^(?:(?:专长|擅长)\s*[:：]?\s*)+", "", text)
            )
            break
    profile_node = container.select_one("div.single-content")
    if profile_node:
        profile_soup = BeautifulSoup(str(profile_node), "html.parser")
        for node in profile_soup.find_all(["h3", "span"]):
            text = clean_text(node.get_text(" ", strip=True))
            if text == "详细介绍" or "此排班仅作参考" in text:
                node.decompose()
        profile_source = clean_text(profile_soup.get_text(" ", strip=True))
    else:
        profile_source = ""
    profile_text, excluded_count = filter_gzbrain_profile_text(profile_source)
    return {
        "name": first_nonempty(
            clean_text(name_node.get_text(" ", strip=True) if name_node else ""),
            fallback.get("name"),
        ),
        "title": first_nonempty(
            clean_text(title_node.get_text(" ", strip=True) if title_node else ""),
            fallback.get("title"),
        ),
        "specialty": first_nonempty(specialty, fallback.get("specialty")),
        "profile_text": profile_text,
        "patient_case_exclusion_count": excluded_count,
    }


def select_gzbrain_trial_doctors(
    doctors: list[dict[str, Any]], max_doctors: int | None
) -> list[dict[str, Any]]:
    if not max_doctors or len(doctors) <= max_doctors:
        return doctors[:]
    normalized = [
        {**doctor, "departments": [clean_text(doctor.get("department"))]}
        for doctor in doctors
    ]
    return select_gykqyy_trial_doctors(normalized, max_doctors)


def collect_gzbrain(
    target: HospitalTarget, today: str, max_doctors: int | None = None
) -> dict[str, Any]:
    entry_status, entry_html, entry_error = fetch_standard_public_get(target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页标准公开 GET 读取失败：{entry_error}")
    page_urls = discover_gzbrain_list_pages(entry_html, target.entry_url)
    if not page_urls:
        raise RuntimeError("官网专家目录未发现静态分页。")

    categories: list[dict[str, Any]] = []
    page_errors: list[dict[str, str]] = []
    relations: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for page_number, page_url in enumerate(page_urls, start=1):
        if page_number == 1:
            status, html, error = entry_status, entry_html, ""
        else:
            status, html, error = fetch_standard_public_get(page_url)
        if status != 200:
            page_errors.append(
                {"page": str(page_number), "url": page_url, "error": error}
            )
            continue
        page_rows = parse_gzbrain_list_page(html, page_url)
        page_ids = [row["id"] for row in page_rows]
        duplicates = sorted(set(page_ids) & seen_ids, key=int)
        if duplicates:
            raise RuntimeError(
                f"官网专家目录分页出现重复详情 ID：{','.join(duplicates)}"
            )
        seen_ids.update(page_ids)
        relations.extend(page_rows)
        categories.append(
            {
                "category_id": str(page_number),
                "category_name": f"专家目录第 {page_number} 页",
                "url": page_url,
                "doctor_relation_count": len(page_rows),
            }
        )
    if page_errors:
        raise RuntimeError(
            "官网专家目录分页读取不完整："
            + "；".join(error["error"] for error in page_errors)
        )
    if not relations:
        raise RuntimeError("官网专家目录未发现严格 itemid 医生详情关系。")

    names_to_ids: dict[str, list[str]] = {}
    for item in relations:
        if item["name"]:
            names_to_ids.setdefault(item["name"], []).append(item["id"])
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in names_to_ids.items()
        if len(set(ids)) > 1
    }
    selected = select_gzbrain_trial_doctors(relations, max_doctors)
    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    excluded_candidates: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, str]] = []
    patient_case_exclusion_count = 0
    for item in selected:
        detail_status, detail_html, detail_error = fetch_standard_public_get(
            item["source_link"]
        )
        if detail_status == 200:
            detail = parse_gzbrain_detail(detail_html, item)
        else:
            detail_errors.append(
                {"source_link": item["source_link"], "error": detail_error}
            )
            detail = {
                "name": item["name"],
                "title": item["title"],
                "specialty": item["specialty"],
                "profile_text": "",
                "patient_case_exclusion_count": 0,
            }
        name = clean_text(str(detail.get("name") or item["name"]))
        title_identity = clean_text(str(detail.get("title") or item["title"]))
        if gyfyyy_nursing_only_identity(title_identity):
            exclusion = {
                "entry_url": target.entry_url,
                "list_title": item["title_department"],
                "source_link": item["source_link"],
                "reason": "官网详情仅标注护理身份，排除医生画像采集范围",
            }
            excluded_candidates.append(exclusion)
            detail_reconciliation.append(
                {
                    "detail_id": item["id"],
                    "source_link": item["source_link"],
                    "name": name,
                    "resolution": "护理排除",
                    "reason": exclusion["reason"],
                }
            )
            continue
        specialty = clean_text(str(detail.get("specialty") or item["specialty"]))
        profile_text = clean_text(str(detail.get("profile_text") or ""))
        patient_case_exclusion_count += int(
            detail.get("patient_case_exclusion_count") or 0
        )
        combined_text = "\n".join(
            [target.hospital, item["department"], title_identity, specialty, profile_text]
        )
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if name != item["name"]:
            warnings.append("列表与详情姓名不一致")
        if not item["department"]:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if name in same_name_groups:
            warnings.append("同名待甄别")
        if warnings:
            groups_found = []
            tags = []
        priority = "普通"
        if not warnings and (
            any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
        ):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": item["department"],
                "科室_列表卡片": item["department"],
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": clip(specialty, 520),
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": clip(profile_text, 1800),
                "来源类型": "医院官网",
                "来源链接": item["source_link"],
                "采集入口": target.entry_url,
                "采集方式": "官网静态专家目录分页+严格 itemid 详情 DOM 结构化抽取",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": (
                    "是" if canonical_url(item["source_link"]) in existing_links else "否"
                ),
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": item["id"],
                "source_link": item["source_link"],
                "name": name,
                "resolution": "正式行",
                "reason": "",
            }
        )
        time.sleep(0.18)

    covered_departments = covered_department_names(rows)
    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(row["科室_分类页"]).split("、")
        if department
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in row["重点关注范围"].split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in row["异常提示"].split("；") if warning
    )
    campus_terms = ["芳村", "白云", "荔湾", "江村", "总部"]
    campus_evidence = {
        term: sum(term in item["schedule"] for item in relations) for term in campus_terms
    }
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #31（与官网入口台账一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(categories),
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(relations),
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(relations),
            "census_named_detail_count": sum(bool(item["name"]) for item in relations),
            "census_blank_name_detail_count": sum(not item["name"] for item in relations),
            "census_unique_nonblank_name_count": len(names_to_ids),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "census_department_count": len(
                {item["department"] for item in relations if item["department"]}
            ),
            "census_group_count": len([value for value in campus_evidence.values() if value]),
            "census_nonempty_department_count": sum(bool(item["department"]) for item in relations),
            "census_empty_department_count": sum(not item["department"] for item in relations),
            "sample_entry_coverage_count": len(covered_departments),
            "sample_entry_categories": covered_departments,
            "pagination_count": len(page_urls),
            "pagination_method": "静态 list_page_N.html 分页，逐页普通公开 GET",
            "standard_public_get": "urllib 默认非浏览器 GET；无 Cookie、代理、指纹模拟或挑战绕过",
            "campus_schedule_evidence": campus_evidence,
            "schedule_field_ingested_count": 0,
            "patient_case_exclusion_count": patient_case_exclusion_count,
            "category_error_count": len(page_errors),
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": 0,
            "excluded_non_doctor_count": len(excluded_candidates),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": categories,
        "entry_reconnaissance": [
            {
                "category_name": "官网专家介绍",
                "entry_url": target.entry_url,
                "page_nature": "医院官网静态专家目录",
                "list_page_count": len(page_urls),
                "raw_detail_relation_count": len(relations),
                "unique_detail_count": len(relations),
                "out_of_scope_detail_count": len(excluded_candidates),
                "affiliation": target.hospital,
                "independent_entity_check": (
                    "同域目录；排班文本仅用于院区存在性普查，不进入医生字段"
                ),
            }
        ],
        "excluded_candidates": excluded_candidates,
        "gzbrain_detail_reconciliation": detail_reconciliation,
        "cross_entry_duplicates": [],
        "category_errors": page_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def gzsys_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gzsys.org.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(r"/(?:node|doctor)/(\d+)/?", parsed.path, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def discover_gzsys_default_pages(html: str, entry_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    declared_pages: set[int] = {0}
    expected = {
        "department_target_id": "All",
        "talent_project": "All",
        "tutor_qualification": "All",
        "doctor_title": "All",
    }
    for anchor in soup.select('a[href*="page="]'):
        absolute = urljoin(entry_url, str(anchor.get("href") or ""))
        parsed = urlparse(absolute)
        query = dict(parse_qsl(parsed.query, keep_blank_values=True))
        if (
            comparable_host(absolute) == "gzsys.org.cn"
            and parsed.path.rstrip("/") == "/doctor/592/search"
            and all(query.get(name) == value for name, value in expected.items())
            and query.get("page", "").isdigit()
        ):
            declared_pages.add(int(query["page"]))
    if len(declared_pages) < 2:
        return []
    return [
        f"{entry_url}?{urlencode({**expected, 'page': page})}"
        for page in range(max(declared_pages) + 1)
    ]


def parse_gzsys_filter_dictionary(html: str) -> dict[str, list[dict[str, str]]]:
    soup = BeautifulSoup(html, "html.parser")
    result: dict[str, list[dict[str, str]]] = {}
    for name in (
        "department_target_id",
        "talent_project",
        "tutor_qualification",
        "doctor_title",
    ):
        select = soup.select_one(f'select[name="{name}"]')
        result[name] = [
            {
                "value": clean_text(str(option.get("value") or "")),
                "label": clean_text(option.get_text(" ", strip=True)),
            }
            for option in (select.select("option") if select else [])
            if clean_text(str(option.get("value") or "")) not in {"", "All"}
        ]
    return result


def parse_gzsys_list_page(html: str, page_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    for card in soup.select(".card-4-0"):
        name_link = card.select_one(".card-title a[href]")
        source_link = urljoin(page_url, str(name_link.get("href") or "")) if name_link else ""
        detail_id = gzsys_detail_id(source_link)
        if not detail_id:
            continue
        title_node = card.select_one(".card-subtitle-content")
        departments = [
            clean_text(node.get_text(" ", strip=True))
            for node in card.select(".card-tag")
            if clean_text(node.get_text(" ", strip=True))
        ]
        rows.append(
            {
                "id": detail_id,
                "name": clean_text(name_link.get_text(" ", strip=True)),
                "title": clean_text(title_node.get_text(" ", strip=True) if title_node else ""),
                "department": normalize_gzsys_department("、".join(dict.fromkeys(departments))),
                "source_link": source_link,
                "page_url": page_url,
            }
        )
    return rows


def merge_gzsys_card_relations(relations: list[dict[str, str]]) -> list[dict[str, str]]:
    merged: dict[str, dict[str, str]] = {}
    for relation in relations:
        detail_id = relation["id"]
        if detail_id not in merged:
            merged[detail_id] = {**relation, "relation_count": "1"}
            continue
        current = merged[detail_id]
        current["relation_count"] = str(int(current["relation_count"]) + 1)
        for field in ("name", "title", "department"):
            if not current[field] and relation[field]:
                current[field] = relation[field]
    return list(merged.values())


def strip_gzsys_forbidden_text(value: str | None) -> str:
    return clean_text(re.sub(r"[\ue000-\uf8ff]", " ", value or ""))


def normalize_gzsys_department(value: str | None) -> str:
    parts = [
        clean_text(part)
        for part in re.split(r"\s*[,，;/；、]+\s*", strip_gzsys_forbidden_text(value))
        if clean_text(part)
    ]
    return "、".join(dict.fromkeys(parts))


GZSYS_SCHEDULE_LABEL_PATTERN = re.compile(
    r"(?:"
    r"(?:院本部|南院区|深汕院区|花都院区)?(?:开诊|出诊|门诊|特诊)时间(?:为)?"
    r"|博济特诊时间"
    r"|减重管理团队教授门诊"
    r")\s*[:：]?\s*.*?"
    r"(?=(?:特长|擅长|专长|简介)\s*[:：]|$)",
    flags=re.IGNORECASE,
)
GZSYS_CAMPUS_SCHEDULE_TAIL_PATTERN = re.compile(
    r"(?:院本部|南院区|南院)\s*(?=(?:逢|每)?周[一二三四五六日天])"
    r"[^。]*(?:。|$)",
    flags=re.IGNORECASE,
)


def strip_gzsys_schedule_text(value: str | None) -> str:
    """Remove only explicit GZSYS clinic-time tails from free-form profiles."""

    text = clean_text(value)
    previous = None
    while text != previous:
        previous = text
        text = GZSYS_SCHEDULE_LABEL_PATTERN.sub(" ", text)
    text = GZSYS_CAMPUS_SCHEDULE_TAIL_PATTERN.sub(" ", text)
    return clean_text(text)


def parse_gzsys_detail(html: str, fallback: dict[str, str]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    container = soup.select_one(".other-2")
    schedule_count = len(soup.select(".calendar-3-1"))
    if not container:
        return {**fallback, "specialty": "", "profile_text": "", "schedule_exclusion_count": schedule_count, "forbidden_segment_count": 0}
    name_node = container.select_one(".other-left-title")
    labeled: dict[str, str] = {}
    for node in container.select(".other-left-text"):
        label_node = node.select_one("span")
        label = clean_text(label_node.get_text(" ", strip=True) if label_node else "")
        clone = BeautifulSoup(str(node), "html.parser")
        for child in clone.select("span"):
            child.decompose()
        value = strip_gzsys_forbidden_text(clone.get_text(" ", strip=True))
        if label.startswith("职称"):
            labeled["title"] = value
        elif label.startswith("科室"):
            labeled["department"] = normalize_gzsys_department(value)
    desc = container.select_one(".desc.line-6")
    leaf_paragraphs = [node for node in desc.select("p") if not node.select("p")] if desc else []
    source_nodes = leaf_paragraphs or ([desc] if desc else [])
    segments: list[str] = []
    inline_schedule_exclusion_count = 0
    for node in source_nodes:
        value = strip_gzsys_forbidden_text(node.get_text(" ", strip=True))
        cleaned_value = strip_gzsys_schedule_text(value)
        if cleaned_value != value:
            inline_schedule_exclusion_count += 1
        if cleaned_value and cleaned_value not in segments:
            segments.append(cleaned_value)
    forbidden_markers = ("好医生榜", "医生排行榜", "患者评价", "患者留言", "问诊记录", "问诊内容")
    sentences = [
        clean_text(sentence)
        for segment in segments
        for sentence in re.split(r"(?<=[。！？；;])\s*", segment)
        if clean_text(sentence)
    ]
    forbidden_markers = (
        *forbidden_markers,
        "好医生",
        "名医录",
        "排行榜",
        "排名",
    )
    kept = [
        sentence
        for sentence in sentences
        if not any(term in sentence for term in forbidden_markers)
        and not contains_gzbrain_patient_case_text(sentence)
    ]
    specialty = ""
    for segment in kept:
        if re.match(r"^(?:专业擅长|擅长|专长|特长)\s*[:：]", segment):
            specialty = clean_text(re.sub(r"^(?:(?:专业擅长|擅长|专长|特长)\s*[:：]?\s*)+", "", segment))
            break
    return {
        "name": first_nonempty(strip_gzsys_forbidden_text(name_node.get_text(" ", strip=True) if name_node else ""), fallback.get("name")),
        "title": first_nonempty(labeled.get("title"), fallback.get("title")),
        "department": first_nonempty(labeled.get("department"), fallback.get("department")),
        "specialty": specialty,
        "profile_text": clip(strip_profile_navigation_text(" ".join(kept)), 1800),
        "schedule_exclusion_count": schedule_count + inline_schedule_exclusion_count,
        "forbidden_segment_count": len(sentences) - len(kept),
    }


def select_gzsys_trial_doctors(doctors: list[dict[str, str]], max_doctors: int | None) -> list[dict[str, str]]:
    if not max_doctors or len(doctors) <= max_doctors:
        return doctors[:]
    buckets: dict[str, list[dict[str, str]]] = {}
    for doctor in doctors:
        buckets.setdefault(clean_text(doctor.get("department")) or "（科室空白）", []).append(doctor)
    selected: list[dict[str, str]] = []
    offset = 0
    while len(selected) < max_doctors:
        added = False
        for values in buckets.values():
            if offset < len(values):
                selected.append(values[offset])
                added = True
                if len(selected) >= max_doctors:
                    return selected
        if not added:
            break
        offset += 1
    return selected


def gdgh_subject_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdghospital.org.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"/Specialistthree/index_subjectid_(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    return match.group(1) if match else ""


def gdgh_detail_id(url: str | None, subject_id: str | None = None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdghospital.org.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"/Expertlistt/info_itemid_(\d+)_subjectid_(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    if not match or (subject_id and match.group(2) != clean_text(subject_id)):
        return ""
    return match.group(1)


def gdgh_detail_subject_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gdghospital.org.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"/Expertlistt/info_itemid_(\d+)_subjectid_(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    return match.group(2) if match else ""


def parse_gdgh_directory(
    html: str, entry_url: str
) -> tuple[list[str], list[dict[str, str]]]:
    soup = BeautifulSoup(html, "html.parser")
    groups = list(
        dict.fromkeys(
            clean_text(node.get_text(" ", strip=True))
            for node in soup.select("h1,h2,h3")
            if clean_text(node.get_text(" ", strip=True))
        )
    )
    departments: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for anchor in soup.select("a[href]"):
        department_url = urljoin(entry_url, str(anchor.get("href") or ""))
        subject_id = gdgh_subject_id(department_url)
        department = clean_text(anchor.get_text(" ", strip=True))
        if not subject_id or subject_id in seen_ids or not department:
            continue
        seen_ids.add(subject_id)
        departments.append(
            {
                "subject_id": subject_id,
                "department": department,
                "department_url": department_url,
            }
        )
    return groups, departments


def gdgh_photo_url(value: str | None, base_url: str) -> str:
    absolute = urljoin(base_url, clean_text(value))
    parsed = urlparse(absolute)
    if (
        parsed.scheme != "https"
        or comparable_host(absolute) != "gdghospital.org.cn"
        or not parsed.path.lower().startswith("/uploadfiles/")
        or parsed.fragment
    ):
        return ""
    return absolute


def gdgh_clean_text(value: str | None) -> str:
    return clean_text(re.sub(r"[\u200b-\u200d\ufeff]", "", value or ""))


def parse_gdgh_department_page(
    html: str,
    department: dict[str, str],
) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    relations: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for anchor in soup.select("a[href]"):
        source_link = urljoin(
            department["department_url"], str(anchor.get("href") or "")
        )
        detail_id = gdgh_detail_id(source_link, department["subject_id"])
        if not detail_id or detail_id in seen_ids:
            continue
        name_node = anchor.find(["h1", "h2", "h3", "h4", "h5"])
        title_node = anchor.find("p")
        image_node = anchor.find("img")
        name = gdgh_clean_text(name_node.get_text(" ", strip=True) if name_node else "")
        title = gdgh_clean_text(title_node.get_text(" ", strip=True) if title_node else "")
        image_url = gdgh_photo_url(
            str(
                (image_node.get("src") or image_node.get("data-src") or image_node.get("data-original") or "")
                if image_node
                else ""
            ),
            department["department_url"],
        )
        if not name:
            continue
        seen_ids.add(detail_id)
        relations.append(
            {
                "id": detail_id,
                "subject_id": department["subject_id"],
                "department": department["department"],
                "department_url": department["department_url"],
                "name": name,
                "list_title": title,
                "source_link": source_link,
                "image_url": image_url,
            }
        )
    return relations


GDGH_SCHEDULE_TAIL_PATTERN = re.compile(
    r"(?:门诊时间地点|门诊时间|出诊时间|开诊时间)\s*[:：]?.*$",
    flags=re.IGNORECASE | re.DOTALL,
)
GDGH_FORBIDDEN_SENTENCE_MARKERS = (
    "好医生榜",
    "医生排行榜",
    "患者评价",
    "患者留言",
    "问诊记录",
    "问诊内容",
    "全美排名",
    "医院排名",
    "专科排名",
)


def parse_gdgh_detail(html: str, fallback: dict[str, str]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    container = soup.select_one(".sub_ex_Info")
    if not container:
        return {
            "name": gdgh_clean_text(fallback.get("name")),
            "title": gdgh_clean_text(fallback.get("list_title")),
            "specialty": "",
            "profile_text": "",
            "photo_url": clean_text(fallback.get("image_url")),
            "schedule_exclusion_count": 0,
            "forbidden_segment_count": 0,
            "patient_case_exclusion_count": 0,
        }

    heading = container.find(["h1", "h2", "h3", "h4", "h5"])
    heading_text = strip_gzsys_forbidden_text(
        heading.get_text(" ", strip=True) if heading else ""
    )
    name = ""
    title = ""
    heading_match = re.match(r"^(.{1,30}?)\s*[—－-]{2,}\s*(.+)$", heading_text)
    if heading_match:
        name, title = gdgh_clean_text(heading_match.group(1)), gdgh_clean_text(
            heading_match.group(2)
        )
    else:
        name = gdgh_clean_text(fallback.get("name"))
        title = gdgh_clean_text(fallback.get("list_title"))

    body = container.select_one(".divmore") or container.select_one(".divtxtbox")
    raw_profile = strip_gzsys_forbidden_text(
        body.get_text(" ", strip=True) if body else ""
    )
    schedule_exclusion_count = int(bool(GDGH_SCHEDULE_TAIL_PATTERN.search(raw_profile)))
    raw_profile = GDGH_SCHEDULE_TAIL_PATTERN.sub(" ", raw_profile)
    kept_sentences: list[str] = []
    forbidden_segment_count = 0
    patient_case_exclusion_count = 0
    for sentence in re.split(r"(?<=[。！？；;])\s*", raw_profile):
        sentence = clean_text(sentence)
        if not sentence:
            continue
        if any(marker in sentence for marker in GDGH_FORBIDDEN_SENTENCE_MARKERS):
            forbidden_segment_count += 1
            continue
        if contains_gzbrain_patient_case_text(sentence):
            patient_case_exclusion_count += 1
            continue
        kept_sentences.append(sentence)
    profile_text = clean_text(" ".join(kept_sentences))
    specialty = ""
    specialty_match = re.search(
        r"(?:擅长|专长|主攻|熟悉)\s*[:：]?\s*([^。！？；;]{4,520})",
        profile_text,
    )
    if specialty_match:
        specialty = clean_text(specialty_match.group(1))

    image_node = container.find("img")
    image_source = ""
    if image_node:
        image_source = str(
            image_node.get("src")
            or image_node.get("data-src")
            or image_node.get("data-original")
            or ""
        )
    photo_url = gdgh_photo_url(image_source, str(fallback.get("source_link") or ""))
    return {
        "name": first_nonempty(name, fallback.get("name")),
        "title": first_nonempty(title, fallback.get("list_title")),
        "specialty": clip(specialty, 520),
        "profile_text": clip(profile_text, 1800),
        "photo_url": first_nonempty(photo_url, fallback.get("image_url")),
        "schedule_exclusion_count": schedule_exclusion_count,
        "forbidden_segment_count": forbidden_segment_count,
        "patient_case_exclusion_count": patient_case_exclusion_count,
    }


def gdgh_primary_title(value: str | None) -> str:
    text = gdgh_clean_text(value)
    for term in (
        "一级主任医师",
        "主任中医师",
        "副主任中医师",
        "主任营养医师",
        "副主任营养医师",
        "主治营养医师",
        "主任医师",
        "副主任医师",
        "主治医师",
        "住院医师",
        "检验医师",
        "医师",
        "主任药师",
        "副主任药师",
        "主管药师",
        "主任技师",
        "副主任技师",
        "主管技师",
        "检验技师",
        "康复治疗师",
        "主管治疗师",
        "心理治疗师",
        "研究员",
        "副研究员",
        "教授",
        "副教授",
    ):
        if term in text:
            return term
    return ""


def gdgh_photo_part(value: str | None) -> str:
    text = gdgh_clean_text(value)
    text = re.sub(r'[\\/:*?"<>|]', "_", text).strip(" .")
    return text or "未标注"


def gdgh_first_atomic_department(value: str | None) -> str:
    return gdgh_photo_part(re.split(r"[、,，;/；|]+", clean_text(value), maxsplit=1)[0])


def gdgh_photo_extension(content: bytes, content_type: str | None) -> str:
    media_type = clean_text(content_type).split(";", 1)[0].lower()
    if not media_type.startswith("image/"):
        return ""
    if content.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if content.startswith((b"GIF87a", b"GIF89a")):
        return "gif"
    if len(content) >= 12 and content.startswith(b"RIFF") and content[8:12] == b"WEBP":
        return "webp"
    return ""


def download_gdgh_photo(
    session: requests.Session,
    photo_url: str,
    output_dir: Path,
    filename_stem: str,
    detail_id: str,
    used_filenames: set[str],
) -> dict[str, Any]:
    official_url = gdgh_photo_url(photo_url, photo_url)
    if not official_url:
        raise RuntimeError(f"照片 URL 不属于官网公开 uploadfiles 路径：{photo_url}")
    response = session.get(official_url, timeout=35)
    content_type = clean_text(response.headers.get("Content-Type"))
    if response.status_code != 200:
        raise RuntimeError(f"照片下载 HTTP {response.status_code}：{official_url}")
    extension = gdgh_photo_extension(response.content, content_type)
    if not extension:
        raise RuntimeError(
            f"照片响应格式不受支持：{content_type or '未声明 Content-Type'} {official_url}"
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{filename_stem}.{extension}"
    path = output_dir / filename
    if filename.casefold() in used_filenames or (path.exists() and path.read_bytes() != response.content):
        filename = f"{filename_stem}-{gdgh_photo_part(detail_id)}.{extension}"
        path = output_dir / filename
    if path.exists() and path.read_bytes() != response.content:
        raise RuntimeError(f"照片目标已存在且内容不同，拒绝覆盖：{path}")
    if not path.exists():
        path.write_bytes(response.content)
    used_filenames.add(filename.casefold())
    return {
        "photo_url": official_url,
        "filename": filename,
        "bytes": len(response.content),
        "sha256": hashlib.sha256(response.content).hexdigest(),
        "disk_path": str(path),
    }


def parse_gdgh_affiliates(html: str, entry_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    affiliates: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select("a[href]"):
        url = urljoin(entry_url, str(anchor.get("href") or ""))
        parsed = urlparse(url)
        if (
            comparable_host(url) != "gdghospital.org.cn"
            or not re.fullmatch(r"/expert/info_itemid_\d+\.html", parsed.path, re.I)
            or parsed.query
            or parsed.fragment
            or url in seen
        ):
            continue
        name_node = anchor.find(["h1", "h2", "h3", "h4", "h5"])
        name = clean_text(
            name_node.get_text(" ", strip=True) if name_node else anchor.get_text(" ", strip=True)
        )
        if name:
            seen.add(url)
            affiliates.append({"name": name, "url": url})
    return affiliates


def select_gdgh_trial_doctors(
    doctors: list[dict[str, Any]],
    max_doctors: int | None,
) -> list[dict[str, Any]]:
    if not max_doctors or len(doctors) <= max_doctors:
        return doctors[:]
    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()
    selected_names: set[str] = set()
    covered_departments: set[str] = set()

    def add(doctor: dict[str, Any]) -> bool:
        identity_id = clean_text(str(doctor.get("id") or ""))
        name = gdzy5413_normalized_name(doctor.get("name"))
        if not identity_id or not name or identity_id in selected_ids or name in selected_names:
            return False
        selected.append(doctor)
        selected_ids.add(identity_id)
        selected_names.add(name)
        covered_departments.update(
            clean_text(str(value))
            for value in doctor.get("departments", [])
            if clean_text(str(value))
        )
        return True

    for doctor in doctors:
        departments = {
            clean_text(str(value))
            for value in doctor.get("departments", [])
            if clean_text(str(value))
        }
        if departments - covered_departments:
            add(doctor)
        if len(selected) >= max_doctors:
            return selected
    for doctor in doctors:
        add(doctor)
        if len(selected) >= max_doctors:
            return selected
    return selected


def merge_gdgh_identity_rows(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(gdzy5413_normalized_name(row.get("姓名")), []).append(row)

    merged_rows: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    longest_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "详情正文摘录"]
    for name, name_rows in by_name.items():
        clusters: list[list[dict[str, Any]]] = []
        for row in name_rows:
            for cluster in clusters:
                if any(gdzy5413_rows_same_identity(row, member) for member in cluster):
                    cluster.append(row)
                    break
            else:
                clusters.append([row])
        distinct_same_name = len(clusters) > 1
        for identity_index, cluster in enumerate(clusters, start=1):
            primary = max(cluster, key=gdzy5413_primary_row_score)
            merged = dict(primary)
            departments: list[str] = []
            for member in cluster:
                for department in clean_text(member.get("科室_分类页")).split("、"):
                    department = clean_text(department)
                    if department and department not in departments:
                        departments.append(department)
                for field in longest_fields:
                    if len(clean_text(member.get(field))) > len(clean_text(merged.get(field))):
                        merged[field] = member.get(field, "")
            merged["科室_分类页"] = "、".join(departments)
            merged["科室_列表卡片"] = "、".join(departments)
            warnings = [
                warning
                for member in cluster
                for warning in clean_text(member.get("异常提示")).split("；")
                if warning
            ]
            titles = {
                clean_text(member.get("职称身份原文"))
                for member in cluster
                if clean_text(member.get("职称身份原文"))
            }
            if len(titles) > 1:
                warnings.append("多详情职称不一致")
            if distinct_same_name:
                warnings.append("同名待甄别")
            merged["异常提示"] = "；".join(dict.fromkeys(warnings))
            if merged["异常提示"]:
                merged["重点优先级"] = "普通"
                merged["重点关注范围"] = ""
                merged["重点疾病标签"] = ""
            merged_rows.append(merged)
            reconciliation.append(
                {
                    "name": name,
                    "identity_index": identity_index,
                    "detail_ids": [str(member.get("_gdgh_item_id") or "") for member in cluster],
                    "resolution": (
                        "同名待甄别"
                        if distinct_same_name
                        else "同一人归并"
                        if len(cluster) > 1
                        else "唯一身份"
                    ),
                    "relation_count": len(cluster),
                    "departments": departments,
                    "primary_source_link": merged.get("来源链接", ""),
                    "merged_source_links": [
                        member.get("来源链接", "")
                        for member in cluster
                        if member.get("来源链接") != merged.get("来源链接")
                    ],
                }
            )
    return merged_rows, reconciliation


def collect_gdgh(
    target: HospitalTarget,
    today: str,
    max_doctors: int | None = None,
    photo_root: Path | None = None,
    full_mode: bool = False,
) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页普通公开 GET 读取失败：{entry_error}")
    groups, departments = parse_gdgh_directory(entry_html, target.entry_url)
    if not departments:
        raise RuntimeError("官网目录未发现严格 Specialistthree 科室入口。")

    relations: list[dict[str, str]] = []
    category_errors: list[dict[str, str]] = []
    entry_reconnaissance: list[dict[str, Any]] = []
    for department in departments:
        status, html, error = fetch(session, department["department_url"])
        if status != 200:
            category_errors.append(
                {
                    "page": department["department"],
                    "url": department["department_url"],
                    "error": error,
                }
            )
            continue
        department_relations = parse_gdgh_department_page(html, department)
        relations.extend(department_relations)
        entry_reconnaissance.append(
            {
                "category_name": department["department"],
                "entry_url": department["department_url"],
                "page_nature": "官网静态科室首页内嵌完整专家卡片",
                "list_page_count": 1,
                "raw_detail_relation_count": len(department_relations),
                "unique_detail_count": len({item["id"] for item in department_relations}),
                "out_of_scope_detail_count": sum(
                    gyfyyy_nursing_only_identity(item["list_title"])
                    for item in department_relations
                ),
                "affiliation": target.hospital,
                "independent_entity_check": "同域科室页；数字 itemid + subjectid 严格详情关系",
            }
        )
        time.sleep(0.05)
    if category_errors:
        raise RuntimeError(
            f"官网科室普查不完整：{len(category_errors)} 个科室读取失败（"
            + "、".join(item["page"] for item in category_errors)
            + "）"
        )

    by_id: dict[str, dict[str, Any]] = {}
    for relation in relations:
        item = by_id.setdefault(
            relation["id"],
            {
                **relation,
                "departments": [],
                "subject_ids": [],
                "department_urls": [],
                "source_links": [],
                "list_titles": [],
                "image_urls": [],
            },
        )
        for field, value in (
            ("departments", relation["department"]),
            ("subject_ids", relation["subject_id"]),
            ("department_urls", relation["department_url"]),
            ("source_links", relation["source_link"]),
            ("list_titles", relation["list_title"]),
            ("image_urls", relation["image_url"]),
        ):
            if value and value not in item[field]:
                item[field].append(value)
    all_doctors = list(by_id.values())

    names_to_ids: dict[str, list[str]] = {}
    for item in all_doctors:
        names_to_ids.setdefault(clean_text(item.get("name")), []).append(item["id"])
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in names_to_ids.items()
        if name and len(set(ids)) > 1
    }

    excluded_candidates: list[dict[str, str]] = []
    eligible_doctors: list[dict[str, Any]] = []
    for item in all_doctors:
        list_title = "、".join(item["list_titles"])
        if gyfyyy_nursing_only_identity(list_title):
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": f"{item['name']} {list_title}",
                    "source_link": item["source_links"][0],
                    "reason": "官网专家卡片仅标注护理身份，排除医生画像采集范围",
                }
            )
            continue
        eligible_doctors.append(item)

    selected_doctors = select_gdgh_trial_doctors(eligible_doctors, max_doctors)
    existing_links = collect_existing_profile_links()
    raw_rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, Any]] = []
    schedule_exclusion_count = 0
    forbidden_segment_count = 0
    patient_case_exclusion_count = 0
    for item in selected_doctors:
        source_link = item["source_links"][0]
        status, html, error = fetch(session, source_link)
        if status == 200:
            detail = parse_gdgh_detail(
                html,
                {
                    "name": item["name"],
                    "list_title": item["list_titles"][0] if item["list_titles"] else "",
                    "image_url": item["image_urls"][0] if item["image_urls"] else "",
                    "source_link": source_link,
                },
            )
        else:
            detail_errors.append({"source_link": source_link, "error": error})
            detail = {
                "name": item["name"],
                "title": item["list_titles"][0] if item["list_titles"] else "",
                "specialty": "",
                "profile_text": "",
                "photo_url": item["image_urls"][0] if item["image_urls"] else "",
                "schedule_exclusion_count": 0,
                "forbidden_segment_count": 0,
                "patient_case_exclusion_count": 0,
            }
        name = gdgh_clean_text(str(detail.get("name") or item["name"]))
        title_identity = gdgh_clean_text(str(detail.get("title") or ""))
        if gyfyyy_nursing_only_identity(title_identity):
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": f"{name} {title_identity}",
                    "source_link": source_link,
                    "reason": "官网详情仅标注护理身份，排除医生画像采集范围",
                }
            )
            continue
        department = "、".join(item["departments"])
        specialty = gdgh_clean_text(str(detail.get("specialty") or ""))
        profile_text = gdgh_clean_text(str(detail.get("profile_text") or ""))
        schedule_exclusion_count += int(detail.get("schedule_exclusion_count") or 0)
        forbidden_segment_count += int(detail.get("forbidden_segment_count") or 0)
        patient_case_exclusion_count += int(detail.get("patient_case_exclusion_count") or 0)
        combined_text = "\n".join(
            [target.hospital, department, title_identity, specialty, profile_text]
        )
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if item["name"] and name != item["name"]:
            warnings.append("列表与详情姓名不一致")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if warnings:
            groups_found, tags = [], []
        priority = "普通"
        if not warnings and (
            any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
        ):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        raw_rows.append(
            {
                "序号": len(raw_rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": department,
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": profile_text,
                "来源类型": "医院官网",
                "来源链接": source_link,
                "照片链接": "",
                "照片文件": "",
                "采集入口": target.entry_url,
                "采集方式": "官网科室树+严格 itemid/subjectid 医生详情 DOM+同域本人职业照",
                "采集日期": today,
                "详情页状态": "200" if status == 200 else "失败",
                "已建画像": "是" if canonical_url(source_link) in existing_links else "否",
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
                "_gdgh_item_id": item["id"],
                "_gdgh_photo_url": clean_text(str(detail.get("photo_url") or "")),
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": item["id"],
                "name": name,
                "departments": item["departments"],
                "source_link": source_link,
                "photo_url": clean_text(str(detail.get("photo_url") or "")),
                "resolution": "详情已读取" if status == 200 else "详情读取失败",
            }
        )
        time.sleep(0.12)

    rows, identity_reconciliation = merge_gdgh_identity_rows(raw_rows)
    photo_dir = photo_root or (VAULT / "01_试点医院" / target.hospital / "照片")
    used_filenames: set[str] = set()
    photo_samples: list[dict[str, Any]] = []
    photo_errors: list[dict[str, str]] = []
    photo_expected_count = 0
    photo_no_source_count = 0
    for index, row in enumerate(rows, start=1):
        row["序号"] = index
        detail_id = clean_text(str(row.pop("_gdgh_item_id", "")))
        photo_url = clean_text(str(row.pop("_gdgh_photo_url", "")))
        if not photo_url:
            photo_no_source_count += 1
            continue
        photo_expected_count += 1
        stem = "-".join(
            [
                gdgh_photo_part(row.get("姓名")),
                gdgh_first_atomic_department(row.get("科室_分类页")),
                gdgh_photo_part(gdgh_primary_title(row.get("职称身份原文"))),
                gdgh_photo_part(target.hospital),
            ]
        )
        try:
            photo = download_gdgh_photo(
                session, photo_url, photo_dir, stem, detail_id, used_filenames
            )
        except Exception as exc:  # noqa: BLE001 - retain per-photo failure evidence
            photo_errors.append(
                {
                    "name": clean_text(str(row.get("姓名") or "")),
                    "detail_id": detail_id,
                    "source_link": clean_text(str(row.get("来源链接") or "")),
                    "photo_url": photo_url,
                    "error": str(exc),
                }
            )
            row["异常提示"] = "；".join(
                dict.fromkeys(
                    [
                        *clean_text(str(row.get("异常提示") or "")).split("；"),
                        "照片获取失败",
                    ]
                )
            ).strip("；")
            row["重点优先级"] = "普通"
            row["重点关注范围"] = ""
            row["重点疾病标签"] = ""
            continue
        relative_path = (
            Path("01_试点医院") / target.hospital / "照片" / photo["filename"]
        ).as_posix()
        row["照片链接"] = photo["photo_url"]
        row["照片文件"] = relative_path
        photo_samples.append(
            {
                "name": row["姓名"],
                "department": gdgh_first_atomic_department(row["科室_分类页"]),
                "title": gdgh_primary_title(row["职称身份原文"]) or "未标注",
                "detail_id": detail_id,
                "source_link": row["来源链接"],
                "photo_url": photo["photo_url"],
                "photo_file": relative_path,
                "filename": photo["filename"],
                "bytes": photo["bytes"],
                "sha256": photo["sha256"],
                "disk_path": photo["disk_path"],
            }
        )

    branch_url = "https://www.gdghospital.org.cn/expert/list.html"
    branch_status, branch_html, branch_error = fetch(session, branch_url)
    if branch_status != 200:
        raise RuntimeError(f"分院/研究所入口读取失败：{branch_error}")
    affiliates = parse_gdgh_affiliates(branch_html, branch_url)
    affiliate_evidence: list[dict[str, str]] = []
    for affiliate in affiliates:
        status, html, error = fetch(session, affiliate["url"])
        if status != 200:
            raise RuntimeError(f"分院/研究所详情读取失败：{affiliate['name']} {error}")
        text = strip_profile_navigation_text(BeautifulSoup(html, "html.parser").get_text(" ", strip=True))
        independent_markers = [
            marker
            for marker in ("独立法人", "法人单位", "统一社会信用代码")
            if marker in text
        ]
        if independent_markers:
            raise RuntimeError(
                f"[FATAL - HUMAN_INTERVENTION_REQUIRED] {affiliate['name']} 出现独立法人证据："
                + "、".join(independent_markers)
            )
        relation = (
            "一套人马、两块牌子；由广东省人民医院代管"
            if "一套人马" in text and "两块牌子" in text
            else "广东省人民医院重要组成部分"
            if "重要组成部分" in text
            else "广东省人民医院所属研究所"
            if "广东省人民医院" in text and "研究所" in affiliate["name"]
            else "官网同域分院/研究所，未发现独立法人标识"
        )
        affiliate_evidence.append(
            {"name": affiliate["name"], "url": affiliate["url"], "relation": relation}
        )

    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(row["科室_分类页"]).split("、")
        if department
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in row["重点关注范围"].split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in row["异常提示"].split("；") if warning
    )
    sample_departments = covered_department_names(rows)
    cross_department = [item for item in all_doctors if len(item["departments"]) > 1]
    average_photo_bytes = (
        round(sum(item["bytes"] for item in photo_samples) / len(photo_samples))
        if photo_samples
        else 0
    )
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #41（与官网入口台账序号 19 一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(departments),
            "census_group_count": len(groups),
            "census_groups": groups,
            "census_department_count": len(departments),
            "census_departments": [item["department"] for item in departments],
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(all_doctors),
            "eligible_candidate_count": len(eligible_doctors),
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(all_doctors),
            "census_named_detail_count": sum(bool(item["name"]) for item in all_doctors),
            "census_blank_name_detail_count": sum(not item["name"] for item in all_doctors),
            "census_unique_nonblank_name_count": len(names_to_ids),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "census_nonempty_department_count": sum(bool(item["departments"]) for item in all_doctors),
            "census_empty_department_count": sum(not item["departments"] for item in all_doctors),
            "sample_entry_coverage_count": len(sample_departments),
            "sample_entry_categories": sample_departments,
            "pagination_count": len(departments),
            "pagination_method": "83 个科室首页均服务端一次性输出完整专家卡片；未发现分页、加载更多或筛选接口",
            "category_error_count": 0,
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(all_doctors),
            "gdgh_cross_department_identity_count": len(cross_department),
            "excluded_non_doctor_count": len(excluded_candidates),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
            "entry_candidate_counts": {target.entry_url: len(relations)},
            "standard_public_session": "requests 常规公开 GET；无登录、Cookie 注入、验证码/挑战求解或非公开接口",
            "schedule_exclusion_count": schedule_exclusion_count,
            "schedule_field_ingested_count": 0,
            "forbidden_segment_exclusion_count": forbidden_segment_count,
            "patient_case_exclusion_count": patient_case_exclusion_count,
            "private_use_character_count": sum(
                len(re.findall(r"[\ue000-\uf8ff]", str(row.get(field) or "")))
                for row in rows
                for field in BASE_HEADERS
            ),
            "photo_sample_count": len(photo_samples),
            "photo_error_count": len(photo_errors),
            "photo_expected_count": photo_expected_count,
            "photo_downloaded_count": len(photo_samples),
            "photo_failed_count": len(photo_errors),
            "photo_no_source_count": photo_no_source_count,
            "photo_average_bytes": average_photo_bytes,
            "photo_estimated_full_count": len(eligible_doctors),
            "photo_estimated_full_bytes": average_photo_bytes * len(eligible_doctors),
            "photo_census_available_count": sum(bool(item["image_urls"]) for item in eligible_doctors),
            "photo_policy_status": (
                "OWNER_APPROVED_ORIGINAL_NO_WIDTH_LIMIT"
                if full_mode
                else "WAITING_OWNER_SIZE_POLICY"
            ),
            "affiliate_count": len(affiliate_evidence),
            "independent_entity_count": 0,
        },
        "categories": departments,
        "entry_reconnaissance": entry_reconnaissance,
        "affiliate_reconnaissance": affiliate_evidence,
        "excluded_candidates": excluded_candidates,
        "cross_entry_duplicates": [
            {
                "name": item["name"],
                "source_link": item["source_links"][0],
                "entry_urls": item["department_urls"],
            }
            for item in cross_department
        ],
        "gdgh_detail_reconciliation": detail_reconciliation,
        "gdgh_identity_reconciliation": identity_reconciliation,
        "photo_samples": photo_samples,
        "photo_errors": photo_errors,
        "category_errors": category_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def gdmch_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "e3861.com" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"/keshizhuanjia/zhuanjiajieshao/(\d+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    return match.group(1) if match else ""


def gdmch_photo_url(value: str | None, base_url: str) -> str:
    absolute = urljoin(base_url, clean_text(value))
    parsed = urlparse(absolute)
    if (
        parsed.scheme != "https"
        or parsed.netloc.lower() != "wx.e3861.com"
        or not parsed.path.lower().startswith("/sfyadmin/images/doctor/")
        or parsed.query
        or parsed.fragment
    ):
        return ""
    return absolute


def gdmch_list_page_url(entry_url: str, page_number: int) -> str:
    if page_number <= 1:
        return entry_url
    parsed = urlparse(entry_url)
    return urlunparse(
        parsed._replace(
            query=urlencode(
                {
                    "searchDoctor": "",
                    "searchDepartment": "",
                    "page": page_number,
                }
            )
        )
    )


def discover_gdmch_page_count(html: str, entry_url: str) -> int:
    soup = BeautifulSoup(html, "html.parser")
    pages = {1}
    for anchor in soup.select(".paged a[href]"):
        candidate = urljoin(entry_url, str(anchor.get("href") or ""))
        parsed = urlparse(candidate)
        if comparable_host(candidate) != "e3861.com":
            continue
        if parsed.path.rstrip("/").lower() != "/keshizhuanjia/zhuanjiajieshao":
            continue
        query = dict(parse_qsl(parsed.query, keep_blank_values=True))
        value = clean_text(query.get("page"))
        if value.isdigit():
            pages.add(int(value))
    return max(pages)


def parse_gdmch_list_page(html: str, page_url: str, page_number: int) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for anchor in soup.select(".expert ul.list a.item[href]"):
        source_link = urljoin(page_url, str(anchor.get("href") or ""))
        detail_id = gdmch_detail_id(source_link)
        if not detail_id or detail_id in seen_ids:
            continue
        name_node = anchor.select_one(".right .title")
        title_node = anchor.select_one(".right .desc")
        photo_node = anchor.select_one(".img-box img")
        name = gdgh_clean_text(name_node.get_text(" ", strip=True) if name_node else "")
        title = gdgh_clean_text(title_node.get_text(" ", strip=True) if title_node else "")
        photo_url = gdmch_photo_url(
            str(photo_node.get("src") or "") if photo_node else "",
            page_url,
        )
        if not name:
            continue
        seen_ids.add(detail_id)
        rows.append(
            {
                "id": detail_id,
                "name": name,
                "list_title": title,
                "source_link": source_link,
                "photo_url": photo_url,
                "list_page": str(page_number),
                "list_page_url": page_url,
            }
        )
    return rows


GDMCH_NON_DOCTOR_NAME_MARKERS = (
    "号",
    "门诊",
    "筛查",
    "管理员",
    "test",
    "咨询医生",
    "手术评估",
    "随访",
    "体重管理",
)
GDMCH_NON_DOCTOR_TITLE_MARKERS = {"收费"}


def gdmch_non_doctor_card(name: str | None, title: str | None) -> bool:
    candidate = gdgh_clean_text(name)
    title_identity = gdgh_clean_text(title)
    if not looks_like_person_name(candidate):
        return True
    if any(marker.casefold() in candidate.casefold() for marker in GDMCH_NON_DOCTOR_NAME_MARKERS):
        return True
    if title_identity in GDMCH_NON_DOCTOR_TITLE_MARKERS:
        return True
    return gyfyyy_nursing_only_identity(title_identity)


GDMCH_SCHEDULE_HEADING_PATTERN = re.compile(
    r"(?:出诊地点及时间|出诊时间|出诊安排|门诊时间|开诊时间|专科门诊)\s*[:：]?"
)
GDMCH_SCHEDULE_ONLY_PATTERN = re.compile(
    r"^周[一二三四五六日天].*(?:院区|门诊|全天|上午|下午|晚上)"
)


def strip_gdmch_schedule_tail(value: str | None) -> tuple[str, bool]:
    text = gdgh_clean_text(value)
    match = GDMCH_SCHEDULE_HEADING_PATTERN.search(text)
    if not match:
        if GDMCH_SCHEDULE_ONLY_PATTERN.search(text):
            return "", True
        return text, False
    return gdgh_clean_text(text[: match.start()]), True


def parse_gdmch_detail(html: str, fallback: dict[str, str]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    container = soup.select_one(".expert-detail")
    if not container:
        return {
            "name": fallback.get("name", ""),
            "title": fallback.get("list_title", ""),
            "departments": [],
            "campuses": [],
            "specialty": "",
            "profile_text": "",
            "photo_url": fallback.get("photo_url", ""),
            "schedule_exclusion_count": 0,
            "forbidden_segment_count": 0,
            "patient_case_exclusion_count": 0,
        }

    labeled: dict[str, str] = {}
    for info in container.select(".detail-head .info"):
        spans = info.find_all("span")
        if len(spans) < 2:
            continue
        label = gdgh_clean_text(spans[0].get_text(" ", strip=True)).rstrip("：:")
        value = gdgh_clean_text(" ".join(span.get_text(" ", strip=True) for span in spans[1:]))
        if label:
            labeled[label] = value

    schedule = labeled.get("出诊安排", "")
    normalized_schedule = schedule.replace("(", "（").replace(")", "）")
    department_campus_pairs = re.findall(
        r"（([^（）]+?)（(番禺|越秀|天河|清远)））",
        normalized_schedule,
    )
    departments: list[str] = []
    campuses: list[str] = []
    for clinic, campus in department_campus_pairs:
        department = gdgh_clean_text(
            re.sub(r"(?:专家门诊|专科门诊|普通门诊|门诊)$", "", clinic)
        )
        campus_name = f"{campus}院区"
        if department and department not in departments:
            departments.append(department)
        if campus_name not in campuses:
            campuses.append(campus_name)

    content_node = container.select_one(".content-box .content")
    raw_segments = (
        [
            gdgh_clean_text(segment)
            for segment in content_node.get_text("\n", strip=True).splitlines()
            if gdgh_clean_text(segment)
        ]
        if content_node
        else []
    )
    kept_segments: list[str] = []
    forbidden_segment_count = 0
    patient_case_exclusion_count = 0
    schedule_content_exclusion_count = 0
    for segment in raw_segments:
        segment, schedule_removed = strip_gdmch_schedule_tail(segment)
        schedule_content_exclusion_count += int(schedule_removed)
        if not segment:
            if schedule_removed:
                break
            continue
        if any(marker in segment for marker in GDGH_FORBIDDEN_SENTENCE_MARKERS):
            forbidden_segment_count += 1
            continue
        if contains_gzbrain_patient_case_text(segment):
            patient_case_exclusion_count += 1
            continue
        cleaned = strip_profile_navigation_text(segment)
        if cleaned:
            kept_segments.append(cleaned)
        if schedule_removed:
            break

    specialty = ""
    for segment in kept_segments:
        match = re.search(
            r"(?:专业专长|专业特长|擅长|专长)\s*[:：]\s*(.+)",
            segment,
        )
        if match:
            specialty = gdgh_clean_text(match.group(1))
            break
    profile_text = clip(" ".join(kept_segments), 1800)
    image_node = container.select_one(".detail-head .img-box img")
    detail_photo = gdmch_photo_url(
        str(image_node.get("src") or "") if image_node else "",
        fallback.get("source_link", ""),
    )
    return {
        "name": first_nonempty(labeled.get("姓名"), fallback.get("name")),
        "title": first_nonempty(labeled.get("职称"), fallback.get("list_title")),
        "departments": departments,
        "campuses": campuses,
        "specialty": specialty,
        "profile_text": profile_text,
        "photo_url": first_nonempty(detail_photo, fallback.get("photo_url")),
        "schedule_exclusion_count": int(bool(schedule)) + schedule_content_exclusion_count,
        "forbidden_segment_count": forbidden_segment_count,
        "patient_case_exclusion_count": patient_case_exclusion_count,
    }


def select_gdmch_trial_doctors(
    doctors: list[dict[str, str]], max_doctors: int | None
) -> list[dict[str, str]]:
    if not max_doctors or len(doctors) <= max_doctors:
        return doctors[:]
    if max_doctors == 1:
        return doctors[:1]
    target_pages = [
        round(1 + index * (GDMCH_EXPECTED_PAGE_COUNT - 1) / (max_doctors - 1))
        for index in range(max_doctors)
    ]
    selected: list[dict[str, str]] = []
    selected_ids: set[str] = set()
    for target_page in target_pages:
        candidates = sorted(
            doctors,
            key=lambda item: (
                abs(int(item["list_page"]) - target_page),
                int(item["list_page"]),
                int(item["id"]),
            ),
        )
        for item in candidates:
            if item["id"] not in selected_ids:
                selected.append(item)
                selected_ids.add(item["id"])
                break
    return selected[:max_doctors]


GDMCH_CAMPUS_SUFFIX_PATTERN = re.compile(
    r"（(?:番禺|越秀|天河|清远)院区"
    r"(?:、(?:番禺|越秀|天河|清远)院区)*）$"
)


def gdmch_covered_department_names(rows: list[dict[str, Any]]) -> list[str]:
    departments: set[str] = set()
    for row in rows:
        value = clean_text(str(row.get("科室_分类页") or ""))
        value = GDMCH_CAMPUS_SUFFIX_PATTERN.sub("", value)
        for department in re.split(r"[、,，;/；|]+", value):
            department = clean_text(department)
            if department:
                departments.add(department)
    return sorted(departments)


def gdmch_photo_dimensions(content: bytes, extension: str) -> tuple[int, int]:
    if extension == "png" and len(content) >= 24:
        return int.from_bytes(content[16:20], "big"), int.from_bytes(content[20:24], "big")
    if extension == "gif" and len(content) >= 10:
        return int.from_bytes(content[6:8], "little"), int.from_bytes(content[8:10], "little")
    if extension == "jpg" and content.startswith(b"\xff\xd8"):
        offset = 2
        sof_markers = {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}
        while offset + 8 < len(content):
            if content[offset] != 0xFF:
                offset += 1
                continue
            marker = content[offset + 1]
            if marker in sof_markers:
                height = int.from_bytes(content[offset + 5 : offset + 7], "big")
                width = int.from_bytes(content[offset + 7 : offset + 9], "big")
                return width, height
            if marker in {0xD8, 0xD9}:
                offset += 2
                continue
            segment_length = int.from_bytes(content[offset + 2 : offset + 4], "big")
            if segment_length < 2:
                break
            offset += 2 + segment_length
    if extension == "webp" and len(content) >= 30 and content[:4] == b"RIFF":
        chunk = content[12:16]
        if chunk == b"VP8X":
            return (
                1 + int.from_bytes(content[24:27], "little"),
                1 + int.from_bytes(content[27:30], "little"),
            )
        if chunk == b"VP8 " and content[23:26] == b"\x9d\x01\x2a":
            return (
                int.from_bytes(content[26:28], "little") & 0x3FFF,
                int.from_bytes(content[28:30], "little") & 0x3FFF,
            )
    return 0, 0


def merge_gdmch_identity_rows(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Apply the verified Issue #43 same-name identity decisions before master append."""

    same_identity_by_id = {
        detail_id: group
        for group in GDMCH_SAME_IDENTITY_DETAIL_GROUPS
        for detail_id in group
    }
    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(gdzy5413_normalized_name(row.get("姓名")), []).append(row)

    merged_rows: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    longest_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "详情正文摘录"]
    for name, name_rows in by_name.items():
        clusters: list[list[dict[str, Any]]] = []
        for row in name_rows:
            detail_id = clean_text(str(row.get("_gdmch_detail_id") or ""))
            expected_group = same_identity_by_id.get(detail_id)
            matched_cluster = next(
                (
                    cluster
                    for cluster in clusters
                    if expected_group
                    and any(
                        clean_text(str(member.get("_gdmch_detail_id") or ""))
                        in expected_group
                        for member in cluster
                    )
                ),
                None,
            )
            if matched_cluster is None:
                clusters.append([row])
            else:
                matched_cluster.append(row)

        distinct_same_name = len(clusters) > 1
        for identity_index, cluster in enumerate(clusters, start=1):
            primary = max(
                cluster,
                key=lambda item: (
                    bool(clean_text(str(item.get("_gdmch_photo_url") or ""))),
                    gdzy5413_primary_row_score(item),
                ),
            )
            merged = dict(primary)
            departments: list[str] = []
            for member in cluster:
                department = clean_text(str(member.get("科室_分类页") or ""))
                if department and department not in departments:
                    departments.append(department)
                for field in longest_fields:
                    if len(clean_text(str(member.get(field) or ""))) > len(
                        clean_text(str(merged.get(field) or ""))
                    ):
                        merged[field] = member.get(field, "")
            merged["科室_分类页"] = "、".join(departments)
            merged["科室_列表卡片"] = ""
            merged["职称_关键词"] = "、".join(
                extract_terms(clean_text(str(merged.get("职称身份原文") or "")), TITLE_TERMS)
            )
            warnings = [
                warning
                for member in cluster
                for warning in clean_text(str(member.get("异常提示") or "")).split("；")
                if warning
            ]
            titles = {
                clean_text(str(member.get("职称身份原文") or ""))
                for member in cluster
                if clean_text(str(member.get("职称身份原文") or ""))
            }
            if len(titles) > 1:
                warnings.append("多详情职称不一致")
            if distinct_same_name:
                warnings.append("同名待甄别")
            merged["异常提示"] = "；".join(dict.fromkeys(warnings))
            if merged["异常提示"]:
                merged["重点优先级"] = "普通"
                merged["重点关注范围"] = ""
                merged["重点疾病标签"] = ""
            merged_rows.append(merged)
            reconciliation.append(
                {
                    "name": name,
                    "identity_index": identity_index,
                    "detail_ids": [
                        clean_text(str(member.get("_gdmch_detail_id") or ""))
                        for member in cluster
                    ],
                    "resolution": (
                        "同名待甄别"
                        if distinct_same_name
                        else "同一人归并"
                        if len(cluster) > 1
                        else "唯一身份"
                    ),
                    "relation_count": len(cluster),
                    "departments": departments,
                    "primary_source_link": merged.get("来源链接", ""),
                    "merged_source_links": [
                        member.get("来源链接", "")
                        for member in cluster
                        if member.get("来源链接") != merged.get("来源链接")
                    ],
                }
            )
    return merged_rows, reconciliation


def download_gdmch_photo(
    session: requests.Session,
    photo_url: str,
    photo_dir: Path,
    filename_stem: str,
    detail_id: str,
    used_filenames: set[str],
) -> dict[str, Any]:
    official_url = gdmch_photo_url(photo_url, photo_url)
    if not official_url:
        raise RuntimeError(f"照片 URL 不属于医院官方 wx.e3861.com 医生图片路径：{photo_url}")
    response = session.get(official_url, timeout=30)
    if response.status_code != 200:
        raise RuntimeError(f"照片下载 HTTP {response.status_code}：{official_url}")
    extension = gdgh_photo_extension(response.content, response.headers.get("Content-Type"))
    if not extension:
        raise RuntimeError(f"照片响应格式不受支持：{official_url}")
    filename = f"{filename_stem}.{extension}"
    if filename.casefold() in used_filenames:
        filename = f"{filename_stem}-{gdgh_photo_part(detail_id)}.{extension}"
    used_filenames.add(filename.casefold())
    photo_dir.mkdir(parents=True, exist_ok=True)
    path = photo_dir / filename
    if path.exists() and path.read_bytes() != response.content:
        raise RuntimeError(f"照片目标已存在且内容不同，拒绝覆盖：{path}")
    path.write_bytes(response.content)
    width, height = gdmch_photo_dimensions(response.content, extension)
    if width <= 0 or height <= 0:
        raise RuntimeError(f"照片尺寸无法从原图解析：{official_url}")
    return {
        "photo_url": official_url,
        "filename": filename,
        "bytes": len(response.content),
        "sha256": hashlib.sha256(response.content).hexdigest(),
        "width": width,
        "height": height,
        "disk_path": str(path),
    }


def parse_gdmch_campuses(html: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    campuses: list[dict[str, str]] = []
    for anchor in soup.select("a[href]"):
        name = gdgh_clean_text(anchor.get_text(" ", strip=True))
        if name not in {"番禺院区", "越秀院区", "天河院区", "清远院区"}:
            continue
        url = urljoin(base_url, str(anchor.get("href") or ""))
        if comparable_host(url) != "e3861.com":
            continue
        if name not in {item["name"] for item in campuses}:
            campuses.append({"name": name, "url": url})
    return campuses


def collect_gdmch(
    target: HospitalTarget,
    today: str,
    max_doctors: int | None = None,
    photo_root: Path | None = None,
    full_mode: bool = False,
) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页普通公开 GET 读取失败：{entry_error}")
    page_count = discover_gdmch_page_count(entry_html, target.entry_url)
    relations: list[dict[str, str]] = []
    categories: list[dict[str, str]] = []
    category_errors: list[dict[str, str]] = []
    for page_number in range(1, page_count + 1):
        page_url = gdmch_list_page_url(target.entry_url, page_number)
        if page_number == 1:
            status, html, error = entry_status, entry_html, ""
        else:
            status, html, error = fetch(session, page_url)
        if status != 200:
            category_errors.append(
                {"page": str(page_number), "url": page_url, "error": error}
            )
            continue
        page_rows = parse_gdmch_list_page(html, page_url, page_number)
        relations.extend(page_rows)
        categories.append(
            {
                "category_id": str(page_number),
                "category_name": f"专家目录第{page_number}页",
                "url": page_url,
                "row_count": str(len(page_rows)),
            }
        )
        if page_number % 20 == 0:
            print(f"GDMCH pages: {page_number}/{page_count}")
        time.sleep(0.08)
    if category_errors:
        raise RuntimeError(f"官网 111 页普查不完整：{len(category_errors)} 页读取失败")

    by_id: dict[str, dict[str, str]] = {}
    for relation in relations:
        by_id.setdefault(relation["id"], relation)
    all_doctors = list(by_id.values())
    excluded_candidates: list[dict[str, str]] = []
    eligible_doctors: list[dict[str, str]] = []
    for item in all_doctors:
        if gdmch_non_doctor_card(item["name"], item["list_title"]):
            excluded_candidates.append(
                {
                    "detail_id": item["id"],
                    "name": item["name"],
                    "list_title": item["list_title"],
                    "entry_url": item["list_page_url"],
                    "source_link": item["source_link"],
                    "reason": "官网目录卡片为号源/系统账号/非姓名或护理身份，排除医生画像范围",
                }
            )
        else:
            eligible_doctors.append(item)

    photo_eligible_doctors = [item for item in eligible_doctors if item["photo_url"]]
    selected_doctors = (
        eligible_doctors[:]
        if full_mode and max_doctors is None
        else select_gdmch_trial_doctors(photo_eligible_doctors, max_doctors)
    )
    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, Any]] = []
    schedule_exclusion_count = 0
    forbidden_segment_count = 0
    patient_case_exclusion_count = 0
    for item in selected_doctors:
        status, html, error = fetch(session, item["source_link"])
        if status == 200:
            detail = parse_gdmch_detail(html, item)
        else:
            detail_errors.append({"source_link": item["source_link"], "error": error})
            detail = {
                "name": item["name"],
                "title": item["list_title"],
                "departments": [],
                "campuses": [],
                "specialty": "",
                "profile_text": "",
                "photo_url": item["photo_url"],
                "schedule_exclusion_count": 0,
                "forbidden_segment_count": 0,
                "patient_case_exclusion_count": 0,
            }
        name = gdgh_clean_text(str(detail.get("name") or item["name"]))
        title_identity = gdgh_clean_text(str(detail.get("title") or item["list_title"]))
        departments = list(detail.get("departments") or [])
        campuses = list(detail.get("campuses") or [])
        department = "、".join(departments)
        department_with_campus = department
        if campuses:
            department_with_campus += f"（{'、'.join(campuses)}）"
        specialty = gdgh_clean_text(str(detail.get("specialty") or ""))
        profile_text = gdgh_clean_text(str(detail.get("profile_text") or ""))
        schedule_exclusion_count += int(detail.get("schedule_exclusion_count") or 0)
        forbidden_segment_count += int(detail.get("forbidden_segment_count") or 0)
        patient_case_exclusion_count += int(detail.get("patient_case_exclusion_count") or 0)
        combined_text = "\n".join(
            [target.hospital, department, title_identity, specialty, profile_text]
        )
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if name != item["name"]:
            warnings.append("列表与详情姓名不一致")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if warnings:
            groups_found, tags = [], []
        priority = "普通"
        if not warnings and (
            any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
        ):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department_with_campus,
                "科室_列表卡片": "",
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": profile_text,
                "来源类型": "医院官网",
                "来源链接": item["source_link"],
                "照片链接": "",
                "照片文件": "",
                "采集入口": target.entry_url,
                "采集方式": "官网 111 页服务端目录+严格数字 ID 详情 DOM+官方子域本人职业照",
                "采集日期": today,
                "详情页状态": "200" if status == 200 else "失败",
                "已建画像": "是" if canonical_url(item["source_link"]) in existing_links else "否",
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
                "_gdmch_detail_id": item["id"],
                "_gdmch_photo_url": gdgh_clean_text(str(detail.get("photo_url") or "")),
                "_gdmch_first_department": departments[0] if departments else "",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": item["id"],
                "name": name,
                "departments": departments,
                "campuses": campuses,
                "source_link": item["source_link"],
                "photo_url": gdgh_clean_text(str(detail.get("photo_url") or "")),
                "resolution": "详情已读取" if status == 200 else "详情读取失败",
            }
        )
        time.sleep(0.1)

    rows, identity_reconciliation = merge_gdmch_identity_rows(rows)
    photo_dir = photo_root or (VAULT / "01_试点医院" / target.hospital / "照片")
    used_filenames: set[str] = set()
    photo_samples: list[dict[str, Any]] = []
    photo_errors: list[dict[str, str]] = []
    photo_no_source_count = 0
    for index, row in enumerate(rows, start=1):
        row["序号"] = index
        detail_id = clean_text(str(row.pop("_gdmch_detail_id", "")))
        photo_url = clean_text(str(row.pop("_gdmch_photo_url", "")))
        first_department = clean_text(str(row.pop("_gdmch_first_department", "")))
        if not photo_url:
            photo_no_source_count += 1
            continue
        stem = "-".join(
            [
                gdgh_photo_part(row.get("姓名")),
                gdgh_photo_part(first_department),
                gdgh_photo_part(gdgh_primary_title(row.get("职称身份原文"))),
                gdgh_photo_part(target.hospital),
            ]
        )
        try:
            photo = download_gdmch_photo(
                session, photo_url, photo_dir, stem, detail_id, used_filenames
            )
        except Exception as exc:  # noqa: BLE001 - retain per-photo failure evidence
            photo_errors.append(
                {
                    "name": row["姓名"],
                    "detail_id": detail_id,
                    "source_link": row["来源链接"],
                    "photo_url": photo_url,
                    "error": str(exc),
                }
            )
            row["异常提示"] = "；".join(
                dict.fromkeys(
                    [
                        *clean_text(str(row.get("异常提示") or "")).split("；"),
                        "照片获取失败",
                    ]
                )
            ).strip("；")
            row["重点优先级"] = "普通"
            row["重点关注范围"] = ""
            row["重点疾病标签"] = ""
            continue
        relative_path = (
            Path("01_试点医院") / target.hospital / "照片" / photo["filename"]
        ).as_posix()
        row["照片链接"] = photo["photo_url"]
        row["照片文件"] = relative_path
        photo_samples.append(
            {
                "name": row["姓名"],
                "department": first_department or "未标注",
                "title": gdgh_primary_title(row["职称身份原文"]) or "未标注",
                "detail_id": detail_id,
                "source_link": row["来源链接"],
                "photo_url": photo["photo_url"],
                "photo_file": relative_path,
                "filename": photo["filename"],
                "bytes": photo["bytes"],
                "sha256": photo["sha256"],
                "width": photo["width"],
                "height": photo["height"],
                "disk_path": photo["disk_path"],
            }
        )

    intro_url = "https://www.e3861.com/guanyuwomen/yiyuanjieshao"
    intro_status, intro_html, intro_error = fetch(session, intro_url)
    if intro_status != 200:
        raise RuntimeError(f"医院/院区归属页读取失败：{intro_error}")
    scope_text = gdgh_clean_text(
        BeautifulSoup(f"{entry_html}\n{intro_html}", "html.parser").get_text(" ", strip=True)
    )
    independent_markers = [
        marker
        for marker in ("独立法人", "法人单位", "统一社会信用代码")
        if marker in scope_text
    ]
    if independent_markers:
        raise RuntimeError(
            "[FATAL - HUMAN_INTERVENTION_REQUIRED] 四院区范围出现独立法人证据："
            + "、".join(independent_markers)
        )
    campuses = parse_gdmch_campuses(entry_html, target.entry_url)
    affiliate_evidence = [
        {
            "name": item["name"],
            "url": item["url"],
            "relation": "官网同一专家目录与统一页脚列示的广东省妇幼保健院院区；未发现独立法人标识",
        }
        for item in campuses
    ]

    names_to_ids: dict[str, list[str]] = {}
    for item in eligible_doctors:
        names_to_ids.setdefault(item["name"], []).append(item["id"])
    same_name_groups = {
        name: sorted(ids, key=int)
        for name, ids in names_to_ids.items()
        if len(ids) > 1
    }
    sample_departments = gdmch_covered_department_names(rows)
    campus_counts = Counter(
        campus
        for item in detail_reconciliation
        for campus in item.get("campuses", [])
    )
    category_counter = Counter(
        department
        for item in detail_reconciliation
        for department in item.get("departments", [])
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in row["重点关注范围"].split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in row["异常提示"].split("；") if warning
    )
    average_photo_bytes = (
        round(sum(item["bytes"] for item in photo_samples) / len(photo_samples))
        if photo_samples
        else 0
    )
    large_photos = [
        item
        for item in photo_samples
        if int(item["bytes"]) > 200 * 1024 or int(item["width"]) > 800
    ]
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #43（与官网入口台账序号 21 一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": page_count,
            "pagination_count": page_count,
            "pagination_method": "111 个服务端公开 GET 分页；仅空白 searchDoctor/searchDepartment，不构造检索词",
            "department_structure": "目录只提供自由文本检索框，无服务端科室分类树；样本科室仅从详情出诊安排括号标签保守提取，日期时段不入库",
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(all_doctors),
            "census_unique_detail_count": len(all_doctors),
            "eligible_candidate_count": len(eligible_doctors),
            "excluded_non_doctor_count": len(excluded_candidates),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "unique_doctor_count": len(rows),
            "gdmch_final_identity_count": len(identity_reconciliation),
            "gdmch_same_identity_merge_group_count": sum(
                int(item.get("relation_count") or 0) > 1
                for item in identity_reconciliation
            ),
            "gdmch_distinct_same_name_group_count": len(
                {
                    clean_text(str(item.get("name") or ""))
                    for item in identity_reconciliation
                    if item.get("resolution") == "同名待甄别"
                }
            ),
            "sample_entry_coverage_count": len(sample_departments),
            "sample_entry_categories": sample_departments,
            "sample_campus_coverage_count": len(campus_counts),
            "campus_relation_counts": dict(campus_counts),
            "category_error_count": len(category_errors),
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(all_doctors),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
            "entry_candidate_counts": {target.entry_url: len(relations)},
            "standard_public_session": "requests 常规公开 GET；无登录、Cookie 注入、验证码/挑战求解或非公开接口",
            "schedule_exclusion_count": schedule_exclusion_count,
            "schedule_field_ingested_count": 0,
            "forbidden_segment_exclusion_count": forbidden_segment_count,
            "patient_case_exclusion_count": patient_case_exclusion_count,
            "private_use_character_count": sum(
                len(re.findall(r"[\ue000-\uf8ff]", str(row.get(field) or "")))
                for row in rows
                for field in BASE_HEADERS
            ),
            "photo_sample_count": len(photo_samples),
            "photo_error_count": len(photo_errors),
            "photo_expected_count": len(rows) - photo_no_source_count,
            "photo_downloaded_count": len(photo_samples),
            "photo_failed_count": len(photo_errors),
            "photo_no_source_count": photo_no_source_count,
            "photo_average_bytes": average_photo_bytes,
            "photo_estimated_full_count": len(photo_eligible_doctors),
            "photo_estimated_full_bytes": average_photo_bytes * len(photo_eligible_doctors),
            "photo_census_available_count": len(photo_eligible_doctors),
            "photo_census_placeholder_count": len(eligible_doctors) - len(photo_eligible_doctors),
            "photo_default_placeholder_count": sum(
                not bool(item["photo_url"]) for item in all_doctors
            ),
            "large_photo_count": len(large_photos),
            "large_photo_threshold": "单张 >200KB 或宽 >800px",
            "photo_policy_status": (
                "WAITING_OWNER_LARGE_IMAGE_POLICY"
                if large_photos
                else "OWNER_APPROVED_ORIGINAL_NO_WIDTH_LIMIT"
            ),
            "affiliate_count": len(affiliate_evidence),
            "independent_entity_count": 0,
        },
        "categories": categories,
        "entry_reconnaissance": [
            {
                "entry_url": target.entry_url,
                "page_nature": "官网服务端专家目录分页；卡片含姓名、职称和官方子域职业照",
                "list_page_count": page_count,
                "raw_detail_relation_count": len(relations),
                "unique_detail_count": len(all_doctors),
                "out_of_scope_detail_count": len(excluded_candidates),
                "affiliation": target.hospital,
                "independent_entity_check": "四院区共用官网专家目录和统一页脚；未发现独立法人标识",
            }
        ],
        "affiliate_reconnaissance": affiliate_evidence,
        "excluded_candidates": excluded_candidates,
        "gdmch_detail_reconciliation": detail_reconciliation,
        "gdmch_identity_reconciliation": identity_reconciliation,
        "photo_samples": photo_samples,
        "photo_errors": photo_errors,
        "category_errors": category_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def fahsysu_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "fahsysu.org.cn" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(r"/node/(\d+)/?", parsed.path, flags=re.IGNORECASE)
    return match.group(1) if match else ""


def parse_fahsysu_directory(html: str, entry_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    relations: list[dict[str, str]] = []
    for group in soup.select(".action-item"):
        group_node = group.select_one(":scope > .action-item-top")
        group_name = clean_text(group_node.get_text(" ", strip=True) if group_node else "")
        for content in group.select(":scope > .action-item-content"):
            specialty_node = content.select_one(":scope > .action-item-left")
            specialty = clean_text(
                specialty_node.get_text(" ", strip=True) if specialty_node else ""
            )
            for title_group in content.select(
                ":scope > .action-item-right > .action-item-list"
            ):
                title_node = title_group.select_one(":scope > .action-item-list-title")
                title_hint = clean_text(
                    title_node.get_text(" ", strip=True) if title_node else ""
                )
                for anchor in title_group.select(
                    ":scope > .action-item-list-text > .action-item-list-tag a[href]"
                ):
                    source_link = urljoin(entry_url, str(anchor.get("href") or ""))
                    detail_id = fahsysu_detail_id(source_link)
                    name = clean_text(anchor.get_text(" ", strip=True))
                    if not (group_name and specialty and title_hint and detail_id and name):
                        continue
                    relations.append(
                        {
                            "id": detail_id,
                            "name": name,
                            "group": group_name,
                            "department": specialty,
                            "title_hint": title_hint,
                            "source_link": source_link,
                        }
                    )
    return relations


def merge_fahsysu_directory_relations(
    relations: list[dict[str, str]],
) -> list[dict[str, str]]:
    merged: dict[str, dict[str, Any]] = {}
    for relation in relations:
        detail_id = relation["id"]
        if detail_id not in merged:
            merged[detail_id] = {
                **relation,
                "groups": [relation["group"]],
                "departments": [relation["department"]],
                "title_hints": [relation["title_hint"]],
                "relation_count": 1,
            }
            continue
        current = merged[detail_id]
        current["relation_count"] += 1
        for source_field, target_field in (
            ("group", "groups"),
            ("department", "departments"),
            ("title_hint", "title_hints"),
        ):
            value = relation[source_field]
            if value not in current[target_field]:
                current[target_field].append(value)
    return [
        {
            **item,
            "group": "、".join(item["groups"]),
            "department": "、".join(item["departments"]),
            "title_hint": "、".join(item["title_hints"]),
            "relation_count": str(item["relation_count"]),
        }
        for item in merged.values()
    ]


def parse_fahsysu_detail(html: str, fallback: dict[str, str]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    article = soup.select_one("article.node--type-doctor")
    schedule_nodes = soup.select('[class*="calendar-"]')
    schedule_count = len(schedule_nodes)
    for node in schedule_nodes:
        node.decompose()
    container = article.select_one(".other-2") if article else None
    if not container:
        return {
            **fallback,
            "title": "",
            "detail_department": "",
            "specialty": "",
            "profile_text": "",
            "schedule_exclusion_count": schedule_count,
            "forbidden_segment_count": 0,
        }

    name_node = container.select_one(".other-left-title")
    labeled: dict[str, str] = {}
    for node in container.select(".other-left-text"):
        label_node = node.select_one("span")
        label = clean_text(label_node.get_text(" ", strip=True) if label_node else "")
        clone = BeautifulSoup(str(node), "html.parser")
        for child in clone.select("span"):
            child.decompose()
        value = strip_gzsys_schedule_text(
            strip_gzsys_forbidden_text(clone.get_text(" ", strip=True))
        )
        if label.startswith("职称"):
            labeled["title"] = value
        elif label.startswith("科室"):
            labeled["department"] = normalize_gzsys_department(value)
        elif label.startswith("简介"):
            labeled["profile"] = value

    showcase_segments: list[str] = []
    for node in (article or soup).select(".showcase-text-content"):
        value = strip_gzsys_schedule_text(
            strip_gzsys_forbidden_text(node.get_text(" ", strip=True))
        )
        if value and value not in showcase_segments:
            showcase_segments.append(value)
    specialty = ""
    for segment in showcase_segments:
        match = re.search(
            r"(?:医疗特长|专业擅长|擅长|专长|特长)\s*[:：]\s*(.*?)(?=\s*【|$)",
            segment,
            flags=re.DOTALL,
        )
        if match:
            specialty = clean_text(match.group(1))
            break

    segments = list(dict.fromkeys([labeled.get("profile", ""), *showcase_segments]))
    forbidden_markers = (
        "好医生榜",
        "医生排行榜",
        "好医生",
        "名医录",
        "排行榜",
        "排名",
        "患者评价",
        "患者留言",
        "问诊记录",
        "问诊内容",
    )
    sentences = [
        clean_text(sentence)
        for segment in segments
        for sentence in re.split(r"(?<=[。！？；;])\s*", segment)
        if clean_text(sentence)
    ]
    kept = [
        sentence
        for sentence in sentences
        if not any(term in sentence for term in forbidden_markers)
        and not contains_gzbrain_patient_case_text(sentence)
    ]
    specialty_sentences = [
        clean_text(sentence)
        for sentence in re.split(r"(?<=[。！？；;])\s*", specialty)
        if clean_text(sentence)
    ]
    clean_specialty = " ".join(
        sentence
        for sentence in specialty_sentences
        if not any(term in sentence for term in forbidden_markers)
        and not contains_gzbrain_patient_case_text(sentence)
    )
    return {
        "name": first_nonempty(
            strip_gzsys_forbidden_text(
                name_node.get_text(" ", strip=True) if name_node else ""
            ),
            fallback.get("name"),
        ),
        "title": strip_gzsys_forbidden_text(labeled.get("title", "")),
        "detail_department": labeled.get("department", ""),
        "specialty": clean_text(
            re.sub(
                r"^(?:(?:医疗特长|专业擅长|擅长|专长|特长)\s*[:：]?\s*)+",
                "",
                clean_specialty,
            )
        ),
        "profile_text": clip(strip_profile_navigation_text(" ".join(kept)), 1800),
        "schedule_exclusion_count": schedule_count,
        "forbidden_segment_count": len(sentences) - len(kept),
    }


def collect_fahsysu(
    target: HospitalTarget, today: str, max_doctors: int | None = None
) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页普通公开 GET 读取失败：{entry_error}")
    relations = parse_fahsysu_directory(entry_html, target.entry_url)
    doctors = merge_fahsysu_directory_relations(relations)
    if not relations or not doctors:
        raise RuntimeError("官网目录未发现严格 action-item 医生关系。")

    names_to_ids: dict[str, list[str]] = {}
    for doctor in doctors:
        names_to_ids.setdefault(doctor["name"], []).append(doctor["id"])
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in names_to_ids.items()
        if len(set(ids)) > 1
    }
    selected = select_gzsys_trial_doctors(doctors, max_doctors)
    if max_doctors:
        selected_ids = {doctor["id"] for doctor in selected}
        selected.extend(doctor for doctor in doctors if doctor["id"] not in selected_ids)

    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    excluded_candidates: list[dict[str, str]] = []
    detail_errors: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, Any]] = []
    identity_reconciliation: list[dict[str, Any]] = []
    schedule_exclusion_count = 0
    forbidden_segment_count = 0
    detail_campus_marker_counts: Counter[str] = Counter()
    detail_campus_evidence: list[dict[str, Any]] = []
    for doctor in selected:
        if max_doctors and len(rows) >= max_doctors:
            break
        detail_status, detail_html, detail_error = fetch(session, doctor["source_link"])
        if detail_status == 200:
            marker_counts = {
                marker: detail_html.count(marker)
                for marker in FAHSYSU_CAMPUS_MARKERS
                if detail_html.count(marker)
            }
            if marker_counts:
                detail_campus_marker_counts.update(marker_counts)
                detail_campus_evidence.append(
                    {
                        "detail_id": doctor["id"],
                        "name": doctor["name"],
                        "source_link": doctor["source_link"],
                        "marker_counts": marker_counts,
                        "evidence_scope": "详情正文履历词，仅作院区存在性证据，不写入结构化科室/院区字段",
                    }
                )
            detail = parse_fahsysu_detail(detail_html, doctor)
        else:
            detail_errors.append({"source_link": doctor["source_link"], "error": detail_error})
            detail = {
                **doctor,
                "title": "",
                "detail_department": "",
                "specialty": "",
                "profile_text": "",
                "schedule_exclusion_count": 0,
                "forbidden_segment_count": 0,
            }
        name = strip_gzsys_forbidden_text(str(detail.get("name") or doctor["name"]))
        title_identity = strip_gzsys_forbidden_text(str(detail.get("title") or ""))
        detail_department = normalize_gzsys_department(
            str(detail.get("detail_department") or "")
        )
        departments = normalize_gzsys_department(doctor["department"])
        specialty = strip_gzsys_forbidden_text(str(detail.get("specialty") or ""))
        profile_text = strip_gzsys_forbidden_text(str(detail.get("profile_text") or ""))
        schedule_exclusion_count += int(detail.get("schedule_exclusion_count") or 0)
        forbidden_segment_count += int(detail.get("forbidden_segment_count") or 0)
        if gyfyyy_nursing_only_identity(title_identity):
            reason = "官网详情仅标注护理身份，排除医生画像采集范围"
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": f"{name} {title_identity}",
                    "source_link": doctor["source_link"],
                    "reason": reason,
                }
            )
            detail_reconciliation.append(
                {
                    "detail_id": doctor["id"],
                    "name": name,
                    "resolution": "护理排除",
                    "departments": doctor["departments"],
                    "relation_count": int(doctor["relation_count"]),
                    "source_link": doctor["source_link"],
                    "reason": reason,
                }
            )
            continue

        combined_text = "\n".join(
            [target.hospital, departments, title_identity, specialty, profile_text]
        )
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if doctor["name"] and name != doctor["name"]:
            warnings.append("列表与详情姓名不一致")
        if not departments:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if name in same_name_groups:
            warnings.append("同名待甄别")
        if warnings:
            groups_found, tags = [], []
        priority = "普通"
        if not warnings and (
            any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
        ):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": departments,
                "科室_列表卡片": detail_department,
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": clip(specialty, 520),
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": clip(profile_text, 1800),
                "来源类型": "医院官网",
                "来源链接": doctor["source_link"],
                "采集入口": target.entry_url,
                "采集方式": "官网单页 action-item 科室树+数字 node ID+严格医生详情 DOM",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": (
                    "是" if canonical_url(doctor["source_link"]) in existing_links else "否"
                ),
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": doctor["id"],
                "name": name,
                "resolution": "同名待甄别" if name in same_name_groups else "正式行",
                "departments": doctor["departments"],
                "groups": doctor["groups"],
                "title_hints": doctor["title_hints"],
                "relation_count": int(doctor["relation_count"]),
                "source_link": doctor["source_link"],
                "reason": "同名不同数字 ID 分行保留" if name in same_name_groups else "",
            }
        )
        identity_reconciliation.append(
            {
                "name": name,
                "detail_ids": [doctor["id"]],
                "resolution": "同名待甄别" if name in same_name_groups else "唯一身份",
                "relation_count": int(doctor["relation_count"]),
                "departments": doctor["departments"],
                "groups": doctor["groups"],
                "primary_source_link": doctor["source_link"],
                "merged_source_links": [],
                "reason": "同名不同数字 ID 分行保留" if name in same_name_groups else "",
            }
        )
        time.sleep(0.12)

    covered_departments = covered_department_names(rows)
    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(str(row["科室_分类页"])).split("、")
        if department
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in row["重点关注范围"].split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in row["异常提示"].split("；") if warning
    )
    campus_marker_counts = {
        marker: entry_html.count(marker) for marker in FAHSYSU_CAMPUS_MARKERS
    }
    title_hint_counts = Counter(relation["title_hint"] for relation in relations)
    specialty_names = list(dict.fromkeys(relation["department"] for relation in relations))
    relationship_group_names = list(dict.fromkeys(relation["group"] for relation in relations))
    directory_soup = BeautifulSoup(entry_html, "html.parser")
    top_level_group_names = [
        clean_text(top.get_text(" ", strip=True))
        for group in directory_soup.select(".action-item")
        if (top := group.select_one(":scope > .action-item-top"))
    ]
    empty_group_names = [
        clean_text(top.get_text(" ", strip=True))
        for group in directory_soup.select(".action-item")
        if (top := group.select_one(":scope > .action-item-top"))
        and not group.select(":scope > .action-item-content")
    ]
    formal_ids_by_name: dict[str, list[str]] = {}
    for item in identity_reconciliation:
        formal_ids_by_name.setdefault(str(item["name"]), []).extend(
            str(value) for value in item["detail_ids"]
        )
    formal_same_name_groups = {
        name: ids for name, ids in formal_ids_by_name.items() if len(ids) > 1
    }
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #37（与官网入口台账一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(specialty_names),
            "census_group_count": len(top_level_group_names),
            "census_groups": top_level_group_names,
            "census_relationship_group_count": len(relationship_group_names),
            "census_relationship_groups": relationship_group_names,
            "census_empty_group_count": len(empty_group_names),
            "census_empty_groups": empty_group_names,
            "census_department_count": len(specialty_names),
            "census_departments": specialty_names,
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(doctors),
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(doctors),
            "census_named_detail_count": sum(bool(item["name"]) for item in doctors),
            "census_blank_name_detail_count": sum(not item["name"] for item in doctors),
            "census_unique_nonblank_name_count": len(names_to_ids),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "census_nonempty_department_count": sum(bool(item["department"]) for item in doctors),
            "census_empty_department_count": sum(not item["department"] for item in doctors),
            "eligible_candidate_count": len(rows),
            "sample_entry_coverage_count": len(covered_departments),
            "sample_entry_categories": covered_departments,
            "pagination_count": 1,
            "pagination_method": "官网服务端完整输出单页长列表；未提交搜索词、未构造筛选组合、未探测接口",
            "title_hint_counts": dict(title_hint_counts),
            "campus_marker_counts": campus_marker_counts,
            "detail_campus_marker_counts": dict(detail_campus_marker_counts),
            "campus_evidence_detail_count": len(detail_campus_evidence),
            "campus_scope_status": (
                "目录页未发现院区词；试采详情仅在履历正文发现院区词，官网未提供统一结构化院区字段，"
                "不能据此为全目录医生推断本部、东院、南沙或黄埔归属"
            ),
            "huangpu_scope_status": (
                "未使用台账序号 8 黄埔院区专属目录；目录与试采详情均无黄埔标记，"
                "仍无法证明或排除未抽样医生中是否混入黄埔归属"
            ),
            "standard_public_session": "requests 常规公开 GET；无搜索提交、挑战求解、指纹模拟或绕过",
            "session_cookie_names": sorted(session.cookies.keys()),
            "schedule_exclusion_count": schedule_exclusion_count,
            "schedule_field_ingested_count": 0,
            "forbidden_segment_exclusion_count": forbidden_segment_count,
            "private_use_character_count": sum(
                len(re.findall(r"[\ue000-\uf8ff]", str(row.get(field) or "")))
                for row in rows
                for field in BASE_HEADERS
            ),
            "category_error_count": 0,
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(doctors),
            "excluded_non_doctor_count": len(excluded_candidates),
            "fahsysu_final_identity_count": len(identity_reconciliation),
            "fahsysu_same_identity_merge_group_count": 0,
            "fahsysu_distinct_same_name_group_count": len(formal_same_name_groups),
            "fahsysu_distinct_same_name_row_count": sum(
                len(ids) for ids in formal_same_name_groups.values()
            ),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": [
            {
                "category_id": str(index),
                "category_name": specialty,
                "url": target.entry_url,
                "doctor_relation_count": sum(
                    relation["department"] == specialty for relation in relations
                ),
            }
            for index, specialty in enumerate(specialty_names, start=1)
        ],
        "entry_reconnaissance": [
            {
                "category_name": "官网专家介绍科室树",
                "entry_url": target.entry_url,
                "page_nature": "医院官网 Drupal 公开专家单页长列表",
                "list_page_count": 1,
                "raw_detail_relation_count": len(relations),
                "unique_detail_count": len(doctors),
                "out_of_scope_detail_count": 0,
                "affiliation": target.hospital,
                "independent_entity_check": "仅 action-item 结构内 /node/<数字ID> 授权；同 ID 跨科室合并，同名不同 ID 分行",
            }
        ],
        "excluded_candidates": excluded_candidates,
        "fahsysu_detail_reconciliation": detail_reconciliation,
        "fahsysu_identity_reconciliation": identity_reconciliation,
        "fahsysu_campus_evidence": detail_campus_evidence,
        "cross_entry_duplicates": [],
        "category_errors": [],
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def collect_gzsys(
    target: HospitalTarget, today: str, max_doctors: int | None = None
) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页普通公开 GET 读取失败：{entry_error}")
    page_urls = discover_gzsys_default_pages(entry_html, target.entry_url)
    if not page_urls:
        raise RuntimeError("官网默认 All 医生目录未声明完整分页范围。")
    filter_dictionary = parse_gzsys_filter_dictionary(entry_html)
    categories: list[dict[str, Any]] = []
    page_errors: list[dict[str, str]] = []
    relations: list[dict[str, str]] = []
    for page_number, page_url in enumerate(page_urls):
        status, html, error = (
            (entry_status, entry_html, "")
            if page_number == 0
            else fetch(session, page_url)
        )
        if status != 200:
            page_errors.append({"page": str(page_number), "url": page_url, "error": error})
            continue
        page_rows = parse_gzsys_list_page(html, page_url)
        if not page_rows:
            page_errors.append(
                {"page": str(page_number), "url": page_url, "error": "严格 .card-4-0 医生卡片为 0"}
            )
            continue
        relations.extend(page_rows)
        categories.append(
            {
                "category_id": str(page_number),
                "category_name": f"默认医生目录第 {page_number + 1} 页",
                "url": page_url,
                "doctor_relation_count": len(page_rows),
            }
        )
    if page_errors:
        raise RuntimeError(
            "官网默认医生目录读取不完整："
            + "；".join(f"page={item['page']} {item['error']}" for item in page_errors)
        )
    doctors = merge_gzsys_card_relations(relations)
    if not doctors:
        raise RuntimeError("官网默认目录未发现严格医生卡片关系。")
    names_to_ids: dict[str, list[str]] = {}
    for doctor in doctors:
        if doctor["name"]:
            names_to_ids.setdefault(doctor["name"], []).append(doctor["id"])
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in names_to_ids.items()
        if len(set(ids)) > 1
    }
    raw_source_path_counts = Counter(
        "node" if urlparse(doctor["source_link"]).path.startswith("/node/") else "doctor"
        for doctor in doctors
    )
    nursing_doctors = [doctor for doctor in doctors if gyfyyy_nursing_only_identity(doctor["title"])]
    eligible_doctors = [doctor for doctor in doctors if doctor not in nursing_doctors]
    selected = select_gzsys_trial_doctors(eligible_doctors, max_doctors)
    if max_doctors:
        selected_ids = {doctor["id"] for doctor in selected}
        selected.extend(doctor for doctor in eligible_doctors if doctor["id"] not in selected_ids)
    existing_links = collect_existing_profile_links()
    exclusion_reason = "官网医生卡片仅标注护理身份，排除医生画像采集范围"
    excluded_candidates = [
        {
            "entry_url": target.entry_url,
            "list_title": f"{doctor['name']} {doctor['title']}",
            "source_link": doctor["source_link"],
            "reason": exclusion_reason,
        }
        for doctor in nursing_doctors
    ]
    detail_reconciliation: list[dict[str, str]] = [
        {
            "detail_id": doctor["id"],
            "source_link": doctor["source_link"],
            "name": doctor["name"],
            "resolution": "护理排除",
            "reason": exclusion_reason,
        }
        for doctor in nursing_doctors
    ]
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    schedule_exclusion_count = 0
    forbidden_segment_count = 0
    for doctor in selected:
        if max_doctors and len(rows) >= max_doctors:
            break
        detail_status, detail_html, detail_error = fetch(session, doctor["source_link"])
        if detail_status == 200:
            detail = parse_gzsys_detail(detail_html, doctor)
        else:
            detail_errors.append({"source_link": doctor["source_link"], "error": detail_error})
            detail = {
                **doctor,
                "specialty": "",
                "profile_text": "",
                "schedule_exclusion_count": 0,
                "forbidden_segment_count": 0,
            }
        name = strip_gzsys_forbidden_text(str(detail.get("name") or doctor["name"]))
        title_identity = strip_gzsys_forbidden_text(str(detail.get("title") or doctor["title"]))
        department = normalize_gzsys_department(str(detail.get("department") or doctor["department"]))
        specialty = strip_gzsys_forbidden_text(str(detail.get("specialty") or ""))
        profile_text = strip_gzsys_forbidden_text(str(detail.get("profile_text") or ""))
        schedule_exclusion_count += int(detail.get("schedule_exclusion_count") or 0)
        forbidden_segment_count += int(detail.get("forbidden_segment_count") or 0)
        if gyfyyy_nursing_only_identity(title_identity):
            exclusion = {
                "entry_url": target.entry_url,
                "list_title": f"{name} {title_identity}",
                "source_link": doctor["source_link"],
                "reason": "官网详情仅标注护理身份，排除医生画像采集范围",
            }
            excluded_candidates.append(exclusion)
            detail_reconciliation.append(
                {
                    "detail_id": doctor["id"],
                    "source_link": doctor["source_link"],
                    "name": name,
                    "resolution": "护理排除",
                    "reason": exclusion["reason"],
                }
            )
            continue
        combined_text = "\n".join([target.hospital, department, title_identity, specialty, profile_text])
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if doctor["name"] and name != doctor["name"]:
            warnings.append("列表与详情姓名不一致")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if name in same_name_groups:
            warnings.append("同名待甄别")
        if warnings:
            groups_found, tags = [], []
        priority = "普通"
        if not warnings and (any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": doctor["department"],
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": clip(specialty, 520),
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": clip(profile_text, 1800),
                "来源类型": "医院官网",
                "来源链接": doctor["source_link"],
                "采集入口": target.entry_url,
                "采集方式": "官网默认 All 静态分页+.card-4-0 医生卡片+严格详情 DOM 抽取",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(doctor["source_link"]) in existing_links else "否",
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": doctor["id"],
                "source_link": doctor["source_link"],
                "name": name,
                "resolution": "正式行",
                "reason": "",
            }
        )
        time.sleep(0.12)
    covered_departments = covered_department_names(rows)
    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(row["科室_分类页"]).split("、")
        if department
    )
    identity_reconciliation = [
        {
            "name": clean_text(str(row.get("姓名") or "")),
            "identity_index": 1,
            "resolution": "唯一身份",
            "detail_ids": [gzsys_detail_id(str(row.get("来源链接") or ""))],
            "primary_source_link": clean_text(str(row.get("来源链接") or "")),
            "merged_source_links": [],
            "departments": [
                department
                for department in clean_text(str(row.get("科室_分类页") or "")).split("、")
                if department
            ],
            "relation_count": 1,
        }
        for row in rows
    ]
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(group for row in rows for group in row["重点关注范围"].split("、") if group)
    warning_counter = Counter(warning for row in rows for warning in row["异常提示"].split("；") if warning)
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #35（与官网入口台账一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(categories),
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(doctors),
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(doctors),
            "census_named_detail_count": sum(bool(item["name"]) for item in doctors),
            "census_blank_name_detail_count": sum(not item["name"] for item in doctors),
            "census_unique_nonblank_name_count": len(names_to_ids),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "census_department_count": len({item["department"] for item in doctors if item["department"]}),
            "census_group_count": 0,
            "census_nonempty_department_count": sum(bool(item["department"]) for item in doctors),
            "census_empty_department_count": sum(not item["department"] for item in doctors),
            "eligible_candidate_count": len(eligible_doctors),
            "sample_entry_coverage_count": len(covered_departments),
            "sample_entry_categories": covered_departments,
            "pagination_count": len(page_urls),
            "pagination_method": "页面声明的默认 All 查询 page=0..末页；不遍历搜索词或筛选组合",
            "filter_dictionary_counts": {name: len(values) for name, values in filter_dictionary.items()},
            "filter_dictionary": filter_dictionary,
            "source_path_counts": dict(raw_source_path_counts),
            "standard_public_session": "requests 常规重定向与站点自设 Cookie；无挑战求解、指纹模拟或绕过",
            "session_cookie_names": sorted(session.cookies.keys()),
            "schedule_exclusion_count": schedule_exclusion_count,
            "schedule_field_ingested_count": 0,
            "forbidden_segment_exclusion_count": forbidden_segment_count,
            "private_use_character_count": sum(
                len(re.findall(r"[\ue000-\uf8ff]", str(row.get(field) or "")))
                for row in rows
                for field in ("擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录")
            ),
            "category_error_count": len(page_errors),
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(doctors),
            "excluded_non_doctor_count": len(excluded_candidates),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "gzsys_final_identity_count": len(rows),
            "gzsys_same_identity_merge_group_count": 0,
            "gzsys_distinct_same_name_group_count": 0,
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": categories,
        "entry_reconnaissance": [
            {
                "category_name": "官网名医名师默认目录",
                "entry_url": target.entry_url,
                "page_nature": "医院官网 Drupal 公开医生目录",
                "list_page_count": len(page_urls),
                "raw_detail_relation_count": len(relations),
                "unique_detail_count": len(doctors),
                "out_of_scope_detail_count": len(excluded_candidates),
                "affiliation": target.hospital,
                "independent_entity_check": "仅 .card-4-0 卡片授权；/node/<ID> 与 /doctor/<ID> 按数字 ID 去重",
            }
        ],
        "excluded_candidates": excluded_candidates,
        "gzsys_detail_reconciliation": detail_reconciliation,
        "gzsys_identity_reconciliation": identity_reconciliation,
        "cross_entry_duplicates": [],
        "category_errors": page_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def gzszyy_detail_id(url: str | None) -> str:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gzszyy.com" or parsed.query or parsed.fragment:
        return ""
    match = re.fullmatch(
        r"/expert/(20\d{2})/([A-Za-z0-9]+)\.html",
        parsed.path,
        flags=re.IGNORECASE,
    )
    return match.group(2) if match else ""


def gzszyy_department_filter(url: str | None) -> tuple[str, str]:
    parsed = urlparse(clean_text(url))
    if comparable_host(parsed.geturl()) != "gzszyy.com" or parsed.query or parsed.fragment:
        return "", ""
    match = re.fullmatch(r"/expert/1/dp/(\d+)/", parsed.path, flags=re.IGNORECASE)
    return (match.group(1), parsed.geturl()) if match else ("", "")


def discover_gzszyy_care_sites(html: str, homepage_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    sites: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select("a[href]"):
        source_url = urljoin(homepage_url, str(anchor.get("href") or ""))
        parsed = urlparse(source_url)
        if (
            comparable_host(source_url) != "gzszyy.com"
            or parsed.query
            or parsed.fragment
            or parsed.path not in GZSZYY_CARE_SITE_PATHS
        ):
            continue
        site_name = GZSZYY_CARE_SITE_PATHS[parsed.path]
        if site_name in seen:
            continue
        seen.add(site_name)
        sites.append({"name": site_name, "source_url": source_url})
    return sites


def discover_gzszyy_department_filters(html: str, entry_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    departments: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select('a[href*="/expert/1/dp/"]'):
        department_url = urljoin(entry_url, str(anchor.get("href") or ""))
        department_id, _ = gzszyy_department_filter(department_url)
        department_name = clean_text(anchor.get_text(" ", strip=True))
        if not department_id or not department_name or department_id in seen:
            continue
        seen.add(department_id)
        departments.append(
            {
                "department_id": department_id,
                "department": department_name,
                "entry_url": department_url,
            }
        )
    return departments


def discover_gzszyy_department_pages(
    html: str, department: dict[str, str]
) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    page_count = 1
    for node in soup.select(".pager [data-all], .pager [attr-pages]"):
        value = clean_text(str(node.get("data-all") or node.get("attr-pages") or ""))
        if value.isdigit():
            page_count = max(page_count, int(value))
    department_id = department["department_id"]
    return [
        f"https://www.gzszyy.com/expert/{page}/dp/{department_id}/"
        for page in range(1, page_count + 1)
    ]


def discover_gzszyy_unfiltered_pages(html: str, entry_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    page_counts = {
        int(value)
        for node in soup.select(".pager [data-all], .pager [attr-pages]")
        if (value := clean_text(str(node.get("data-all") or node.get("attr-pages") or ""))).isdigit()
    }
    if len(page_counts) != 1:
        return []
    page_count = page_counts.pop()
    return [
        entry_url if page == 1 else urljoin(entry_url, f"{page}/")
        for page in range(1, page_count + 1)
    ]


def parse_gzszyy_department_page(
    html: str, page_url: str, department: dict[str, str]
) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for card in soup.select("ul.doctor-list > li"):
        anchor = card.select_one("h2 a[href]")
        source_link = urljoin(page_url, str(anchor.get("href") or "")) if anchor else ""
        detail_id = gzszyy_detail_id(source_link)
        if not detail_id or detail_id in seen:
            continue
        seen.add(detail_id)
        name = clean_text(anchor.get_text(" ", strip=True) if anchor else "")
        title = ""
        for node in card.select(".info > div"):
            text = clean_text(node.get_text(" ", strip=True))
            if text.startswith("职称："):
                title = clean_text(text.partition("：")[2])
                break
        card_department_node = card.select_one(".depart-info a[title], .depart-info a")
        card_department = clean_text(
            str(card_department_node.get("title") or card_department_node.get_text(" ", strip=True))
            if card_department_node
            else ""
        )
        specialty_node = card.select_one("p")
        specialty = clean_text(
            specialty_node.get_text(" ", strip=True) if specialty_node else ""
        )
        specialty = clean_text(re.sub(r"^(?:(?:专长|擅长)\s*[:：]?\s*)+", "", specialty))
        rows.append(
            {
                "id": detail_id,
                "name": name,
                "title": title,
                "department": department["department"],
                "card_department": card_department,
                "specialty": specialty,
                "source_link": source_link,
                "list_page": page_url,
            }
        )
    return rows


def merge_gzszyy_relations(relations: list[dict[str, str]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for relation in relations:
        grouped.setdefault(relation["id"], []).append(relation)
    doctors: list[dict[str, Any]] = []
    for detail_id, items in grouped.items():
        first = items[0]
        names = list(dict.fromkeys(clean_text(item["name"]) for item in items if clean_text(item["name"])))
        titles = list(dict.fromkeys(clean_text(item["title"]) for item in items if clean_text(item["title"])))
        specialties = list(
            dict.fromkeys(
                clean_text(item["specialty"])
                for item in items
                if clean_text(item["specialty"])
            )
        )
        departments = list(
            dict.fromkeys(
                clean_text(item["department"])
                for item in items
                if clean_text(item["department"])
            )
        )
        card_departments = list(
            dict.fromkeys(
                clean_text(item["card_department"])
                for item in items
                if clean_text(item["card_department"])
            )
        )
        doctors.append(
            {
                "id": detail_id,
                "name": names[0] if names else "",
                "names": names,
                "title": "、".join(titles),
                "titles": titles,
                "specialty": max(specialties, key=len) if specialties else "",
                "departments": departments,
                "card_departments": card_departments,
                "source_link": first["source_link"],
                "relation_count": len(items),
                "list_pages": list(dict.fromkeys(item["list_page"] for item in items)),
            }
        )
    return doctors


def parse_gzszyy_detail(html: str, fallback: dict[str, Any]) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    resume = soup.select_one(".doctor-resume")
    intro = soup.select_one(".doctor-items-intro")
    name_node = resume.select_one("h1") if resume else None
    departments = [
        clean_text(re.sub(r"[\ue000-\uf8ff]", " ", anchor.get_text(" ", strip=True)))
        for anchor in (resume.select('a[href*="/department_"]') if resume else [])
        if clean_text(re.sub(r"[\ue000-\uf8ff]", " ", anchor.get_text(" ", strip=True)))
    ]
    campuses: list[str] = []
    for node in soup.select(".doctor-code .qr-img span[title]"):
        raw_label = clean_text(str(node.get("title") or node.get_text(" ", strip=True)))
        labels = [label for label in GZSZYY_CAMPUS_LABELS if label in raw_label]
        for label in sorted(labels, key=raw_label.find):
            if label not in campuses:
                campuses.append(label)
    resume_text = clean_text(resume.get_text(" ", strip=True) if resume else "")
    title_match = re.search(r"职称\s*[:：]\s*(.*?)(?=级别\s*[:：]|擅长\s*[:：]|$)", resume_text)
    title = clean_text(title_match.group(1)) if title_match else ""
    specialty_node = resume.select_one("p.good-at") if resume else None
    specialty = clean_text(
        specialty_node.get_text(" ", strip=True) if specialty_node else ""
    )
    specialty = clean_text(re.sub(r"^(?:(?:专长|擅长)\s*[:：]?\s*)+", "", specialty))
    profile_text = strip_gyfyyy_schedule_text(
        clean_text(intro.get_text(" ", strip=True) if intro else "")
    )
    profile_text, _ = filter_gzbrain_profile_text(profile_text)
    return {
        "name": first_nonempty(
            clean_text(name_node.get_text(" ", strip=True) if name_node else ""),
            str(fallback.get("name") or ""),
        ),
        "title": first_nonempty(title, str(fallback.get("title") or "")),
        "departments": departments,
        "campuses": campuses,
        "specialty": first_nonempty(specialty, str(fallback.get("specialty") or "")),
        "profile_text": clip(profile_text, 1800),
    }


def select_gzszyy_trial_doctors(
    doctors: list[dict[str, Any]], max_doctors: int | None
) -> list[dict[str, Any]]:
    return select_gykqyy_trial_doctors(doctors, max_doctors)


def collect_gzszyy(
    target: HospitalTarget, today: str, max_doctors: int | None = None
) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页普通公开 GET 读取失败：{entry_error}")
    homepage_status, homepage_html, homepage_error = fetch(session, target.homepage)
    if homepage_status != 200:
        raise RuntimeError(f"官网首页院区普查读取失败：{homepage_error}")
    care_sites = discover_gzszyy_care_sites(homepage_html, target.homepage)
    if len(care_sites) != len(GZSZYY_CARE_SITE_PATHS):
        raise RuntimeError(
            "官网首页未发现完整院区/门诊部范围："
            + "、".join(item["name"] for item in care_sites)
        )
    unfiltered_page_urls = discover_gzszyy_unfiltered_pages(entry_html, target.entry_url)
    if not unfiltered_page_urls:
        raise RuntimeError("官网顶层专家目录未发现唯一、完整的公开分页范围。")
    unfiltered_relations: list[dict[str, str]] = []
    for page_number, page_url in enumerate(unfiltered_page_urls, start=1):
        if page_number == 1:
            page_status, html, page_error = entry_status, entry_html, ""
        else:
            page_status, html, page_error = fetch(session, page_url)
        if page_status != 200:
            raise RuntimeError(f"官网顶层专家目录第 {page_number} 页读取失败：{page_error}")
        page_rows = parse_gzszyy_department_page(
            html,
            page_url,
            {"department": "", "department_id": "", "entry_url": page_url},
        )
        if not page_rows:
            raise RuntimeError(f"官网顶层专家目录第 {page_number} 页严格医生卡片为 0。")
        unfiltered_relations.extend(page_rows)
    unfiltered_by_id = {
        relation["id"]: relation for relation in unfiltered_relations
    }
    filter_counts = {
        mode: len(
            {
                match.group(1)
                for match in re.finditer(
                    rf"/expert/1/{mode}/(\d+)/",
                    entry_html,
                    flags=re.IGNORECASE,
                )
            }
        )
        for mode in ("dp", "pr", "le")
    }
    departments = discover_gzszyy_department_filters(entry_html, target.entry_url)
    if not departments:
        raise RuntimeError("官网专家入口未发现严格 /expert/1/dp/<ID>/ 科室树。")

    categories: list[dict[str, Any]] = []
    category_errors: list[dict[str, str]] = []
    relations: list[dict[str, str]] = []
    for department in departments:
        status, first_html, error = fetch(session, department["entry_url"])
        if status != 200:
            category_errors.append(
                {"page": department["department"], "url": department["entry_url"], "error": error}
            )
            continue
        page_urls = discover_gzszyy_department_pages(first_html, department)
        department_relation_start = len(relations)
        for page_number, page_url in enumerate(page_urls, start=1):
            if page_number == 1:
                page_status, html, page_error = status, first_html, ""
            else:
                page_status, html, page_error = fetch(session, page_url)
            if page_status != 200:
                category_errors.append(
                    {"page": f"{department['department']} 第 {page_number} 页", "url": page_url, "error": page_error}
                )
                continue
            page_rows = parse_gzszyy_department_page(html, page_url, department)
            if not page_rows:
                category_errors.append(
                    {"page": f"{department['department']} 第 {page_number} 页", "url": page_url, "error": "严格医生卡片关系为 0"}
                )
                continue
            relations.extend(page_rows)
            categories.append(
                {
                    "category_id": f"{department['department_id']}:{page_number}",
                    "category_name": department["department"],
                    "url": page_url,
                    "doctor_relation_count": len(page_rows),
                }
            )
        if len(relations) == department_relation_start:
            category_errors.append(
                {"page": department["department"], "url": department["entry_url"], "error": "科室全部分页无医生关系"}
            )
    if category_errors:
        raise RuntimeError(
            "官网科室专家目录读取不完整："
            + "；".join(f"{item['page']} {item['error']}" for item in category_errors)
        )
    dp_relation_count = len(relations)
    dp_detail_ids = {relation["id"] for relation in relations}
    unfiltered_detail_ids = set(unfiltered_by_id)
    dp_only_detail_ids = sorted(dp_detail_ids - unfiltered_detail_ids)
    if dp_only_detail_ids:
        raise RuntimeError(
            "dp 科室树存在顶层全院目录未授权的详情 ID："
            + "、".join(dp_only_detail_ids)
        )
    unfiltered_only_detail_ids = sorted(unfiltered_detail_ids - dp_detail_ids)
    for detail_id in unfiltered_only_detail_ids:
        relation = dict(unfiltered_by_id[detail_id])
        relation["department"] = relation["card_department"]
        relations.append(relation)
    doctors = merge_gzszyy_relations(relations)
    if not doctors:
        raise RuntimeError("官网科室专家树未发现严格医生详情关系。")

    names_to_ids: dict[str, list[str]] = {}
    for doctor in doctors:
        if doctor["name"]:
            names_to_ids.setdefault(doctor["name"], []).append(doctor["id"])
    same_name_groups = {
        name: sorted(set(ids))
        for name, ids in names_to_ids.items()
        if len(set(ids)) > 1
    }
    nursing_doctors = [
        doctor for doctor in doctors if gyfyyy_nursing_only_identity(doctor["title"])
    ]
    eligible_doctors = [doctor for doctor in doctors if doctor not in nursing_doctors]
    selected = select_gzszyy_trial_doctors(eligible_doctors, max_doctors)
    existing_links = collect_existing_profile_links()
    excluded_candidates = [
        {
            "entry_url": target.entry_url,
            "list_title": f"{doctor['name']} {doctor['title']}",
            "source_link": doctor["source_link"],
            "reason": "官网科室目录仅标注护理身份，排除医生画像采集范围",
        }
        for doctor in nursing_doctors
    ]
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, Any]] = []
    for doctor in selected:
        detail_status, detail_html, detail_error = fetch(session, doctor["source_link"])
        if detail_status == 200:
            detail = parse_gzszyy_detail(detail_html, doctor)
        else:
            detail_errors.append({"source_link": doctor["source_link"], "error": detail_error})
            detail = {
                "name": doctor["name"],
                "title": doctor["title"],
                "departments": [],
                "campuses": [],
                "specialty": doctor["specialty"],
                "profile_text": "",
            }
        name = clean_text(str(detail.get("name") or doctor["name"]))
        title_identity = clean_text(str(detail.get("title") or doctor["title"]))
        departments_for_doctor = list(
            dict.fromkeys(
                clean_text(str(value))
                for value in [*doctor["departments"], *detail.get("departments", [])]
                if clean_text(str(value))
            )
        )
        card_departments = list(
            dict.fromkeys(
                clean_text(str(value))
                for value in doctor["card_departments"]
                if clean_text(str(value))
            )
        )
        campuses = list(
            dict.fromkeys(
                clean_text(str(value))
                for value in detail.get("campuses", [])
                if clean_text(str(value))
            )
        )
        department_and_sites = list(dict.fromkeys([*card_departments, *campuses]))
        specialty = clean_text(str(detail.get("specialty") or doctor["specialty"]))
        profile_text = clean_text(str(detail.get("profile_text") or ""))
        combined_text = "\n".join(
            [target.hospital, "、".join(departments_for_doctor), title_identity, specialty, profile_text]
        )
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if doctor["names"] and name not in doctor["names"]:
            warnings.append("列表与详情姓名不一致")
        if not departments_for_doctor:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if name in same_name_groups:
            warnings.append("同名待甄别")
        if warnings:
            groups_found = []
            tags = []
        priority = "普通"
        if not warnings and (
            any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
        ):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": "、".join(departments_for_doctor),
                "科室_列表卡片": "、".join(department_and_sites),
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": clip(specialty, 520),
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": clip(profile_text, 1800),
                "来源类型": "医院官网",
                "来源链接": doctor["source_link"],
                "采集入口": target.entry_url,
                "采集方式": "官网 dp 科室树静态分页+严格专家详情 DOM 结构化抽取",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(doctor["source_link"]) in existing_links else "否",
                "异常提示": "；".join(dict.fromkeys(warnings)),
                "复核状态": "待人工复核",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": doctor["id"],
                "name": name,
                "resolution": "正式行",
                "relation_count": doctor["relation_count"],
                "departments": departments_for_doctor,
                "campuses": campuses,
                "source_link": doctor["source_link"],
            }
        )
        time.sleep(0.12)

    raw_rows = rows
    if max_doctors is None:
        rows, identity_reconciliation = merge_gzszyy_identity_rows(
            raw_rows, detail_reconciliation
        )
    else:
        identity_reconciliation = [
            {
                "name": clean_text(str(row.get("姓名") or "")),
                "identity_index": 1,
                "resolution": "TRIAL 样本逐详情保留",
                "detail_ids": [
                    gzszyy_detail_id(str(row.get("来源链接") or ""))
                ],
                "primary_source_link": row.get("来源链接", ""),
                "merged_source_links": [],
                "departments": clean_text(row.get("科室_分类页")).split("、"),
                "campuses": detail_reconciliation[index].get("campuses", []),
                "relation_count": int(
                    detail_reconciliation[index].get("relation_count") or 1
                ),
            }
            for index, row in enumerate(rows)
        ]
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index
    identity_count_by_name = Counter(
        clean_text(str(item.get("name") or "")) for item in identity_reconciliation
    )
    distinct_same_name_groups = {
        name for name, count in identity_count_by_name.items() if name and count > 1
    }
    same_identity_merge_count = sum(
        1
        for item in identity_reconciliation
        if str(item.get("resolution") or "") == "同一人归并"
    )

    covered_departments = covered_department_names(rows)
    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(row["科室_分类页"]).split("、")
        if department
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in row["重点关注范围"].split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in row["异常提示"].split("；") if warning
    )
    campus_counter = Counter(
        campus
        for item in detail_reconciliation
        for campus in item.get("campuses", [])
        if campus
    )
    campus_tagged_details = [
        item for item in detail_reconciliation if item.get("campuses")
    ]
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #33（与官网入口台账一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(categories),
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(doctors),
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(doctors),
            "census_named_detail_count": sum(bool(doctor["name"]) for doctor in doctors),
            "census_blank_name_detail_count": sum(not doctor["name"] for doctor in doctors),
            "census_unique_nonblank_name_count": len(names_to_ids),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "gzszyy_final_identity_count": len(rows),
            "gzszyy_same_identity_merge_group_count": same_identity_merge_count,
            "gzszyy_distinct_same_name_group_count": len(distinct_same_name_groups),
            "gzszyy_distinct_same_name_row_count": sum(
                identity_count_by_name[name] for name in distinct_same_name_groups
            ),
            "census_department_count": len(departments),
            "census_group_count": len(care_sites),
            "census_nonempty_department_count": sum(bool(doctor["departments"]) for doctor in doctors),
            "census_empty_department_count": sum(not doctor["departments"] for doctor in doctors),
            "sample_entry_coverage_count": len(covered_departments),
            "sample_entry_categories": covered_departments,
            "pagination_count": len(categories),
            "pagination_method": "顶层全院目录 18 页校验身份覆盖；35 个 dp 科室筛选入口共 37 页提供科室关系；pr/le 仅为职称/级别筛选证据",
            "filter_semantics": "dp=科室、pr=职称、le=专家级别；顶层目录与 dp 科室树取并集，dp 缺失的顶层医生沿用其官方卡片科室",
            "filter_link_counts": filter_counts,
            "gzszyy_unfiltered_page_count": len(unfiltered_page_urls),
            "gzszyy_unfiltered_unique_detail_count": len(unfiltered_detail_ids),
            "gzszyy_dp_unique_detail_count": len(dp_detail_ids),
            "gzszyy_unfiltered_only_detail_count": len(unfiltered_only_detail_ids),
            "gzszyy_unfiltered_only_detail_ids": unfiltered_only_detail_ids,
            "gzszyy_dp_only_detail_count": len(dp_only_detail_ids),
            "gzszyy_multi_department_detail_count": sum(
                len(doctor["departments"]) > 1 for doctor in doctors
            ),
            "campus_relation_counts": dict(campus_counter),
            "cross_campus_detail_count": sum(
                len(item.get("campuses", [])) > 1 for item in detail_reconciliation
            ),
            "gzszyy_official_care_site_count": len(care_sites),
            "gzszyy_sample_detail_count": len(detail_reconciliation),
            "gzszyy_campus_tagged_sample_count": len(campus_tagged_details),
            "gzszyy_campus_untagged_sample_count": (
                len(detail_reconciliation) - len(campus_tagged_details)
            ),
            "schedule_field_ingested_count": 0,
            "category_error_count": len(category_errors),
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(doctors),
            "excluded_non_doctor_count": len(excluded_candidates),
            "eligible_candidate_count": len(eligible_doctors),
            "existing_profile_count": sum(row["已建画像"] == "是" for row in rows),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": categories,
        "entry_reconnaissance": [
            {
                "category_name": "官网顶层全院专家目录",
                "entry_url": target.entry_url,
                "page_nature": "医院官网名医名家未筛选目录；用于校验全院详情 ID 覆盖",
                "list_page_count": len(unfiltered_page_urls),
                "raw_detail_relation_count": len(unfiltered_detail_ids),
                "unique_detail_count": len(unfiltered_detail_ids),
                "out_of_scope_detail_count": 0,
                "affiliation": target.hospital,
                "independent_entity_check": "同域公开目录；与 dp 科室树逐 ID 对账",
            },
            {
                "category_name": "官网 dp 科室专家树",
                "entry_url": target.entry_url,
                "page_nature": "医院官网名医名家目录；dp 科室筛选为全院普查入口",
                "list_page_count": len(categories),
                "raw_detail_relation_count": dp_relation_count,
                "unique_detail_count": len(dp_detail_ids),
                "out_of_scope_detail_count": len(excluded_candidates),
                "affiliation": target.hospital,
                "independent_entity_check": "同域单一医院；首页院区/门诊部与详情二维码标签独立留痕",
            }
        ],
        "gzszyy_campus_reconnaissance": care_sites,
        "excluded_candidates": excluded_candidates,
        "gzszyy_detail_reconciliation": detail_reconciliation,
        "gzszyy_identity_reconciliation": identity_reconciliation,
        "cross_entry_duplicates": [
            {
                "name": doctor["name"],
                "source_link": doctor["source_link"],
                "entry_urls": doctor["departments"],
            }
            for doctor in doctors
            if doctor["relation_count"] > 1
        ],
        "category_errors": category_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def discover_gy3y_directory(html: str, entry_url: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    area = soup.select_one("section.areatab.tab")
    if not area:
        return {"campuses": [], "categories": [], "relations": []}
    campus_labels = [
        clean_text(node.get_text(" ", strip=True))
        for node in area.select(":scope > div.tabnav > span")
    ]
    campus_tabs = area.select(":scope > div.tabcontent > div.tabsingle")
    if campus_labels != ["荔湾院区", "黄埔院区"] or len(campus_tabs) != 2:
        return {"campuses": [], "categories": [], "relations": []}

    categories: list[dict[str, Any]] = []
    relations: list[dict[str, str]] = []
    campuses: list[dict[str, Any]] = []
    for campus, campus_tab in zip(campus_labels, campus_tabs, strict=True):
        campus_relation_start = len(relations)
        campus_category_start = len(categories)
        for section in campus_tab.select("section.ksdoclist"):
            title_node = section.find_previous_sibling("div", class_="title")
            system_name = clean_text(title_node.get_text(" ", strip=True)) if title_node else ""
            for block in section.find_all("dl", recursive=False):
                department_node = block.select_one(":scope > dt")
                department_name = (
                    clean_text(department_node.get_text(" ", strip=True)) if department_node else ""
                )
                if not department_name:
                    continue
                scoped_department = f"{campus}{department_name}"
                relation_start = len(relations)
                seen_ids: set[str] = set()
                for anchor in block.select(":scope > dd > a[href]"):
                    source_link = urljoin(entry_url, str(anchor.get("href") or ""))
                    doctor_id = gy3y_detail_id(source_link)
                    if not doctor_id or doctor_id in seen_ids:
                        continue
                    seen_ids.add(doctor_id)
                    department_url = urljoin(source_link, "./")
                    relations.append(
                        {
                            "id": doctor_id,
                            "name": clean_text(anchor.get_text(" ", strip=True)),
                            "campus": campus,
                            "system": system_name,
                            "department": scoped_department,
                            "department_url": department_url,
                            "source_link": source_link,
                            "list_title": clean_text(anchor.get_text(" ", strip=True)),
                        }
                    )
                categories.append(
                    {
                        "name": scoped_department,
                        "campus": campus,
                        "system": system_name,
                        "department": department_name,
                        "entry_url": entry_url,
                        "doctor_relation_count": len(relations) - relation_start,
                    }
                )
        campus_relations = relations[campus_relation_start:]
        campuses.append(
            {
                "name": campus,
                "department_count": len(categories) - campus_category_start,
                "doctor_relation_count": len(campus_relations),
                "unique_detail_count": len({item["id"] for item in campus_relations}),
            }
        )
    return {"campuses": campuses, "categories": categories, "relations": relations}


def select_gy3y_trial_doctors(
    doctors: list[dict[str, Any]],
    max_doctors: int | None,
) -> list[dict[str, Any]]:
    if not max_doctors:
        return doctors[:]
    huangpu_only = next(
        (
            item
            for item in doctors
            if item.get("campuses") == ["黄埔院区"]
            or (
                not item.get("campuses")
                and any("/ks/hp/" in str(link) for link in item.get("source_links", []))
            )
        ),
        None,
    )
    selected: list[dict[str, Any]] = [huangpu_only] if huangpu_only else []
    remaining = [item for item in doctors if item is not huangpu_only]
    selected.extend(
        select_gyfyyy_trial_doctors(
            remaining,
            min(max_doctors - len(selected), len(remaining)),
        )
    )
    selected_ids = {str(item["id"]) for item in selected}
    return selected + [item for item in doctors if str(item["id"]) not in selected_ids]


def collect_gy3y(target: HospitalTarget, today: str, max_doctors: int | None = None) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页读取失败：{entry_error}")
    directory = discover_gy3y_directory(entry_html, target.entry_url)
    campuses = directory["campuses"]
    categories = directory["categories"]
    relations = directory["relations"]
    if len(campuses) != 2 or not categories or not relations:
        raise RuntimeError("官网静态总目录未发现完整的荔湾院区与黄埔院区科室关系。")

    by_id: dict[str, dict[str, Any]] = {}
    for relation in relations:
        item = by_id.setdefault(
            relation["id"],
            {
                "id": relation["id"],
                "name": relation["name"],
                "departments": [],
                "campuses": [],
                "department_urls": [],
                "source_links": [],
                "list_titles": [],
            },
        )
        for field, value in (
            ("departments", relation["department"]),
            ("campuses", relation["campus"]),
            ("department_urls", relation["department_url"]),
            ("source_links", relation["source_link"]),
            ("list_titles", relation["list_title"]),
        ):
            if value and value not in item[field]:
                item[field].append(value)

    all_doctors = sorted(by_id.values(), key=lambda item: int(str(item["id"])))
    names_to_ids: dict[str, list[str]] = {}
    for item in all_doctors:
        name = clean_text(str(item.get("name") or ""))
        if name:
            names_to_ids.setdefault(name, []).append(str(item["id"]))
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in names_to_ids.items()
        if len(set(ids)) > 1
    }

    trial_candidates = all_doctors
    if max_doctors:
        trial_candidates = []
        seen_names: set[str] = set()
        for item in all_doctors:
            name = clean_text(str(item.get("name") or ""))
            if name and name in seen_names:
                continue
            if name:
                seen_names.add(name)
            trial_candidates.append(item)
    selected_doctors = select_gy3y_trial_doctors(trial_candidates, max_doctors)

    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    excluded_candidates: list[dict[str, str]] = []
    detail_errors: list[dict[str, str]] = []
    detail_reconciliation: list[dict[str, Any]] = []
    for item in selected_doctors:
        if max_doctors and len(rows) >= max_doctors:
            break
        source_link = item["source_links"][0]
        detail_status, detail_html, detail_error = fetch(session, source_link)
        if detail_status != 200:
            detail_errors.append({"source_link": source_link, "error": detail_error})
            detail: dict[str, str] = {}
        else:
            detail = parse_gyfyyy_detail(
                detail_html,
                {
                    "name": item.get("name", ""),
                    "list_title": item["list_titles"][0] if item["list_titles"] else "",
                },
            )
        name = clean_text(str(detail.get("name") or item.get("name") or ""))
        title_identity = clean_text(str(detail.get("title") or ""))
        if gyfyyy_nursing_only_identity(title_identity):
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": title_identity,
                    "source_link": source_link,
                    "reason": "官网详情仅标注护理身份，排除医生画像采集范围",
                }
            )
            continue
        department = "、".join(item["departments"])
        specialty = clean_text(str(detail.get("specialty") or ""))
        profile_text = clean_text(str(detail.get("profile_text") or ""))
        combined_text = "\n".join([target.hospital, department, title_identity, specialty, profile_text])
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not name or not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if warnings:
            groups_found = []
            tags = []
        priority = "普通"
        if not warnings and (any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": department,
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": profile_text,
                "来源类型": "医院官网",
                "来源链接": source_link,
                "采集入口": target.entry_url,
                "采集方式": "官网静态两院区总目录+同院区科室路径静态医生详情页",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(source_link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        detail_reconciliation.append(
            {
                "detail_id": item["id"],
                "name": name,
                "resolution": "按唯一 doctor ID 合并跨院区/跨科室归属",
                "relation_count": len(item["source_links"]),
                "campuses": item["campuses"],
                "departments": item["departments"],
                "primary_source_link": source_link,
                "merged_source_links": item["source_links"][1:],
            }
        )

    rows, identity_reconciliation = merge_gyfyyy_identity_rows(rows)
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index
    category_counter = Counter(
        department
        for row in rows
        for department in clean_text(str(row["科室_分类页"])).split("、")
        if department
    )
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in str(row["重点关注范围"]).split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in str(row["异常提示"]).split("；") if warning
    )
    sampled_departments = sorted(
        {
            department
            for row in rows
            for department in clean_text(str(row["科室_分类页"])).split("、")
            if department
        }
    )
    multi_relation = [item for item in all_doctors if len(item["source_links"]) > 1]
    cross_campus = [item for item in multi_relation if len(item["campuses"]) > 1]
    entry_reconnaissance = [
        {
            "category_name": campus["name"],
            "entry_url": target.entry_url,
            "page_nature": "静态全院区医生总目录",
            "list_page_count": 1,
            "raw_detail_relation_count": campus["doctor_relation_count"],
            "unique_detail_count": campus["unique_detail_count"],
            "out_of_scope_detail_count": "需逐详情核验",
            "affiliation": target.hospital,
            "independent_entity_check": "owner 已裁决两院区均属同一法人；官网同域静态目录",
        }
        for campus in campuses
    ]
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #27（入口台账主表与 owner 人工复核裁决一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(categories),
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(all_doctors),
            "sample_entry_coverage_count": len(sampled_departments),
            "sample_entry_categories": sampled_departments,
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(all_doctors),
            "census_department_count": len(categories),
            "census_nonempty_department_count": sum(
                1 for item in categories if int(item["doctor_relation_count"]) > 0
            ),
            "census_empty_department_count": sum(
                1 for item in categories if int(item["doctor_relation_count"]) == 0
            ),
            "census_group_count": len(campuses),
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "gy3y_final_identity_count": len(rows),
            "excluded_non_doctor_count": len(excluded_candidates),
            "census_nursing_identity_status": (
                "静态总目录只展示姓名，不展示职称身份；"
                f"{'TRIAL 仅核验样本详情' if max_doctors else f'FULL 已逐详情核验 {len(all_doctors)} 个唯一 ID'}，"
                f"其中纯护理身份排除 {len(excluded_candidates)} 位"
            ),
            "pagination_count": 0,
            "pagination_method": "单个 team.html 一次性列出两院区全部科室关系，无下一页或加载更多",
            "category_error_count": 0,
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(all_doctors),
            "gy3y_multi_relation_identity_count": len(multi_relation),
            "gy3y_cross_campus_identity_count": len(cross_campus),
            "campus_relation_counts": {
                campus["name"]: campus["doctor_relation_count"] for campus in campuses
            },
            "campus_unique_detail_counts": {
                campus["name"]: campus["unique_detail_count"] for campus in campuses
            },
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
            "entry_candidate_counts": {target.entry_url: len(relations)},
        },
        "categories": categories,
        "entry_reconnaissance": entry_reconnaissance,
        "excluded_candidates": excluded_candidates,
        "cross_entry_duplicates": [
            {
                "name": clean_text(str(item.get("name") or f"doctor_{item['id']}")),
                "source_link": item["source_links"][0],
                "entry_urls": item["departments"],
            }
            for item in multi_relation
        ],
        "gy3y_detail_reconciliation": detail_reconciliation,
        "gy3y_identity_reconciliation": identity_reconciliation,
        "category_errors": [],
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def collect_gyfyyy(target: HospitalTarget, today: str, max_doctors: int | None = None) -> dict[str, Any]:
    session = create_official_session()
    entry_status, entry_html, entry_error = fetch(session, target.entry_url)
    if entry_status != 200:
        raise RuntimeError(f"入口页读取失败：{entry_error}")
    departments = discover_gyfyyy_departments(entry_html, target.entry_url)
    if not departments:
        raise RuntimeError("官网科室树未发现严格同域科室入口。")

    relations: list[dict[str, str]] = []
    category_errors: list[dict[str, str]] = []
    entry_reconnaissance: list[dict[str, Any]] = []
    for department in departments:
        team_status, team_html, team_error = fetch(session, department["team_url"])
        if team_status != 200:
            category_errors.append(
                {"page": department["name"], "url": department["team_url"], "error": team_error}
            )
            continue
        department_relations = discover_gyfyyy_doctor_relations(team_html, department)
        relations.extend(department_relations)
        entry_reconnaissance.append(
            {
                "category_name": department["name"],
                "entry_url": department["team_url"],
                "page_nature": "静态科室专家团队页",
                "list_page_count": 1,
                "raw_detail_relation_count": len(department_relations),
                "unique_detail_count": len({item["id"] for item in department_relations}),
                "out_of_scope_detail_count": sum(
                    1 for item in department_relations if gyfyyy_nursing_only_identity(item["list_title"])
                ),
                "affiliation": target.hospital,
                "independent_entity_check": "官网同域科室路径内静态 doctor_<id>.html",
            }
        )
    if category_errors:
        failed = "、".join(item["page"] for item in category_errors)
        raise RuntimeError(f"官网科室专家团队普查不完整：{len(category_errors)} 个科室读取失败（{failed}）")

    by_id: dict[str, dict[str, Any]] = {}
    for relation in relations:
        item = by_id.setdefault(
            relation["id"],
            {
                "id": relation["id"],
                "departments": [],
                "department_urls": [],
                "source_links": [],
                "list_titles": [],
            },
        )
        for field, value in (
            ("departments", relation["department"]),
            ("department_urls", relation["department_url"]),
            ("source_links", relation["source_link"]),
            ("list_titles", relation["list_title"]),
        ):
            if value and value not in item[field]:
                item[field].append(value)

    all_doctors = sorted(by_id.values(), key=lambda item: int(str(item["id"])))
    excluded_candidates: list[dict[str, str]] = []
    eligible_doctors: list[dict[str, Any]] = []
    for item in all_doctors:
        list_title = "、".join(item["list_titles"])
        if gyfyyy_nursing_only_identity(list_title):
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": list_title,
                    "source_link": item["source_links"][0],
                    "reason": "官网团队卡片仅标注护理身份，排除医生画像采集范围",
                }
            )
            continue
        eligible_doctors.append(item)

    selected_doctors = select_gyfyyy_trial_doctors(eligible_doctors, max_doctors)
    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    identity_reconciliation: list[dict[str, Any]] = []
    for item in selected_doctors:
        source_link = item["source_links"][0]
        detail_status, detail_html, detail_error = fetch(session, source_link)
        if detail_status != 200:
            detail_errors.append({"source_link": source_link, "error": detail_error})
            detail = {}
        else:
            detail = parse_gyfyyy_detail(
                detail_html,
                {"list_title": item["list_titles"][0] if item["list_titles"] else ""},
            )
        name = clean_text(str(detail.get("name") or ""))
        title_identity = clean_text(str(detail.get("title") or ""))
        if gyfyyy_nursing_only_identity(title_identity):
            excluded_candidates.append(
                {
                    "entry_url": target.entry_url,
                    "list_title": title_identity,
                    "source_link": source_link,
                    "reason": "官网详情仅标注护理身份，排除医生画像采集范围",
                }
            )
            continue
        department = "、".join(item["departments"])
        specialty = clean_text(str(detail.get("specialty") or ""))
        profile_text = clean_text(str(detail.get("profile_text") or ""))
        combined_text = "\n".join([target.hospital, department, title_identity, specialty, profile_text])
        title_hits = extract_terms(title_identity, TITLE_TERMS)
        groups_found, tags = group_tags(combined_text)
        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not name or not looks_like_person_name(name):
            warnings.append("非医生页面或姓名异常")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not specialty and not profile_text:
            warnings.append("详情正文为空或未识别")
        if warnings:
            groups_found = []
            tags = []
        priority = "普通"
        if not warnings and (any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found):
            priority = "高"
        elif not warnings and any(term != "医师" for term in title_hits):
            priority = "中"
        rows.append(
            {
                "序号": len(rows) + 1,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": department,
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(title_identity, 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups_found),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": extract_clean_highlights(profile_text),
                "列表简介": "",
                "详情正文摘录": profile_text,
                "来源类型": "医院官网",
                "来源链接": source_link,
                "采集入口": target.entry_url,
                "采集方式": "官网静态科室树+同科室路径静态医生详情页",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(source_link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        identity_reconciliation.append(
            {
                "detail_id": item["id"],
                "name": name,
                "resolution": "按唯一 doctor ID 合并跨科室归属",
                "relation_count": len(item["source_links"]),
                "departments": item["departments"],
                "primary_source_link": source_link,
                "merged_source_links": item["source_links"][1:],
            }
        )

    raw_rows = rows
    detail_reconciliation = identity_reconciliation
    ids_by_name: dict[str, list[str]] = {}
    for row in raw_rows:
        name = clean_text(str(row.get("姓名") or ""))
        detail_id = gyfyyy_detail_id(str(row.get("来源链接") or ""))
        if name and detail_id:
            ids_by_name.setdefault(name, []).append(detail_id)
    same_name_groups = {
        name: sorted(set(ids), key=int)
        for name, ids in ids_by_name.items()
        if len(set(ids)) > 1
    }
    rows, identity_reconciliation = merge_gyfyyy_identity_rows(raw_rows)
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index
    identity_count_by_name = Counter(
        clean_text(str(item.get("name") or "")) for item in identity_reconciliation
    )
    distinct_same_name_groups = {
        name for name, count in identity_count_by_name.items() if name and count > 1
    }
    same_identity_merge_count = sum(
        1 for item in identity_reconciliation if int(item.get("relation_count") or 0) > 1
    )

    category_counter = Counter(row["科室_分类页"] for row in rows if row["科室_分类页"])
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter(
        group for row in rows for group in str(row["重点关注范围"]).split("、") if group
    )
    warning_counter = Counter(
        warning for row in rows for warning in str(row["异常提示"]).split("；") if warning
    )
    sampled_departments = sorted(
        {department for item in selected_doctors for department in item["departments"] if department}
    )
    cross_department = [item for item in all_doctors if len(item["departments"]) > 1]
    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "entry_url_source": "GitHub Issue #25（与官网入口台账人工复核裁决一致）",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(departments),
            "raw_card_rows": len(relations),
            "candidate_membership_count": len(relations),
            "unique_candidate_count": len(all_doctors),
            "sample_entry_coverage_count": len(sampled_departments),
            "sample_entry_categories": sampled_departments,
            "unique_doctor_count": len(rows),
            "census_unique_detail_count": len(all_doctors),
            "census_department_count": len(departments),
            "census_group_count": 0,
            "census_same_name_group_count": len(same_name_groups),
            "census_same_name_groups": same_name_groups,
            "gyfyyy_final_identity_count": len(rows),
            "gyfyyy_same_identity_merge_group_count": same_identity_merge_count,
            "gyfyyy_distinct_same_name_group_count": len(distinct_same_name_groups),
            "gyfyyy_distinct_same_name_row_count": sum(
                identity_count_by_name[name] for name in distinct_same_name_groups
            ),
            "excluded_non_doctor_count": len(excluded_candidates),
            "pagination_count": len(departments),
            "pagination_method": "每个科室单个 doctorList.html 一次性列出团队，无分页参数或加载更多",
            "category_error_count": len(category_errors),
            "detail_error_count": len(detail_errors),
            "cross_entry_duplicate_count": len(relations) - len(all_doctors),
            "gyfyyy_cross_department_identity_count": len(cross_department),
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
            "entry_candidate_counts": {target.entry_url: len(relations)},
        },
        "categories": departments,
        "entry_reconnaissance": entry_reconnaissance,
        "excluded_candidates": excluded_candidates,
        "cross_entry_duplicates": [
            {
                "name": "doctor_" + item["id"],
                "source_link": item["source_links"][0],
                "entry_urls": item["department_urls"],
            }
            for item in cross_department
        ],
        "gyfyyy_detail_reconciliation": detail_reconciliation,
        "gyfyyy_identity_reconciliation": identity_reconciliation,
        "category_errors": category_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def parse_gzzoc_detail(html: str, fallback: dict[str, str]) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
    title_name = title.split("|", 1)[0].strip() if "|" in title else ""
    remove_profile_noise_elements(soup)
    main = soup.select_one(".main") or soup
    lines = [clean_text(line) for line in main.get_text("\n", strip=True).splitlines() if clean_text(line)]

    specialty = clean_text((soup.select_one(".field-adept") or "").get_text(" ", strip=True)) if soup.select_one(".field-adept") else ""
    profile_text = clean_text((soup.select_one(".field-body") or "").get_text(" ", strip=True)) if soup.select_one(".field-body") else ""
    department = extract_label_from_lines(lines, "专业", ["职称", "擅长", "简介", "预约挂号"])
    title_field = extract_label_from_lines(lines, "职称", ["擅长", "简介", "预约挂号"])

    if not specialty:
        specialty = extract_label_from_lines(lines, "擅长", ["简介", "预约挂号", "在线预约"])
    if not profile_text:
        profile_text = extract_label_from_lines(lines, "简介", ["预约挂号", "在线预约", "快速连接"])

    return {
        "name": first_nonempty(title_name, fallback.get("name", "")),
        "department": first_nonempty(department, fallback.get("department", "")),
        "title_field": first_nonempty(title_field, fallback.get("list_title", "")),
        "specialty": specialty,
        "profile_text": profile_text,
        "title_text": title_name,
    }


def collect_gzzoc(target: HospitalTarget, today: str, max_doctors: int | None = None) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36; "
                "public official-site collection"
            )
        }
    )
    status, entry_html, entry_error = fetch(session, target.entry_url)
    if status != 200:
        raise RuntimeError(f"入口页读取失败：{entry_error}")

    page_urls = gzzoc_list_pages(target.entry_url, entry_html)
    page_errors: list[dict[str, str]] = []
    raw_rows: list[dict[str, str]] = []
    for index, page_url in enumerate(page_urls, start=1):
        if index == 1 and canonical_url(page_url) == canonical_url(with_query_param(target.entry_url, "page", 0)):
            html = entry_html
            page_status = status
            page_error = ""
        else:
            page_status, html, page_error = fetch(session, page_url)
        if page_status != 200:
            page_errors.append({"page": str(index), "url": page_url, "error": page_error})
            continue
        page_rows = parse_gzzoc_list_page(html, page_url)
        raw_rows.extend(page_rows)
        print(f"[{index}/{len(page_urls)}] list rows: {len(page_rows)}")
        time.sleep(0.2)

    by_link: dict[str, dict[str, str]] = {}
    for row in raw_rows:
        key = generic_detail_identity(row["source_link"])
        item = by_link.setdefault(
            key,
            {
                "source_link": row["source_link"],
                "name": row["name"],
                "list_title": row["list_title"],
                "department": row["department"],
                "description": row["description"],
                "list_pages": "",
            },
        )
        for field in ["name", "list_title", "department", "description"]:
            if row[field] and len(row[field]) > len(item.get(field, "")):
                item[field] = row[field]
        pages = [value for value in item["list_pages"].split("；") if value]
        if row["list_page"] not in pages:
            pages.append(row["list_page"])
            item["list_pages"] = "；".join(pages)

    detail_items = sorted(by_link.values(), key=lambda item: item["source_link"])
    if max_doctors:
        detail_items = detail_items[:max_doctors]

    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    for index, item in enumerate(detail_items, start=1):
        link = item["source_link"]
        detail_status, detail_html, detail_error = fetch(session, link)
        detail = {
            "name": item.get("name", ""),
            "department": item.get("department", ""),
            "title_field": item.get("list_title", ""),
            "specialty": "",
            "profile_text": "",
            "title_text": "",
        }
        if detail_status == 200:
            detail = parse_gzzoc_detail(detail_html, item)
        else:
            detail_errors.append({"source_link": link, "error": detail_error})

        name = first_nonempty(detail.get("name"), item.get("name"))
        department = first_nonempty(detail.get("department"), item.get("department"))
        combined_text = "\n".join(
            [
                target.hospital,
                department,
                item.get("list_title", ""),
                item.get("description", ""),
                detail.get("title_field", ""),
                detail.get("specialty", ""),
                detail.get("profile_text", ""),
            ]
        )
        title_source = first_nonempty(detail.get("title_field"), item.get("list_title"))
        title_hits = extract_terms(title_source, TITLE_TERMS) or extract_terms(combined_text, TITLE_TERMS)
        groups, tags = group_tags(combined_text)
        specialty = clip(
            first_nonempty(
                detail.get("specialty"),
                item.get("description"),
                extract_sentences(
                    combined_text,
                    ["擅长", "研究方向", "诊治", "治疗", "诊断", "手术", "专长", "复杂", "疑难"],
                    limit=4,
                    max_len=520,
                ),
            ),
            520,
        )
        highlights = extract_sentences(combined_text, HIGHLIGHT_TERMS)

        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not name:
            warnings.append("姓名需人工复核")
        if not department:
            warnings.append("专业/科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not detail.get("specialty") and not detail.get("profile_text") and not item.get("description"):
            warnings.append("详情页正文为空或未识别")

        priority = "普通"
        if any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups:
            priority = "高"
        elif any(term != "医师" for term in title_hits):
            priority = "中"

        rows.append(
            {
                "序号": index,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": item.get("department", ""),
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(first_nonempty(detail.get("title_field"), item.get("list_title")), 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": highlights,
                "列表简介": clip(item.get("description", ""), 700),
                "详情正文摘录": clip(detail.get("profile_text", ""), 1800),
                "来源类型": "医院官网",
                "来源链接": link,
                "采集入口": target.entry_url,
                "采集方式": "官网专家列表页+官网医生详情页",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        if index % 20 == 0:
            print(f"details: {index}/{len(detail_items)}")
        time.sleep(0.18)

    rows.sort(key=lambda row: (row["重点优先级"] != "高", row["科室_分类页"], row["姓名"]))
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index

    category_counter = Counter()
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter()
    warning_counter = Counter()
    for row in rows:
        if row["科室_分类页"]:
            category_counter[row["科室_分类页"]] += 1
        for group in row["重点关注范围"].split("、"):
            if group:
                group_counter[group] += 1
        for warning in row["异常提示"].split("；"):
            if warning:
                warning_counter[warning] += 1

    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(page_urls),
            "raw_card_rows": len(raw_rows),
            "unique_doctor_count": len(rows),
            "category_error_count": len(page_errors),
            "detail_error_count": len(detail_errors),
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": [{"category_id": str(index), "category_name": f"专家列表第{index}页", "url": url} for index, url in enumerate(page_urls, start=1)],
        "category_errors": page_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def nbkj_list_pages(entry_url: str, html: str) -> list[str]:
    page_numbers = [int(match.group(1)) for match in re.finditer(r"(?:/expert/)?list_(\d+)\.html", html)]
    max_page = max(page_numbers) if page_numbers else 1
    pages = [entry_url]
    for page in range(2, max_page + 1):
        pages.append(urljoin(entry_url.rstrip("/") + "/", f"list_{page}.html"))
    return pages


def parse_nbkj_list_page(html: str, page_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, str]] = []
    for anchor in soup.find_all("a", href=True):
        href = urljoin(page_url, anchor["href"])
        if not re.search(r"/expert/\d+\.html$", urlparse(href).path):
            continue
        text = clean_text(anchor.get_text(" ", strip=True))
        image = anchor.find("img")
        image_alt = clean_text(image.get("alt")) if image else ""
        list_title = first_nonempty(text, image_alt)
        rows.append(
            {
                "source_link": href,
                "list_title": list_title,
                "list_page": page_url,
            }
        )
    return rows


def extract_labeled_value(text: str, labels: list[str], stop_labels: list[str]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = "|".join(re.escape(label) for label in stop_labels)
    pattern = rf"(?:{label_pattern})\s*[:：]\s*(.*?)(?=(?:{stop_pattern})\s*[:：]?|上一篇[:：]|下一篇[:：]|相关文章[:：]|南部战区空军医院|$)"
    match = re.search(pattern, text, flags=re.S)
    return clean_text(match.group(1)) if match else ""


def parse_nbkj_detail(html: str, fallback_title: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
    title_name = title.split("_", 1)[0].strip() if "_" in title else ""
    remove_profile_noise_elements(soup)
    cont = soup.find("div", class_="cont")
    paragraph_scope = cont if cont else soup
    paragraphs = [
        clean_text(paragraph.get_text(" ", strip=True))
        for paragraph in paragraph_scope.find_all("p")
        if clean_text(paragraph.get_text(" ", strip=True))
    ]
    body_text = strip_article_tail(" ".join(paragraphs))
    text = "\n".join(
        clean_text(line)
        for line in soup.get_text("\n", strip=True).splitlines()
        if clean_text(line)
    )
    compact = clean_text(text)
    name = extract_labeled_value(compact, ["姓名", "医生名称"], ["科室", "所在科室", "职称", "专家职称", "专业擅长", "专家专长"])
    department = extract_labeled_value(compact, ["科室", "所在科室"], ["职称", "专家职称", "专家学历", "专业擅长", "专家专长"])
    title_field = extract_labeled_value(compact, ["职称", "专家职称"], ["专家学历", "专业擅长", "专家专长"])
    specialty = strip_article_tail(extract_labeled_value(compact, ["专业擅长", "专家专长"], ["上一篇", "下一篇", "相关文章"]))
    title_context = title_name
    if len(clean_text(fallback_title)) > len(clean_text(title_context)):
        title_context = clean_text(fallback_title)
    title_context = first_nonempty(title_context, fallback_title)
    if not name and title_context:
        name = clean_text(re.split(r"[\s，,、；;]", title_context, maxsplit=1)[0])
    if not name:
        name = title_name.split(" ", 1)[0] if title_name else ""

    title_rest = title_context
    if name and title_rest.startswith(name):
        title_rest = clean_text(title_rest[len(name) :])
    if not department:
        department = infer_department(title_rest) or infer_department(body_text[:180])
    if not title_field:
        title_field = "、".join(extract_terms(f"{title_context} {body_text[:260]}", TITLE_TERMS))
    if not specialty:
        specialty = body_text
    profile_text = strip_article_tail(specialty or body_text)
    return {
        "name": name,
        "department": department,
        "title_field": title_field,
        "specialty": specialty,
        "profile_text": profile_text,
        "title_text": title_name,
    }


def collect_nbkj(target: HospitalTarget, today: str, max_doctors: int | None = None) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36; "
                "public official-site collection"
            )
        }
    )
    status, entry_html, entry_error = fetch(session, target.entry_url)
    if status != 200:
        raise RuntimeError(f"入口页读取失败：{entry_error}")

    page_urls = nbkj_list_pages(target.entry_url, entry_html)
    page_errors: list[dict[str, str]] = []
    raw_rows: list[dict[str, str]] = []
    for index, page_url in enumerate(page_urls, start=1):
        if index == 1:
            html = entry_html
            page_status = status
            page_error = ""
        else:
            page_status, html, page_error = fetch(session, page_url)
        if page_status != 200:
            page_errors.append({"page": str(index), "url": page_url, "error": page_error})
            continue
        page_rows = parse_nbkj_list_page(html, page_url)
        raw_rows.extend(page_rows)
        print(f"[{index}/{len(page_urls)}] list rows: {len(page_rows)}")
        time.sleep(0.25)

    by_link: dict[str, dict[str, str]] = {}
    for row in raw_rows:
        key = canonical_url(row["source_link"])
        item = by_link.setdefault(
            key,
            {
                "source_link": row["source_link"],
                "list_title": row["list_title"],
                "list_pages": "",
            },
        )
        if row["list_title"] and len(row["list_title"]) > len(item.get("list_title", "")):
            item["list_title"] = row["list_title"]
        pages = [value for value in item["list_pages"].split("；") if value]
        if row["list_page"] not in pages:
            pages.append(row["list_page"])
            item["list_pages"] = "；".join(pages)

    detail_items = sorted(by_link.values(), key=lambda item: item["source_link"])
    if max_doctors:
        detail_items = detail_items[:max_doctors]

    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    for index, item in enumerate(detail_items, start=1):
        link = item["source_link"]
        detail_status, detail_html, detail_error = fetch(session, link)
        detail = {
            "name": "",
            "department": "",
            "title_field": "",
            "specialty": "",
            "profile_text": "",
            "title_text": "",
        }
        if detail_status == 200:
            detail = parse_nbkj_detail(detail_html, item.get("list_title", ""))
        else:
            detail_errors.append({"source_link": link, "error": detail_error})

        name = first_nonempty(detail.get("name"), item.get("list_title"))
        department = clean_text(detail.get("department"))
        combined_text = "\n".join(
            [
                target.hospital,
                department,
                item.get("list_title", ""),
                detail.get("title_field", ""),
                detail.get("specialty", ""),
                detail.get("profile_text", ""),
            ]
        )
        title_hits = extract_terms(combined_text, TITLE_TERMS)
        groups, tags = group_tags(combined_text)
        specialty = clip(
            first_nonempty(
                detail.get("specialty"),
                extract_sentences(
                    combined_text,
                    ["擅长", "研究方向", "诊治", "治疗", "诊断", "手术", "放疗", "介入", "专长"],
                    limit=4,
                    max_len=520,
                ),
            ),
            520,
        )
        highlights = extract_sentences(combined_text, HIGHLIGHT_TERMS)

        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        if not name:
            warnings.append("姓名需人工复核")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not detail.get("specialty") and not detail.get("profile_text"):
            warnings.append("详情页正文为空或未识别")

        priority = "普通"
        if any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups:
            priority = "高"
        elif any(term != "医师" for term in title_hits):
            priority = "中"

        rows.append(
            {
                "序号": index,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": item.get("list_title", ""),
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(first_nonempty(detail.get("title_field"), item.get("list_title")), 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": highlights,
                "列表简介": clip(item.get("list_title", ""), 700),
                "详情正文摘录": clip(detail.get("profile_text", ""), 1800),
                "来源类型": "医院官网",
                "来源链接": link,
                "采集入口": target.entry_url,
                "采集方式": "官网专家列表页+官网医生详情页",
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        if index % 20 == 0:
            print(f"details: {index}/{len(detail_items)}")
        time.sleep(0.22)

    rows.sort(key=lambda row: (row["重点优先级"] != "高", row["科室_分类页"], row["姓名"]))
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index

    category_counter = Counter()
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter()
    warning_counter = Counter()
    for row in rows:
        if row["科室_分类页"]:
            category_counter[row["科室_分类页"]] += 1
        for group in row["重点关注范围"].split("、"):
            if group:
                group_counter[group] += 1
        for warning in row["异常提示"].split("；"):
            if warning:
                warning_counter[warning] += 1

    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(page_urls),
            "raw_card_rows": len(raw_rows),
            "unique_doctor_count": len(rows),
            "category_error_count": len(page_errors),
            "detail_error_count": len(detail_errors),
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
        },
        "categories": [{"category_id": str(index), "category_name": f"专家列表第{index}页", "url": url} for index, url in enumerate(page_urls, start=1)],
        "category_errors": page_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def round_robin_generic_items(
    items: list[dict[str, str]],
    entry_urls: list[str],
    max_items: int | None,
    spread: bool = False,
) -> list[dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = {canonical_url(url): [] for url in entry_urls}
    for item in items:
        grouped.setdefault(canonical_url(item.get("entry_url", "")), []).append(item)
    for values in grouped.values():
        if spread and max_items and len(values) > 1:
            sample_count = min(max_items, len(values))
            spread_indices = {
                round(index * (len(values) - 1) / (sample_count - 1))
                for index in range(sample_count)
            }
            values[:] = [values[index] for index in sorted(spread_indices)] + [
                item for index, item in enumerate(values) if index not in spread_indices
            ]
        else:
            values.sort(key=lambda item: item["source_link"])

    ordered: list[dict[str, str]] = []
    offsets = {key: 0 for key in grouped}
    while True:
        added = False
        for entry_url in entry_urls:
            key = canonical_url(entry_url)
            values = grouped.get(key, [])
            offset = offsets.get(key, 0)
            if offset >= len(values):
                continue
            ordered.append(values[offset])
            offsets[key] = offset + 1
            added = True
            if max_items and len(ordered) >= max_items:
                return ordered
        if not added:
            return ordered


def gdzy5413_normalized_name(value: str | None) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def select_gdzy5413_trial2_items(
    items: list[dict[str, str]],
    max_doctors: int | None,
) -> list[dict[str, str]]:
    """Select auditable 852 name groups, including one merge case and one Baiyun record."""

    if not max_doctors:
        return items
    groups: dict[str, list[dict[str, str]]] = {}
    for item in items:
        name = gdzy5413_normalized_name(item.get("name"))
        if name:
            groups.setdefault(name, []).append(item)

    duplicate_groups = [values for values in groups.values() if len(values) > 1]
    duplicate_groups.sort(
        key=lambda values: (
            len({clean_text(item.get("department")) for item in values}) != 1,
        )
    )
    selected_names: list[str] = []
    if duplicate_groups:
        selected_names.append(gdzy5413_normalized_name(duplicate_groups[0][0].get("name")))

    baiyun_names = [
        name
        for name, values in groups.items()
        if len(values) == 1 and "白云院区" in clean_text(values[0].get("department"))
    ]
    if baiyun_names and baiyun_names[0] not in selected_names:
        selected_names.append(baiyun_names[0])

    covered_departments = {
        clean_text(item.get("department"))
        for name in selected_names
        for item in groups[name]
        if clean_text(item.get("department"))
    }
    unique_groups = [(name, values) for name, values in groups.items() if len(values) == 1]
    for name, values in unique_groups:
        if len(selected_names) >= max_doctors:
            break
        department = clean_text(values[0].get("department"))
        if name in selected_names or not department or department in covered_departments:
            continue
        selected_names.append(name)
        covered_departments.add(department)
    for name, _values in unique_groups:
        if len(selected_names) >= max_doctors:
            break
        if name not in selected_names:
            selected_names.append(name)

    selected: list[dict[str, str]] = []
    for name in selected_names[:max_doctors]:
        selected.extend(groups[name])
    return selected


def expand_gdzy5413_full_detail_items(items: list[dict[str, str]]) -> list[dict[str, str]]:
    """FULL must fetch every same-name link so the owner identity rule has complete evidence."""

    groups: dict[str, list[dict[str, str]]] = {}
    order: list[str] = []
    for item in items:
        name = gdzy5413_normalized_name(item.get("name"))
        if name not in groups:
            groups[name] = []
            order.append(name)
        groups[name].append(item)
    return [item for name in order for item in groups[name]]


def gdzy5413_identity_text(value: str | None) -> str:
    return re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]+", "", clean_text(value)).lower()


def gdzy5413_text_similarity(left: str | None, right: str | None) -> float:
    first = gdzy5413_identity_text(left)
    second = gdzy5413_identity_text(right)
    if not first or not second:
        return 0.0
    if first in second or second in first:
        return 1.0
    return SequenceMatcher(None, first, second).ratio()


def gdzy5413_rows_same_identity(left: dict[str, Any], right: dict[str, Any]) -> bool:
    if gdzy5413_normalized_name(left.get("姓名")) != gdzy5413_normalized_name(right.get("姓名")):
        return False
    left_titles = set(extract_terms(clean_text(left.get("职称身份原文")), TITLE_TERMS))
    right_titles = set(extract_terms(clean_text(right.get("职称身份原文")), TITLE_TERMS))
    title_consistent = bool(left_titles & right_titles) or not left_titles or not right_titles
    specialty_similarity = gdzy5413_text_similarity(
        left.get("擅长诊疗方向摘录"), right.get("擅长诊疗方向摘录")
    )
    profile_similarity = gdzy5413_text_similarity(
        left.get("详情正文摘录"), right.get("详情正文摘录")
    )
    specialty_evidence = bool(
        gdzy5413_identity_text(left.get("擅长诊疗方向摘录"))
        and gdzy5413_identity_text(right.get("擅长诊疗方向摘录"))
    )
    if specialty_evidence:
        return title_consistent and specialty_similarity >= 0.55
    return title_consistent and profile_similarity >= 0.7


def gdzy5413_primary_row_score(row: dict[str, Any]) -> int:
    """Prefer the detail with the richest clinical biography over noisy title repetition."""

    return (
        len(clean_text(row.get("详情正文摘录"))) * 5
        + len(clean_text(row.get("擅长诊疗方向摘录"))) * 2
        + len(clean_text(row.get("职称身份原文")))
        + len(clean_text(row.get("亮眼经历线索")))
    )


def merge_gdzy5413_identity_rows(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Apply the owner rule: merge materially consistent names, preserve distinct identities."""

    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(gdzy5413_normalized_name(row.get("姓名")), []).append(row)

    merged_rows: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    longest_fields = [
        "擅长诊疗方向摘录",
        "亮眼经历线索",
        "列表简介",
        "详情正文摘录",
    ]
    for name, name_rows in by_name.items():
        clusters: list[list[dict[str, Any]]] = []
        for row in name_rows:
            for cluster in clusters:
                if any(gdzy5413_rows_same_identity(row, member) for member in cluster):
                    cluster.append(row)
                    break
            else:
                clusters.append([row])

        distinct_same_name = len(clusters) > 1
        for identity_index, cluster in enumerate(clusters, start=1):
            primary = max(cluster, key=gdzy5413_primary_row_score)
            merged = dict(primary)
            departments: list[str] = []
            for member in cluster:
                for field in ["科室_分类页", "科室_列表卡片"]:
                    for department in clean_text(member.get(field)).split("、"):
                        department = clean_text(department)
                        if department and department not in departments:
                            departments.append(department)
                for field in longest_fields:
                    if len(clean_text(member.get(field))) > len(clean_text(merged.get(field))):
                        merged[field] = member.get(field, "")
            merged["科室_分类页"] = "、".join(departments)
            merged["科室_列表卡片"] = "、".join(departments)
            merged["职称_关键词"] = "、".join(
                extract_terms(clean_text(primary.get("职称身份原文")), TITLE_TERMS)
            )
            warnings = [
                warning
                for member in cluster
                for warning in clean_text(member.get("异常提示")).split("；")
                if warning
            ]
            distinct_titles = {
                clean_text(member.get("职称身份原文"))
                for member in cluster
                if clean_text(member.get("职称身份原文"))
            }
            if len(distinct_titles) > 1:
                warnings.append("多详情职称不一致")
            if distinct_same_name:
                warnings.append("同名待甄别")
            merged["异常提示"] = "；".join(dict.fromkeys(warnings))
            merged_rows.append(merged)
            reconciliation.append(
                {
                    "name": name,
                    "identity_index": identity_index,
                    "resolution": (
                        "同名待甄别"
                        if distinct_same_name
                        else "同一人归并"
                        if len(cluster) > 1
                        else "唯一身份"
                    ),
                    "primary_source_link": merged.get("来源链接", ""),
                    "merged_source_links": [
                        member.get("来源链接", "")
                        for member in cluster
                        if member.get("来源链接") != merged.get("来源链接")
                    ],
                    "departments": departments,
                    "relation_count": len(cluster),
                }
            )

    return merged_rows, reconciliation


def merge_gyfyyy_identity_rows(
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Cluster same-name GYFYYY detail IDs using official identity fields."""

    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(gdzy5413_normalized_name(row.get("姓名")), []).append(row)

    merged_rows: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    longest_fields = [
        "擅长诊疗方向摘录",
        "亮眼经历线索",
        "列表简介",
        "详情正文摘录",
    ]
    for name, name_rows in by_name.items():
        clusters: list[list[dict[str, Any]]] = []
        for row in name_rows:
            for cluster in clusters:
                if any(gdzy5413_rows_same_identity(row, member) for member in cluster):
                    cluster.append(row)
                    break
            else:
                clusters.append([row])

        distinct_same_name = len(clusters) > 1
        for identity_index, cluster in enumerate(clusters, start=1):
            primary = max(cluster, key=gdzy5413_primary_row_score)
            merged = dict(primary)
            departments: list[str] = []
            for member in cluster:
                for field in ["科室_分类页", "科室_列表卡片"]:
                    for department in clean_text(member.get(field)).split("、"):
                        department = clean_text(department)
                        if department and department not in departments:
                            departments.append(department)
                for field in longest_fields:
                    if len(clean_text(member.get(field))) > len(clean_text(merged.get(field))):
                        merged[field] = member.get(field, "")

            merged["科室_分类页"] = "、".join(departments)
            merged["科室_列表卡片"] = "、".join(departments)
            merged["职称_关键词"] = "、".join(
                extract_terms(clean_text(primary.get("职称身份原文")), TITLE_TERMS)
            )
            warnings = [
                warning
                for member in cluster
                for warning in clean_text(member.get("异常提示")).split("；")
                if warning
            ]
            distinct_titles = {
                clean_text(member.get("职称身份原文"))
                for member in cluster
                if clean_text(member.get("职称身份原文"))
            }
            if len(distinct_titles) > 1:
                warnings.append("多详情职称不一致")
            if distinct_same_name:
                warnings.append("同名待甄别")
            merged["异常提示"] = "；".join(dict.fromkeys(warnings))

            combined_text = "\n".join(
                clean_text(str(merged.get(field) or ""))
                for field in [
                    "医院",
                    "科室_分类页",
                    "职称身份原文",
                    "擅长诊疗方向摘录",
                    "详情正文摘录",
                ]
            )
            groups_found, tags = group_tags(combined_text)
            if merged["异常提示"]:
                groups_found = []
                tags = []
            merged["重点关注范围"] = "、".join(groups_found)
            merged["重点疾病标签"] = "、".join(tags)
            merged["重点优先级"] = "普通"
            if not merged["异常提示"] and (
                any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
            ):
                merged["重点优先级"] = "高"
            elif not merged["异常提示"] and any(
                term != "医师" for term in extract_terms(merged["职称身份原文"], TITLE_TERMS)
            ):
                merged["重点优先级"] = "中"

            primary_source = clean_text(str(merged.get("来源链接") or ""))
            source_links = [
                clean_text(str(member.get("来源链接") or "")) for member in cluster
            ]
            merged_rows.append(merged)
            reconciliation.append(
                {
                    "name": name,
                    "identity_index": identity_index,
                    "resolution": (
                        "同名待甄别"
                        if distinct_same_name
                        else "同一人归并"
                        if len(cluster) > 1
                        else "唯一身份"
                    ),
                    "detail_ids": [
                        gyfyyy_detail_id(source) or gy3y_detail_id(source)
                        for source in source_links
                    ],
                    "primary_source_link": primary_source,
                    "merged_source_links": [
                        source for source in source_links if source != primary_source
                    ],
                    "departments": departments,
                    "relation_count": len(cluster),
                }
            )

    return merged_rows, reconciliation


def merge_gzszyy_identity_rows(
    rows: list[dict[str, Any]],
    detail_reconciliation: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Apply the four owner-audited same-name decisions without heuristic over-merging."""

    detail_evidence = {
        str(item.get("detail_id") or ""): item for item in detail_reconciliation
    }
    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_name.setdefault(gdzy5413_normalized_name(row.get("姓名")), []).append(row)

    merged_rows: list[dict[str, Any]] = []
    reconciliation: list[dict[str, Any]] = []
    longest_fields = [
        "擅长诊疗方向摘录",
        "亮眼经历线索",
        "列表简介",
        "详情正文摘录",
    ]
    for name, name_rows in by_name.items():
        ids_for_name = [
            gzszyy_detail_id(str(row.get("来源链接") or "")) for row in name_rows
        ]
        id_group = frozenset(detail_id for detail_id in ids_for_name if detail_id)
        if id_group in GZSZYY_SAME_IDENTITY_DETAIL_GROUPS:
            clusters = [name_rows]
        else:
            # 王健是已审计的实质不同身份；任何未来未知同名也按安全侧分行。
            clusters = [[row] for row in name_rows]

        distinct_same_name = len(clusters) > 1
        for identity_index, cluster in enumerate(clusters, start=1):
            primary = max(cluster, key=gdzy5413_primary_row_score)
            merged = dict(primary)
            category_departments: list[str] = []
            card_departments_and_sites: list[str] = []
            detail_ids: list[str] = []
            campuses: list[str] = []
            relation_count = 0
            for member in cluster:
                detail_id = gzszyy_detail_id(str(member.get("来源链接") or ""))
                if detail_id and detail_id not in detail_ids:
                    detail_ids.append(detail_id)
                evidence = detail_evidence.get(detail_id, {})
                relation_count += int(evidence.get("relation_count") or 1)
                for campus in evidence.get("campuses", []):
                    campus = clean_text(str(campus))
                    if campus and campus not in campuses:
                        campuses.append(campus)
                for department in clean_text(member.get("科室_分类页")).split("、"):
                    department = clean_text(department)
                    if department and department not in category_departments:
                        category_departments.append(department)
                for department in clean_text(member.get("科室_列表卡片")).split("、"):
                    department = clean_text(department)
                    if department and department not in card_departments_and_sites:
                        card_departments_and_sites.append(department)
                for field in longest_fields:
                    if len(clean_text(member.get(field))) > len(clean_text(merged.get(field))):
                        merged[field] = member.get(field, "")

            merged["科室_分类页"] = "、".join(category_departments)
            merged["科室_列表卡片"] = "、".join(card_departments_and_sites)
            merged["职称_关键词"] = "、".join(
                extract_terms(clean_text(primary.get("职称身份原文")), TITLE_TERMS)
            )
            warnings = [
                warning
                for member in cluster
                for warning in clean_text(member.get("异常提示")).split("；")
                if warning
                and not (warning == "同名待甄别" and len(name_rows) > 1)
            ]
            distinct_titles = {
                clean_text(member.get("职称身份原文"))
                for member in cluster
                if clean_text(member.get("职称身份原文"))
            }
            if len(distinct_titles) > 1:
                warnings.append("多详情职称不一致")
            if distinct_same_name:
                warnings.append("同名待甄别")
            merged["异常提示"] = "；".join(dict.fromkeys(warnings))

            combined_text = "\n".join(
                clean_text(str(merged.get(field) or ""))
                for field in [
                    "医院",
                    "科室_分类页",
                    "职称身份原文",
                    "擅长诊疗方向摘录",
                    "详情正文摘录",
                ]
            )
            groups_found, tags = group_tags(combined_text)
            if merged["异常提示"]:
                groups_found = []
                tags = []
            merged["重点关注范围"] = "、".join(groups_found)
            merged["重点疾病标签"] = "、".join(tags)
            merged["重点优先级"] = "普通"
            if not merged["异常提示"] and (
                any(term in combined_text for term in PRIORITY_DEPARTMENTS) or groups_found
            ):
                merged["重点优先级"] = "高"
            elif not merged["异常提示"] and any(
                term != "医师"
                for term in extract_terms(merged["职称身份原文"], TITLE_TERMS)
            ):
                merged["重点优先级"] = "中"

            primary_source = clean_text(str(merged.get("来源链接") or ""))
            source_links = [
                clean_text(str(member.get("来源链接") or "")) for member in cluster
            ]
            merged_rows.append(merged)
            reconciliation.append(
                {
                    "name": name,
                    "identity_index": identity_index,
                    "resolution": (
                        "同名待甄别"
                        if distinct_same_name
                        else "同一人归并"
                        if len(cluster) > 1
                        else "唯一身份"
                    ),
                    "detail_ids": detail_ids,
                    "primary_source_link": primary_source,
                    "merged_source_links": [
                        source for source in source_links if source != primary_source
                    ],
                    "departments": category_departments,
                    "campuses": campuses,
                    "relation_count": relation_count,
                }
            )

    return merged_rows, reconciliation


def collect_generic(
    target: HospitalTarget,
    today: str,
    max_doctors: int | None = None,
    max_pages: int = GENERIC_MAX_PAGES_DEFAULT,
    gdzy5413_trial2: bool = False,
) -> dict[str, Any]:
    session = create_official_session()
    entry_urls = effective_entry_urls(target)
    if not entry_urls:
        raise RuntimeError("通用模板未提供医生目录入口。")

    categories: list[dict[str, str]] = []
    page_errors: list[dict[str, str]] = []
    raw_rows: list[dict[str, str]] = []
    excluded_candidates_by_link: dict[str, dict[str, str]] = {}
    entry_candidate_links: dict[str, set[str]] = {entry_url: set() for entry_url in entry_urls}
    entry_raw_relation_counts: dict[str, int] = {entry_url: 0 for entry_url in entry_urls}
    accessible_entry_count = 0
    for entry_index, entry_url in enumerate(entry_urls, start=1):
        status, entry_html, entry_error = fetch(session, entry_url)
        if status != 200:
            page_errors.append({"page": "1", "entry_url": entry_url, "url": entry_url, "error": entry_error})
            continue
        accessible_entry_count += 1
        if gdskin_entry_id(entry_url):
            page_documents, postback_errors = discover_gdskin_postback_documents(
                session,
                entry_url,
                entry_html,
                max_pages=max_pages,
            )
            page_errors.extend(postback_errors)
        else:
            page_documents = []
            for page_index, page_url in enumerate(
                discover_generic_list_pages(entry_url, entry_html, max_pages=max_pages),
                start=1,
            ):
                if page_index == 1:
                    html = entry_html
                    page_status = status
                    page_error = ""
                else:
                    page_status, html, page_error = fetch(session, page_url)
                if page_status != 200:
                    page_errors.append(
                        {
                            "page": str(page_index),
                            "entry_url": entry_url,
                            "url": page_url,
                            "error": page_error,
                        }
                    )
                    continue
                page_documents.append({"page": str(page_index), "url": page_url, "html": html})

        for page_document in page_documents:
            page_number = page_document["page"]
            page_url = page_document["url"]
            categories.append(
                {
                    "category_id": f"{entry_index}-{page_number}",
                    "category_name": f"通用模板入口{entry_index}列表第{page_number}页",
                    "url": page_url,
                    "entry_url": entry_url,
                }
            )
            page_rows = discover_generic_detail_links(page_document["html"], page_url, entry_url)
            entry_raw_relation_counts[entry_url] += gdzy5413_raw_detail_relation_count(
                page_document["html"], page_url, entry_url
            )
            for excluded in discover_gdskin_excluded_links(page_document["html"], page_url, entry_url):
                excluded_candidates_by_link.setdefault(
                    generic_detail_identity(excluded["source_link"]),
                    excluded,
                )
            for row in page_rows:
                row["entry_url"] = entry_url
                entry_candidate_links[entry_url].add(generic_detail_identity(row["source_link"]))
            raw_rows.extend(page_rows)
            print(
                f"[entry {entry_index}/{len(entry_urls)} page {page_number}/{len(page_documents)}] "
                f"generic list rows: {len(page_rows)}"
            )
            time.sleep(0.2)

    entry_candidate_counts = {
        entry_url: len(candidate_links)
        for entry_url, candidate_links in entry_candidate_links.items()
    }

    if accessible_entry_count == 0:
        errors = "；".join(error.get("error", "") for error in page_errors if error.get("error"))
        raise RuntimeError(f"全部医生目录入口读取失败：{errors}")

    by_link: dict[str, dict[str, str]] = {}
    for row in raw_rows:
        key = generic_detail_identity(row["source_link"])
        item = by_link.setdefault(
            key,
            {
                "source_link": row["source_link"],
                "name": row.get("name", ""),
                "list_title": row.get("list_title", ""),
                "department": row.get("department", ""),
                "description": row.get("description", ""),
                "list_pages": "",
                "score": row.get("score", ""),
                "entry_url": row.get("entry_url", ""),
                "all_entry_urls": row.get("entry_url", ""),
            },
        )
        for field in ["name", "list_title", "description"]:
            if row.get(field, "") and len(row.get(field, "")) > len(item.get(field, "")):
                item[field] = row[field]
        row_department = row.get("department", "")
        item_department = item.get("department", "")
        if row_department and (
            not item_department
            or (
                item_department in (GDSKIN_GENERAL_CATEGORIES | NY5Y_GENERAL_CATEGORIES)
                and row_department not in (GDSKIN_GENERAL_CATEGORIES | NY5Y_GENERAL_CATEGORIES)
            )
            or len(row_department) > len(item_department)
        ):
            item["department"] = row_department
        if int(row.get("score") or 0) > int(item.get("score") or 0):
            item["score"] = row.get("score", "")
        entry_members = [value for value in item["all_entry_urls"].split("；") if value]
        if row.get("entry_url", "") and row["entry_url"] not in entry_members:
            entry_members.append(row["entry_url"])
            item["all_entry_urls"] = "；".join(entry_members)
        pages = [value for value in item["list_pages"].split("；") if value]
        if row["list_page"] not in pages:
            pages.append(row["list_page"])
            item["list_pages"] = "；".join(pages)

    sampling_items = list(by_link.values())
    sampling_entry_urls = entry_urls
    if gdzy5413_trial2:
        sampling_items = [
            item
            for item in sampling_items
            if gdzy5413_entry_kind(item.get("entry_url")) == "852"
            and gdzy5413_ksdoctor_detail_id(item.get("source_link"))
        ]
        sampling_items = select_gdzy5413_trial2_items(sampling_items, max_doctors)
        sampling_entry_urls = [
            entry_url for entry_url in entry_urls if gdzy5413_entry_kind(entry_url) == "852"
        ]
    elif target.adapter_id == GDZY5413_ADAPTER_ID and max_doctors is None:
        sampling_items = expand_gdzy5413_full_detail_items(sampling_items)
    detail_items = round_robin_generic_items(
        sampling_items,
        sampling_entry_urls,
        None if gdzy5413_trial2 else max_doctors,
        spread=target.adapter_id in {NY5Y_ADAPTER_ID, GDZY5413_ADAPTER_ID},
    )
    if target.adapter_id == GDSKIN_ADAPTER_ID:
        for item in detail_items:
            if item.get("department") in GDSKIN_GENERAL_CATEGORIES:
                # 首席/知名专家是官网荣誉分组，不是院内科室；官网未给出真实科室时保持空白。
                item["department"] = ""
    elif target.adapter_id == NY5Y_ADAPTER_ID:
        for item in detail_items:
            if item.get("department") in NY5Y_GENERAL_CATEGORIES:
                # 专家风采/岭南名医是官网栏目或荣誉分组，不是院内科室。
                item["department"] = ""

    existing_links = collect_existing_profile_links()
    rows: list[dict[str, Any]] = []
    detail_errors: list[dict[str, str]] = []
    for index, item in enumerate(detail_items, start=1):
        link = item["source_link"]
        detail_status, detail_html, detail_error = fetch(session, link)
        detail = {
            "name": item.get("name", ""),
            "department": item.get("department", ""),
            "title_field": item.get("list_title", ""),
            "specialty": "",
            "profile_text": "",
            "title_text": "",
            "department_raw": item.get("department", ""),
            "department_polluted": "no",
            "specialty_raw": "",
            "specialty_navigation_polluted": "no",
            "highlight_navigation_polluted": "no",
        }
        if detail_status == 200:
            detail = (
                parse_gdskin_detail(detail_html, item)
                if gdskin_detail_id(link)
                else parse_ny5y_detail(detail_html, item)
                if ny5y_detail_id(link)
                else parse_gdzy5413_detail(detail_html, item)
                if gdzy5413_detail_id(link)
                else parse_gdzy5413_ksdoctor_detail(detail_html, item)
                if gdzy5413_ksdoctor_detail_id(link)
                else parse_generic_detail(detail_html, item)
            )
        else:
            detail_errors.append({"source_link": link, "error": detail_error})

        name = first_nonempty(detail.get("name"), item.get("name"))
        department = first_nonempty(detail.get("department"), item.get("department"))
        combined_text = "\n".join(
            [
                target.hospital,
                department,
                item.get("list_title", ""),
                item.get("description", ""),
                detail.get("title_field", ""),
                detail.get("specialty", ""),
                detail.get("profile_text", ""),
            ]
        )
        title_source = first_nonempty(detail.get("title_field"), item.get("list_title"))
        title_hits = extract_terms(title_source, TITLE_TERMS) or extract_terms(combined_text, TITLE_TERMS)
        valid_doctor_record, quality_warnings = generic_record_quality(
            name,
            link,
            item.get("entry_url", target.entry_url),
            detail,
            item,
        )
        priority, groups, tags = classify_generic_record(valid_doctor_record, combined_text, title_hits)
        specialty = clip(detail.get("specialty"), 520) if valid_doctor_record else ""
        highlight_source = (
            detail.get("profile_text", "")
            if ny5y_detail_id(link) or gdzy5413_detail_id(link) or gdzy5413_ksdoctor_detail_id(link)
            else combined_text
        )
        highlights = extract_clean_highlights(highlight_source)
        gdzy5413_honor = ""
        if gdzy5413_detail_id(link) and any(
            term in item.get("list_title", "")
            for term in ["国务院特殊津贴", "名中医", "名老中医", "学科带头人", "优秀中医临床人才"]
        ):
            gdzy5413_honor = item.get("list_title", "")

        warnings: list[str] = []
        if detail_status != 200:
            warnings.append("详情页读取失败")
        warnings.extend(quality_warnings)
        if not name:
            warnings.append("姓名需人工复核")
        if not department:
            warnings.append("科室需人工复核")
        if not title_hits:
            warnings.append("职称/身份需人工复核")
        if not detail.get("specialty") and not detail.get("profile_text") and not item.get("description"):
            warnings.append("详情页正文为空或未识别")
        if int(item.get("score") or 0) < 8:
            warnings.append("通用模板低置信度")
        if detail.get("highlight_navigation_polluted") == "yes":
            warnings.append(
                "亮眼经历含导航文本，已清洗" if highlights else "亮眼经历含导航文本，已清空"
            )

        rows.append(
            {
                "序号": index,
                "医院": target.hospital,
                "姓名": name,
                "科室_分类页": department,
                "科室_列表卡片": item.get("department", ""),
                "职称_关键词": "、".join(title_hits),
                "职称身份原文": clip(first_nonempty(detail.get("title_field"), item.get("list_title")), 500),
                "重点优先级": priority,
                "重点关注范围": "、".join(groups),
                "重点疾病标签": "、".join(tags),
                "擅长诊疗方向摘录": specialty,
                "亮眼经历线索": clip(
                    clean_text(
                        " ".join(
                            part
                            for part in [
                                "岭南名医"
                                if any(
                                    ny5y_entry_kind(entry_url) == "162"
                                    for entry_url in item.get("all_entry_urls", "").split("；")
                                    if entry_url
                                )
                                else "",
                                gdzy5413_honor,
                                highlights,
                            ]
                            if part
                        )
                    ),
                    520,
                ),
                "列表简介": clip(item.get("description", ""), 700),
                "详情正文摘录": clip(detail.get("profile_text", ""), 1800),
                "来源类型": "医院官网",
                "来源链接": link,
                "采集入口": item.get("entry_url", target.entry_url),
                "采集方式": (
                    "医院官网 ASP.NET 专家团队：GridView 列表+详情页结构化抽取"
                    if gdskin_detail_id(link)
                    else "医院官网专家栏目：指定入口普查+医生详情 DOM 结构化抽取"
                    if ny5y_detail_id(link)
                    else "医院官网名医栏目：严格 specialist typeid 过滤+详情 DOM 结构化抽取"
                    if gdzy5413_detail_id(link)
                    else "医院官网各科专家栏目：严格 ksdoctorinfo 参数过滤+详情 DOM 结构化抽取"
                    if gdzy5413_ksdoctor_detail_id(link)
                    else "医院官网通用模板：列表页自动发现+详情页文本抽取"
                ),
                "采集日期": today,
                "详情页状态": "200" if detail_status == 200 else "失败",
                "已建画像": "是" if canonical_url(link) in existing_links else "否",
                "异常提示": "；".join(warnings),
                "复核状态": "待人工复核",
            }
        )
        if index % 20 == 0:
            print(f"generic details: {index}/{len(detail_items)}")
        time.sleep(0.18)

    gdzy5413_identity_reconciliation: list[dict[str, Any]] = []
    if target.adapter_id == GDZY5413_ADAPTER_ID:
        rows, gdzy5413_identity_reconciliation = merge_gdzy5413_identity_rows(rows)
    rows.sort(key=lambda row: (row["重点优先级"] != "高", row["科室_分类页"], row["姓名"]))
    for new_index, row in enumerate(rows, start=1):
        row["序号"] = new_index

    category_counter = Counter()
    priority_counter = Counter(row["重点优先级"] for row in rows)
    group_counter = Counter()
    warning_counter = Counter()
    for row in rows:
        if row["科室_分类页"]:
            category_counter[row["科室_分类页"]] += 1
        for group in row["重点关注范围"].split("、"):
            if group:
                group_counter[group] += 1
        for warning in row["异常提示"].split("；"):
            if warning:
                warning_counter[warning] += 1

    sample_entry_urls: list[str] = []
    seen_sample_entries: set[str] = set()
    for row in rows:
        entry_url = clean_text(row.get("采集入口"))
        entry_key = canonical_url(entry_url)
        if entry_url and entry_key not in seen_sample_entries:
            seen_sample_entries.add(entry_key)
            sample_entry_urls.append(entry_url)
    sample_entry_categories = [
        (
            GDSKIN_ENTRY_METADATA[entry_id]["category_name"]
            if (entry_id := gdskin_entry_id(entry_url))
            else NY5Y_ENTRY_METADATA[ny5y_id]["category_name"]
            if (ny5y_id := ny5y_entry_kind(entry_url))
            else GDZY5413_ENTRY_METADATA[gdzy5413_id]["category_name"]
        )
        for entry_url in sample_entry_urls
        if (entry_id := gdskin_entry_id(entry_url))
        or (ny5y_id := ny5y_entry_kind(entry_url))
        or (gdzy5413_id := gdzy5413_entry_kind(entry_url))
    ]
    entry_page_counts = Counter(category.get("entry_url", "") for category in categories)
    gdzy5413_names_by_mode: dict[str, dict[str, dict[str, str]]] = {"851": {}, "852": {}}
    for row in raw_rows:
        entry_kind = gdzy5413_entry_kind(row.get("entry_url"))
        name = re.sub(r"\s+", "", clean_text(row.get("name")))
        if entry_kind in gdzy5413_names_by_mode and name:
            gdzy5413_names_by_mode[entry_kind].setdefault(
                name,
                {"name": name, "source_link": row.get("source_link", ""), "department": row.get("department", "")},
            )
    gdzy5413_cross_mode_matches = [
        {
            "name": name,
            "specialist_source_link": gdzy5413_names_by_mode["851"][name]["source_link"],
            "specialist_department": gdzy5413_names_by_mode["851"][name]["department"],
            "ksdoctor_source_link": gdzy5413_names_by_mode["852"][name]["source_link"],
            "ksdoctor_department": gdzy5413_names_by_mode["852"][name]["department"],
        }
        for name in sorted(set(gdzy5413_names_by_mode["851"]) & set(gdzy5413_names_by_mode["852"]))
    ]
    entry_reconnaissance: list[dict[str, Any]] = []
    for entry_url in entry_urls:
        entry_id = gdskin_entry_id(entry_url)
        ny5y_id = ny5y_entry_kind(entry_url)
        candidate_count = entry_candidate_counts.get(entry_url, 0)
        metadata = (
            GDSKIN_ENTRY_METADATA.get(entry_id, {})
            if entry_id
            else NY5Y_ENTRY_METADATA.get(ny5y_id, {})
            if ny5y_id
            else GDZY5413_ENTRY_METADATA.get(gdzy5413_entry_kind(entry_url), {})
        )
        gdzy5413_id = gdzy5413_entry_kind(entry_url)
        out_of_scope_count = sum(
            1
            for item in excluded_candidates_by_link.values()
            if item.get("entry_url") == entry_url and "另一详情模板" in item.get("reason", "")
        )
        list_page_count = entry_page_counts.get(entry_url, 0)
        entry_reconnaissance.append(
            {
                "entry_url": entry_url,
                "category_name": metadata.get("category_name", f"入口 {entry_id or entry_url}"),
                "page_nature": (
                    "医院官网 ASP.NET GridView 专家列表"
                    if entry_id and candidate_count
                    else "医院官网单页专家名单"
                    if ny5y_id and candidate_count
                    else "医院官网名医名家单页名单"
                    if gdzy5413_id == "851"
                    else "医院官网各科专家单页主目录（ksdoctorinfo 已获 TRIAL-2 授权）"
                    if gdzy5413_id == "852"
                    else "医院官网专家团队分类页（当前未列出可采医生详情）"
                ),
                "raw_detail_relation_count": entry_raw_relation_counts.get(entry_url, 0),
                "unique_detail_count": candidate_count,
                "out_of_scope_detail_count": out_of_scope_count,
                "list_page_count": list_page_count,
                "pagination_evidence": "未发现分页证据" if list_page_count == 1 else f"发现 {list_page_count} 个列表文档",
                "affiliation": metadata.get("affiliation", target.hospital),
                "independent_entity_check": (
                    "未发现独立院区归属；页面保持在本院官网同站专家栏目"
                    if ny5y_id
                    else "owner 已裁决院区/门诊均属同一法人实体授权范围"
                    if gdzy5413_id
                    else "未发现独立挂牌机构；页面保持在医院官网同站专家团队栏目"
                ),
            }
        )

    cross_entry_duplicates = [
        {
            "source_link": item["source_link"],
            "name": item.get("name", ""),
            "entry_urls": item.get("all_entry_urls", "").split("；"),
        }
        for item in by_link.values()
        if len([value for value in item.get("all_entry_urls", "").split("；") if value]) > 1
    ]
    candidate_membership_count = sum(entry_candidate_counts.values())
    out_of_scope_candidate_count = sum(
        1
        for item in excluded_candidates_by_link.values()
        if "另一详情模板" in item.get("reason", "")
    )

    return {
        "meta": {
            "city": target.city,
            "hospital": target.hospital,
            "homepage": target.homepage,
            "entry_url": " ".join(entry_urls),
            "entry_url_count": len(entry_urls),
            "entry_candidate_counts": entry_candidate_counts,
            "entry_raw_relation_counts": entry_raw_relation_counts,
            "entry_url_source": "Claude owner Issue 显式多入口" if target.entry_urls else "官网入口台账",
            "ledger_entry_url": target.ledger_entry_url or target.entry_url,
            "adapter_id": target.adapter_id,
            "collected_at": today,
            "category_count": len(categories),
            "raw_card_rows": len(raw_rows),
            "candidate_membership_count": candidate_membership_count,
            "unique_candidate_count": len(by_link),
            "cross_entry_duplicate_count": candidate_membership_count - len(by_link),
            "unique_doctor_count": len(rows),
            "category_error_count": len(page_errors),
            "detail_error_count": len(detail_errors),
            "existing_profile_count": sum(1 for row in rows if row["已建画像"] == "是"),
            "ledger_review": target.review,
            "ledger_difficulty": target.difficulty,
            "generic_template": "yes" if target.adapter_id == GENERIC_ADAPTER_ID else "no",
            "site_adapter_profile": (
                "gdskin_aspnet_gridview" if target.adapter_id == GDSKIN_ADAPTER_ID else "generic"
                if target.adapter_id not in {NY5Y_ADAPTER_ID, GDZY5413_ADAPTER_ID}
                else "ny5y_strict_expert_dom"
                if target.adapter_id == NY5Y_ADAPTER_ID
                else "gdzy5413_strict_specialist_dom"
            ),
            "generic_max_pages": max_pages,
            "sample_entry_coverage_count": len(sample_entry_urls),
            "sample_entry_categories": sample_entry_categories,
            "excluded_non_doctor_count": len(excluded_candidates_by_link) - out_of_scope_candidate_count,
            "out_of_scope_candidate_count": out_of_scope_candidate_count,
            "gdzy5413_trial2": gdzy5413_trial2,
            "gdzy5413_851_unique_name_count": len(gdzy5413_names_by_mode["851"]),
            "gdzy5413_852_unique_name_count": len(gdzy5413_names_by_mode["852"]),
            "gdzy5413_cross_mode_name_match_count": len(gdzy5413_cross_mode_matches),
            "gdzy5413_detail_relation_count": len(detail_items),
            "gdzy5413_final_identity_count": len(rows),
            "gdzy5413_trial2_sample_relation_count": len(detail_items),
            "gdzy5413_trial2_sample_identity_count": len(rows),
            "gdzy5413_trial2_baiyun_sample_count": sum(
                1 for row in rows if "白云院区" in clean_text(row.get("科室_分类页"))
            ),
            "gdzy5413_trial2_merged_identity_count": sum(
                1
                for item in gdzy5413_identity_reconciliation
                if item.get("resolution") == "同一人归并" and int(item.get("relation_count") or 0) > 1
            ),
        },
        "categories": categories,
        "entry_reconnaissance": entry_reconnaissance,
        "cross_entry_duplicates": cross_entry_duplicates,
        "gdzy5413_cross_mode_matches": gdzy5413_cross_mode_matches,
        "gdzy5413_identity_reconciliation": gdzy5413_identity_reconciliation,
        "excluded_candidates": list(excluded_candidates_by_link.values()),
        "category_errors": page_errors,
        "detail_errors": detail_errors,
        "category_counts": category_counter.most_common(),
        "priority_counts": dict(priority_counter),
        "group_counts": dict(group_counter),
        "warning_counts": dict(warning_counter),
        "rows": rows,
    }


def validate_gdskin_full_append(payload: dict[str, Any]) -> None:
    """Fail before master writes when the owner-audited GDSKIN census has drifted."""

    errors: list[str] = []
    meta = payload.get("meta", {})
    expected_total = sum(GDSKIN_EXPECTED_ENTRY_COUNTS.values())
    if int(meta.get("unique_candidate_count") or 0) != expected_total:
        errors.append(
            f"唯一候选预期 {expected_total}，实际 {meta.get('unique_candidate_count', 0)}"
        )
    if int(meta.get("unique_doctor_count") or 0) != expected_total:
        errors.append(f"医生记录预期 {expected_total}，实际 {meta.get('unique_doctor_count', 0)}")
    if int(meta.get("candidate_membership_count") or 0) != expected_total:
        errors.append(
            f"入口候选关系预期 {expected_total}，实际 {meta.get('candidate_membership_count', 0)}"
        )
    if int(meta.get("cross_entry_duplicate_count") or 0) != 0:
        errors.append(f"跨入口重复预期 0，实际 {meta.get('cross_entry_duplicate_count', 0)}")
    if int(meta.get("category_error_count") or 0) != 0:
        errors.append(f"列表读取失败 {meta.get('category_error_count', 0)} 条")
    if int(meta.get("detail_error_count") or 0) != 0:
        errors.append(f"详情读取失败 {meta.get('detail_error_count', 0)} 条")
    if int(meta.get("excluded_non_doctor_count") or 0) != 1:
        errors.append(f"非医生排除预期 1，实际 {meta.get('excluded_non_doctor_count', 0)}")

    reconnaissance_by_id = {
        entry_id: item
        for item in payload.get("entry_reconnaissance", [])
        if (entry_id := gdskin_entry_id(str(item.get("entry_url") or "")))
    }
    for entry_id, expected_count in GDSKIN_EXPECTED_ENTRY_COUNTS.items():
        item = reconnaissance_by_id.get(entry_id)
        if item is None:
            errors.append(f"入口 {entry_id} 缺少普查记录")
            continue
        actual_count = int(item.get("unique_detail_count") or 0)
        if actual_count != expected_count:
            errors.append(f"入口 {entry_id} 唯一详情预期 {expected_count}，实际 {actual_count}")
        expected_pages = GDSKIN_EXPECTED_PAGE_COUNTS[entry_id]
        actual_pages = int(item.get("list_page_count") or 0)
        if actual_pages != expected_pages:
            errors.append(f"入口 {entry_id} 列表页预期 {expected_pages}，实际 {actual_pages}")

    general_department_rows = [
        str(row.get("姓名") or "未命名")
        for row in payload.get("rows", [])
        if str(row.get("科室_分类页") or "") in GDSKIN_GENERAL_CATEGORIES
    ]
    if general_department_rows:
        errors.append(
            "科室仍为首席/知名专家类目：" + "、".join(general_department_rows)
        )

    if errors:
        raise RuntimeError("GDSKIN 全量写入前门禁失败：" + "；".join(errors))


def validate_gdzy5413_trial2(payload: dict[str, Any], expected_identities: int = 10) -> None:
    """Protect the owner-audited TRIAL-2 census, sample composition, and URL scope."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    errors: list[str] = []
    entry_counts = meta.get("entry_candidate_counts", {})
    if sorted(int(value) for value in entry_counts.values()) != [21, 346]:
        errors.append(f"851/852 唯一详情应为 21/346，实际 {entry_counts}")
    if int(meta.get("gdzy5413_852_unique_name_count") or 0) != 289:
        errors.append(f"852 唯一姓名应为 289，实际 {meta.get('gdzy5413_852_unique_name_count')}")
    if int(meta.get("gdzy5413_cross_mode_name_match_count") or 0) != 20:
        errors.append(
            "851/852 同名应为 20，实际 "
            f"{meta.get('gdzy5413_cross_mode_name_match_count')}"
        )
    if len(rows) != expected_identities:
        errors.append(f"最终身份应为 {expected_identities}，实际 {len(rows)}")
    if int(meta.get("gdzy5413_trial2_baiyun_sample_count") or 0) < 1:
        errors.append("样本缺少白云院区条目")
    if int(meta.get("gdzy5413_trial2_merged_identity_count") or 0) < 1:
        errors.append("样本缺少多链接同一人归并案例")
    if int(meta.get("gdzy5413_trial2_sample_relation_count") or 0) <= len(rows):
        errors.append("样本详情关系数未体现多链接归并")
    if int(meta.get("department_coverage_count") or 0) < 3:
        errors.append("样本科室覆盖不足 3 个")
    if int(meta.get("category_error_count") or 0) or int(meta.get("detail_error_count") or 0):
        errors.append("列表或详情存在读取失败")
    relation_links = [
        link
        for item in payload.get("gdzy5413_identity_reconciliation", [])
        for link in [item.get("primary_source_link"), *item.get("merged_source_links", [])]
        if link
    ]
    if len(relation_links) != int(meta.get("gdzy5413_trial2_sample_relation_count") or 0):
        errors.append("归并对账中的详情关系数与抽样关系数不一致")
    if any(not gdzy5413_ksdoctor_detail_id(link) for link in relation_links):
        errors.append("样本归并对账含非授权 ksdoctorinfo 来源")
    if errors:
        raise RuntimeError("GDZY5413 TRIAL-2 门禁失败：" + "；".join(errors))


def validate_gdzy5413_full_append(payload: dict[str, Any]) -> None:
    """Fail before master writes when the owner-audited FULL census or identity audit drifts."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    reconciliation = payload.get("gdzy5413_identity_reconciliation", [])
    errors: list[str] = []
    entry_counts = {
        gdzy5413_entry_kind(str(entry_url)): int(count)
        for entry_url, count in meta.get("entry_candidate_counts", {}).items()
        if gdzy5413_entry_kind(str(entry_url))
    }
    if entry_counts != {"851": 21, "852": 346}:
        errors.append(f"851/852 唯一详情应为 21/346，实际 {entry_counts}")
    if int(meta.get("gdzy5413_851_unique_name_count") or 0) != 21:
        errors.append(f"851 唯一姓名应为 21，实际 {meta.get('gdzy5413_851_unique_name_count')}")
    if int(meta.get("gdzy5413_852_unique_name_count") or 0) != 289:
        errors.append(f"852 唯一姓名应为 289，实际 {meta.get('gdzy5413_852_unique_name_count')}")
    if int(meta.get("gdzy5413_cross_mode_name_match_count") or 0) != 20:
        errors.append(
            "851/852 同名应为 20，实际 "
            f"{meta.get('gdzy5413_cross_mode_name_match_count')}"
        )
    if int(meta.get("category_error_count") or 0) or int(meta.get("detail_error_count") or 0):
        errors.append("列表或详情存在读取失败")
    if not 290 <= len(rows) <= 367:
        errors.append(f"最终身份数应在姓名并集 290 与详情关系 367 之间，实际 {len(rows)}")
    if len(reconciliation) != len(rows):
        errors.append(f"归并对账行数应等于最终身份数，实际 {len(reconciliation)}/{len(rows)}")

    relation_links = [
        str(link)
        for item in reconciliation
        for link in [item.get("primary_source_link"), *item.get("merged_source_links", [])]
        if link
    ]
    relation_count = sum(int(item.get("relation_count") or 0) for item in reconciliation)
    if len(relation_links) != 367 or relation_count != 367:
        errors.append(f"归并详情关系应为 367，实际链接 {len(relation_links)} / 对账 {relation_count}")
    if len(set(relation_links)) != len(relation_links):
        errors.append("归并对账存在重复详情链接")
    specialist_links = [link for link in relation_links if gdzy5413_detail_id(link)]
    ksdoctor_links = [link for link in relation_links if gdzy5413_ksdoctor_detail_id(link)]
    if len(specialist_links) != 21 or len(ksdoctor_links) != 346:
        errors.append(
            f"归并对账授权链接应为 specialist 21 / ksdoctorinfo 346，实际 "
            f"{len(specialist_links)} / {len(ksdoctor_links)}"
        )
    if len(specialist_links) + len(ksdoctor_links) != len(relation_links):
        errors.append("归并对账含非授权详情链接")

    for row in rows:
        if clean_text(row.get("医院")) != "广东省第二中医院":
            errors.append("存在医院字段未统一为广东省第二中医院")
            break
        source_link = str(row.get("来源链接") or "")
        if not (gdzy5413_detail_id(source_link) or gdzy5413_ksdoctor_detail_id(source_link)):
            errors.append("最终行含非授权主详情链接")
            break
        expected_titles = "、".join(extract_terms(clean_text(row.get("职称身份原文")), TITLE_TERMS))
        if clean_text(row.get("职称_关键词")) != expected_titles:
            errors.append(f"{row.get('姓名', '未命名')} 的职称关键词未严格取自主详情")
            break
        if re.match(r"^\s*擅长\s*[:：]", str(row.get("擅长诊疗方向摘录") or "")):
            errors.append(f"{row.get('姓名', '未命名')} 的擅长字段仍保留前缀")
            break

    if errors:
        raise RuntimeError("GDZY5413 FULL 写入前门禁失败：" + "；".join(errors))


def validate_ny5y_full_append(payload: dict[str, Any]) -> None:
    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    errors: list[str] = []

    if int(meta.get("unique_doctor_count") or 0) != NY5Y_EXPECTED_UNIQUE_COUNT:
        errors.append(
            f"医生记录预期 {NY5Y_EXPECTED_UNIQUE_COUNT}，实际 {meta.get('unique_doctor_count', 0)}"
        )
    if len(rows) != NY5Y_EXPECTED_UNIQUE_COUNT:
        errors.append(f"结果行数预期 {NY5Y_EXPECTED_UNIQUE_COUNT}，实际 {len(rows)}")
    if int(meta.get("unique_candidate_count") or 0) != NY5Y_EXPECTED_UNIQUE_COUNT:
        errors.append(
            f"去重候选预期 {NY5Y_EXPECTED_UNIQUE_COUNT}，实际 {meta.get('unique_candidate_count', 0)}"
        )
    expected_memberships = sum(NY5Y_EXPECTED_ENTRY_COUNTS.values())
    if int(meta.get("candidate_membership_count") or 0) != expected_memberships:
        errors.append(
            f"入口候选关系预期 {expected_memberships}，实际 {meta.get('candidate_membership_count', 0)}"
        )
    if int(meta.get("cross_entry_duplicate_count") or 0) != NY5Y_EXPECTED_CROSS_ENTRY_DUPLICATES:
        errors.append(
            "跨入口重复预期 "
            f"{NY5Y_EXPECTED_CROSS_ENTRY_DUPLICATES}，实际 {meta.get('cross_entry_duplicate_count', 0)}"
        )
    if int(meta.get("category_error_count") or 0) != 0:
        errors.append(f"列表读取失败 {meta.get('category_error_count', 0)} 条")
    if int(meta.get("detail_error_count") or 0) != 0:
        errors.append(f"详情读取失败 {meta.get('detail_error_count', 0)} 条")
    if int(meta.get("excluded_non_doctor_count") or 0) != 0:
        errors.append(f"非医生排除预期 0，实际 {meta.get('excluded_non_doctor_count', 0)}")

    actual_entry_counts: dict[str, int] = {}
    for entry_url, count in meta.get("entry_candidate_counts", {}).items():
        if entry_id := ny5y_entry_kind(str(entry_url)):
            actual_entry_counts[entry_id] = int(count)
    for entry_id, expected_count in NY5Y_EXPECTED_ENTRY_COUNTS.items():
        actual_count = actual_entry_counts.get(entry_id)
        if actual_count != expected_count:
            errors.append(f"入口 {entry_id} 唯一详情预期 {expected_count}，实际 {actual_count}")

    source_ids = [ny5y_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not source_id for source_id in source_ids):
        errors.append("存在非授权 yisheng_xq.php?id=<数字> 来源")
    if len({source_id for source_id in source_ids if source_id}) != len(rows):
        errors.append("结果中存在重复来源详情 ID")

    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(r"^\s*擅长\s*[:：]", str(row.get("擅长诊疗方向摘录") or ""))
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))

    huang_rows = [row for row in rows if clean_text(str(row.get("姓名") or "")) == "黄艺洪"]
    if len(huang_rows) != 1:
        errors.append(f"黄艺洪记录预期 1，实际 {len(huang_rows)}")
    else:
        huang = huang_rows[0]
        if clean_text(str(huang.get("科室_分类页") or "")):
            errors.append("黄艺洪真实科室应保持空白")
        if "科室需人工复核" not in str(huang.get("异常提示") or ""):
            errors.append("黄艺洪缺少科室需人工复核标记")
        if "岭南名医" not in " ".join(
            [str(huang.get("职称身份原文") or ""), str(huang.get("亮眼经历线索") or "")]
        ):
            errors.append("黄艺洪缺少岭南名医官方荣誉证据")

    if errors:
        raise RuntimeError("NY5Y 全量写入前门禁失败：" + "；".join(errors))


def validate_gykqyy_full_append(payload: dict[str, Any]) -> None:
    meta = payload["meta"]
    rows = payload["rows"]
    reconciliation = payload.get("gykqyy_identity_reconciliation", [])
    errors: list[str] = []
    if int(meta.get("candidate_membership_count") or 0) != 317:
        errors.append(f"医生-科室关系应为 317，实际 {meta.get('candidate_membership_count', 0)}")
    if int(meta.get("census_unique_detail_count") or 0) != 297:
        errors.append(f"唯一详情 ID 应为 297，实际 {meta.get('census_unique_detail_count', 0)}")
    named_count = int(meta.get("census_named_detail_count") or 0)
    blank_count = int(meta.get("census_blank_name_detail_count") or 0)
    if named_count + blank_count != 297:
        errors.append("有姓名与空姓名详情 ID 之和不等于 297")
    if int(meta.get("detail_error_count") or 0) != 0:
        errors.append(f"详情接口失败应为 0，实际 {meta.get('detail_error_count', 0)}")
    if len(rows) != named_count or int(meta.get("gykqyy_final_row_count") or 0) != named_count:
        errors.append(f"正式行应等于有姓名详情 ID 数 {named_count}，实际 {len(rows)}")
    if len(reconciliation) != len(rows):
        errors.append(f"逐 ID 对账行应等于正式行，实际 {len(reconciliation)}/{len(rows)}")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")
    if len({canonical_url(str(row.get("来源链接") or "")) for row in rows}) != len(rows):
        errors.append("正式行来源链接不唯一")
    expected_same_name_groups = {"方颖": ["128", "307"], "赵稚宁": ["29", "323"]}
    if meta.get("census_same_name_groups") != expected_same_name_groups:
        errors.append(f"同名不同 ID 组不符合审计基线：{meta.get('census_same_name_groups', {})}")
    if int(meta.get("gykqyy_same_name_separate_row_count") or 0) != 4:
        errors.append(
            f"方颖/赵稚宁应共保留 4 行，实际 {meta.get('gykqyy_same_name_separate_row_count', 0)}"
        )
    same_name_rows = [row for row in rows if clean_text(str(row.get("姓名") or "")) in expected_same_name_groups]
    if len(same_name_rows) != 4 or any(
        "同名待甄别" not in str(row.get("异常提示") or "") for row in same_name_rows
    ):
        errors.append("同名不同 ID 行未全部保留“同名待甄别”")
    blank_excluded_count = sum(
        1
        for item in payload.get("excluded_candidates", [])
        if "核心追溯字段缺失" in str(item.get("reason") or "")
    )
    if blank_excluded_count != blank_count:
        errors.append(f"空姓名详情应全部进入排除对账，实际 {blank_excluded_count}/{blank_count}")
    if errors:
        raise RuntimeError("GYKQYY FULL 写入前门禁失败：" + "；".join(errors))


def validate_gyfyyy_full_append(payload: dict[str, Any]) -> None:
    meta = payload["meta"]
    rows = payload["rows"]
    excluded = payload.get("excluded_candidates", [])
    detail_reconciliation = payload.get("gyfyyy_detail_reconciliation", [])
    identity_reconciliation = payload.get("gyfyyy_identity_reconciliation", [])
    errors: list[str] = []
    expected = {
        "category_count": 59,
        "candidate_membership_count": 650,
        "census_unique_detail_count": 646,
        "cross_entry_duplicate_count": 4,
        "gyfyyy_cross_department_identity_count": 4,
        "excluded_non_doctor_count": 9,
        "category_error_count": 0,
        "detail_error_count": 0,
    }
    for field, expected_value in expected.items():
        actual = int(meta.get(field) or 0)
        if actual != expected_value:
            errors.append(f"{field} 应为 {expected_value}，实际 {actual}")
    expected_detail_rows = 646 - 9
    expected_final_identities = 616
    if len(detail_reconciliation) != expected_detail_rows:
        errors.append(
            f"合规详情 ID 对账应为 {expected_detail_rows} 行，实际 {len(detail_reconciliation)}"
        )
    if (
        len(rows) != expected_final_identities
        or int(meta.get("unique_doctor_count") or 0) != expected_final_identities
        or int(meta.get("gyfyyy_final_identity_count") or 0) != expected_final_identities
    ):
        errors.append(f"最终身份应为 {expected_final_identities}，实际 {len(rows)}")
    if len(identity_reconciliation) != len(rows):
        errors.append(
            f"身份聚类对账应为 {len(rows)} 行，实际 {len(identity_reconciliation)}"
        )
    nursing_exclusions = [
        item for item in excluded if "仅标注护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != 9:
        errors.append(f"纯护理身份排除应为 9，实际 {len(nursing_exclusions)}")
    primary_source_ids = [gyfyyy_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not source_id for source_id in primary_source_ids):
        errors.append("存在非授权 gyfyyy.cn doctor_<数字ID>.html 来源")
    if len(set(primary_source_ids)) != len(rows):
        errors.append("最终身份主详情存在重复 doctor ID")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")
    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(r"^\s*擅长\s*[:：]?", str(row.get("擅长诊疗方向摘录") or ""))
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))
    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))
    detail_ids = {
        str(item.get("detail_id") or "") for item in detail_reconciliation
        if str(item.get("detail_id") or "")
    }
    mapped_ids = {
        str(detail_id)
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if str(detail_id)
    }
    if len(detail_ids) != expected_detail_rows or mapped_ids != detail_ids:
        errors.append(
            f"身份聚类未完整映射合规详情 ID：对账 {len(detail_ids)} / 映射 {len(mapped_ids)}"
        )
    cross_ids = {
        str(item.get("detail_id") or "")
        for item in detail_reconciliation
        if int(item.get("relation_count") or 0) > 1
    }
    if cross_ids != {"101", "549", "607", "618"}:
        errors.append(f"跨科室详情 ID 不符合审计基线：{sorted(cross_ids)}")
    merged_groups = [
        item for item in identity_reconciliation if int(item.get("relation_count") or 0) > 1
    ]
    if len(merged_groups) != 21 or int(meta.get("gyfyyy_same_identity_merge_group_count") or 0) != 21:
        errors.append(f"同一人归并组应为 21，实际 {len(merged_groups)}")
    identity_name_counts = Counter(
        clean_text(str(item.get("name") or "")) for item in identity_reconciliation
    )
    distinct_names = {name for name, count in identity_name_counts.items() if name and count > 1}
    distinct_rows = sum(identity_name_counts[name] for name in distinct_names)
    if (
        len(distinct_names) != 4
        or distinct_rows != 8
        or int(meta.get("gyfyyy_distinct_same_name_group_count") or 0) != 4
        or int(meta.get("gyfyyy_distinct_same_name_row_count") or 0) != 8
    ):
        errors.append(f"实质不同同名身份应为 4 组 8 行，实际 {len(distinct_names)} 组 {distinct_rows} 行")
    distinct_sources = {
        clean_text(str(item.get("primary_source_link") or ""))
        for item in identity_reconciliation
        if clean_text(str(item.get("name") or "")) in distinct_names
    }
    distinct_rows_by_source = {
        clean_text(str(row.get("来源链接") or "")): row for row in rows
    }
    if any(
        "同名待甄别" not in str(distinct_rows_by_source.get(source, {}).get("异常提示") or "")
        for source in distinct_sources
    ):
        errors.append("实质不同的同名身份未全部保留“同名待甄别”")
    if int(meta.get("census_same_name_group_count") or 0) != 25:
        errors.append(
            f"同名详情组应为 25，实际 {meta.get('census_same_name_group_count', 0)}"
        )
    if errors:
        raise RuntimeError("GYFYYY FULL 写入前门禁失败：" + "；".join(errors))


def validate_gy3y_full_append(payload: dict[str, Any]) -> None:
    """Block the GY3Y master write unless the owner-audited census fully reconciles."""

    meta = payload["meta"]
    rows = payload["rows"]
    excluded = payload.get("excluded_candidates", [])
    detail_reconciliation = payload.get("gy3y_detail_reconciliation", [])
    identity_reconciliation = payload.get("gy3y_identity_reconciliation", [])
    errors: list[str] = []
    expected = {
        "category_count": 104,
        "candidate_membership_count": 580,
        "census_unique_detail_count": 438,
        "cross_entry_duplicate_count": 142,
        "gy3y_multi_relation_identity_count": 126,
        "gy3y_cross_campus_identity_count": 117,
        "census_nonempty_department_count": 99,
        "census_empty_department_count": 5,
        "category_error_count": 0,
        "detail_error_count": 0,
    }
    for field, expected_value in expected.items():
        actual = int(meta.get(field) or 0)
        if actual != expected_value:
            errors.append(f"{field} 应为 {expected_value}，实际 {actual}")

    if meta.get("campus_relation_counts") != {"荔湾院区": 390, "黄埔院区": 190}:
        errors.append(f"两院区关系数不符合审计基线：{meta.get('campus_relation_counts', {})}")

    nursing_exclusions = [
        item for item in excluded if "仅标注护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != len(excluded):
        errors.append("存在未按纯护理身份规则留痕的排除候选")
    if int(meta.get("excluded_non_doctor_count") or 0) != len(nursing_exclusions):
        errors.append(
            "护理排除计数与 meta 不一致："
            f"{len(nursing_exclusions)}/{meta.get('excluded_non_doctor_count', 0)}"
        )

    reconciled_ids = {
        str(item.get("detail_id") or "")
        for item in detail_reconciliation
        if str(item.get("detail_id") or "")
    }
    excluded_ids = {
        gy3y_detail_id(str(item.get("source_link") or ""))
        for item in nursing_exclusions
        if gy3y_detail_id(str(item.get("source_link") or ""))
    }
    if len(reconciled_ids) != len(detail_reconciliation):
        errors.append("逐 ID 对账存在空或重复详情 ID")
    if reconciled_ids & excluded_ids:
        errors.append("正式详情与护理排除详情 ID 重叠")
    if len(reconciled_ids | excluded_ids) != 438:
        errors.append(
            f"逐 ID 对账未覆盖 438 个唯一详情：正式 {len(reconciled_ids)} / "
            f"护理排除 {len(excluded_ids)}"
        )

    mapped_ids = {
        str(detail_id)
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if str(detail_id)
    }
    if mapped_ids != reconciled_ids:
        errors.append(
            f"身份聚类未完整映射正式详情 ID：对账 {len(reconciled_ids)} / 映射 {len(mapped_ids)}"
        )
    if len(identity_reconciliation) != len(rows):
        errors.append(
            f"身份聚类对账应与最终正式行一致：{len(identity_reconciliation)}/{len(rows)}"
        )
    if (
        int(meta.get("unique_doctor_count") or 0) != len(rows)
        or int(meta.get("gy3y_final_identity_count") or 0) != len(rows)
    ):
        errors.append(f"最终身份计数与正式行不一致：{len(rows)}")

    primary_ids = [gy3y_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not detail_id for detail_id in primary_ids):
        errors.append("存在非授权 gy3y.cn doctor_<数字ID>.html 来源")
    if len(set(primary_ids)) != len(rows):
        errors.append("最终身份主详情存在重复 doctor ID")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")

    invalid_departments = [
        str(item.get("name") or item.get("detail_id") or "未命名")
        for item in detail_reconciliation
        if any(
            clean_text(str(department))
            and not clean_text(str(department)).startswith(("荔湾院区", "黄埔院区"))
            for department in item.get("departments", [])
        )
    ]
    if invalid_departments:
        errors.append("科室字段丢失院区前缀：" + "、".join(invalid_departments[:10]))

    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(r"^\s*擅长\s*[:：]?", str(row.get("擅长诊疗方向摘录") or ""))
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))

    schedule_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            strip_gyfyyy_schedule_text(str(row.get(field) or ""))
            != clean_text(str(row.get(field) or ""))
            for field in schedule_fields
        )
    ]
    if schedule_rows:
        errors.append("四正式文本字段仍含排班片段：" + "、".join(schedule_rows[:10]))

    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))

    if errors:
        raise RuntimeError("GY3Y FULL 写入前门禁失败：" + "；".join(errors))


def validate_gzbrain_full_append(payload: dict[str, Any]) -> None:
    """Block the GZBRAIN master write unless all 183 audited detail IDs reconcile."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    reconciliation = payload.get("gzbrain_detail_reconciliation", [])
    errors: list[str] = []
    expected = {
        "category_count": 31,
        "pagination_count": 31,
        "candidate_membership_count": 183,
        "unique_candidate_count": 183,
        "census_unique_detail_count": 183,
        "census_named_detail_count": 183,
        "census_blank_name_detail_count": 0,
        "census_unique_nonblank_name_count": 181,
        "census_same_name_group_count": 2,
        "category_error_count": 0,
        "schedule_field_ingested_count": 0,
    }
    for field, expected_value in expected.items():
        actual = int(meta.get(field) or 0)
        if actual != expected_value:
            errors.append(f"{field} 应为 {expected_value}，实际 {actual}")

    if int(meta.get("unique_doctor_count") or 0) != len(rows):
        errors.append(
            f"最终正式行计数不一致：meta {meta.get('unique_doctor_count', 0)} / rows {len(rows)}"
        )

    nursing_exclusions = [
        item for item in excluded if "仅标注护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != len(excluded):
        errors.append("存在未按明确纯护理身份规则留痕的排除候选")
    if int(meta.get("excluded_non_doctor_count") or 0) != len(nursing_exclusions):
        errors.append(
            "护理排除计数与 meta 不一致："
            f"{len(nursing_exclusions)}/{meta.get('excluded_non_doctor_count', 0)}"
        )

    formal_ids = [gzbrain_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not detail_id for detail_id in formal_ids):
        errors.append("正式行存在非授权 /myzj/info_itemid_<数字>.html 来源")
    if len({detail_id for detail_id in formal_ids if detail_id}) != len(rows):
        errors.append("正式行来源详情 ID 不唯一")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")

    excluded_ids = [
        gzbrain_detail_id(str(item.get("source_link") or "")) for item in nursing_exclusions
    ]
    if any(not detail_id for detail_id in excluded_ids):
        errors.append("护理排除含非授权详情来源")
    formal_id_set = {detail_id for detail_id in formal_ids if detail_id}
    excluded_id_set = {detail_id for detail_id in excluded_ids if detail_id}
    if len(excluded_id_set) != len(excluded_ids):
        errors.append("护理排除存在重复详情 ID")
    if formal_id_set & excluded_id_set:
        errors.append("正式行与护理排除详情 ID 重叠")
    if len(formal_id_set | excluded_id_set) != 183:
        errors.append(
            f"逐 ID 对账未覆盖 183 个唯一详情：正式 {len(formal_id_set)} / "
            f"护理排除 {len(excluded_id_set)}"
        )

    reconciliation_ids = [
        str(item.get("detail_id") or "") for item in reconciliation
    ]
    if (
        len(reconciliation_ids) != 183
        or any(not detail_id for detail_id in reconciliation_ids)
        or len(set(reconciliation_ids)) != 183
        or set(reconciliation_ids) != formal_id_set | excluded_id_set
    ):
        errors.append(
            f"逐 ID 对账工件不完整或重复：{len(reconciliation_ids)} 行 / "
            f"{len(set(reconciliation_ids) - {''})} 个非空唯一 ID"
        )
    reconciliation_resolution = {
        str(item.get("detail_id") or ""): str(item.get("resolution") or "")
        for item in reconciliation
    }
    if any(reconciliation_resolution.get(detail_id) != "正式行" for detail_id in formal_id_set):
        errors.append("逐 ID 对账中的正式行裁决与输出不一致")
    if any(reconciliation_resolution.get(detail_id) != "护理排除" for detail_id in excluded_id_set):
        errors.append("逐 ID 对账中的护理排除裁决与排除清单不一致")

    expected_same_name_groups = {
        "沈峰": ["551", "102037"],
        "王丹逢": ["990", "1231"],
    }
    if meta.get("census_same_name_groups") != expected_same_name_groups:
        errors.append(f"同名不同 ID 组不符合审计基线：{meta.get('census_same_name_groups', {})}")
    same_name_rows = [
        row for row in rows if clean_text(str(row.get("姓名") or "")) in expected_same_name_groups
    ]
    if len(same_name_rows) != 4 or any(
        "同名待甄别" not in str(row.get("异常提示") or "") for row in same_name_rows
    ):
        errors.append("同名不同 ID 行未全部保留“同名待甄别”")

    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(
            r"^\s*(?:(?:擅长|专长)\s*[:：]?\s*)+",
            str(row.get("擅长诊疗方向摘录") or ""),
        )
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            strip_gyfyyy_schedule_text(str(row.get(field) or ""))
            != clean_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if schedule_rows:
        errors.append("四正式文本字段仍含排班片段：" + "、".join(schedule_rows[:10]))
    patient_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            contains_gzbrain_patient_case_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if patient_rows:
        errors.append("四正式文本字段仍含患者案例或可识别信息：" + "、".join(patient_rows[:10]))

    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))

    if errors:
        raise RuntimeError("GZBRAIN FULL 写入前门禁失败：" + "；".join(errors))


def validate_gzszyy_full_append(payload: dict[str, Any]) -> None:
    """Block the GZSZYY write unless all 423 audited IDs and four name decisions reconcile."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    detail_reconciliation = payload.get("gzszyy_detail_reconciliation", [])
    identity_reconciliation = payload.get("gzszyy_identity_reconciliation", [])
    errors: list[str] = []
    expected = {
        "candidate_membership_count": 434,
        "unique_candidate_count": 423,
        "census_unique_detail_count": 423,
        "census_named_detail_count": 423,
        "census_blank_name_detail_count": 0,
        "census_unique_nonblank_name_count": 419,
        "census_same_name_group_count": 4,
        "census_department_count": 35,
        "census_nonempty_department_count": 422,
        "census_empty_department_count": 1,
        "gzszyy_unfiltered_page_count": 18,
        "gzszyy_unfiltered_unique_detail_count": 423,
        "gzszyy_dp_unique_detail_count": 422,
        "gzszyy_unfiltered_only_detail_count": 1,
        "gzszyy_dp_only_detail_count": 0,
        "gzszyy_official_care_site_count": 5,
        "excluded_non_doctor_count": 5,
        "eligible_candidate_count": 418,
        "category_error_count": 0,
        "detail_error_count": 0,
        "schedule_field_ingested_count": 0,
    }
    for field, expected_value in expected.items():
        actual = int(meta.get(field) or 0)
        if actual != expected_value:
            errors.append(f"{field} 应为 {expected_value}，实际 {actual}")
    if meta.get("gzszyy_unfiltered_only_detail_ids") != ["lNbWW4by"]:
        errors.append(
            "顶层专属详情应仅为 lNbWW4by，实际 "
            f"{meta.get('gzszyy_unfiltered_only_detail_ids', [])}"
        )
    campus_labels = {
        clean_text(str(campus))
        for item in detail_reconciliation
        for campus in item.get("campuses", [])
        if clean_text(str(campus))
    }
    if not campus_labels or not campus_labels <= GZSZYY_CAMPUS_LABELS:
        errors.append(f"详情院区/出诊点存在非规范二维码标题：{sorted(campus_labels)}")
    expected_same_name_groups = {
        "林少贞": ["ELe31Mb6", "JxboyNeg"],
        "唐瑾秋": ["4QbYVOdz", "X7ax9byv"],
        "王健": ["3YaOggax", "WZdP6yaK"],
        "高三德": ["LDdwkmd1", "QBeXY8ay"],
    }
    if meta.get("census_same_name_groups") != expected_same_name_groups:
        errors.append(
            "同名详情组不符合逐 ID 审计证据："
            f"{meta.get('census_same_name_groups', {})}"
        )

    nursing_exclusions = [
        item for item in excluded if "仅标注护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != 5 or len(nursing_exclusions) != len(excluded):
        errors.append(f"纯护理排除应为 5，实际 {len(nursing_exclusions)}")
    excluded_ids = {
        gzszyy_detail_id(str(item.get("source_link") or ""))
        for item in nursing_exclusions
        if gzszyy_detail_id(str(item.get("source_link") or ""))
    }
    formal_detail_ids = {
        str(item.get("detail_id") or "")
        for item in detail_reconciliation
        if str(item.get("detail_id") or "")
    }
    if len(formal_detail_ids) != len(detail_reconciliation):
        errors.append("合规详情逐 ID 对账存在空或重复 ID")
    if formal_detail_ids & excluded_ids:
        errors.append("合规详情与护理排除详情 ID 重叠")
    if len(excluded_ids) != 5 or len(formal_detail_ids | excluded_ids) != 423:
        errors.append(
            f"逐 ID 对账未覆盖 423 个唯一详情：合规 {len(formal_detail_ids)} / "
            f"护理排除 {len(excluded_ids)}"
        )

    mapped_ids = {
        str(detail_id)
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if str(detail_id)
    }
    if mapped_ids != formal_detail_ids:
        errors.append(
            f"身份聚类未完整映射 418 个合规详情 ID：对账 {len(formal_detail_ids)} / "
            f"映射 {len(mapped_ids)}"
        )
    if len(identity_reconciliation) != len(rows):
        errors.append(
            f"身份聚类对账应与最终正式行一致：{len(identity_reconciliation)}/{len(rows)}"
        )
    if (
        int(meta.get("unique_doctor_count") or 0) != len(rows)
        or int(meta.get("gzszyy_final_identity_count") or 0) != len(rows)
    ):
        errors.append(f"最终身份计数与正式行不一致：{len(rows)}")

    primary_ids = [gzszyy_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not detail_id for detail_id in primary_ids):
        errors.append("正式行存在非授权 gzszyy.com 专家详情来源")
    if len(set(primary_ids)) != len(rows):
        errors.append("最终身份主详情 ID 不唯一")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")

    actual_same_name_groups = {
        frozenset(str(detail_id) for detail_id in item.get("detail_ids", []))
        for item in identity_reconciliation
        if len(item.get("detail_ids", [])) > 1
    }
    if actual_same_name_groups != GZSZYY_SAME_IDENTITY_DETAIL_GROUPS:
        errors.append(
            "同一身份归并组不符合 owner 审计裁决："
            + "、".join(",".join(sorted(group)) for group in actual_same_name_groups)
        )
    if int(meta.get("gzszyy_same_identity_merge_group_count") or 0) != 3:
        errors.append(
            "同一身份归并组应为 3，实际 "
            f"{meta.get('gzszyy_same_identity_merge_group_count', 0)}"
        )
    distinct_groups = {
        frozenset(
            detail_id
            for item in identity_reconciliation
            if clean_text(str(item.get("name") or "")) == name
            for detail_id in item.get("detail_ids", [])
        )
        for name in {
            clean_text(str(item.get("name") or ""))
            for item in identity_reconciliation
            if str(item.get("resolution") or "") == "同名待甄别"
        }
    }
    if distinct_groups != GZSZYY_DISTINCT_SAME_NAME_DETAIL_GROUPS:
        errors.append("实质不同同名身份裁决不完整或出现未知同名分行")
    if (
        int(meta.get("gzszyy_distinct_same_name_group_count") or 0) != 1
        or int(meta.get("gzszyy_distinct_same_name_row_count") or 0) != 2
    ):
        errors.append("实质不同同名身份应为 1 组 2 行")
    distinct_rows = [
        row for row in rows if clean_text(str(row.get("姓名") or "")) == "王健"
    ]
    if len(distinct_rows) != 2 or any(
        "同名待甄别" not in str(row.get("异常提示") or "") for row in distinct_rows
    ):
        errors.append("王健两种实质不同身份未分行保留“同名待甄别”")
    title_conflict_names = {
        clean_text(str(item.get("name") or ""))
        for item in identity_reconciliation
        if len(item.get("detail_ids", [])) > 1
        and "多详情职称不一致"
        in str(
            next(
                (
                    row.get("异常提示", "")
                    for row in rows
                    if row.get("来源链接") == item.get("primary_source_link")
                ),
                "",
            )
        )
    }
    if title_conflict_names != {"唐瑾秋", "高三德"}:
        errors.append(f"多详情职称不一致标记不符合证据：{sorted(title_conflict_names)}")

    top_only_rows = [
        row
        for row in rows
        if gzszyy_detail_id(str(row.get("来源链接") or "")) == "lNbWW4by"
    ]
    top_only_detail = next(
        (
            item
            for item in detail_reconciliation
            if str(item.get("detail_id") or "") == "lNbWW4by"
        ),
        {},
    )
    if (
        len(top_only_rows) != 1
        or not clean_text(str(top_only_rows[0].get("姓名") or ""))
        or clean_text(str(top_only_rows[0].get("姓名") or ""))
        != clean_text(str(top_only_detail.get("name") or ""))
    ):
        errors.append("顶层专属详情 lNbWW4by 未按当前官网姓名单行保留")
    elif any(
        clean_text(str(top_only_rows[0].get(field) or ""))
        for field in ["科室_分类页", "职称身份原文", "擅长诊疗方向摘录"]
    ):
        errors.append("顶层专属详情的官网缺失科室/显式职称/擅长字段被推断或补造")

    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(
            r"^\s*(?:(?:擅长|专长)\s*[:：]?\s*)+",
            str(row.get("擅长诊疗方向摘录") or ""),
        )
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))
    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            strip_gyfyyy_schedule_text(str(row.get(field) or ""))
            != clean_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if schedule_rows:
        errors.append("四正式文本字段仍含排班片段：" + "、".join(schedule_rows[:10]))
    patient_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            contains_gzbrain_patient_case_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if patient_rows:
        errors.append("四正式文本字段仍含患者案例或可识别信息：" + "、".join(patient_rows[:10]))
    private_use_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            re.search(r"[\ue000-\uf8ff]", str(row.get(field) or ""))
            for field in BASE_HEADERS
        )
    ]
    if private_use_rows:
        errors.append("正式字段仍含 iconfont 私用区字符：" + "、".join(private_use_rows[:10]))
    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))

    if errors:
        raise RuntimeError("GZSZYY FULL 写入前门禁失败：" + "；".join(errors))


def validate_gzsys_trial(payload: dict[str, Any], expected_rows: int) -> None:
    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    errors: list[str] = []
    expected_counts = {
        "category_count": 23,
        "raw_card_rows": 664,
        "candidate_membership_count": 664,
        "unique_candidate_count": 664,
        "census_unique_detail_count": 664,
        "census_named_detail_count": 664,
        "census_blank_name_detail_count": 0,
        "census_nonempty_department_count": 664,
        "census_empty_department_count": 0,
        "excluded_non_doctor_count": 6,
        "eligible_candidate_count": 658,
        "category_error_count": 0,
        "detail_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        if int(meta.get(field) or 0) != expected:
            errors.append(f"{field} 应为 {expected}，实际 {meta.get(field, 0)}")
    if int(meta.get("census_same_name_group_count") or 0) != 0:
        errors.append("全目录出现未裁决的同名不同 ID")
    if int(meta.get("cross_entry_duplicate_count") or 0) != 0:
        errors.append("默认分页出现重复数字 ID")
    if len(rows) != expected_rows or int(meta.get("unique_doctor_count") or 0) != expected_rows:
        errors.append(f"TRIAL 正式行应为 {expected_rows}，实际 {len(rows)}")
    source_ids = [gzsys_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not detail_id for detail_id in source_ids):
        errors.append("TRIAL 存在非授权 /node/<ID> 或 /doctor/<ID> 来源")
    if len(set(source_ids)) != len(source_ids):
        errors.append("TRIAL 来源数字 ID 不唯一")
    if len({clean_text(str(row.get("姓名") or "")) for row in rows}) != len(rows):
        errors.append("TRIAL 姓名为空或不唯一")
    if any(not clean_text(str(row.get("科室_分类页") or "")) for row in rows):
        errors.append("TRIAL 存在空科室")
    if int(meta.get("sample_entry_coverage_count") or 0) < 3:
        errors.append("TRIAL 科室覆盖少于 3")
    excluded = payload.get("excluded_candidates", [])
    if len(excluded) != 6 or any("护理身份" not in str(item.get("reason") or "") for item in excluded):
        errors.append("6 个纯护理排除未完整留痕")
    excluded_ids = [gzsys_detail_id(str(item.get("source_link") or "")) for item in excluded]
    if len(set(excluded_ids)) != 6 or set(excluded_ids) & set(source_ids):
        errors.append("护理排除 ID 无效、重复或与正式行重叠")
    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    forbidden_terms = [
        "好医生",
        "名医录",
        "排行榜",
        "排名",
        "患者评价",
        "患者留言",
        "患者案例",
        "病例详情",
    ]
    if any(
        any(term in str(row.get(field) or "") for term in forbidden_terms)
        for row in rows
        for field in formal_text_fields
    ):
        errors.append("四正式文本字段仍含排名或患者信息")
    if any(
        strip_gyfyyy_schedule_text(str(row.get(field) or ""))
        != clean_text(str(row.get(field) or ""))
        for row in rows
        for field in formal_text_fields
    ):
        errors.append("四正式文本字段仍含排班片段")
    if any(
        re.search(r"[\ue000-\uf8ff]", str(row.get(field) or ""))
        for row in rows
        for field in BASE_HEADERS
    ):
        errors.append("正式字段仍含私用区字符")
    if any(
        re.match(r"^\s*(?:(?:专业擅长|擅长|专长|特长)\s*[:：]?\s*)+", str(row.get("擅长诊疗方向摘录") or ""))
        for row in rows
    ):
        errors.append("擅长字段仍保留前缀")
    if any(
        clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
        for row in rows
    ):
        errors.append("异常行仍被打标签或提升优先级")
    if errors:
        raise RuntimeError("GZSYS TRIAL 写出前门禁失败：" + "；".join(errors))


def validate_fahsysu_trial(payload: dict[str, Any], expected_rows: int) -> None:
    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    errors: list[str] = []
    expected_counts = {
        "census_group_count": 42,
        "census_relationship_group_count": 32,
        "census_empty_group_count": 10,
        "category_count": 90,
        "census_department_count": 90,
        "raw_card_rows": 881,
        "candidate_membership_count": 881,
        "unique_candidate_count": 860,
        "census_unique_detail_count": 860,
        "census_named_detail_count": 860,
        "census_blank_name_detail_count": 0,
        "census_nonempty_department_count": 860,
        "census_empty_department_count": 0,
        "census_same_name_group_count": 8,
        "cross_entry_duplicate_count": 21,
        "pagination_count": 1,
        "category_error_count": 0,
        "detail_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        if int(meta.get(field) or 0) != expected:
            errors.append(f"{field} 应为 {expected}，实际 {meta.get(field, 0)}")
    if meta.get("title_hint_counts") != {"正高": 447, "副高": 434}:
        errors.append(f"正高/副高关系基线漂移：{meta.get('title_hint_counts', {})}")
    if meta.get("census_same_name_groups") != FAHSYSU_EXPECTED_SAME_NAME_GROUPS:
        errors.append(
            f"同名不同数字 ID 组不符合普查基线：{meta.get('census_same_name_groups', {})}"
        )
    if len(rows) != expected_rows or int(meta.get("unique_doctor_count") or 0) != expected_rows:
        errors.append(f"TRIAL 正式行应为 {expected_rows}，实际 {len(rows)}")
    source_ids = [fahsysu_detail_id(str(row.get("来源链接") or "")) for row in rows]
    if any(not detail_id for detail_id in source_ids):
        errors.append("TRIAL 存在非授权 fahsysu.org.cn /node/<数字ID> 来源")
    if len(set(source_ids)) != len(source_ids):
        errors.append("TRIAL 来源数字 ID 不唯一")
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("TRIAL 存在空姓名")
    if any(not clean_text(str(row.get("科室_分类页") or "")) for row in rows):
        errors.append("TRIAL 存在空科室")
    if int(meta.get("sample_entry_coverage_count") or 0) < 3:
        errors.append("TRIAL 科室覆盖少于 3")
    if any(
        clean_text(str(row.get("职称身份原文") or "")) in {"正高", "副高", ""}
        for row in rows
    ):
        errors.append("正式职称为空或误用了目录正高/副高线索")
    same_name_rows = [
        row
        for row in rows
        if clean_text(str(row.get("姓名") or "")) in FAHSYSU_EXPECTED_SAME_NAME_GROUPS
    ]
    if any("同名待甄别" not in str(row.get("异常提示") or "") for row in same_name_rows):
        errors.append("样本中的同名不同 ID 未保留待甄别标记")
    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    forbidden_terms = [
        "好医生",
        "名医录",
        "排行榜",
        "排名",
        "患者评价",
        "患者留言",
        "患者案例",
        "病例详情",
    ]
    if any(
        any(term in str(row.get(field) or "") for term in forbidden_terms)
        for row in rows
        for field in formal_text_fields
    ):
        errors.append("四正式文本字段仍含排名或患者信息")
    if any(
        strip_gzsys_schedule_text(str(row.get(field) or ""))
        != clean_text(str(row.get(field) or ""))
        for row in rows
        for field in formal_text_fields
    ):
        errors.append("四正式文本字段仍含排班片段")
    if any(
        re.search(r"[\ue000-\uf8ff]", str(row.get(field) or ""))
        for row in rows
        for field in BASE_HEADERS
    ):
        errors.append("正式字段仍含私用区字符")
    if any(
        re.match(
            r"^\s*(?:(?:医疗特长|专业擅长|擅长|专长|特长)\s*[:：]?\s*)+",
            str(row.get("擅长诊疗方向摘录") or ""),
        )
        for row in rows
    ):
        errors.append("擅长字段仍保留多重前缀")
    if any(
        clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
        for row in rows
    ):
        errors.append("异常行仍被打标签或提升优先级")
    if errors:
        raise RuntimeError("FAHSYSU TRIAL 写出前门禁失败：" + "；".join(errors))


def validate_gdgh_trial(payload: dict[str, Any], expected_rows: int) -> None:
    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    photo_samples = payload.get("photo_samples", [])
    errors: list[str] = []
    expected_counts = {
        "census_group_count": GDGH_EXPECTED_GROUP_COUNT,
        "census_department_count": GDGH_EXPECTED_DEPARTMENT_COUNT,
        "candidate_membership_count": GDGH_EXPECTED_RELATION_COUNT,
        "census_unique_detail_count": GDGH_EXPECTED_RELATION_COUNT,
        "excluded_non_doctor_count": GDGH_EXPECTED_NURSING_COUNT,
        "eligible_candidate_count": GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT,
        "photo_census_available_count": GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT,
        "independent_entity_count": 0,
        "category_error_count": 0,
        "detail_error_count": 0,
        "photo_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field}={actual}，预期 {expected}")

    if expected_rows != 10 or len(rows) != expected_rows:
        errors.append(f"试采结果 {len(rows)} 行，必须固定为 10 行")
    if int(meta.get("sample_entry_coverage_count") or 0) < 3:
        errors.append("试采未覆盖至少 3 个科室")
    if len(photo_samples) != expected_rows or int(meta.get("photo_sample_count") or 0) != expected_rows:
        errors.append("试采未形成 10 张本人职业照样本")
    if meta.get("photo_policy_status") != "WAITING_OWNER_SIZE_POLICY":
        errors.append("照片压缩/宽度方案未保持 owner 裁决等待状态")

    expected_affiliates = {
        "广东省心血管病研究所",
        "广东省老年医学研究所",
        "惠福分院",
        "广东省肺癌研究所",
    }
    affiliate_rows = payload.get("affiliate_reconnaissance", [])
    actual_affiliates = {
        clean_text(str(item.get("name") or "")) for item in affiliate_rows
    }
    if not expected_affiliates.issubset(actual_affiliates):
        errors.append("分院/研究所归属证据不完整")
    if any(not clean_text(str(item.get("relation") or "")) for item in affiliate_rows):
        errors.append("分院/研究所存在未记录归属结论")

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_terms = ("门诊时间地点", "门诊时间", "出诊时间", "开诊时间")
    filenames: set[str] = set()
    sample_by_source = {
        clean_text(str(item.get("source_link") or "")): item for item in photo_samples
    }
    byte_sizes: list[int] = []
    for row in rows:
        name = clean_text(str(row.get("姓名") or ""))
        source = clean_text(str(row.get("来源链接") or ""))
        photo_url = clean_text(str(row.get("照片链接") or ""))
        photo_file = clean_text(str(row.get("照片文件") or ""))
        if not looks_like_person_name(name):
            errors.append(f"姓名格式异常：{name or '空'}")
        if not gdgh_detail_id(source):
            errors.append(f"非严格官网医生详情链接：{source or '空'}")
        if not gdgh_photo_url(photo_url, photo_url):
            errors.append(f"非官网同域 uploadfiles 照片：{photo_url or '空'}")
        sample = sample_by_source.get(source)
        if not sample:
            errors.append(f"缺少照片命名对照：{name or source}")
            continue
        filename = clean_text(str(sample.get("filename") or ""))
        expected_relative = (
            Path("01_试点医院") / "广东省人民医院" / "照片" / filename
        ).as_posix()
        if photo_file != expected_relative or sample.get("photo_file") != expected_relative:
            errors.append(f"照片相对路径不符合约定：{name}")
        if not filename or re.search(r'[\\/:*?"<>|]', filename):
            errors.append(f"照片文件名为空或含非法字符：{name}")
        if filename.casefold() in filenames:
            errors.append(f"照片文件名发生覆盖冲突：{filename}")
        filenames.add(filename.casefold())
        disk_path = Path(str(sample.get("disk_path") or ""))
        if not disk_path.is_file():
            errors.append(f"照片文件不存在：{disk_path}")
            continue
        content = disk_path.read_bytes()
        extension = disk_path.suffix.lower().lstrip(".")
        media_type = {
            "jpg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
        }.get(extension, "")
        detected = gdgh_photo_extension(content, media_type)
        if detected != extension:
            errors.append(f"照片扩展名与魔数不一致：{filename}")
        if int(sample.get("bytes") or 0) != len(content):
            errors.append(f"照片字节数对账失败：{filename}")
        if clean_text(str(sample.get("sha256") or "")) != hashlib.sha256(content).hexdigest():
            errors.append(f"照片 SHA-256 对账失败：{filename}")
        byte_sizes.append(len(content))

        formal_text = "\n".join(clean_text(str(row.get(field) or "")) for field in formal_text_fields)
        if any(term in formal_text for term in schedule_terms):
            errors.append(f"正式字段仍含排班片段：{name}")
        if any(marker in formal_text for marker in GDGH_FORBIDDEN_SENTENCE_MARKERS):
            errors.append(f"正式字段仍含排名/患者片段：{name}")
        if contains_gzbrain_patient_case_text(formal_text):
            errors.append(f"正式字段仍含患者案例或可识别信息：{name}")
        if contains_navigation_text(formal_text):
            errors.append(f"正式字段仍含导航片段：{name}")
        if re.search(r"[\ue000-\uf8ff]", "\n".join(str(row.get(field) or "") for field in BASE_HEADERS)):
            errors.append(f"正式字段仍含私用区字符：{name}")
        if clean_text(str(row.get("异常提示") or "")) and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        ):
            errors.append(f"异常行仍被打标签或提升优先级：{name}")

    if byte_sizes:
        average_bytes = round(sum(byte_sizes) / len(byte_sizes))
        if int(meta.get("photo_average_bytes") or 0) != average_bytes:
            errors.append("平均单张照片大小对账失败")
        expected_full_count = GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT
        if int(meta.get("photo_estimated_full_count") or 0) != expected_full_count:
            errors.append("全院照片候选估算数量不是护理排除后的 1,334")
        if int(meta.get("photo_estimated_full_bytes") or 0) != average_bytes * expected_full_count:
            errors.append("全院照片容量估算对账失败")

    if errors:
        raise RuntimeError("GDGH TRIAL 写出前门禁失败：" + "；".join(dict.fromkeys(errors)))


def validate_gdmch_trial(payload: dict[str, Any], expected_rows: int) -> None:
    """Validate the fixed Issue #43 census and ten auditable photo rows."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    details = payload.get("gdmch_detail_reconciliation", [])
    photos = payload.get("photo_samples", [])
    errors: list[str] = []
    expected_counts = {
        "category_count": GDMCH_EXPECTED_PAGE_COUNT,
        "pagination_count": GDMCH_EXPECTED_PAGE_COUNT,
        "raw_card_rows": GDMCH_EXPECTED_RELATION_COUNT,
        "candidate_membership_count": GDMCH_EXPECTED_RELATION_COUNT,
        "unique_candidate_count": GDMCH_EXPECTED_RELATION_COUNT,
        "census_unique_detail_count": GDMCH_EXPECTED_RELATION_COUNT,
        "excluded_non_doctor_count": GDMCH_EXPECTED_NON_DOCTOR_COUNT,
        "eligible_candidate_count": GDMCH_EXPECTED_RELATION_COUNT
        - GDMCH_EXPECTED_NON_DOCTOR_COUNT,
        "photo_census_available_count": GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT,
        "photo_census_placeholder_count": GDMCH_EXPECTED_RELATION_COUNT
        - GDMCH_EXPECTED_NON_DOCTOR_COUNT
        - GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT,
        "photo_default_placeholder_count": GDMCH_EXPECTED_DEFAULT_PHOTO_COUNT,
        "cross_entry_duplicate_count": 0,
        "category_error_count": 0,
        "detail_error_count": 0,
        "photo_error_count": 0,
        "photo_failed_count": 0,
        "photo_no_source_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
        "independent_entity_count": 0,
        "affiliate_count": 4,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field}={actual}，预期 {expected}")

    if expected_rows != 10 or len(rows) != expected_rows:
        errors.append(f"试采结果 {len(rows)} 行，必须固定为 10 行")
    if int(meta.get("department_coverage_count") or 0) < 3:
        errors.append("试采未覆盖至少 3 个科室")
    if (
        len(photos) != expected_rows
        or int(meta.get("photo_sample_count") or 0) != expected_rows
        or int(meta.get("photo_expected_count") or 0) != expected_rows
        or int(meta.get("photo_downloaded_count") or 0) != expected_rows
    ):
        errors.append("试采未形成 10 张可逐项对账的本人职业照样本")

    excluded_ids = [clean_text(str(item.get("detail_id") or "")) for item in excluded]
    if (
        len(excluded) != GDMCH_EXPECTED_NON_DOCTOR_COUNT
        or any(not value for value in excluded_ids)
        or len(set(excluded_ids)) != len(excluded_ids)
        or any(
            not gdmch_non_doctor_card(item.get("name"), item.get("list_title"))
            for item in excluded
        )
    ):
        errors.append(
            f"{GDMCH_EXPECTED_NON_DOCTOR_COUNT} 个号源/系统账号/非医生候选的逐 ID 排除表不完整"
        )

    detail_ids = [clean_text(str(item.get("detail_id") or "")) for item in details]
    if (
        len(details) != expected_rows
        or any(not value for value in detail_ids)
        or len(set(detail_ids)) != len(detail_ids)
        or any(item.get("resolution") != "详情已读取" for item in details)
    ):
        errors.append("10 个样本详情 ID 未完整、唯一且成功读取")

    expected_campuses = {
        "番禺院区": "/keshizhuanjia/panyuyuanqu",
        "越秀院区": "/keshizhuanjia/yuexiuyuanqu",
        "天河院区": "/keshizhuanjia/tianheyuanqu",
        "清远院区": "/keshizhuanjia/qingyuanyuanqu",
    }
    affiliates = payload.get("affiliate_reconnaissance", [])
    affiliate_map = {
        clean_text(str(item.get("name") or "")): clean_text(str(item.get("url") or ""))
        for item in affiliates
    }
    for campus, expected_path in expected_campuses.items():
        url = affiliate_map.get(campus, "")
        parsed = urlparse(url)
        if (
            comparable_host(url) != "e3861.com"
            or parsed.path.rstrip("/").lower() != expected_path
        ):
            errors.append(f"{campus}官方归属链接缺失或不准确")
    if any(not clean_text(str(item.get("relation") or "")) for item in affiliates):
        errors.append("四院区归属结论存在空白")

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_pattern = re.compile(
        r"(?:出诊安排|门诊时间|出诊时间|开诊时间|周[一二三四五六日天]|星期[一二三四五六日天]|上午|下午)"
    )
    photo_by_source = {
        clean_text(str(item.get("source_link") or "")): item for item in photos
    }
    filenames: set[str] = set()
    byte_sizes: list[int] = []
    large_photo_count = 0
    for row in rows:
        name = clean_text(str(row.get("姓名") or ""))
        source = clean_text(str(row.get("来源链接") or ""))
        photo_url = clean_text(str(row.get("照片链接") or ""))
        photo_file = clean_text(str(row.get("照片文件") or ""))
        if not looks_like_person_name(name) or gdmch_non_doctor_card(
            name, row.get("职称身份原文")
        ):
            errors.append(f"样本不是明确医生身份：{name or source}")
        if not gdmch_detail_id(source):
            errors.append(f"非严格官网数字 ID 详情链接：{source or '空'}")
        if not gdmch_photo_url(photo_url, photo_url):
            errors.append(f"非 wx.e3861.com 官方医生照片路径：{photo_url or '空'}")

        sample = photo_by_source.get(source)
        if not sample:
            errors.append(f"缺少照片命名与校验对照：{name or source}")
            continue
        filename = clean_text(str(sample.get("filename") or ""))
        expected_relative = (
            Path("01_试点医院") / "广东省妇幼保健院" / "照片" / filename
        ).as_posix()
        if photo_file != expected_relative or sample.get("photo_file") != expected_relative:
            errors.append(f"照片相对路径不符合约定：{name}")
        if not filename or re.search(r'[\\/:*?"<>|]', filename):
            errors.append(f"照片文件名为空或含非法字符：{name}")
        if filename.casefold() in filenames:
            errors.append(f"照片文件名发生覆盖冲突：{filename}")
        filenames.add(filename.casefold())

        disk_path = Path(str(sample.get("disk_path") or ""))
        if not disk_path.is_file():
            errors.append(f"照片文件不存在：{disk_path}")
            continue
        content = disk_path.read_bytes()
        extension = disk_path.suffix.lower().lstrip(".")
        media_type = {
            "jpg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
        }.get(extension, "")
        if gdgh_photo_extension(content, media_type) != extension:
            errors.append(f"照片扩展名与魔数不一致：{filename}")
        if int(sample.get("bytes") or 0) != len(content):
            errors.append(f"照片字节数对账失败：{filename}")
        if clean_text(str(sample.get("sha256") or "")) != hashlib.sha256(content).hexdigest():
            errors.append(f"照片 SHA-256 对账失败：{filename}")
        width, height = gdmch_photo_dimensions(content, extension)
        if (
            width <= 0
            or height <= 0
            or int(sample.get("width") or 0) != width
            or int(sample.get("height") or 0) != height
        ):
            errors.append(f"照片宽高对账失败：{filename}")
        byte_sizes.append(len(content))
        if len(content) > 200 * 1024 or width > 800:
            large_photo_count += 1

        formal_text = "\n".join(
            clean_text(str(row.get(field) or "")) for field in formal_text_fields
        )
        if schedule_pattern.search(formal_text):
            errors.append(f"正式字段仍含排班日期/时段：{name}")
        if any(marker in formal_text for marker in GDGH_FORBIDDEN_SENTENCE_MARKERS):
            errors.append(f"正式字段仍含排名/患者片段：{name}")
        if contains_gzbrain_patient_case_text(formal_text):
            errors.append(f"正式字段仍含患者案例或可识别信息：{name}")
        if contains_navigation_text(formal_text):
            errors.append(f"正式字段仍含导航片段：{name}")
        if re.search(
            r"[\ue000-\uf8ff]",
            "\n".join(str(row.get(field) or "") for field in BASE_HEADERS),
        ):
            errors.append(f"正式字段仍含私用区字符：{name}")
        if clean_text(str(row.get("异常提示") or "")) and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        ):
            errors.append(f"异常行仍被打标签或提升优先级：{name}")

    if byte_sizes:
        average_bytes = round(sum(byte_sizes) / len(byte_sizes))
        photo_count = GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT
        if int(meta.get("photo_average_bytes") or 0) != average_bytes:
            errors.append("平均单张照片大小对账失败")
        if int(meta.get("photo_estimated_full_count") or 0) != photo_count:
            errors.append("全院本人职业照候选估算数量不是 658")
        if int(meta.get("photo_estimated_full_bytes") or 0) != average_bytes * photo_count:
            errors.append("全院照片容量估算对账失败")
    if int(meta.get("large_photo_count") or 0) != large_photo_count:
        errors.append("大图阈值计数对账失败")
    expected_policy = (
        "WAITING_OWNER_LARGE_IMAGE_POLICY"
        if large_photo_count
        else "OWNER_APPROVED_ORIGINAL_NO_WIDTH_LIMIT"
    )
    if meta.get("photo_policy_status") != expected_policy:
        errors.append("照片政策状态与 >200KB 或宽 >800px 的现场结果不一致")

    if errors:
        raise RuntimeError("GDMCH TRIAL 写出前门禁失败：" + "；".join(dict.fromkeys(errors)))


def validate_gdmch_full_append(payload: dict[str, Any] | None = None) -> None:
    """Block the master write unless the owner-approved FULL census fully reconciles."""

    if payload is None:
        raise RuntimeError("GDMCH FULL 写出前门禁失败：缺少全量 payload")
    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    details = payload.get("gdmch_detail_reconciliation", [])
    identities = payload.get("gdmch_identity_reconciliation", [])
    photos = payload.get("photo_samples", [])
    photo_errors = payload.get("photo_errors", [])
    errors: list[str] = []
    expected_counts = {
        "category_count": GDMCH_EXPECTED_PAGE_COUNT,
        "pagination_count": GDMCH_EXPECTED_PAGE_COUNT,
        "raw_card_rows": GDMCH_EXPECTED_RELATION_COUNT,
        "candidate_membership_count": GDMCH_EXPECTED_RELATION_COUNT,
        "unique_candidate_count": GDMCH_EXPECTED_RELATION_COUNT,
        "census_unique_detail_count": GDMCH_EXPECTED_RELATION_COUNT,
        "excluded_non_doctor_count": GDMCH_EXPECTED_NON_DOCTOR_COUNT,
        "eligible_candidate_count": GDMCH_EXPECTED_RELATION_COUNT
        - GDMCH_EXPECTED_NON_DOCTOR_COUNT,
        "photo_census_available_count": GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT,
        "photo_census_placeholder_count": GDMCH_EXPECTED_RELATION_COUNT
        - GDMCH_EXPECTED_NON_DOCTOR_COUNT
        - GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT,
        "photo_default_placeholder_count": GDMCH_EXPECTED_DEFAULT_PHOTO_COUNT,
        "cross_entry_duplicate_count": 0,
        "category_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
        "independent_entity_count": 0,
        "affiliate_count": 4,
        "gdmch_final_identity_count": GDMCH_EXPECTED_FINAL_IDENTITY_COUNT,
        "gdmch_same_identity_merge_group_count": 1,
        "gdmch_distinct_same_name_group_count": 3,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field}={actual}，预期 {expected}")

    if meta.get("census_same_name_groups") != GDMCH_EXPECTED_SAME_NAME_GROUPS:
        errors.append(f"同名详情组不符合现场基线：{meta.get('census_same_name_groups', {})}")
    excluded_ids = [clean_text(str(item.get("detail_id") or "")) for item in excluded]
    if (
        len(excluded_ids) != GDMCH_EXPECTED_NON_DOCTOR_COUNT
        or any(not detail_id for detail_id in excluded_ids)
        or len(set(excluded_ids)) != len(excluded_ids)
        or any(
            not gdmch_non_doctor_card(item.get("name"), item.get("list_title"))
            for item in excluded
        )
    ):
        errors.append(
            f"{GDMCH_EXPECTED_NON_DOCTOR_COUNT} 个号源/系统账号/非医生候选的逐 ID 排除表不完整"
        )

    detail_ids = [clean_text(str(item.get("detail_id") or "")) for item in details]
    if (
        len(detail_ids) != GDMCH_EXPECTED_RELATION_COUNT - GDMCH_EXPECTED_NON_DOCTOR_COUNT
        or any(not detail_id for detail_id in detail_ids)
        or len(set(detail_ids)) != len(detail_ids)
    ):
        errors.append(
            f"逐 ID 对账应覆盖 "
            f"{GDMCH_EXPECTED_RELATION_COUNT - GDMCH_EXPECTED_NON_DOCTOR_COUNT} "
            f"个唯一合规详情，实际 {len(detail_ids)} 行/"
            f"{len(set(detail_ids) - {''})} 个非空唯一 ID"
        )
    mapped_ids = [
        clean_text(str(detail_id))
        for item in identities
        for detail_id in item.get("detail_ids", [])
        if clean_text(str(detail_id))
    ]
    if len(mapped_ids) != len(detail_ids) or set(mapped_ids) != set(detail_ids):
        errors.append(
            "身份聚类未完整且唯一映射 "
            f"{GDMCH_EXPECTED_RELATION_COUNT - GDMCH_EXPECTED_NON_DOCTOR_COUNT} "
            "个合规详情 ID"
        )
    if len(rows) != GDMCH_EXPECTED_FINAL_IDENTITY_COUNT or len(identities) != len(rows):
        errors.append(
            f"最终身份应为 {GDMCH_EXPECTED_FINAL_IDENTITY_COUNT} 行，实际 rows/对账 "
            f"{len(rows)}/{len(identities)}"
        )
    actual_merged_groups = {
        frozenset(clean_text(str(value)) for value in item.get("detail_ids", []) if value)
        for item in identities
        if int(item.get("relation_count") or 0) > 1
    }
    if actual_merged_groups != GDMCH_SAME_IDENTITY_DETAIL_GROUPS:
        errors.append("同一人归并详情组不符合逐页证据裁决")

    expected_photos = int(meta.get("photo_expected_count") or 0)
    downloaded_photos = int(meta.get("photo_downloaded_count") or 0)
    failed_photos = int(meta.get("photo_failed_count") or 0)
    no_source_photos = int(meta.get("photo_no_source_count") or 0)
    if expected_photos + no_source_photos != len(rows):
        errors.append("照片应采与占位留空未覆盖全部最终身份")
    if downloaded_photos + failed_photos != expected_photos:
        errors.append("照片实采与失败之和不等于应采数")
    if downloaded_photos != len(photos) or failed_photos != len(photo_errors):
        errors.append("照片四数与照片对账清单不一致")
    if expected_photos != GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT:
        errors.append(f"最终本人职业照应采数不是 {GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT}")
    if no_source_photos != len(rows) - GDMCH_EXPECTED_PHOTO_AVAILABLE_COUNT:
        errors.append("最终占位留空数与身份归并后的照片基线不一致")
    if int(meta.get("large_photo_count") or 0) != 0 or meta.get(
        "photo_policy_status"
    ) != "OWNER_APPROVED_ORIGINAL_NO_WIDTH_LIMIT":
        errors.append("发现未获 owner 裁决的大图，禁止写入总底表")

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    schedule_pattern = re.compile(
        r"(?:出诊安排|门诊时间|出诊时间|开诊时间|周[一二三四五六日天]|星期[一二三四五六日天]|上午|下午)"
    )
    photo_by_source = {
        clean_text(str(item.get("source_link") or "")): item for item in photos
    }
    failed_photo_sources = {
        clean_text(str(item.get("source_link") or "")) for item in photo_errors
    }
    filenames: set[str] = set()
    for row in rows:
        name = clean_text(str(row.get("姓名") or ""))
        source = clean_text(str(row.get("来源链接") or ""))
        photo_url = clean_text(str(row.get("照片链接") or ""))
        photo_file = clean_text(str(row.get("照片文件") or ""))
        if not looks_like_person_name(name) or gdmch_non_doctor_card(
            name, row.get("职称身份原文")
        ):
            errors.append(f"正式行不是明确医生身份：{name or source}")
        if not gdmch_detail_id(source):
            errors.append(f"非严格官网数字 ID 详情链接：{source or '空'}")
        formal_text = "\n".join(
            clean_text(str(row.get(field) or "")) for field in formal_text_fields
        )
        if schedule_pattern.search(formal_text):
            errors.append(f"正式字段仍含排班日期/时段：{name}")
        if any(marker in formal_text for marker in GDGH_FORBIDDEN_SENTENCE_MARKERS):
            errors.append(f"正式字段仍含排名/患者片段：{name}")
        if contains_gzbrain_patient_case_text(formal_text):
            errors.append(f"正式字段仍含患者案例或可识别信息：{name}")
        if contains_navigation_text(formal_text):
            errors.append(f"正式字段仍含导航片段：{name}")
        if clean_text(str(row.get("异常提示") or "")) and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        ):
            errors.append(f"异常行仍被打标签或提升优先级：{name}")

        if photo_url or photo_file:
            if not gdmch_photo_url(photo_url, photo_url):
                errors.append(f"非官方医生照片路径：{photo_url or source}")
            sample = photo_by_source.get(source)
            if not sample:
                errors.append(f"有照片正式行缺少照片对账：{name or source}")
                continue
            filename = clean_text(str(sample.get("filename") or ""))
            if filename.casefold() in filenames:
                errors.append(f"照片文件名发生覆盖冲突：{filename}")
            filenames.add(filename.casefold())
            disk_path = Path(str(sample.get("disk_path") or ""))
            if not disk_path.is_file():
                errors.append(f"照片文件不存在：{disk_path}")
                continue
            content = disk_path.read_bytes()
            extension = disk_path.suffix.lower().lstrip(".")
            media_type = {
                "jpg": "image/jpeg",
                "png": "image/png",
                "gif": "image/gif",
                "webp": "image/webp",
            }.get(extension, "")
            width, height = gdmch_photo_dimensions(content, extension)
            if (
                gdgh_photo_extension(content, media_type) != extension
                or int(sample.get("bytes") or 0) != len(content)
                or clean_text(str(sample.get("sha256") or ""))
                != hashlib.sha256(content).hexdigest()
                or int(sample.get("width") or 0) != width
                or int(sample.get("height") or 0) != height
            ):
                errors.append(f"照片字节/魔数/SHA-256/宽高对账失败：{filename}")
        elif source in failed_photo_sources:
            if "照片获取失败" not in clean_text(str(row.get("异常提示") or "")):
                errors.append(f"照片失败行未保留异常提示：{name or source}")
        elif "照片获取失败" in clean_text(str(row.get("异常提示") or "")):
            errors.append(f"无照片占位行被误标为下载失败：{name or source}")

    if errors:
        raise RuntimeError("GDMCH FULL 写出前门禁失败：" + "；".join(dict.fromkeys(errors)))


def validate_gdgh_full_append(payload: dict[str, Any]) -> None:
    """Block the GDGH master write unless census, identities and photos reconcile."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    detail_reconciliation = payload.get("gdgh_detail_reconciliation", [])
    identity_reconciliation = payload.get("gdgh_identity_reconciliation", [])
    photo_samples = payload.get("photo_samples", [])
    photo_errors = payload.get("photo_errors", [])
    errors: list[str] = []
    expected_counts = {
        "census_group_count": GDGH_EXPECTED_GROUP_COUNT,
        "census_department_count": GDGH_EXPECTED_DEPARTMENT_COUNT,
        "candidate_membership_count": GDGH_EXPECTED_RELATION_COUNT,
        "census_unique_detail_count": GDGH_EXPECTED_RELATION_COUNT,
        "unique_candidate_count": GDGH_EXPECTED_RELATION_COUNT,
        "excluded_non_doctor_count": GDGH_EXPECTED_NURSING_COUNT,
        "eligible_candidate_count": GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT,
        "photo_census_available_count": GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT,
        "category_error_count": 0,
        "independent_entity_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field}={actual}，预期 {expected}")

    excluded = payload.get("excluded_candidates", [])
    if len(excluded) != GDGH_EXPECTED_NURSING_COUNT or any(
        "护理身份" not in str(item.get("reason") or "") for item in excluded
    ):
        errors.append("9 个纯护理身份排除清单不完整或混入其他排除理由")

    detail_ids = [clean_text(str(item.get("detail_id") or "")) for item in detail_reconciliation]
    if (
        len(detail_ids) != GDGH_EXPECTED_RELATION_COUNT - GDGH_EXPECTED_NURSING_COUNT
        or any(not detail_id for detail_id in detail_ids)
        or len(set(detail_ids)) != len(detail_ids)
    ):
        errors.append(
            f"逐 ID 对账应覆盖 1,334 个唯一合规详情，实际 {len(detail_ids)} 行/"
            f"{len(set(detail_ids) - {''})} 个非空唯一 ID"
        )
    mapped_ids = [
        clean_text(str(detail_id))
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if clean_text(str(detail_id))
    ]
    if len(mapped_ids) != len(detail_ids) or set(mapped_ids) != set(detail_ids):
        errors.append("身份聚类未完整且唯一映射 1,334 个合规详情 ID")
    if len(identity_reconciliation) != len(rows):
        errors.append(f"身份归并对账与最终正式行不一致：{len(identity_reconciliation)}/{len(rows)}")
    if int(meta.get("unique_doctor_count") or 0) != len(rows):
        errors.append("最终身份计数与正式行不一致")

    if meta.get("photo_policy_status") != "OWNER_APPROVED_ORIGINAL_NO_WIDTH_LIMIT":
        errors.append("照片政策未记录 owner 已批准的不压缩、不限宽裁决")
    expected_photos = int(meta.get("photo_expected_count") or 0)
    downloaded_photos = int(meta.get("photo_downloaded_count") or 0)
    failed_photos = int(meta.get("photo_failed_count") or 0)
    no_source_photos = int(meta.get("photo_no_source_count") or 0)
    if expected_photos + no_source_photos != len(rows):
        errors.append("照片应采与无照片四数未覆盖全部最终身份")
    if downloaded_photos + failed_photos != expected_photos:
        errors.append("照片实采与失败四数不等于应采数")
    if downloaded_photos != len(photo_samples) or failed_photos != len(photo_errors):
        errors.append("照片实采/失败 meta 与对账清单不一致")
    if int(meta.get("photo_sample_count") or 0) != downloaded_photos:
        errors.append("照片样本计数与全量实采数不一致")

    sample_by_source = {
        clean_text(str(item.get("source_link") or "")): item for item in photo_samples
    }
    failed_sources = {
        clean_text(str(item.get("source_link") or "")) for item in photo_errors
    }
    if "" in failed_sources:
        errors.append("照片失败清单存在空来源链接")
    for row in rows:
        name = clean_text(str(row.get("姓名") or ""))
        source = clean_text(str(row.get("来源链接") or ""))
        photo_url = clean_text(str(row.get("照片链接") or ""))
        photo_file = clean_text(str(row.get("照片文件") or ""))
        warning = clean_text(str(row.get("异常提示") or ""))
        if not looks_like_person_name(name) or not gdgh_detail_id(source):
            errors.append(f"正式行姓名或官网详情来源异常：{name or source}")
        if source in failed_sources:
            if photo_url or photo_file or "照片获取失败" not in warning:
                errors.append(f"照片失败行未按要求留空并标记：{name}")
        elif photo_file:
            sample = sample_by_source.get(source)
            if not sample or not gdgh_photo_url(photo_url, photo_url):
                errors.append(f"照片成功行缺少官网 URL 或对账项：{name}")
                continue
            disk_path = Path(str(sample.get("disk_path") or ""))
            expected_prefix = "01_试点医院/广东省人民医院/照片/"
            if not photo_file.startswith(expected_prefix) or not disk_path.is_file():
                errors.append(f"照片文件路径不存在或越界：{name}")
                continue
            content = disk_path.read_bytes()
            detected = gdgh_photo_extension(
                content,
                {
                    ".jpg": "image/jpeg",
                    ".png": "image/png",
                    ".gif": "image/gif",
                    ".webp": "image/webp",
                }.get(disk_path.suffix.lower(), ""),
            )
            if detected != disk_path.suffix.lower().lstrip("."):
                errors.append(f"照片扩展名与魔数不一致：{name}")
            if int(sample.get("bytes") or 0) != len(content):
                errors.append(f"照片字节数对账失败：{name}")
            if clean_text(str(sample.get("sha256") or "")) != hashlib.sha256(content).hexdigest():
                errors.append(f"照片 SHA-256 对账失败：{name}")
        elif source not in failed_sources and photo_url:
            errors.append(f"照片列未成对写入：{name}")

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    if any(
        contains_navigation_text(str(row.get(field) or ""))
        or contains_gzbrain_patient_case_text(str(row.get(field) or ""))
        or strip_gzsys_schedule_text(str(row.get(field) or ""))
        != clean_text(str(row.get(field) or ""))
        for row in rows
        for field in formal_text_fields
    ):
        errors.append("四正式文本字段仍含导航、排班或患者案例片段")
    if any(
        re.search(r"[\ue000-\uf8ff]", str(row.get(field) or ""))
        for row in rows
        for field in BASE_HEADERS
    ):
        errors.append("正式字段仍含私用区字符")
    if any(
        re.search(r"[\u200b-\u200d\ufeff]", str(row.get(field) or ""))
        for row in rows
        for field in BASE_HEADERS
    ):
        errors.append("正式字段仍含零宽格式字符")
    if any(
        clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
        for row in rows
    ):
        errors.append("异常行仍被打标签或提升优先级")

    if errors:
        raise RuntimeError("GDGH FULL 写入前门禁失败：" + "；".join(dict.fromkeys(errors)))


def validate_fahsysu_full_append(payload: dict[str, Any]) -> None:
    """Block the FAHSYSU master write unless all 860 directory IDs reconcile."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    detail_reconciliation = payload.get("fahsysu_detail_reconciliation", [])
    identity_reconciliation = payload.get("fahsysu_identity_reconciliation", [])
    errors: list[str] = []
    expected_counts = {
        "census_group_count": 42,
        "census_relationship_group_count": 32,
        "census_empty_group_count": 10,
        "category_count": 90,
        "census_department_count": 90,
        "raw_card_rows": 881,
        "candidate_membership_count": 881,
        "unique_candidate_count": 860,
        "census_unique_detail_count": 860,
        "census_named_detail_count": 860,
        "census_blank_name_detail_count": 0,
        "census_nonempty_department_count": 860,
        "census_empty_department_count": 0,
        "census_same_name_group_count": 8,
        "cross_entry_duplicate_count": 21,
        "pagination_count": 1,
        "category_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field} 应为 {expected}，实际 {actual}")
    if meta.get("title_hint_counts") != {"正高": 447, "副高": 434}:
        errors.append(f"正高/副高关系基线漂移：{meta.get('title_hint_counts', {})}")
    if meta.get("census_same_name_groups") != FAHSYSU_EXPECTED_SAME_NAME_GROUPS:
        errors.append(
            f"同名不同数字 ID 组不符合审计基线：{meta.get('census_same_name_groups', {})}"
        )
    if int(meta.get("detail_error_count") or 0) != len(payload.get("detail_errors", [])):
        errors.append("详情读取失败计数与失败清单不一致")

    nursing_exclusions = [
        item for item in excluded if "护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != len(excluded):
        errors.append("存在未按明确护理身份规则留痕的排除候选")
    if int(meta.get("excluded_non_doctor_count") or 0) != len(nursing_exclusions):
        errors.append(
            "护理排除计数与 meta 不一致："
            f"{len(nursing_exclusions)}/{meta.get('excluded_non_doctor_count', 0)}"
        )

    formal_ids = [fahsysu_detail_id(str(row.get("来源链接") or "")) for row in rows]
    excluded_ids = [
        fahsysu_detail_id(str(item.get("source_link") or ""))
        for item in nursing_exclusions
    ]
    failed_ids = {
        fahsysu_detail_id(str(item.get("source_link") or ""))
        for item in payload.get("detail_errors", [])
    }
    if any(not detail_id for detail_id in formal_ids):
        errors.append("正式行存在非授权 fahsysu.org.cn /node/<数字ID> 来源")
    if any(not detail_id for detail_id in excluded_ids):
        errors.append("护理排除含非授权详情来源")
    if "" in failed_ids:
        errors.append("详情失败清单含非授权详情来源")
    formal_id_set = {detail_id for detail_id in formal_ids if detail_id}
    excluded_id_set = {detail_id for detail_id in excluded_ids if detail_id}
    if len(formal_id_set) != len(rows):
        errors.append("正式行来源数字 ID 为空或重复")
    if len(excluded_id_set) != len(excluded_ids):
        errors.append("护理排除来源数字 ID 为空或重复")
    if formal_id_set & excluded_id_set:
        errors.append("正式行与护理排除数字 ID 重叠")
    if len(formal_id_set | excluded_id_set) != 860:
        errors.append(
            f"逐 ID 对账未覆盖 860 个唯一详情：正式 {len(formal_id_set)} / "
            f"护理排除 {len(excluded_id_set)}"
        )
    if not failed_ids <= formal_id_set:
        errors.append("详情失败 ID 未按列表证据保守保留为正式行")

    reconciliation_ids = [
        str(item.get("detail_id") or "") for item in detail_reconciliation
    ]
    if (
        len(reconciliation_ids) != 860
        or any(not detail_id for detail_id in reconciliation_ids)
        or len(set(reconciliation_ids)) != 860
        or set(reconciliation_ids) != formal_id_set | excluded_id_set
    ):
        errors.append(
            f"逐 ID 对账工件不完整或重复：{len(reconciliation_ids)} 行 / "
            f"{len(set(reconciliation_ids) - {''})} 个非空唯一 ID"
        )
    if sum(int(item.get("relation_count") or 0) for item in detail_reconciliation) != 881:
        errors.append("逐 ID 对账关系总数未覆盖 881 条医生—专科关系")
    resolutions = {
        str(item.get("detail_id") or ""): str(item.get("resolution") or "")
        for item in detail_reconciliation
    }
    if any(
        resolutions.get(detail_id) not in {"正式行", "同名待甄别"}
        for detail_id in formal_id_set
    ):
        errors.append("逐 ID 对账中的正式行裁决与输出不一致")
    if any(resolutions.get(detail_id) != "护理排除" for detail_id in excluded_id_set):
        errors.append("逐 ID 对账中的护理排除裁决与排除清单不一致")

    mapped_ids = [
        str(detail_id)
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if str(detail_id)
    ]
    if set(mapped_ids) != formal_id_set or len(mapped_ids) != len(formal_ids):
        errors.append(
            f"身份归并未完整且唯一映射正式详情 ID：正式 {len(formal_id_set)} / "
            f"映射 {len(set(mapped_ids))}"
        )
    if len(identity_reconciliation) != len(rows):
        errors.append(
            f"身份归并对账应与最终正式行一致：{len(identity_reconciliation)}/{len(rows)}"
        )
    if (
        int(meta.get("unique_doctor_count") or 0) != len(rows)
        or int(meta.get("eligible_candidate_count") or 0) != len(rows)
        or int(meta.get("fahsysu_final_identity_count") or 0) != len(rows)
    ):
        errors.append(f"最终身份计数与正式行不一致：{len(rows)}")
    if int(meta.get("fahsysu_same_identity_merge_group_count") or 0) != 0:
        errors.append("不得按姓名启发式归并不同数字 ID")
    rows_by_id = {
        fahsysu_detail_id(str(row.get("来源链接") or "")): row for row in rows
    }
    formal_same_name_groups = {
        name: [detail_id for detail_id in ids if detail_id in formal_id_set]
        for name, ids in FAHSYSU_EXPECTED_SAME_NAME_GROUPS.items()
    }
    formal_same_name_groups = {
        name: ids for name, ids in formal_same_name_groups.items() if len(ids) > 1
    }
    if int(meta.get("fahsysu_distinct_same_name_group_count") or 0) != len(
        formal_same_name_groups
    ):
        errors.append("实质不同同名组计数与护理排除后的正式行不一致")
    if int(meta.get("fahsysu_distinct_same_name_row_count") or 0) != sum(
        len(ids) for ids in formal_same_name_groups.values()
    ):
        errors.append("实质不同同名行计数与护理排除后的正式行不一致")
    if any(
        clean_text(str(rows_by_id[detail_id].get("姓名") or "")) != name
        or "同名待甄别" not in str(rows_by_id[detail_id].get("异常提示") or "")
        for name, ids in formal_same_name_groups.items()
        for detail_id in ids
    ):
        errors.append("护理排除后的同名不同 ID 未完整分行保留并标记")

    invalid_core_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if row.get("医院") != "中山大学附属第一医院"
        or row.get("来源类型") != "医院官网"
        or row.get("采集入口") != "https://www.fahsysu.org.cn/page/6945"
        or (
            row.get("详情页状态") != "200"
            and not (
                fahsysu_detail_id(str(row.get("来源链接") or "")) in failed_ids
                and row.get("详情页状态") == "失败"
                and "详情页读取失败" in str(row.get("异常提示") or "")
            )
        )
    ]
    if invalid_core_rows:
        errors.append("正式行核心来源字段或详情状态异常：" + "、".join(invalid_core_rows[:10]))
    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")
    if any(not clean_text(str(row.get("科室_分类页") or "")) for row in rows):
        errors.append("正式行存在空目录科室")
    if any(
        fahsysu_detail_id(str(row.get("来源链接") or "")) not in failed_ids
        and clean_text(str(row.get("职称身份原文") or "")) in {"正高", "副高"}
        for row in rows
    ):
        errors.append("详情成功行误用了目录正高/副高线索")
    if any(
        fahsysu_detail_id(str(row.get("来源链接") or "")) not in failed_ids
        and not clean_text(str(row.get("职称身份原文") or ""))
        and "职称/身份需人工复核" not in str(row.get("异常提示") or "")
        for row in rows
    ):
        errors.append("官网未展示职称的详情成功行未保守留空并标记复核")

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    forbidden_terms = [
        "好医生",
        "名医录",
        "排行榜",
        "排名",
        "患者评价",
        "患者留言",
        "患者案例",
        "病例详情",
    ]
    forbidden_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            term in str(row.get(field) or "")
            for field in formal_text_fields
            for term in forbidden_terms
        )
        or any(
            contains_gzbrain_patient_case_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if forbidden_rows:
        errors.append("四正式文本字段仍含排名、患者或病例信息：" + "、".join(forbidden_rows[:10]))
    schedule_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            strip_gzsys_schedule_text(str(row.get(field) or ""))
            != clean_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if schedule_rows:
        errors.append("四正式文本字段仍含排班片段：" + "、".join(schedule_rows[:10]))
    navigation_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(contains_navigation_text(str(row.get(field) or "")) for field in formal_text_fields)
    ]
    if navigation_rows:
        errors.append("四正式文本字段仍含导航片段：" + "、".join(navigation_rows[:10]))
    private_use_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(re.search(r"[\ue000-\uf8ff]", str(row.get(field) or "")) for field in BASE_HEADERS)
    ]
    if private_use_rows:
        errors.append("正式字段仍含 iconfont 私用区字符：" + "、".join(private_use_rows[:10]))
    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(
            r"^\s*(?:(?:医疗特长|专业擅长|擅长|专长|特长)\s*[:：]?\s*)+",
            str(row.get("擅长诊疗方向摘录") or ""),
        )
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留多重前缀：" + "、".join(prefixed_specialties[:10]))
    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))

    if errors:
        raise RuntimeError("FAHSYSU FULL 写入前门禁失败：" + "；".join(errors))


def validate_gzsys_full_append(payload: dict[str, Any]) -> None:
    """Block the GZSYS master write unless all 664 audited IDs reconcile."""

    meta = payload.get("meta", {})
    rows = payload.get("rows", [])
    excluded = payload.get("excluded_candidates", [])
    detail_reconciliation = payload.get("gzsys_detail_reconciliation", [])
    identity_reconciliation = payload.get("gzsys_identity_reconciliation", [])
    errors: list[str] = []
    expected_counts = {
        "category_count": 23,
        "pagination_count": 23,
        "raw_card_rows": 664,
        "candidate_membership_count": 664,
        "unique_candidate_count": 664,
        "census_unique_detail_count": 664,
        "census_named_detail_count": 664,
        "census_blank_name_detail_count": 0,
        "census_unique_nonblank_name_count": 664,
        "census_same_name_group_count": 0,
        "census_department_count": 65,
        "census_nonempty_department_count": 664,
        "census_empty_department_count": 0,
        "eligible_candidate_count": 658,
        "cross_entry_duplicate_count": 0,
        "category_error_count": 0,
        "schedule_field_ingested_count": 0,
        "private_use_character_count": 0,
    }
    for field, expected in expected_counts.items():
        actual = int(meta.get(field) or 0)
        if actual != expected:
            errors.append(f"{field} 应为 {expected}，实际 {actual}")

    allowed_failed_detail_ids = {"25208"}
    actual_failed_detail_ids = {
        gzsys_detail_id(str(item.get("source_link") or ""))
        for item in payload.get("detail_errors", [])
    }
    if "" in actual_failed_detail_ids or not actual_failed_detail_ids <= allowed_failed_detail_ids:
        errors.append(
            "存在未授权放行的详情读取失败 ID："
            + "、".join(sorted(actual_failed_detail_ids - allowed_failed_detail_ids) or ["无效来源"])
        )
    if int(meta.get("detail_error_count") or 0) != len(payload.get("detail_errors", [])):
        errors.append("详情读取失败计数与失败清单不一致")

    if meta.get("filter_dictionary_counts") != {
        "department_target_id": 96,
        "talent_project": 4,
        "tutor_qualification": 5,
        "doctor_title": 33,
    }:
        errors.append(f"筛选字典计数偏离审计基线：{meta.get('filter_dictionary_counts', {})}")
    if meta.get("source_path_counts") != {"node": 432, "doctor": 232}:
        errors.append(f"node/doctor 来源关系计数偏离审计基线：{meta.get('source_path_counts', {})}")

    nursing_exclusions = [
        item for item in excluded if "护理身份" in str(item.get("reason") or "")
    ]
    if len(nursing_exclusions) != len(excluded):
        errors.append("存在未按明确护理身份规则留痕的排除候选")
    if len(nursing_exclusions) < 6:
        errors.append(f"官网列表基线的纯护理排除至少应为 6，实际 {len(nursing_exclusions)}")
    if int(meta.get("excluded_non_doctor_count") or 0) != len(excluded):
        errors.append(
            "排除计数与 meta 不一致："
            f"{len(excluded)}/{meta.get('excluded_non_doctor_count', 0)}"
        )

    formal_ids = [gzsys_detail_id(str(row.get("来源链接") or "")) for row in rows]
    excluded_ids = [
        gzsys_detail_id(str(item.get("source_link") or "")) for item in nursing_exclusions
    ]
    if any(not detail_id for detail_id in formal_ids):
        errors.append("正式行存在非授权 /node/<ID> 或 /doctor/<ID> 来源")
    if any(not detail_id for detail_id in excluded_ids):
        errors.append("护理排除含非授权详情来源")
    formal_id_set = {detail_id for detail_id in formal_ids if detail_id}
    excluded_id_set = {detail_id for detail_id in excluded_ids if detail_id}
    if len(formal_id_set) != len(rows):
        errors.append("正式行来源数字 ID 为空或重复")
    if len(excluded_id_set) != len(excluded_ids):
        errors.append("护理排除来源数字 ID 为空或重复")
    if formal_id_set & excluded_id_set:
        errors.append("正式行与护理排除数字 ID 重叠")
    if len(formal_id_set | excluded_id_set) != 664:
        errors.append(
            f"逐 ID 对账未覆盖 664 个唯一详情：正式 {len(formal_id_set)} / "
            f"护理排除 {len(excluded_id_set)}"
        )

    reconciliation_ids = [
        str(item.get("detail_id") or "") for item in detail_reconciliation
    ]
    if (
        len(reconciliation_ids) != 664
        or any(not detail_id for detail_id in reconciliation_ids)
        or len(set(reconciliation_ids)) != 664
        or set(reconciliation_ids) != formal_id_set | excluded_id_set
    ):
        errors.append(
            f"逐 ID 对账工件不完整或重复：{len(reconciliation_ids)} 行 / "
            f"{len(set(reconciliation_ids) - {''})} 个非空唯一 ID"
        )
    resolutions = {
        str(item.get("detail_id") or ""): str(item.get("resolution") or "")
        for item in detail_reconciliation
    }
    if any(resolutions.get(detail_id) != "正式行" for detail_id in formal_id_set):
        errors.append("逐 ID 对账中的正式行裁决与输出不一致")
    if any(resolutions.get(detail_id) != "护理排除" for detail_id in excluded_id_set):
        errors.append("逐 ID 对账中的护理排除裁决与排除清单不一致")

    mapped_ids = [
        str(detail_id)
        for item in identity_reconciliation
        for detail_id in item.get("detail_ids", [])
        if str(detail_id)
    ]
    if set(mapped_ids) != formal_id_set or len(mapped_ids) != len(formal_ids):
        errors.append(
            f"身份归并未完整且唯一映射正式详情 ID：正式 {len(formal_id_set)} / "
            f"映射 {len(set(mapped_ids))}"
        )
    if len(identity_reconciliation) != len(rows):
        errors.append(
            f"身份归并对账应与最终正式行一致：{len(identity_reconciliation)}/{len(rows)}"
        )
    if (
        int(meta.get("unique_doctor_count") or 0) != len(rows)
        or int(meta.get("gzsys_final_identity_count") or 0) != len(rows)
    ):
        errors.append(f"最终身份计数与正式行不一致：{len(rows)}")
    if int(meta.get("gzsys_same_identity_merge_group_count") or 0) != 0:
        errors.append("无同名基线下不应出现启发式身份归并")
    if int(meta.get("gzsys_distinct_same_name_group_count") or 0) != 0:
        errors.append("无同名基线下不应出现同名分行裁决")

    if any(not clean_text(str(row.get("姓名") or "")) for row in rows):
        errors.append("正式行存在空姓名")
    if len({clean_text(str(row.get("姓名") or "")) for row in rows}) != len(rows):
        errors.append("正式行出现未裁决的同名身份")
    invalid_core_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if row.get("医院") != "中山大学孙逸仙纪念医院"
        or row.get("来源类型") != "医院官网"
        or row.get("采集入口") != "https://www.gzsys.org.cn/doctor/592/search"
        or (
            row.get("详情页状态") != "200"
            and not (
                gzsys_detail_id(str(row.get("来源链接") or "")) in actual_failed_detail_ids
                and row.get("详情页状态") == "失败"
                and "详情页读取失败" in str(row.get("异常提示") or "")
            )
        )
    ]
    if invalid_core_rows:
        errors.append("正式行核心来源字段或详情状态异常：" + "、".join(invalid_core_rows[:10]))

    prefixed_specialties = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if re.match(
            r"^\s*(?:(?:专业擅长|擅长|专长|特长)\s*[:：]?\s*)+",
            str(row.get("擅长诊疗方向摘录") or ""),
        )
    ]
    if prefixed_specialties:
        errors.append("擅长字段仍保留前缀：" + "、".join(prefixed_specialties[:10]))

    formal_text_fields = ["擅长诊疗方向摘录", "亮眼经历线索", "列表简介", "详情正文摘录"]
    forbidden_terms = [
        "好医生",
        "名医录",
        "排行榜",
        "排名",
        "患者评价",
        "患者留言",
        "患者案例",
        "病例详情",
    ]
    forbidden_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            term in str(row.get(field) or "")
            for field in formal_text_fields
            for term in forbidden_terms
        )
    ]
    if forbidden_rows:
        errors.append("四正式文本字段仍含排名或患者信息：" + "、".join(forbidden_rows[:10]))
    schedule_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            strip_gyfyyy_schedule_text(str(row.get(field) or ""))
            != clean_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if schedule_rows:
        errors.append("四正式文本字段仍含排班片段：" + "、".join(schedule_rows[:10]))
    patient_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(
            contains_gzbrain_patient_case_text(str(row.get(field) or ""))
            for field in formal_text_fields
        )
    ]
    if patient_rows:
        errors.append("四正式文本字段仍含患者案例或可识别信息：" + "、".join(patient_rows[:10]))
    private_use_rows = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if any(re.search(r"[\ue000-\uf8ff]", str(row.get(field) or "")) for field in BASE_HEADERS)
    ]
    if private_use_rows:
        errors.append("正式字段仍含 iconfont 私用区字符：" + "、".join(private_use_rows[:10]))
    tagged_abnormal = [
        str(row.get("姓名") or "未命名")
        for row in rows
        if clean_text(str(row.get("异常提示") or ""))
        and (
            clean_text(str(row.get("重点关注范围") or ""))
            or clean_text(str(row.get("重点疾病标签") or ""))
            or clean_text(str(row.get("重点优先级") or "")) != "普通"
        )
    ]
    if tagged_abnormal:
        errors.append("异常行仍被打标签或提升优先级：" + "、".join(tagged_abnormal[:10]))

    if errors:
        raise RuntimeError("GZSYS FULL 写入前门禁失败：" + "；".join(errors))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=BASE_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def render_counter_table(counter: dict[str, int] | list[tuple[str, int]], empty: str = "| 无 | 0 |") -> str:
    items = counter.items() if isinstance(counter, dict) else counter
    lines = [f"| {name} | {count} |" for name, count in items if name]
    return "\n".join(lines) if lines else empty


def markdown_table_cell(value: Any) -> str:
    return clean_text(str(value or "")).replace("|", "\\|")


def write_report(path: Path, payload: dict[str, Any], csv_path: Path, xlsx_path: Path) -> None:
    meta = payload["meta"]
    full_append = meta.get("execution_mode") == "full_append"
    report_label = "全量采集归并审计报告" if full_append else "自动采集试跑报告"
    report_title = "官方医生全量采集归并审计报告" if full_append else "官方医生自动采集试跑报告"
    run_label = "全量采集" if full_append else "试采"
    xlsx_output = f"`{xlsx_path}`" if xlsx_path.exists() else "未生成（本轮使用 --no-xlsx）"
    top_departments = render_counter_table(payload["category_counts"][:20])
    group_lines = render_counter_table(dict(sorted(payload["group_counts"].items())))
    warning_lines = render_counter_table(payload["warning_counts"])
    category_error_lines = "\n".join(
        f"| {err.get('page', '')} | {err.get('url', '')} | {err.get('error', '')} |"
        for err in payload["category_errors"]
    )
    if not category_error_lines:
        category_error_lines = "| 无 | 无 | 无 |"
    detail_error_lines = "\n".join(
        f"| {err.get('source_link', '')} | {err.get('error', '')} |"
        for err in payload["detail_errors"][:50]
    )
    if not detail_error_lines:
        detail_error_lines = "| 无 | 无 |"
    entry_recon_lines = "\n".join(
        (
            f"| {item.get('category_name', '')} | {item.get('entry_url', '')} | "
            f"{item.get('page_nature', '')} | {item.get('list_page_count', 0)} | "
            f"{item.get('raw_detail_relation_count', item.get('unique_detail_count', 0))} | "
            f"{item.get('unique_detail_count', 0)} | {item.get('out_of_scope_detail_count', 0)} | "
            f"{item.get('affiliation', '')} | "
            f"{item.get('independent_entity_check', '')} |"
        )
        for item in payload.get("entry_reconnaissance", [])
    ) or "| 无 | 无 | 无 | 0 | 0 | 0 | 0 | 无 | 无 |"
    excluded_lines = "\n".join(
        (
            f"| {item.get('entry_url', '')} | {markdown_table_cell(item.get('list_title', ''))} | "
            f"{item.get('source_link', '')} | {item.get('reason', '')} |"
        )
        for item in payload.get("excluded_candidates", [])
    ) or "| 无 | 无 | 无 | 无 |"
    duplicate_lines = "\n".join(
        (
            f"| {item.get('name', '')} | {item.get('source_link', '')} | "
            f"{'；'.join(item.get('entry_urls', []))} |"
        )
        for item in payload.get("cross_entry_duplicates", [])
    ) or "| 无 | 无 | 无 |"
    same_name_lines = "\n".join(
        f"| {markdown_table_cell(name)} | {','.join(ids)} |"
        for name, ids in meta.get("census_same_name_groups", {}).items()
    ) or "| 无 | 无 |"
    identity_reconciliation_lines = "\n".join(
        (
            f"| {item.get('name', '')} | {item.get('resolution', '')} | "
            f"{item.get('relation_count', 0)} | {'、'.join(item.get('departments', []))} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gdzy5413_identity_reconciliation", [])
    ) or "| 无 | 无 | 0 | 无 | 无 | 无 |"
    gykqyy_reconciliation_lines = "\n".join(
        (
            f"| {item.get('detail_id', '')} | {markdown_table_cell(item.get('name', ''))} | "
            f"{item.get('resolution', '')} | {'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('source_link', '')} | {item.get('reason', '')} |"
        )
        for item in payload.get("gykqyy_identity_reconciliation", [])
    ) or "| 无 | 无 | 无 | 无 | 无 | 无 |"
    gyfyyy_reconciliation_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gyfyyy_identity_reconciliation", [])
        if int(item.get("relation_count") or 0) > 1
        or item.get("resolution") == "同名待甄别"
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 |"
    gy3y_reconciliation_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gy3y_identity_reconciliation", [])
        if int(item.get("relation_count") or 0) > 1
        or item.get("resolution") == "同名待甄别"
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 |"
    gzszyy_reconciliation_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | {item.get('detail_id', '')} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{'、'.join(item.get('campuses', [])) or '官网详情未标注'} | "
            f"{item.get('source_link', '')} |"
        )
        for item in payload.get("gzszyy_detail_reconciliation", [])
    ) or "| 无 | 无 | 无 | 无 | 无 |"
    gzszyy_identity_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{'、'.join(item.get('campuses', [])) or '官网详情未标注'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gzszyy_identity_reconciliation", [])
        if len(item.get("detail_ids", [])) > 1
        or item.get("resolution") == "同名待甄别"
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 | 无 |"
    gzszyy_care_site_lines = "\n".join(
        f"| {item.get('name', '')} | {item.get('source_url', '')} |"
        for item in payload.get("gzszyy_campus_reconnaissance", [])
    ) or "| 无 | 无 |"
    gzsys_reconciliation_lines = "\n".join(
        (
            f"| {item.get('detail_id', '')} | {markdown_table_cell(item.get('name', ''))} | "
            f"{item.get('resolution', '')} | {item.get('source_link', '')} | "
            f"{markdown_table_cell(item.get('reason', '')) or '无'} |"
        )
        for item in payload.get("gzsys_detail_reconciliation", [])
    ) or "| 无 | 无 | 无 | 无 | 无 |"
    fahsysu_reconciliation_lines = "\n".join(
        (
            f"| {item.get('detail_id', '')} | {markdown_table_cell(item.get('name', ''))} | "
            f"{item.get('resolution', '')} | {'、'.join(item.get('groups', [])) or '无'} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{'、'.join(item.get('title_hints', [])) or '无'} | "
            f"{item.get('relation_count', 0)} | {item.get('source_link', '')} | "
            f"{markdown_table_cell(item.get('reason', '')) or '无'} |"
        )
        for item in payload.get("fahsysu_detail_reconciliation", [])
    ) or "| 无 | 无 | 无 | 无 | 无 | 无 | 0 | 无 | 无 |"
    fahsysu_identity_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{markdown_table_cell(item.get('reason', '')) or '无'} |"
        )
        for item in payload.get("fahsysu_identity_reconciliation", [])
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 |"
    gdgh_identity_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gdgh_identity_reconciliation", [])
        if int(item.get("relation_count") or 0) > 1
        or item.get("resolution") == "同名待甄别"
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 |"
    gdgh_affiliate_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | {item.get('url', '')} | "
            f"{markdown_table_cell(item.get('relation', ''))} |"
        )
        for item in payload.get("affiliate_reconnaissance", [])
    ) or "| 无 | 无 | 无 |"
    gdgh_photo_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{markdown_table_cell(item.get('department', ''))} | "
            f"{markdown_table_cell(item.get('title', ''))} | "
            f"{markdown_table_cell(item.get('filename', ''))} | {item.get('bytes', 0)} | "
            f"`{item.get('sha256', '')}` | {item.get('photo_url', '')} |"
        )
        for item in payload.get("photo_samples", [])
    ) or "| 无 | 无 | 无 | 无 | 0 | 无 | 无 |"
    gdmch_excluded_lines = "\n".join(
        (
            f"| {item.get('detail_id', '')} | {markdown_table_cell(item.get('name', ''))} | "
            f"{markdown_table_cell(item.get('list_title', ''))} | {item.get('source_link', '')} | "
            f"{markdown_table_cell(item.get('reason', ''))} |"
        )
        for item in payload.get("excluded_candidates", [])
    ) or "| 无 | 无 | 无 | 无 | 无 |"
    gdmch_detail_lines = "\n".join(
        (
            f"| {item.get('detail_id', '')} | {markdown_table_cell(item.get('name', ''))} | "
            f"{'、'.join(item.get('departments', [])) or '官网详情未标注'} | "
            f"{'、'.join(item.get('campuses', [])) or '官网详情未标注'} | "
            f"{item.get('resolution', '')} | {item.get('source_link', '')} |"
        )
        for item in payload.get("gdmch_detail_reconciliation", [])
    ) or "| 无 | 无 | 无 | 无 | 无 | 无 |"
    gdmch_affiliate_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | {item.get('url', '')} | "
            f"{markdown_table_cell(item.get('relation', ''))} |"
        )
        for item in payload.get("affiliate_reconnaissance", [])
    ) or "| 无 | 无 | 无 |"
    gdmch_photo_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{markdown_table_cell(item.get('department', ''))} | "
            f"{markdown_table_cell(item.get('title', ''))} | "
            f"{markdown_table_cell(item.get('filename', ''))} | {item.get('bytes', 0)} | "
            f"{item.get('width', 0)}×{item.get('height', 0)} | "
            f"`{item.get('sha256', '')}` | {item.get('photo_url', '')} |"
        )
        for item in payload.get("photo_samples", [])
    ) or "| 无 | 无 | 无 | 无 | 0 | 0×0 | 无 | 无 |"
    gdmch_identity_lines = "\n".join(
        (
            f"| {markdown_table_cell(item.get('name', ''))} | "
            f"{','.join(str(value) for value in item.get('detail_ids', []))} | "
            f"{item.get('resolution', '')} | {item.get('relation_count', 0)} | "
            f"{'、'.join(item.get('departments', [])) or '无'} | "
            f"{item.get('primary_source_link', '')} | "
            f"{'；'.join(item.get('merged_source_links', [])) or '无'} |"
        )
        for item in payload.get("gdmch_identity_reconciliation", [])
        if int(item.get("relation_count") or 0) > 1
        or item.get("resolution") == "同名待甄别"
    ) or "| 无 | 无 | 无 | 0 | 无 | 无 | 无 |"
    campus_relation_summary = "；".join(
        f"{name} {count} 条"
        for name, count in meta.get("campus_relation_counts", {}).items()
    ) or "无"
    adapter_specific_sections: list[str] = []
    if payload.get("gdzy5413_identity_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广东省第二中医院同名归并对账

- 详情关系：{meta.get('gdzy5413_detail_relation_count', meta.get('gdzy5413_trial2_sample_relation_count', 0))}
- 最终身份：{meta.get('gdzy5413_final_identity_count', meta.get('gdzy5413_trial2_sample_identity_count', 0))}
- 白云院区样本：{meta.get('gdzy5413_trial2_baiyun_sample_count', 0)}
- 多链接同一人归并样本：{meta.get('gdzy5413_trial2_merged_identity_count', 0)}

| 姓名 | 裁决 | 详情关系 | 合并科室 | 主详情 | 其余详情 |
|---|---|---:|---|---|---|
{identity_reconciliation_lines}"""
        )
    if payload.get("gykqyy_identity_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广医口腔逐 ID 归并/排除对账

- 目录详情 ID：{meta.get('census_unique_detail_count', meta.get('unique_candidate_count', 0))}
- 有姓名详情 ID / 正式行：{meta.get('census_named_detail_count', 0)} / {meta.get('gykqyy_final_row_count', 0)}
- 空姓名详情 ID：{meta.get('census_blank_name_detail_count', 0)}
- 同名不同 ID 分行：{meta.get('gykqyy_same_name_separate_row_count', 0)}

| 详情 ID | 姓名 | 处置 | 科室 | 来源链接 | 理由 |
|---|---|---|---|---|---|
{gykqyy_reconciliation_lines}"""
        )
    if payload.get("gyfyyy_identity_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广医一院同名详情身份聚类对账

- 合规详情 ID：{len(payload.get('gyfyyy_detail_reconciliation', []))}
- 最终身份：{meta.get('gyfyyy_final_identity_count', 0)}
- 同一人归并组：{meta.get('gyfyyy_same_identity_merge_group_count', 0)}
- 实质不同同名身份：{meta.get('gyfyyy_distinct_same_name_group_count', 0)} 组 / {meta.get('gyfyyy_distinct_same_name_row_count', 0)} 行

| 姓名 | 详情 ID | 裁决 | 详情关系 | 合并科室 | 主详情 | 其余详情 |
|---|---|---|---:|---|---|---|
{gyfyyy_reconciliation_lines}"""
        )
    if payload.get("gy3y_detail_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广医三院两院区详情身份对账

- 静态目录详情 ID：{meta.get('census_unique_detail_count', 0)}
- 本轮已读取详情 ID：{len(payload.get('gy3y_detail_reconciliation', []))}
- 多院区/多科室详情 ID：{meta.get('gy3y_multi_relation_identity_count', 0)}
- 跨院区详情 ID：{meta.get('gy3y_cross_campus_identity_count', 0)}
- 护理身份核验：{meta.get('census_nursing_identity_status', '未单独记录')}

| 姓名 | 详情 ID | 裁决 | 详情关系 | 合并院区科室 | 主详情 | 其余详情 |
|---|---|---|---:|---|---|---|
{gy3y_reconciliation_lines}"""
        )
    if payload.get("gzszyy_detail_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广州市中医院院区/出诊点证据

- 顶层全院目录 / dp 科室树：{meta.get('gzszyy_unfiltered_page_count', 0)} 页、{meta.get('gzszyy_unfiltered_unique_detail_count', 0)} ID / {meta.get('pagination_count', 0)} 页、{meta.get('gzszyy_dp_unique_detail_count', 0)} ID
- 顶层目录专属详情：{meta.get('gzszyy_unfiltered_only_detail_count', 0)} 个（{'、'.join(meta.get('gzszyy_unfiltered_only_detail_ids', [])) or '无'}）；dp 树专属详情：{meta.get('gzszyy_dp_only_detail_count', 0)} 个
- 筛选链接：dp {meta.get('filter_link_counts', {}).get('dp', 0)} 个（科室）、pr {meta.get('filter_link_counts', {}).get('pr', 0)} 个（职称）、le {meta.get('filter_link_counts', {}).get('le', 0)} 个（专家级别）；pr/le 不重复采集
- 纯护理排除后合规候选：{meta.get('eligible_candidate_count', 0)} 个
- 最终身份：{meta.get('gzszyy_final_identity_count', len(payload.get('rows', [])))}；同一人归并 {meta.get('gzszyy_same_identity_merge_group_count', 0)} 组；实质不同同名 {meta.get('gzszyy_distinct_same_name_group_count', 0)} 组 / {meta.get('gzszyy_distinct_same_name_row_count', 0)} 行
- 官网公开院区/门诊部范围：{meta.get('gzszyy_official_care_site_count', 0)} 个
- 试采详情：{meta.get('gzszyy_sample_detail_count', 0)} 个；有二维码院区/出诊点标签 {meta.get('gzszyy_campus_tagged_sample_count', 0)} 个；未标注 {meta.get('gzszyy_campus_untagged_sample_count', 0)} 个
- 多院区/出诊点标签详情：{meta.get('cross_campus_detail_count', 0)} 个
- 详情标签计数：{campus_relation_summary}
- 字段处理：详情页明确标签与列表卡片科室共同保留在 `科室_列表卡片`；不推断院区与科室之间未由官网明示的组合关系。

| 官网公开院区/门诊部 | 官方链接 |
|---|---|
{gzszyy_care_site_lines}

| 姓名 | 详情 ID | 科室归属 | 详情二维码院区/出诊点 | 来源链接 |
|---|---|---|---|---|
{gzszyy_reconciliation_lines}

### 同名身份聚类裁决

| 姓名 | 详情 ID | 裁决 | 原详情关系 | 合并科室 | 院区/出诊点 | 主详情 | 其余详情 |
|---|---|---|---:|---|---|---|---|
{gzszyy_identity_lines}"""
        )
    if payload.get("gzsys_detail_reconciliation"):
        adapter_specific_sections.append(
            f"""## 中山大学孙逸仙纪念医院默认目录范围门禁

- 默认 All 目录：{meta.get('pagination_count', 0)} 页、{meta.get('candidate_membership_count', 0)} 张严格 `.card-4-0` 卡片、{meta.get('census_unique_detail_count', 0)} 个唯一数字 ID。
- 身份别名：`/node/<ID>` 与 `/doctor/<ID>` 只按同一数字 ID 去重；非卡片链接不构成授权详情。
- 页面筛选字典：科室 {meta.get('filter_dictionary_counts', {}).get('department_target_id', 0)}、人才项目 {meta.get('filter_dictionary_counts', {}).get('talent_project', 0)}、导师资格 {meta.get('filter_dictionary_counts', {}).get('tutor_qualification', 0)}、职称 {meta.get('filter_dictionary_counts', {}).get('doctor_title', 0)}；仅解析留痕，不遍历组合或构造关键词。
- 纯护理排除：{meta.get('excluded_non_doctor_count', 0)} 个；排除后合规候选 {meta.get('eligible_candidate_count', 0)} 个。
- 详情清洗：排班 DOM 排除 {meta.get('schedule_exclusion_count', 0)} 个；排名/患者片段排除 {meta.get('forbidden_segment_exclusion_count', 0)} 个；正式字段排班写入 {meta.get('schedule_field_ingested_count', 0)}、私用区字符 {meta.get('private_use_character_count', 0)}。
- 普通公开会话：{meta.get('standard_public_session', '未记录')}；本轮最终 Cookie 名称仅留痕为 `{'、'.join(meta.get('session_cookie_names', [])) or '无'}`。

### 逐 ID / 身份归并对账表

- 对账范围：664 个唯一数字 ID；正式身份 {meta.get('gzsys_final_identity_count', len(payload.get('rows', [])))} 行；护理排除 {meta.get('excluded_non_doctor_count', 0)} 行。
- 同一人归并：{meta.get('gzsys_same_identity_merge_group_count', 0)} 组；实质不同同名：{meta.get('gzsys_distinct_same_name_group_count', 0)} 组。

| 详情 ID | 姓名 | 裁决 | 来源链接 | 理由 |
|---|---|---|---|---|
{gzsys_reconciliation_lines}
"""
        )
    if payload.get("fahsysu_detail_reconciliation"):
        directory_campus_counts = "、".join(
            f"{marker}={count}"
            for marker, count in meta.get("campus_marker_counts", {}).items()
        ) or "无"
        detail_campus_counts = "、".join(
            f"{marker}={count}"
            for marker, count in meta.get("detail_campus_marker_counts", {}).items()
        ) or "无"
        adapter_specific_sections.append(
            f"""## 中山大学附属第一医院目录范围与 ID 门禁

- 官网服务端单页目录：顶层容器 {meta.get('census_group_count', 0)}（其中含医生关系 {meta.get('census_relationship_group_count', 0)}、空容器 {meta.get('census_empty_group_count', 0)}）、下级专科 {meta.get('census_department_count', 0)}、医生—专科关系 {meta.get('candidate_membership_count', 0)}、唯一数字 ID {meta.get('census_unique_detail_count', 0)}。
- 空顶层容器：{'、'.join(meta.get('census_empty_groups', [])) or '无'}。计入页面结构普查，但不构造医生或专科关系。
- 跨专科重复：{meta.get('cross_entry_duplicate_count', 0)} 条关系增量；同一数字 ID 的科室以顿号合并，不按姓名归并。
- 同名不同 ID：{meta.get('census_same_name_group_count', 0)} 组，全部按数字 ID 保持独立；样本命中时标记“同名待甄别”。
- 目录职级线索：正高 {meta.get('title_hint_counts', {}).get('正高', 0)}、副高 {meta.get('title_hint_counts', {}).get('副高', 0)}；正式职称只取详情页显式字段，不拼接正高/副高。
- 分页/交互：{meta.get('pagination_method', '未记录')}。
- 院区词扫描：目录页 {directory_campus_counts}；本轮详情 {detail_campus_counts}，涉及 {meta.get('campus_evidence_detail_count', 0)} 位。{meta.get('campus_scope_status', '')}
- 黄埔边界：{meta.get('huangpu_scope_status', '')}
- 黄埔去重预案：执行台账序号 8 时，必须以其目录数字 node ID 与本轮 860-ID 对账；命中本轮 ID 的医生不得重复入库。
- 详情清洗：排班 DOM 排除 {meta.get('schedule_exclusion_count', 0)} 个；排名/患者片段排除 {meta.get('forbidden_segment_exclusion_count', 0)} 个；正式字段排班写入 {meta.get('schedule_field_ingested_count', 0)}、私用区字符 {meta.get('private_use_character_count', 0)}。

### 逐 ID 对账

| 详情 ID | 姓名 | 裁决 | 顶层分组 | 科室 | 目录职级线索 | 原关系数 | 来源链接 | 理由 |
|---|---|---|---|---|---|---:|---|---|
{fahsysu_reconciliation_lines}

### 身份归并对账

- 正式身份：{meta.get('fahsysu_final_identity_count', len(payload.get('rows', [])))} 行；同一人归并 {meta.get('fahsysu_same_identity_merge_group_count', 0)} 组；实质不同同名 {meta.get('fahsysu_distinct_same_name_group_count', 0)} 组 / {meta.get('fahsysu_distinct_same_name_row_count', 0)} 行。

| 姓名 | 详情 ID | 裁决 | 原关系数 | 合并科室 | 主详情 | 理由 |
|---|---|---|---:|---|---|---|
{fahsysu_identity_lines}
"""
        )
    if payload.get("gdgh_detail_reconciliation"):
        adapter_specific_sections.append(
            f"""## 广东省人民医院官网目录、归属与照片{run_label}对账

- 顶层分组 / 科室：{meta.get('census_group_count', 0)} / {meta.get('census_department_count', 0)}。
- 医生—科室关系 / 唯一数字详情 ID：{meta.get('candidate_membership_count', 0)} / {meta.get('census_unique_detail_count', 0)}；跨科室复用 ID：{meta.get('gdgh_cross_department_identity_count', 0)}。
- 纯护理排除：{meta.get('excluded_non_doctor_count', 0)}；护理排除后合规候选：{meta.get('eligible_candidate_count', 0)}。
- 分页结论：{meta.get('pagination_method', '未记录')}。
- 普通公开会话：{meta.get('standard_public_session', '未记录')}。
- 详情清洗：排班片段排除 {meta.get('schedule_exclusion_count', 0)}，排名/患者片段排除 {meta.get('forbidden_segment_exclusion_count', 0)}，患者案例排除 {meta.get('patient_case_exclusion_count', 0)}；正式字段排班写入 {meta.get('schedule_field_ingested_count', 0)}、私用区字符 {meta.get('private_use_character_count', 0)}。
- 照片四数：应采 {meta.get('photo_expected_count', meta.get('photo_sample_count', 0))} / 实采 {meta.get('photo_downloaded_count', meta.get('photo_sample_count', 0))} / 失败 {meta.get('photo_failed_count', meta.get('photo_error_count', 0))} / 无照片 {meta.get('photo_no_source_count', 0)}。
- 平均照片大小：{meta.get('photo_average_bytes', 0)} bytes；按护理排除后 {meta.get('photo_estimated_full_count', 0)} 人估算全院照片容量：{meta.get('photo_estimated_full_bytes', 0)} bytes。
- 照片方案状态：`{meta.get('photo_policy_status', '未记录')}`；owner 已裁决原图不压缩、标准 Markdown 直接嵌入且不限宽，大图站点阈值为单张 >200KB 或宽 >800px 时另行回报。

### 分院/研究所归属证据

| 名称 | 官方链接 | 归属结论 |
|---|---|---|
{gdgh_affiliate_lines}

### {run_label}照片命名与校验对照

| 姓名 | 首个原子科室 | 主职称 | 文件名 | 字节数 | SHA-256 | 官网照片 |
|---|---|---|---|---:|---|---|
{gdgh_photo_lines}

### 同名身份聚类裁决

| 姓名 | 详情 ID | 裁决 | 原关系数 | 合并科室 | 主详情 | 其余详情 |
|---|---|---|---:|---|---|---|
{gdgh_identity_lines}
"""
        )
    if payload.get("gdmch_detail_reconciliation"):
        sample_departments = "、".join(meta.get("covered_departments", [])) or "无"
        adapter_specific_sections.append(
            f"""## 广东省妇幼保健院目录、四院区与照片{run_label}对账

- 官网服务端分页：{meta.get('pagination_count', 0)} 页；原始卡片 / 唯一数字详情 ID：{meta.get('candidate_membership_count', 0)} / {meta.get('census_unique_detail_count', 0)}。
- 重复关系：{meta.get('cross_entry_duplicate_count', 0)}；号源/系统账号/非医生排除：{meta.get('excluded_non_doctor_count', 0)}；排除后合规候选：{meta.get('eligible_candidate_count', 0)}。
- 科室结构：{meta.get('department_structure', '未记录')}。
- {run_label}覆盖科室：{meta.get('department_coverage_count', 0)} 个（{sample_departments}）；详情失败：{meta.get('detail_error_count', 0)}。
- 最终身份：{meta.get('gdmch_final_identity_count', len(payload.get('rows', [])))}；同一人归并 {meta.get('gdmch_same_identity_merge_group_count', 0)} 组；实质不同同名 {meta.get('gdmch_distinct_same_name_group_count', 0)} 组。
- 四院区官网归属证据：{meta.get('affiliate_count', 0)} 条；独立实体信号：{meta.get('independent_entity_count', 0)}。
- 详情清洗：排班片段排除 {meta.get('schedule_exclusion_count', 0)}，排名/患者片段排除 {meta.get('forbidden_segment_exclusion_count', 0)}，患者案例排除 {meta.get('patient_case_exclusion_count', 0)}；正式字段排班写入 {meta.get('schedule_field_ingested_count', 0)}、私用区字符 {meta.get('private_use_character_count', 0)}。
- 照片四数：应采 {meta.get('photo_expected_count', 0)} / 实采 {meta.get('photo_downloaded_count', 0)} / 失败 {meta.get('photo_failed_count', 0)} / 无照片 {meta.get('photo_no_source_count', 0)}。
- 本人职业照可得 / 官网默认占位图：{meta.get('photo_census_available_count', 0)} / {meta.get('photo_census_placeholder_count', 0)}（全目录默认占位图 {meta.get('photo_default_placeholder_count', 0)}）。
- 平均照片大小：{meta.get('photo_average_bytes', 0)} bytes；按 {meta.get('photo_estimated_full_count', 0)} 位有本人职业照候选估算全院照片容量：{meta.get('photo_estimated_full_bytes', 0)} bytes。
- 大图阈值：`{meta.get('large_photo_threshold', '单张 >200KB 或宽 >800px')}`；命中 {meta.get('large_photo_count', 0)} 张；照片政策状态：`{meta.get('photo_policy_status', '未记录')}`。原图不压缩；命中阈值时必须等待 owner 裁决本院 FULL 策略。
- 普通公开会话：{meta.get('standard_public_session', '未记录')}。

### 四院区官方归属证据

| 院区 | 官方链接 | 归属结论 |
|---|---|---|
{gdmch_affiliate_lines}

### {len(payload.get('gdmch_detail_reconciliation', []))} 个合规详情逐 ID 对账

| 详情 ID | 姓名 | 科室 | 院区 | 详情状态 | 来源链接 |
|---|---|---|---|---|---|
{gdmch_detail_lines}

### {meta.get('excluded_non_doctor_count', 0)} 个非医生候选逐 ID 排除表

| 详情 ID | 名称 | 列表身份 | 来源链接 | 排除理由 |
|---|---|---|---|---|
{gdmch_excluded_lines}

### {run_label}照片命名、字节、魔数、SHA-256 与尺寸对照

| 姓名 | 科室 | 主职称 | 文件名 | 字节数 | 宽×高 | SHA-256 | 官网照片 |
|---|---|---|---|---:|---:|---|---|
{gdmch_photo_lines}

### 同名身份聚类裁决

| 姓名 | 详情 ID | 裁决 | 原关系数 | 合并科室 | 主详情 | 其余详情 |
|---|---|---|---:|---|---|---|
{gdmch_identity_lines}
"""
        )
    adapter_specific_text = "\n\n".join(adapter_specific_sections)

    report = f"""---
类型: {report_label}
医院: {meta['hospital']}
城市: {meta['city']}
采集日期: {meta['collected_at']}
来源范围: 医院官网
采集入口: {meta['entry_url']}
适配器: {meta['adapter_id']}
---

# {meta['hospital']} {report_title}

## 结论

本次{run_label}只读取医院官网公开专家列表页和医生详情页。系统没有使用第三方平台、没有绕过登录或验证码、没有采集私人联系方式或患者信息。

本轮生成医生自动采集{run_label}底表，共 {meta['unique_doctor_count']} 位唯一医生；官网列表页原始卡片记录 {meta['raw_card_rows']} 条；读取入口分类 {meta['category_count']} 个；覆盖 {meta.get('department_coverage_count', 0)} 个科室；详情页失败 {meta['detail_error_count']} 条。

## 台账来源

| 项目 | 内容 |
|---|---|
| 城市 | {meta['city']} |
| 医院 | {meta['hospital']} |
| 官网首页 | {meta['homepage']} |
| 本轮医生入口 | {meta['entry_url']} |
| 入口来源 | {meta.get('entry_url_source', '官网入口台账')} |
| 原台账医生入口 | {meta.get('ledger_entry_url', meta['entry_url'])} |
| 台账人工复核 | {meta['ledger_review']} |
| 采集难度初判 | {meta['ledger_difficulty']} |

## 入口普查表

| 分类 | 官方入口 | 页面性质 | 列表分页 | 授权详情关系 | 唯一授权详情 | 范围外详情 | 归属医院/中心 | 独立实体核验 |
|---|---|---|---:|---:|---:|---:|---|---|
{entry_recon_lines}

### 动态目录专项证据

- 医生分页/载入方式：{meta.get('pagination_method', '按列表页分页读取')}
- 医生目录公开接口：{meta.get('directory_api', '不适用')}
- 医生详情公开接口：{meta.get('detail_api', '不适用')}
- 接口出处证据：{meta.get('api_source_evidence', '不适用')}
- 院区/分组：{meta.get('census_group_count', 0)} 个；科室分类：{meta.get('census_department_count', 0)} 个
- 医生-科室关系：{meta.get('candidate_membership_count', meta['raw_card_rows'])} 条
- 唯一详情 ID：{meta.get('census_unique_detail_count', meta.get('unique_candidate_count', meta['unique_doctor_count']))} 个
- 有姓名详情 ID：{meta.get('census_named_detail_count', meta.get('census_unique_nonblank_name_count', meta['unique_doctor_count']))} 个
- 空姓名详情 ID：{meta.get('census_blank_name_detail_count', 0)} 个
- 去重后的非空姓名值：{meta.get('census_unique_nonblank_name_count', meta['unique_doctor_count'])} 个
- 同名不同详情 ID：{meta.get('census_same_name_group_count', 0)} 组
- 非空/空科室块：{meta.get('census_nonempty_department_count', meta.get('census_department_count', 0))} / {meta.get('census_empty_department_count', 0)}
- 院区/出诊点标签关系：{campus_relation_summary}
- 跨院区/出诊点详情 ID：{meta.get('cross_campus_detail_count', meta.get('gy3y_cross_campus_identity_count', 0))} 个

| 同名 | 详情 ID |
|---|---|
{same_name_lines}

## 跨入口去重与试采覆盖

- 各入口唯一医生详情 URL 关系数：{meta.get('candidate_membership_count', meta['raw_card_rows'])}
- 跨入口去重后唯一候选：{meta.get('unique_candidate_count', meta['unique_doctor_count'])}
- 跨入口重复关系：{meta.get('cross_entry_duplicate_count', 0)}
- 试采覆盖入口分类：{meta.get('sample_entry_coverage_count', 0)} 个（{'、'.join(meta.get('sample_entry_categories', [])) or '无'}）

| 重复医生 | 唯一详情 URL | 出现入口 |
|---|---|---|
{duplicate_lines}

{adapter_specific_text}

## 已排除或范围外候选

| 入口 | 列表身份 | 来源链接 | 排除原因 |
|---|---|---|---|
{excluded_lines}

## 输出文件

- Excel 底表：{xlsx_output}
- CSV 底表：`{csv_path}`

## 采集统计

| 指标 | 数量 |
|---|---:|
| 读取列表文档数 | {meta['category_count']} |
| 原始医生卡片记录 | {meta['raw_card_rows']} |
| 跨入口去重前候选关系 | {meta.get('candidate_membership_count', meta['raw_card_rows'])} |
| 跨入口去重后唯一候选 | {meta.get('unique_candidate_count', meta['unique_doctor_count'])} |
| 排除非医生候选 | {meta.get('excluded_non_doctor_count', 0)} |
| 合规医生详情页 | {meta.get('census_unique_detail_count', meta['unique_doctor_count']) - meta.get('excluded_non_doctor_count', 0)} |
| 最终医生身份 | {meta['unique_doctor_count']} |
| 覆盖科室数 | {meta.get('department_coverage_count', 0)} |
| 列表页失败数 | {meta['category_error_count']} |
| 详情页失败数 | {meta['detail_error_count']} |
| 已建画像匹配数 | {meta['existing_profile_count']} |

## 重点关注范围统计

| 范围 | 医生数 |
|---|---:|
{group_lines}

## 科室数量 Top 20

| 科室 | 医生数 |
|---|---:|
{top_departments}

## 异常与复核提示

| 提示 | 数量 |
|---|---:|
{warning_lines}

## 列表页读取异常

| 页码 | URL | 错误 |
|---|---|---|
{category_error_lines}

## 详情页读取异常

| 来源链接 | 错误 |
|---|---|
{detail_error_lines}

## 人工复核建议

1. 优先复核“异常提示”不为空的医生。
2. “亮眼经历线索”只作为官方证据线索，不直接改写为对外宣传语。
3. 官网没有展示的擅长、经历、疾病标签保持空白，不补造。
4. 标记为试采门禁的适配器必须先完成小样本复核；只有取得 Claude 明确通过指令后才可全量追加。

## 合规边界

- 仅使用医院官网公开网页。
- 不采集私人电话、私人微信、家庭住址、患者隐私或非公开排班信息。
- 不使用第三方医疗平台评价、排名、患者评论。
- 不写“保证治愈”“包治疑难杂症”“疗效第一”等无法由官网证明的表达。
"""
    path.write_text(report, encoding="utf-8")


def write_master_report(path: Path, payload: dict[str, Any], csv_path: Path, xlsx_path: Path) -> None:
    meta = payload["meta"]
    hospital_lines = "\n".join(
        (
            f"| {row.get('医院', '')} | {row.get('医生数', 0)} | {row.get('采集日期', '')} | "
            f"{row.get('待复核数', 0)} | {row.get('已建画像数', 0)} | {row.get('采集入口', '')} |"
        )
        for row in payload.get("hospital_batches", [])
    )
    if not hospital_lines:
        hospital_lines = "| 无 | 0 |  | 0 | 0 |  |"

    group_lines = render_counter_table(dict(sorted(payload["group_counts"].items())))
    warning_lines = render_counter_table(payload["warning_counts"])

    report = f"""---
类型: 自动采集总底表更新报告
范围: 珠三角三甲医院
采集日期: {meta['collected_at']}
来源范围: 医院官网
适配器: {meta['adapter_id']}
---

# 珠三角三甲医院医生画像自动采集总底表更新报告

## 结论

总底表当前包含 {meta['hospital_count']} 家医院、{meta['unique_doctor_count']} 位医生。后续新增医院医生将继续写入同一张总底表，不再默认生成单院 Excel/CSV。

本次批次医院：{meta.get('current_batch_hospital') or '无，本次仅重建总表'}；本次批次原始医生数 {meta.get('current_batch_rows', 0)}；新增写入 {meta.get('new_rows_added', 0)}；重复跳过 {meta.get('duplicate_rows_skipped', 0)}；显式刷新既有记录 {meta.get('existing_rows_refreshed', 0)}；初始化合并时识别并折叠既有重复 {meta.get('existing_duplicate_rows', 0)}。

## 输出文件

- Excel 总底表：`{xlsx_path}`
- CSV 总底表：`{csv_path}`

## 医院批次说明

| 医院 | 医生数 | 采集日期 | 待复核数 | 已建画像数 | 采集入口 |
|---|---:|---|---:|---:|---|
{hospital_lines}

## 重点关注范围统计

| 范围 | 医生数 |
|---|---:|
{group_lines}

## 异常与复核提示

| 提示 | 数量 |
|---|---:|
{warning_lines}

## 合规边界

- 仅使用医院官网公开网页。
- 不采集私人电话、私人微信、家庭住址、患者隐私或非公开排班信息。
- 不使用第三方医疗平台评价、排名、患者评论。
- 官网没有展示的信息保持空白，不推断、不补造。
"""
    path.write_text(report, encoding="utf-8")


def find_node() -> str:
    if BUNDLED_NODE.exists():
        return str(BUNDLED_NODE)
    node = shutil.which("node")
    if node:
        return node
    raise RuntimeError("未找到 Node.js，无法生成 Excel 底表。")


def build_workbook(json_path: Path, xlsx_path: Path, preview_path: Path) -> None:
    if not WORKBOOK_BUILDER.exists():
        raise RuntimeError(f"Excel 生成脚本不存在：{WORKBOOK_BUILDER}")
    command = [
        find_node(),
        str(WORKBOOK_BUILDER),
        "--json",
        str(json_path),
        "--xlsx",
        str(xlsx_path),
        "--preview",
        str(preview_path),
    ]
    completed = subprocess.run(command, cwd=str(WORK_DIR), check=False)
    if completed.returncode != 0 and not xlsx_path.exists():
        raise RuntimeError(f"Excel 生成失败，退出码：{completed.returncode}")
    if completed.returncode != 0:
        print(f"[WARN] Excel 生成器返回非零退出码 {completed.returncode}，但输出文件已存在，继续生成报告。")
    inspect_path = Path(f"{xlsx_path}.inspect.ndjson")
    if inspect_path.exists():
        inspect_path.unlink()


def select_target(rows: list[dict[str, Any]], hospital: str | None) -> HospitalTarget:
    targets = confirmed_a_targets(rows, include_generic=bool(hospital))
    if not targets:
        generic_targets = confirmed_a_targets(rows, include_generic=True)
        if not hospital and generic_targets:
            supported = "、".join(target.hospital for target in generic_targets)
            raise RuntimeError(f"当前只有通用模板候选医院，请用 --hospital 指定医院后先试采。当前候选：{supported}")
        raise RuntimeError("台账中没有找到已确认可采集且入口完整的医院；请先补齐官网入口台账。")
    if hospital:
        for target in targets:
            if target.hospital == hospital:
                return target
        supported = "、".join(target.hospital for target in targets)
        raise RuntimeError(f"未找到指定医院的已确认采集入口。当前可测试：{supported}")

    dedicated_targets = [target for target in targets if target.adapter_id != GENERIC_ADAPTER_ID]
    if dedicated_targets:
        return dedicated_targets[0]

    supported = "、".join(target.hospital for target in confirmed_a_targets(rows, include_generic=True))
    raise RuntimeError(f"当前只有通用模板候选医院，请用 --hospital 指定医院后先试采。当前候选：{supported}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从已人工确认的医院官网入口批量采集医生基础画像底表。")
    parser.add_argument("--ledger", default=str(LEDGER_PATH), help="官网入口台账路径")
    parser.add_argument("--hospital", default="", help="指定医院名称；为空则选择首个已确认A级且已有适配器的医院")
    parser.add_argument(
        "--entry-url",
        action="append",
        default=[],
        help="显式覆盖医生目录入口；多入口时重复传入，且必须与医院官网同域",
    )
    parser.add_argument("--today", default=date.today().isoformat(), help="采集日期")
    parser.add_argument("--max-doctors", type=int, default=0, help="仅测试前 N 位医生；0 表示全量")
    parser.add_argument("--min-departments", type=int, default=0, help="要求结果至少覆盖的非空科室数；0 表示不设门禁")
    parser.add_argument(
        "--min-entry-categories",
        type=int,
        default=0,
        help="要求试采结果至少覆盖的显式入口分类数；0 表示不设门禁",
    )
    parser.add_argument("--max-pages", type=int, default=GENERIC_MAX_PAGES_DEFAULT, help="通用模板最多读取的列表分页数")
    parser.add_argument("--trial-only", action="store_true", help="仅试采并输出临时底表/报告，不追加统一总底表；未指定 --max-doctors 时默认试采 10 位")
    parser.add_argument("--force-generic", action="store_true", help="即使已有专用适配器，也强制使用通用模板试采")
    parser.add_argument("--allow-generic-append", action="store_true", help="允许通用模板结果追加统一总底表；建议先完成 --trial-only 人工复核")
    parser.add_argument(
        "--refresh-existing-hospital",
        action="store_true",
        help="仅在显式指定医院并通过追加门禁时，用本轮官网结果刷新同来源既有记录。",
    )
    parser.add_argument("--no-xlsx", action="store_true", help="只生成 JSON/CSV/报告，不生成 Excel")
    parser.add_argument(
        "--gdzy5413-trial2",
        action="store_true",
        help="Issue #17 TRIAL-2：仅从 cid=852 的 ksdoctorinfo 主目录抽样，仍普查 851 用于跨模式去重说明",
    )
    parser.add_argument("--single-output", action="store_true", help="保留旧模式：按单家医院生成单独底表")
    parser.add_argument("--rebuild-master-only", action="store_true", help="不联网采集，仅用已有底表重建总底表")
    parser.add_argument("--list-targets", action="store_true", help="列出当前已确认A级医院；无专用适配器的医院标记为通用模板候选")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    WORK_DIR.mkdir(parents=True, exist_ok=True)

    ledger_rows = read_ledger(Path(args.ledger))
    if args.list_targets:
        targets = confirmed_a_targets(ledger_rows, include_generic=True)
        print(json.dumps([target.__dict__ for target in targets], ensure_ascii=False, indent=2))
        return

    if args.refresh_existing_hospital and (not args.hospital or not args.allow_generic_append):
        raise RuntimeError("刷新既有医院记录必须同时指定 --hospital 和 --allow-generic-append。")

    if args.rebuild_master_only:
        previous_batch_meta: dict[str, Any] = {}
        if MASTER_JSON_PATH.exists():
            try:
                previous_payload = json.loads(MASTER_JSON_PATH.read_text(encoding="utf-8"))
                previous_batch_meta = previous_payload.get("meta", {})
            except (OSError, ValueError, TypeError):
                previous_batch_meta = {}
        payload = build_master_payload(args.today, batch_meta_override=previous_batch_meta)
        MASTER_JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        write_csv(MASTER_CSV_PATH, payload["rows"])
        if not args.no_xlsx:
            build_workbook(MASTER_JSON_PATH, MASTER_XLSX_PATH, MASTER_PREVIEW_PATH)
        write_master_report(MASTER_REPORT_PATH, payload, MASTER_CSV_PATH, MASTER_XLSX_PATH)
        print(
            json.dumps(
                {
                    "mode": "master_rebuild",
                    "hospitals": payload["meta"]["hospital_count"],
                    "rows": payload["meta"]["unique_doctor_count"],
                    "xlsx": str(MASTER_XLSX_PATH),
                    "csv": str(MASTER_CSV_PATH),
                    "report": str(MASTER_REPORT_PATH),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    target = select_target(ledger_rows, args.hospital or None)
    if args.gdzy5413_trial2 and (target.hospital != "广东省第二中医院" or not args.trial_only):
        raise RuntimeError("--gdzy5413-trial2 只允许广东省第二中医院的 --trial-only 补充试采。")
    if args.entry_url:
        if not args.hospital:
            raise RuntimeError("使用 --entry-url 覆盖入口时必须同时指定 --hospital。")
        override_urls: list[str] = []
        seen_override_urls: set[str] = set()
        official_host = comparable_host(target.homepage)
        for raw_url in args.entry_url:
            value = clean_text(raw_url)
            parsed = urlparse(value)
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                raise RuntimeError(f"无效的医生目录入口：{value}")
            if comparable_host(value) != official_host:
                raise RuntimeError(f"医生目录入口与医院官网不同域，拒绝执行：{value}")
            key = canonical_url(value)
            if key not in seen_override_urls:
                seen_override_urls.add(key)
                override_urls.append(value)
        target = replace(
            target,
            entry_url=override_urls[0],
            entry_urls=tuple(override_urls),
            ledger_entry_url=target.entry_url,
            adapter_id=adapter_for(override_urls[0], include_generic=True),
        )
    if args.force_generic:
        target = replace(target, adapter_id=GENERIC_ADAPTER_ID)
    print(
        f"selected: {target.city} {target.hospital} "
        f"{' '.join(effective_entry_urls(target))} adapter={target.adapter_id}"
    )

    if (
        target.adapter_id
        in {
            GENERIC_ADAPTER_ID,
            GDSKIN_ADAPTER_ID,
            NY5Y_ADAPTER_ID,
            GDZY5413_ADAPTER_ID,
            GYFYYY_ADAPTER_ID,
            GY3Y_ADAPTER_ID,
            GZBRAIN_ADAPTER_ID,
            GZSZYY_ADAPTER_ID,
            GZSYS_ADAPTER_ID,
            FAHSYSU_ADAPTER_ID,
            GDGH_ADAPTER_ID,
            GDMCH_ADAPTER_ID,
        }
        and not args.trial_only
        and not args.single_output
        and not args.allow_generic_append
    ):
        raise RuntimeError(
            "试采门禁适配器结果存在误识别风险。请先运行 --trial-only --max-doctors 10 试采复核；"
            "确认质量可接受后，再增加 --allow-generic-append 全量追加统一总底表。"
        )

    max_doctors = args.max_doctors or (10 if args.trial_only else None)
    if target.adapter_id == "gzzoc_drupal_doctor":
        payload = collect_gzzoc(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == "nbkjyy_static_expert":
        payload = collect_nbkj(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GYKQYY_ADAPTER_ID:
        payload = collect_gykqyy(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GYFYYY_ADAPTER_ID:
        payload = collect_gyfyyy(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GY3Y_ADAPTER_ID:
        payload = collect_gy3y(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GZBRAIN_ADAPTER_ID:
        payload = collect_gzbrain(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GZSZYY_ADAPTER_ID:
        payload = collect_gzszyy(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GZSYS_ADAPTER_ID:
        payload = collect_gzsys(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == FAHSYSU_ADAPTER_ID:
        payload = collect_fahsysu(target, args.today, max_doctors=max_doctors)
    elif target.adapter_id == GDGH_ADAPTER_ID:
        payload = collect_gdgh(
            target,
            args.today,
            max_doctors=max_doctors,
            full_mode=not args.trial_only,
        )
    elif target.adapter_id == GDMCH_ADAPTER_ID:
        payload = collect_gdmch(
            target,
            args.today,
            max_doctors=max_doctors,
            full_mode=not args.trial_only,
        )
    elif target.adapter_id in {GENERIC_ADAPTER_ID, GDSKIN_ADAPTER_ID, NY5Y_ADAPTER_ID, GDZY5413_ADAPTER_ID}:
        payload = collect_generic(
            target,
            args.today,
            max_doctors=max_doctors,
            max_pages=args.max_pages,
            gdzy5413_trial2=args.gdzy5413_trial2,
        )
    else:
        raise RuntimeError(f"暂不支持的适配器：{target.adapter_id}")

    if args.entry_url:
        failed_entries = [
            entry_url
            for entry_url, count in payload["meta"].get("entry_candidate_counts", {}).items()
            if int(count) == 0
        ]
        empty_entries_blocking = [
            entry_url
            for entry_url in failed_entries
            if not (
                (target.adapter_id == GDSKIN_ADAPTER_ID and gdskin_entry_id(entry_url) == "924")
                or (target.adapter_id == GDZY5413_ADAPTER_ID and gdzy5413_entry_kind(entry_url) == "852")
            )
        ]
        if payload["meta"].get("category_error_count") or empty_entries_blocking:
            raise RuntimeError(
                "显式多入口采集不完整："
                f"列表错误 {payload['meta'].get('category_error_count', 0)} 条，"
                f"无候选入口 {'、'.join(empty_entries_blocking) or '无'}"
            )

    if args.min_entry_categories:
        coverage_count = int(payload["meta"].get("sample_entry_coverage_count") or 0)
        if coverage_count < args.min_entry_categories:
            raise RuntimeError(
                "入口分类覆盖门禁不满足："
                f"要求至少 {args.min_entry_categories} 个，实际 {coverage_count} 个："
                f"{'、'.join(payload['meta'].get('sample_entry_categories', [])) or '无'}"
            )

    covered_departments = (
        gdmch_covered_department_names(payload["rows"])
        if target.adapter_id == GDMCH_ADAPTER_ID
        else covered_department_names(payload["rows"])
    )
    payload["meta"]["department_coverage_count"] = len(covered_departments)
    payload["meta"]["covered_departments"] = covered_departments
    if args.min_departments and len(covered_departments) < args.min_departments:
        raise RuntimeError(
            f"科室覆盖门禁不满足：要求至少 {args.min_departments} 个，实际 {len(covered_departments)} 个："
            f"{'、'.join(covered_departments) or '无'}"
        )

    if args.gdzy5413_trial2:
        validate_gdzy5413_trial2(payload, expected_identities=max_doctors or 10)
    if target.adapter_id == GZSYS_ADAPTER_ID and args.trial_only:
        validate_gzsys_trial(payload, expected_rows=max_doctors or 10)
    if target.adapter_id == FAHSYSU_ADAPTER_ID and args.trial_only:
        validate_fahsysu_trial(payload, expected_rows=max_doctors or 10)
    if target.adapter_id == GDGH_ADAPTER_ID and args.trial_only:
        validate_gdgh_trial(payload, expected_rows=max_doctors or 10)
    if target.adapter_id == GDMCH_ADAPTER_ID and args.trial_only:
        validate_gdmch_trial(payload, expected_rows=max_doctors or 10)
    if target.adapter_id == GDGH_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gdgh_full_append(payload)
    if target.adapter_id == GDMCH_ADAPTER_ID and not args.trial_only:
        validate_gdmch_full_append(payload)

    if target.adapter_id == FAHSYSU_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_fahsysu_full_append(payload)

    if target.adapter_id == GZSYS_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gzsys_full_append(payload)

    if target.adapter_id == GDZY5413_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gdzy5413_full_append(payload)
    if target.adapter_id == GYKQYY_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gykqyy_full_append(payload)
    if target.adapter_id == GYFYYY_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gyfyyy_full_append(payload)
    if target.adapter_id == GY3Y_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gy3y_full_append(payload)
    if target.adapter_id == GZBRAIN_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gzbrain_full_append(payload)
    if target.adapter_id == GZSZYY_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gzszyy_full_append(payload)
    if target.adapter_id == GDSKIN_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_gdskin_full_append(payload)
    if target.adapter_id == NY5Y_ADAPTER_ID and not args.trial_only and not args.single_output:
        validate_ny5y_full_append(payload)

    safe_name = safe_file_part(target.hospital)
    json_path = WORK_DIR / f"{safe_name}_official_doctors_payload.json"
    preview_path = WORK_DIR / f"{safe_name}_official_doctors_preview.png"

    if args.trial_only:
        payload["meta"]["execution_mode"] = "trial"
        json_path = WORK_DIR / f"{safe_name}_trial_payload.json"
        csv_path = WORK_DIR / f"{safe_name}_trial_doctors.csv"
        xlsx_path = WORK_DIR / f"{safe_name}_trial_doctors.xlsx"
        report_path = WORK_DIR / f"{safe_name}_trial_report.md"

        payload["meta"]["json_path"] = str(json_path)
        payload["meta"]["csv_path"] = str(csv_path)
        payload["meta"]["xlsx_path"] = "" if args.no_xlsx else str(xlsx_path)
        payload["meta"]["preview_path"] = "" if args.no_xlsx else str(preview_path)

        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        write_csv(csv_path, payload["rows"])

        if not args.no_xlsx:
            build_workbook(json_path, xlsx_path, preview_path)

        write_report(report_path, payload, csv_path, xlsx_path)
        print(
            json.dumps(
                {
                    "mode": "trial_only",
                    "hospital": target.hospital,
                    "adapter_id": target.adapter_id,
                    "rows": payload["meta"]["unique_doctor_count"],
                    "detail_errors": payload["meta"]["detail_error_count"],
                    "csv": str(csv_path),
                    "xlsx": "" if args.no_xlsx else str(xlsx_path),
                    "report": str(report_path),
                    "master_updated": False,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    if args.single_output:
        csv_path = SOURCE_DIR / f"{target.hospital}_全院医生自动采集底表.csv"
        xlsx_path = SOURCE_DIR / f"{target.hospital}_全院医生自动采集底表.xlsx"
        report_path = SOURCE_DIR / f"{target.hospital}_全院医生自动采集试跑报告.md"

        payload["meta"]["json_path"] = str(json_path)
        payload["meta"]["csv_path"] = str(csv_path)
        payload["meta"]["xlsx_path"] = str(xlsx_path)
        payload["meta"]["preview_path"] = str(preview_path)

        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        write_csv(csv_path, payload["rows"])

        if not args.no_xlsx:
            build_workbook(json_path, xlsx_path, preview_path)

        write_report(report_path, payload, csv_path, xlsx_path)
        print(
            json.dumps(
                {
                    "mode": "single_output",
                    "hospital": target.hospital,
                    "rows": payload["meta"]["unique_doctor_count"],
                    "detail_errors": payload["meta"]["detail_error_count"],
                    "csv": str(csv_path),
                    "xlsx": str(xlsx_path),
                    "report": str(report_path),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    payload["meta"]["execution_mode"] = "full_append"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    master_payload = build_master_payload(
        args.today,
        incoming_payload=payload,
        refresh_incoming=args.refresh_existing_hospital,
        replace_incoming_hospital=target.adapter_id == GDZY5413_ADAPTER_ID,
    )
    MASTER_JSON_PATH.write_text(json.dumps(master_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(MASTER_CSV_PATH, master_payload["rows"])

    if not args.no_xlsx:
        build_workbook(MASTER_JSON_PATH, MASTER_XLSX_PATH, MASTER_PREVIEW_PATH)

    write_master_report(MASTER_REPORT_PATH, master_payload, MASTER_CSV_PATH, MASTER_XLSX_PATH)
    full_report_path = WORK_DIR / f"{safe_name}_official_doctors_report.md"
    write_report(full_report_path, payload, MASTER_CSV_PATH, MASTER_XLSX_PATH)
    print(
        json.dumps(
            {
                "mode": "master_append",
                "hospital": target.hospital,
                "batch_rows": payload["meta"]["unique_doctor_count"],
                "new_rows_added": master_payload["meta"]["new_rows_added"],
                "duplicate_rows_skipped": master_payload["meta"]["duplicate_rows_skipped"],
                "existing_rows_refreshed": master_payload["meta"]["existing_rows_refreshed"],
                "master_rows": master_payload["meta"]["unique_doctor_count"],
                "detail_errors": payload["meta"]["detail_error_count"],
                "csv": str(MASTER_CSV_PATH),
                "xlsx": str(MASTER_XLSX_PATH),
                "report": str(MASTER_REPORT_PATH),
                "hospital_audit_report": str(full_report_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        raise SystemExit("用户中断")
    except Exception as exc:  # noqa: BLE001 - CLI should report exact blocker
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
