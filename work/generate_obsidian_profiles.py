from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET


ROOT = Path(r"D:\workspace\信息收集整理")
VAULT = ROOT / "医生画像仓库"
SOURCE_DIR = VAULT / "99_资料来源"
PROFILE_ROOT = VAULT / "01_试点医院"
DEFAULT_MASTER_CSV = SOURCE_DIR / "珠三角三甲医院_医生画像自动采集总底表.csv"
DEFAULT_MASTER_XLSX = SOURCE_DIR / "珠三角三甲医院_医生画像自动采集总底表.xlsx"
DEFAULT_REPORT = SOURCE_DIR / "珠三角三甲医院_Obsidian画像全量生成报告.md"
DEFAULT_SUPPLEMENT_REPORT = SOURCE_DIR / "珠三角三甲医院_Obsidian画像补充生成报告.md"
DEFAULT_MISSING_REPORT = SOURCE_DIR / "珠三角三甲医院_Obsidian缺失画像补充生成报告.md"
TEMPLATE_PATH = VAULT / "模板" / "医生画像模板.md"

AUTO_MARKER = "<!-- AUTO-GENERATED-BY: work/generate_obsidian_profiles.py -->"
DEFAULT_SKIP_HOSPITALS = {"中山大学附属第五医院"}

GROUP_ORDER = [
    "慢性病",
    "肿瘤",
    "生殖疾病",
    "免疫/风湿/感染",
    "术后恢复/康复",
    "疑难重症",
]

CORE_FIELDS = ("医院", "姓名", "来源链接")


def clean(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def clip(value: object | None, limit: int) -> str:
    text = clean(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip("，。；;、 ") + "…"


def yaml_value(value: object | None) -> str:
    return json.dumps(clean(value), ensure_ascii=False)


def md_escape(value: object | None) -> str:
    return clean(value).replace("|", "\\|")


def safe_filename(value: object | None, max_len: int = 80) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", clean(value))
    if len(name) > max_len:
        name = name[:max_len].rstrip(" .，,。；;、")
    return name.strip(" .") or "未命名"


def split_terms(value: object | None) -> list[str]:
    return [term for term in re.split(r"[、,，；;]\s*", clean(value)) if term]


def split_sentences(value: object | None, limit: int = 6, max_each: int = 160) -> list[str]:
    parts = [clean(part) for part in re.split(r"(?<=[。！？；;])|\n", clean(value)) if clean(part)]
    seen: set[str] = set()
    result: list[str] = []
    for part in parts:
        item = clip(part, max_each)
        key = re.sub(r"[\s，,。；;：:、（）()]", "", item)
        if item and key and key not in seen:
            seen.add(key)
            result.append(item)
        if len(result) >= limit:
            break
    return result


def first_value(row: dict[str, str], *fields: str) -> str:
    for field in fields:
        value = clean(row.get(field))
        if value:
            return value
    return ""


def department(row: dict[str, str]) -> str:
    return first_value(row, "科室", "科室_分类页", "科室_列表卡片")


def title(row: dict[str, str]) -> str:
    return first_value(row, "职称", "职称身份原文", "职称_关键词")


def specialty(row: dict[str, str]) -> str:
    return first_value(row, "简介/擅长", "擅长诊疗方向摘录", "列表简介")


def row_number(row: dict[str, str], fallback: int) -> str:
    return clean(row.get("序号")) or str(fallback)


def source_hash(source: str) -> str:
    return hashlib.sha1(source.encode("utf-8")).hexdigest()[:8]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def cell_to_text(value: object | None) -> str:
    return clean(value)


def read_xlsx_rows_with_openpyxl(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_to_text(value) for value in rows[0]]
    result: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        if not any(cell_to_text(value) for value in raw_row):
            continue
        result.append(
            {
                headers[index]: cell_to_text(value)
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
        )
    return result


def column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for char in letters:
        index = index * 26 + ord(char) - ord("A") + 1
    return index - 1


def read_shared_strings(package: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(package.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    shared: list[str] = []
    for item in root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
        shared.append("".join(text.text or "" for text in item.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")))
    return shared


def workbook_first_sheet_path(package: zipfile.ZipFile) -> str:
    workbook = ET.fromstring(package.read("xl/workbook.xml"))
    rels = ET.fromstring(package.read("xl/_rels/workbook.xml.rels"))
    rel_by_id = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
    }
    sheet = workbook.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet")
    if sheet is None:
        raise ValueError("XLSX 中未找到工作表")
    rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
    target = rel_by_id[rel_id]
    if target.startswith("/"):
        return target.lstrip("/")
    return "xl/" + target


def read_xlsx_rows_basic(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as package:
        shared = read_shared_strings(package)
        sheet_path = workbook_first_sheet_path(package)
        root = ET.fromstring(package.read(sheet_path))
    rows: list[list[str]] = []
    for row in root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"):
        values: list[str] = []
        for cell in row.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
            index = column_index(cell.attrib.get("r", "A1"))
            while len(values) <= index:
                values.append("")
            cell_type = cell.attrib.get("t")
            value = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
            if cell_type == "s" and value is not None and value.text is not None:
                values[index] = shared[int(value.text)]
            elif cell_type == "inlineStr":
                values[index] = "".join(
                    text.text or ""
                    for text in cell.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
                )
            elif value is not None and value.text is not None:
                values[index] = clean(value.text)
        rows.append(values)
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    result: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        if not any(clean(value) for value in raw_row):
            continue
        result.append(
            {
                headers[index]: clean(value)
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
        )
    return result


def read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        return read_xlsx_rows_with_openpyxl(path)
    except ImportError:
        return read_xlsx_rows_basic(path)


def read_master_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    raise ValueError(f"不支持的总底表格式：{path}")


def choose_source(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
    elif DEFAULT_MASTER_CSV.exists():
        path = DEFAULT_MASTER_CSV
    else:
        path = DEFAULT_MASTER_XLSX
    if not path.exists():
        raise FileNotFoundError(f"总底表不存在：{path}")
    return path


def validate_source_rows(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        return ["总底表没有可读取记录"]
    fields = set(rows[0].keys())
    return [field for field in CORE_FIELDS if field not in fields]


def extract_existing_sources(hospital_dir: Path) -> dict[str, Path]:
    sources: dict[str, Path] = {}
    if not hospital_dir.exists():
        return sources
    for path in hospital_dir.glob("*.md"):
        if path.name == "_索引.md":
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        for match in re.finditer(r"https?://[^\s)>\"]+", text):
            sources.setdefault(match.group(0).rstrip("，。；;"), path)
    return sources


def is_auto_generated_profile(path: Path) -> bool:
    if not path.exists():
        return False
    text = path.read_text(encoding="utf-8", errors="ignore")
    return AUTO_MARKER in text


def candidate_names(row: dict[str, str], fallback_index: int) -> list[str]:
    name = safe_filename(row.get("姓名"))
    dept = safe_filename(department(row))
    seq = safe_filename(row_number(row, fallback_index))
    source = clean(row.get("来源链接"))
    candidates = [f"{name}.md"]
    if dept and dept != "未命名":
        candidates.append(f"{name}_{dept}.md")
    candidates.append(f"{name}_{seq}.md")
    if source:
        candidates.append(f"{name}_{source_hash(source)}.md")
    result: list[str] = []
    for candidate in candidates:
        if candidate not in result:
            result.append(candidate)
    return result


def select_profile_path(
    row: dict[str, str],
    hospital_dir: Path,
    used_names: set[str],
    existing_names: set[str],
    existing_sources: dict[str, Path],
    fallback_index: int,
    allow_overwrite: bool,
) -> tuple[Path | None, str]:
    source = clean(row.get("来源链接"))
    if source in existing_sources:
        existing_path = existing_sources[source]
        if allow_overwrite:
            return existing_path, ""
        return None, "已有同来源画像，默认保护不覆盖"

    for candidate in candidate_names(row, fallback_index):
        path = hospital_dir / candidate
        if candidate in existing_names:
            if allow_overwrite:
                return path, ""
            return None, f"已有同名画像文件，默认保护不覆盖：{candidate}"
        if candidate not in used_names:
            used_names.add(candidate)
            return path, ""

    return None, "无法生成不冲突的文件名"


def yaml_list(key: str, values: Iterable[str]) -> list[str]:
    values = [clean(value) for value in values if clean(value)]
    if not values:
        return [f"{key}: []"]
    return [f"{key}:"] + [f"  - {yaml_value(value)}" for value in values]


def add_field_section(lines: list[str], title_text: str, value: str, limit: int = 700) -> None:
    lines.extend(["", f"## {title_text}", ""])
    clipped = clip(value, limit)
    if clipped:
        lines.append(clipped)


def add_term_section(lines: list[str], title_text: str, values: list[str]) -> None:
    lines.extend(["", f"## {title_text}", ""])
    lines.extend(f"- {value}" for value in values)


def unique_terms(*groups: Iterable[str]) -> list[str]:
    result: list[str] = []
    for group in groups:
        for item in group:
            item = clean(item)
            if item and item not in result:
                result.append(item)
    return result


def group_tag_lines(focus: list[str], disease_tags: list[str]) -> list[str]:
    lines: list[str] = []
    for group in GROUP_ORDER:
        label = "免疫/风湿" if group == "免疫/风湿/感染" else group
        tags = disease_tags if group in focus else []
        lines.append(f"- {label}：{'、'.join(tags)}")
    return lines


def html_to_markdown_text(value: object | None) -> str:
    text = clean(value)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def official_profile_text(row: dict[str, str]) -> str:
    return html_to_markdown_text(row.get("_报告修正正文")) or specialty(row)


def build_bd_summary(row: dict[str, str], focus: list[str]) -> str:
    name = clean(row.get("姓名"))
    hospital = clean(row.get("医院"))
    dept = department(row)
    focus_text = "、".join(focus)
    if focus_text:
        return (
            f"{name}医生可作为{focus_text}方向的基础画像候选。"
            "后续 BD 使用应继续以官网来源和人工复核为准，不添加疗效承诺或无来源包装。"
        )
    if dept:
        return (
            f"{name}医生可作为{hospital}{dept}的基础医生资料入库。"
            "当前重点方向以总底表标签为准，官网未展示的信息保持空白。"
        )
    return "当前画像仅作基础资料归档，后续使用需先完成人工复核。"


def build_profile(row: dict[str, str], generated_at: str) -> str:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"医生画像模板不存在：{TEMPLATE_PATH}")

    name = clean(row.get("姓名"))
    hospital = clean(row.get("医院"))
    dept = department(row)
    doctor_title = title(row)
    source = clean(row.get("来源链接"))
    focus = split_terms(row.get("重点关注范围"))
    disease_tags = split_terms(row.get("重点疾病标签"))
    specialty_text = official_profile_text(row)
    detail_excerpt = clean(row.get("详情正文摘录"))
    highlight_text = clean(row.get("亮眼经历线索"))
    review_status = clean(row.get("复核状态"))
    direction = unique_terms(focus, disease_tags)[:12]
    evidence_items = []
    if doctor_title:
        evidence_items.append(f"官网身份字段：{clip(doctor_title, 120)}")
    if specialty_text:
        evidence_items.append(f"官网简介/擅长：{clip(specialty_text, 220)}")
    if highlight_text:
        evidence_items.append(f"官网亮眼经历线索：{clip(highlight_text, 220)}")
    if clean(row.get("_原异常提示")):
        evidence_items.append(f"原异常提示：{clip(row.get('_原异常提示'), 160)}")
    evidence_items.append(f"来源链接：{source}")

    yaml_lines = [
        "---",
        "类型: 医生画像",
        "采集状态: 正式画像",
        f"医院: {yaml_value(hospital)}",
        f"科室: {yaml_value(dept)}",
        f"姓名: {yaml_value(name)}",
        f"职称: {yaml_value(doctor_title)}",
    ]
    yaml_lines.extend(yaml_list("重点方向", direction))
    yaml_lines.extend(
        [
            "来源类型: 医院官网",
            f"来源链接: {yaml_value(source)}",
            f"采集日期: {yaml_value(row.get('采集日期'))}",
            f"复核状态: {yaml_value(review_status or '待复核')}",
            "---",
            AUTO_MARKER,
            "",
            f"<!-- TEMPLATE: {TEMPLATE_PATH} -->",
            "",
        ]
    )

    lines: list[str] = []
    lines.extend(yaml_lines)
    lines.extend(
        [
            f"# {name}",
            "",
            "## 基础信息",
            "",
            "| 字段 | 内容 |",
            "|---|---|",
            f"| 姓名 | {md_escape(name)} |",
            f"| 医院 | {md_escape(hospital)} |",
            f"| 科室 | {md_escape(dept)} |",
            f"| 职称/身份 | {md_escape(doctor_title)} |",
            f"| 来源链接 | [官网原文]({source}) |",
            f"| 采集日期 | {md_escape(row.get('采集日期'))} |",
            "",
            "## 简介/擅长",
            "",
        ]
    )
    if specialty_text:
        lines.append(clip(specialty_text, 1100))
    elif detail_excerpt:
        lines.append(clip(detail_excerpt, 900))

    lines.extend(["", "## 可包装亮点", ""])
    highlight_items = split_sentences(highlight_text or specialty_text, limit=5, max_each=180)
    if highlight_items:
        lines.extend(f"- {item}" for item in highlight_items)
    else:
        lines.append("- ")

    lines.extend(["", "## 重点疾病/人群标签", ""])
    lines.extend(group_tag_lines(focus, disease_tags))

    lines.extend(["", "## 证据摘录", ""])
    lines.append("> 只摘录官网公开信息中的短句或关键字段，不复制大段原文。")
    lines.append("")
    lines.extend(f"- {item}" for item in evidence_items)

    lines.extend(["", "## BD 画像摘要", ""])
    lines.append(build_bd_summary(row, focus))

    lines.extend(
        [
            "",
            "## 合规边界",
            "",
            "- 仅使用医院官网、官方公众号、官方小程序等公开渠道。",
            "- 不采集私人电话、私人微信、家庭住址、患者隐私或非公开排班信息。",
            "- 不写“保证治愈”“疗效第一”“包治疑难杂症”等无法由官网证明的表述。",
            "",
        ]
    )
    return "\n".join(lines)


def build_index(hospital: str, created_rows: list[dict[str, str]], skipped_rows: list[dict[str, str]], path_by_source: dict[str, Path], generated_at: str) -> str:
    focus_counter: Counter[str] = Counter()
    priority_counter: Counter[str] = Counter()
    anomaly_count = 0
    for row in created_rows:
        focus_counter.update(split_terms(row.get("重点关注范围")))
        priority_counter.update([clean(row.get("重点优先级"))])
        if clean(row.get("异常提示")):
            anomaly_count += 1

    lines = [
        "---",
        '类型: "医院医生画像索引"',
        f"医院: {yaml_value(hospital)}",
        f"生成时间: {yaml_value(generated_at)}",
        f"画像数量: {len(created_rows)}",
        f"跳过数量: {len(skipped_rows)}",
        "---",
        AUTO_MARKER,
        "",
        f"# {hospital} 医生画像索引",
        "",
        "## 生成概览",
        "",
        f"- 本次生成医生画像：{len(created_rows)}",
        f"- 本院跳过记录：{len(skipped_rows)}",
        f"- 异常提示不为空：{anomaly_count}",
        f"- 复核状态：沿用统一总底表字段",
        "",
        "## 重点优先级统计",
        "",
        "| 重点优先级 | 医生数 |",
        "|---|---:|",
    ]
    for priority, count in priority_counter.most_common():
        lines.append(f"| {md_escape(priority)} | {count} |")
    lines.extend(["", "## 重点关注范围统计", "", "| 重点关注范围 | 医生数 |", "|---|---:|"])
    for group in GROUP_ORDER:
        lines.append(f"| {md_escape(group)} | {focus_counter.get(group, 0)} |")
    for group, count in focus_counter.most_common():
        if group not in GROUP_ORDER:
            lines.append(f"| {md_escape(group)} | {count} |")

    lines.extend(
        [
            "",
            "## 医生画像列表",
            "",
            "| 序号 | 医生 | 科室 | 职称/身份 | 重点优先级 | 重点关注范围 | 异常提示 | 来源 |",
            "|---:|---|---|---|---|---|---|---|",
        ]
    )
    for row in created_rows:
        source = clean(row.get("来源链接"))
        path = path_by_source[source]
        lines.append(
            "| {seq} | [[{wiki}]] | {dept} | {title} | {priority} | {focus} | {anomaly} | [官网]({source}) |".format(
                seq=md_escape(row.get("_profile_seq")),
                wiki=path.stem,
                dept=md_escape(department(row)),
                title=md_escape(title(row)),
                priority=md_escape(row.get("重点优先级")),
                focus=md_escape("、".join(split_terms(row.get("重点关注范围")))),
                anomaly=md_escape(row.get("异常提示")),
                source=source,
            )
        )

    if skipped_rows:
        lines.extend(
            [
                "",
                "## 未生成记录",
                "",
                "| 序号 | 姓名 | 科室 | 未生成原因 |",
                "|---:|---|---|---|",
            ]
        )
        for row in skipped_rows:
            lines.append(
                "| {seq} | {name} | {dept} | {reason} |".format(
                    seq=md_escape(row.get("序号")),
                    name=md_escape(row.get("姓名")),
                    dept=md_escape(department(row)),
                    reason=md_escape(row.get("_skip_reason")),
                )
            )

    lines.extend(
        [
            "",
            "## 使用边界",
            "",
            "- 本索引只汇总统一总底表中的医院官网字段。",
            "- 医生画像为待人工复核资料，不作为对外宣传成稿。",
            "",
        ]
    )
    return "\n".join(lines)


def skip_reason_for_core_fields(row: dict[str, str]) -> str:
    missing = [field for field in CORE_FIELDS if not clean(row.get(field))]
    if missing:
        return "核心字段缺失：" + "、".join(missing)
    return ""


def build_report(
    rows: list[dict[str, str]],
    source_path: Path,
    report_path: Path,
    output_root: Path,
    hospital_stats: dict[str, dict[str, object]],
    skipped_rows: list[dict[str, str]],
    generated_at: str,
) -> str:
    total = len(rows)
    success = sum(int(stats["generated"]) for stats in hospital_stats.values())
    skipped = len(skipped_rows)
    anomalies = [row for row in rows if clean(row.get("异常提示"))]
    suspicious_names = [row for row in rows if len(clean(row.get("姓名"))) > 30]
    reason_counter = Counter(row.get("_skip_reason", "未记录原因") for row in skipped_rows)

    lines = [
        "---",
        '类型: "画像生成报告"',
        '数据来源: "统一总底表"',
        f"生成时间: {yaml_value(generated_at)}",
        f"总底表: {yaml_value(str(source_path))}",
        f"输出目录: {yaml_value(str(output_root))}",
        "---",
        AUTO_MARKER,
        "",
        "# 珠三角三甲医院 Obsidian 医生画像全量生成报告",
        "",
        "## 总览",
        "",
        f"- 总医生数：{total}",
        f"- 成功生成数：{success}",
        f"- 跳过数：{skipped}",
        f"- 异常提示不为空医生数：{len(anomalies)}",
        f"- 总底表来源：`{source_path}`",
        f"- 画像输出根目录：`{output_root}`",
        "",
        "## 按医院统计",
        "",
        "| 医院 | 底表记录数 | 成功生成数 | 跳过数 | 异常提示数 | 索引页 |",
        "|---|---:|---:|---:|---:|---|",
    ]
    for hospital in sorted(hospital_stats):
        stats = hospital_stats[hospital]
        index_path = clean(stats.get("index_path"))
        index_display = f"`{index_path}`" if index_path else "未生成"
        lines.append(
            "| {hospital} | {source} | {generated} | {skipped} | {anomalies} | {index} |".format(
                hospital=md_escape(hospital),
                source=stats["source"],
                generated=stats["generated"],
                skipped=stats["skipped"],
                anomalies=stats["anomalies"],
                index=index_display,
            )
        )

    lines.extend(["", "## 未生成原因", "", "| 原因 | 数量 |", "|---|---:|"])
    for reason, count in reason_counter.most_common():
        lines.append(f"| {md_escape(reason)} | {count} |")
    if not reason_counter:
        lines.append("| 无 | 0 |")

    lines.extend(
        [
            "",
            "## 异常提示不为空的医生清单",
            "",
            "| 序号 | 医院 | 姓名 | 科室 | 职称/身份 | 异常提示 | 来源 |",
            "|---:|---|---|---|---|---|---|",
        ]
    )
    for row in anomalies:
        source = clean(row.get("来源链接"))
        lines.append(
            "| {seq} | {hospital} | {name} | {dept} | {title} | {anomaly} | [官网]({source}) |".format(
                seq=md_escape(row.get("序号")),
                hospital=md_escape(row.get("医院")),
                name=md_escape(row.get("姓名")),
                dept=md_escape(department(row)),
                title=md_escape(title(row)),
                anomaly=md_escape(row.get("异常提示")),
                source=source,
            )
        )

    lines.extend(
        [
            "",
            "## 生成过程质量提示",
            "",
            "以下记录的 `姓名` 字段异常过长，但总底表 `异常提示` 为空。本脚本不修改底表，已按总底表字段生成画像，并使用截断文件名避免 Windows 路径超限；建议后续回到底表复核。",
            "",
            "| 序号 | 医院 | 姓名字段摘录 | 科室 | 来源 |",
            "|---:|---|---|---|---|",
        ]
    )
    for row in suspicious_names:
        source = clean(row.get("来源链接"))
        lines.append(
            "| {seq} | {hospital} | {name} | {dept} | [官网]({source}) |".format(
                seq=md_escape(row.get("序号")),
                hospital=md_escape(row.get("医院")),
                name=md_escape(clip(row.get("姓名"), 120)),
                dept=md_escape(department(row)),
                source=source,
            )
        )
    if not suspicious_names:
        lines.append("| 无 |  |  |  |  |")

    lines.extend(
        [
            "",
            "## 未生成明细",
            "",
            "| 序号 | 医院 | 姓名 | 科室 | 未生成原因 |",
            "|---:|---|---|---|---|",
        ]
    )
    for row in skipped_rows:
        lines.append(
            "| {seq} | {hospital} | {name} | {dept} | {reason} |".format(
                seq=md_escape(row.get("序号")),
                hospital=md_escape(row.get("医院")),
                name=md_escape(row.get("姓名")),
                dept=md_escape(department(row)),
                reason=md_escape(row.get("_skip_reason")),
            )
        )

    lines.extend(
        [
            "",
            "## 内容边界",
            "",
            "- 本次只读取统一总底表中的医院官网字段。",
            "- 未访问第三方平台，未采集患者评价、排名或隐私信息。",
            "- 医生画像不写疗效承诺、治愈保证或夸大宣传。",
            "- 已存在的医生 Markdown 默认视为可能人工精修，未获管理员授权不覆盖。",
            "",
        ]
    )
    return "\n".join(lines)


def extract_markdown_url(value: str) -> str:
    match = re.search(r"\((https?://[^)]+)\)", value)
    if match:
        return clean(match.group(1))
    return clean(value)


def split_markdown_table_row(line: str) -> list[str]:
    value = line.strip()
    if value.startswith("|"):
        value = value[1:]
    if value.endswith("|"):
        value = value[:-1]
    cells: list[str] = []
    current: list[str] = []
    bracket_depth = 0
    for char in value:
        if char == "[":
            bracket_depth += 1
        elif char == "]" and bracket_depth:
            bracket_depth -= 1
        if char == "|" and bracket_depth == 0:
            cells.append(clean("".join(current)))
            current = []
        else:
            current.append(char)
    cells.append(clean("".join(current)))
    return cells


def report_section(text: str, heading: str, next_heading: str) -> str:
    start_marker = f"## {heading}"
    next_marker = f"## {next_heading}"
    start = text.find(start_marker)
    if start == -1:
        return ""
    end = text.find(next_marker, start + len(start_marker))
    if end == -1:
        end = len(text)
    return text[start:end]


def parse_markdown_table(section_text: str) -> list[dict[str, str]]:
    lines = [line for line in section_text.splitlines() if line.strip().startswith("|")]
    if len(lines) < 3:
        return []
    headers = split_markdown_table_row(lines[0])
    rows: list[dict[str, str]] = []
    for line in lines[2:]:
        if re.match(r"^\|\s*-+", line):
            continue
        cells = split_markdown_table_row(line)
        if len(cells) != len(headers):
            continue
        rows.append(dict(zip(headers, cells)))
    return rows


def parse_report_corrections(report_path: Path) -> dict[str, dict[str, str]]:
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    corrections: dict[str, dict[str, str]] = {}

    abnormal_section = report_section(text, "异常提示不为空的医生清单", "生成过程质量提示")
    for item in parse_markdown_table(abnormal_section):
        source = extract_markdown_url(item.get("来源", ""))
        if not source:
            continue
        official_text = html_to_markdown_text(item.get("异常提示", ""))
        corrections[source] = {
            "kind": "异常提示官方内容补充",
            "序号": clean(item.get("序号")),
            "医院": clean(item.get("医院")),
            "姓名": clean(item.get("姓名")),
            "科室": clean(item.get("科室")),
            "职称": clean(item.get("职称/身份")),
            "官方正文": official_text,
        }

    quality_section = report_section(text, "生成过程质量提示", "未生成明细")
    for item in parse_markdown_table(quality_section):
        source = extract_markdown_url(item.get("来源", ""))
        if not source:
            continue
        existing = corrections.setdefault(source, {"kind": "姓名字段修正"})
        existing["kind"] = "；".join(unique_terms([existing.get("kind", "")], ["姓名字段修正"]))
        existing["序号"] = clean(item.get("序号")) or existing.get("序号", "")
        existing["医院"] = clean(item.get("医院")) or existing.get("医院", "")
        existing["姓名"] = clean(item.get("姓名字段摘录")) or existing.get("姓名", "")
        existing["科室"] = clean(item.get("科室")) or existing.get("科室", "")

    return corrections


def corrected_row(row: dict[str, str], correction: dict[str, str]) -> dict[str, str]:
    updated = dict(row)
    if clean(correction.get("姓名")):
        updated["姓名"] = clean(correction.get("姓名"))
    if clean(correction.get("科室")):
        updated["科室_分类页"] = clean(correction.get("科室"))
        updated["科室_列表卡片"] = clean(correction.get("科室"))
    if clean(correction.get("职称")):
        updated["职称身份原文"] = clean(correction.get("职称"))
    if clean(correction.get("官方正文")):
        updated["_报告修正正文"] = clean(correction.get("官方正文"))
        updated["擅长诊疗方向摘录"] = clean(correction.get("官方正文"))
        updated["详情正文摘录"] = clean(correction.get("官方正文"))
        updated["_原异常提示"] = clean(row.get("异常提示"))
        updated["异常提示"] = ""
    updated["_报告修正类型"] = clean(correction.get("kind"))
    return updated


def map_rows_by_source(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    return {clean(row.get("来源链接")): row for row in rows if clean(row.get("来源链接"))}


def choose_corrected_profile_path(row: dict[str, str], hospital_dir: Path, existing_path: Path | None) -> Path:
    for candidate in candidate_names(row, int(clean(row.get("序号")) or "0")):
        path = hospital_dir / candidate
        if existing_path and path == existing_path:
            return path
        if not path.exists():
            return path
    if existing_path:
        return existing_path
    return hospital_dir / f"{safe_filename(row.get('姓名'))}_{source_hash(clean(row.get('来源链接')))}.md"


def prepare_corrected_profile_path(
    row: dict[str, str],
    output_root: Path,
    allow_overwrite: bool,
) -> tuple[Path | None, str, Path | None]:
    hospital_dir = output_root / safe_filename(row.get("医院"))
    hospital_dir.mkdir(parents=True, exist_ok=True)
    existing_sources = extract_existing_sources(hospital_dir)
    source = clean(row.get("来源链接"))
    existing_path = existing_sources.get(source)

    if existing_path and not allow_overwrite and not is_auto_generated_profile(existing_path):
        return None, "已有同来源人工画像，默认保护不覆盖", existing_path

    target_path = choose_corrected_profile_path(row, hospital_dir, existing_path)
    if target_path.exists() and target_path != existing_path and not allow_overwrite:
        return None, f"目标文件已存在，默认保护不覆盖：{target_path.name}", existing_path

    if existing_path and existing_path != target_path:
        if not is_auto_generated_profile(existing_path) and not allow_overwrite:
            return None, "已有画像不是本脚本自动生成，默认不重命名", existing_path
        existing_path.rename(target_path)
    return target_path, "", existing_path


def build_index_from_existing_profiles(
    hospital: str,
    hospital_rows: list[dict[str, str]],
    output_root: Path,
    generated_at: str,
) -> tuple[str, Path]:
    hospital_dir = output_root / safe_filename(hospital)
    existing_sources = extract_existing_sources(hospital_dir)
    rows_for_index: list[dict[str, str]] = []
    skipped_rows: list[dict[str, str]] = []
    path_by_source: dict[str, Path] = {}
    for row in hospital_rows:
        source = clean(row.get("来源链接"))
        path = existing_sources.get(source)
        if path:
            row["_profile_seq"] = str(len(rows_for_index) + 1)
            rows_for_index.append(row)
            path_by_source[source] = path
        else:
            skipped = dict(row)
            skipped["_skip_reason"] = "未找到对应画像文件"
            skipped_rows.append(skipped)
    index_path = hospital_dir / "_索引.md"
    return build_index(hospital, rows_for_index, skipped_rows, path_by_source, generated_at), index_path


def build_supplement_report(
    report_path: Path,
    source_report: Path,
    updated_rows: list[dict[str, str]],
    skipped: list[dict[str, str]],
    index_paths: list[Path],
    generated_at: str,
) -> str:
    lines = [
        "---",
        '类型: "画像补充生成报告"',
        f"生成时间: {yaml_value(generated_at)}",
        f"修正来源报告: {yaml_value(str(source_report))}",
        f"模板: {yaml_value(str(TEMPLATE_PATH))}",
        "---",
        AUTO_MARKER,
        "",
        "# Obsidian 医生画像补充生成报告",
        "",
        "## 总览",
        "",
        f"- 修正来源报告：`{source_report}`",
        f"- 沿用模板：`{TEMPLATE_PATH}`",
        f"- 已刷新画像：{len(updated_rows)}",
        f"- 跳过记录：{len(skipped)}",
        f"- 重建索引：{len(index_paths)}",
        "",
        "## 已刷新画像",
        "",
        "| 序号 | 医院 | 姓名 | 科室 | 修正类型 | 来源 |",
        "|---:|---|---|---|---|---|",
    ]
    for row in updated_rows:
        source = clean(row.get("来源链接"))
        lines.append(
            "| {seq} | {hospital} | {name} | {dept} | {kind} | [官网]({source}) |".format(
                seq=md_escape(row.get("序号")),
                hospital=md_escape(row.get("医院")),
                name=md_escape(row.get("姓名")),
                dept=md_escape(department(row)),
                kind=md_escape(row.get("_报告修正类型")),
                source=source,
            )
        )
    lines.extend(["", "## 跳过记录", "", "| 序号 | 医院 | 姓名 | 原因 |", "|---:|---|---|---|"])
    if skipped:
        for row in skipped:
            lines.append(
                "| {seq} | {hospital} | {name} | {reason} |".format(
                    seq=md_escape(row.get("序号")),
                    hospital=md_escape(row.get("医院")),
                    name=md_escape(row.get("姓名")),
                    reason=md_escape(row.get("_skip_reason")),
                )
            )
    else:
        lines.append("|  |  |  | 无 |")
    lines.extend(["", "## 重建索引", ""])
    lines.extend(f"- `{path}`" for path in index_paths)
    lines.append("")
    return "\n".join(lines)


def apply_report_corrections(
    rows: list[dict[str, str]],
    output_root: Path,
    report_path: Path,
    supplement_report_path: Path,
    allow_overwrite: bool,
) -> dict[str, object]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    corrections = parse_report_corrections(report_path)
    rows_by_source = map_rows_by_source(rows)
    corrected_by_source: dict[str, dict[str, str]] = {}
    skipped: list[dict[str, str]] = []
    updated_rows: list[dict[str, str]] = []
    moved_files: list[dict[str, str]] = []

    for source, correction in corrections.items():
        source_row = rows_by_source.get(source)
        if source_row is None:
            skipped.append({"序号": correction.get("序号", ""), "医院": correction.get("医院", ""), "姓名": correction.get("姓名", ""), "_skip_reason": "总底表未找到同来源记录"})
            continue
        row = corrected_row(source_row, correction)
        target_path, reason, previous_path = prepare_corrected_profile_path(row, output_root, allow_overwrite)
        if target_path is None:
            row["_skip_reason"] = reason
            skipped.append(row)
            continue
        target_path.write_text(build_profile(row, generated_at), encoding="utf-8", newline="\n")
        if previous_path and previous_path != target_path:
            moved_files.append({"from": str(previous_path), "to": str(target_path)})
        corrected_by_source[source] = row
        updated_rows.append(row)

    rows_with_corrections = []
    for row in rows:
        source = clean(row.get("来源链接"))
        rows_with_corrections.append(corrected_by_source.get(source, row))

    affected_hospitals = sorted({clean(row.get("医院")) for row in updated_rows})
    index_paths: list[Path] = []
    for hospital in affected_hospitals:
        hospital_rows = [row for row in rows_with_corrections if clean(row.get("医院")) == hospital]
        index_text, index_path = build_index_from_existing_profiles(hospital, hospital_rows, output_root, generated_at)
        index_path.write_text(index_text, encoding="utf-8", newline="\n")
        index_paths.append(index_path)

    supplement_report_path.write_text(
        build_supplement_report(
            supplement_report_path,
            report_path,
            updated_rows,
            skipped,
            index_paths,
            generated_at,
        ),
        encoding="utf-8",
        newline="\n",
    )

    return {
        "corrections_found": len(corrections),
        "profiles_refreshed": len(updated_rows),
        "skipped": len(skipped),
        "indexes_rebuilt": len(index_paths),
        "moved_files": moved_files,
        "supplement_report": str(supplement_report_path),
    }


def build_missing_generation_report(
    report_path: Path,
    generated_rows: list[dict[str, str]],
    skipped_rows: list[dict[str, str]],
    index_paths: list[Path],
    generated_at: str,
) -> str:
    hospital_counter = Counter(clean(row.get("医院")) for row in generated_rows)
    anomaly_counter = Counter(clean(row.get("医院")) for row in generated_rows if clean(row.get("异常提示")))
    reason_counter = Counter(clean(row.get("_skip_reason")) for row in skipped_rows)
    lines = [
        "---",
        '类型: "缺失画像补充生成报告"',
        f"生成时间: {yaml_value(generated_at)}",
        f"模板: {yaml_value(str(TEMPLATE_PATH))}",
        "---",
        AUTO_MARKER,
        "",
        "# Obsidian 缺失医生画像补充生成报告",
        "",
        "## 总览",
        "",
        f"- 新生成缺失画像：{len(generated_rows)}",
        f"- 跳过记录：{len(skipped_rows)}",
        f"- 重建索引：{len(index_paths)}",
        f"- 沿用模板：`{TEMPLATE_PATH}`",
        "",
        "## 按医院统计",
        "",
        "| 医院 | 新生成画像 | 异常提示不为空 |",
        "|---|---:|---:|",
    ]
    for hospital in sorted(hospital_counter):
        lines.append(f"| {md_escape(hospital)} | {hospital_counter[hospital]} | {anomaly_counter[hospital]} |")
    lines.extend(["", "## 跳过原因", "", "| 原因 | 数量 |", "|---|---:|"])
    if reason_counter:
        for reason, count in reason_counter.most_common():
            lines.append(f"| {md_escape(reason)} | {count} |")
    else:
        lines.append("| 无 | 0 |")
    lines.extend(["", "## 重建索引", ""])
    lines.extend(f"- `{path}`" for path in index_paths)
    lines.append("")
    return "\n".join(lines)


def generate_missing_profiles(
    rows: list[dict[str, str]],
    output_root: Path,
    skip_hospitals: set[str],
    report_path: Path,
) -> dict[str, object]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_by_hospital: dict[str, list[dict[str, str]]] = defaultdict(list)
    for index, row in enumerate(rows, start=1):
        row.setdefault("序号", str(index))
        rows_by_hospital[clean(row.get("医院"))].append(row)

    generated_rows: list[dict[str, str]] = []
    skipped_rows: list[dict[str, str]] = []
    generated_paths: list[Path] = []
    affected_hospitals: set[str] = set()

    for hospital, hospital_rows in sorted(rows_by_hospital.items()):
        if hospital in skip_hospitals:
            continue
        hospital_dir = output_root / safe_filename(hospital)
        hospital_dir.mkdir(parents=True, exist_ok=True)
        existing_sources = extract_existing_sources(hospital_dir)
        used_names = {path.name for path in hospital_dir.glob("*.md") if path.name != "_索引.md"}

        for fallback_index, row in enumerate(hospital_rows, start=1):
            reason = skip_reason_for_core_fields(row)
            if reason:
                skipped = dict(row)
                skipped["_skip_reason"] = reason
                skipped_rows.append(skipped)
                continue
            source = clean(row.get("来源链接"))
            if source in existing_sources:
                continue

            path: Path | None = None
            for candidate in candidate_names(row, fallback_index):
                if candidate in used_names:
                    continue
                candidate_path = hospital_dir / candidate
                if candidate_path.exists():
                    continue
                path = candidate_path
                used_names.add(candidate)
                break
            if path is None:
                path = hospital_dir / f"{safe_filename(row.get('姓名'))}_{source_hash(source)}.md"
                if path.name in used_names or path.exists():
                    skipped = dict(row)
                    skipped["_skip_reason"] = "无法生成不冲突的文件名"
                    skipped_rows.append(skipped)
                    continue
                used_names.add(path.name)

            row["_profile_seq"] = str(len(generated_rows) + 1)
            path.write_text(build_profile(row, generated_at), encoding="utf-8", newline="\n")
            existing_sources[source] = path
            generated_rows.append(row)
            generated_paths.append(path)
            affected_hospitals.add(hospital)

    index_paths: list[Path] = []
    for hospital in sorted(affected_hospitals):
        hospital_rows = [row for row in rows if clean(row.get("医院")) == hospital]
        index_text, index_path = build_index_from_existing_profiles(hospital, hospital_rows, output_root, generated_at)
        index_path.write_text(index_text, encoding="utf-8", newline="\n")
        index_paths.append(index_path)

    report_path.write_text(
        build_missing_generation_report(report_path, generated_rows, skipped_rows, index_paths, generated_at),
        encoding="utf-8",
        newline="\n",
    )

    return {
        "generated_missing_profiles": len(generated_paths),
        "skipped": len(skipped_rows),
        "indexes_rebuilt": len(index_paths),
        "affected_hospitals": sorted(affected_hospitals),
        "report": str(report_path),
    }


def hospital_stat_template(source_count: int) -> dict[str, object]:
    return {
        "source": source_count,
        "generated": 0,
        "skipped": 0,
        "anomalies": 0,
        "index_path": "",
    }


def generate_profiles(
    rows: list[dict[str, str]],
    source_path: Path,
    output_root: Path,
    report_path: Path,
    skip_hospitals: set[str],
    allow_overwrite: bool,
) -> dict[str, object]:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_by_hospital: dict[str, list[dict[str, str]]] = defaultdict(list)
    for index, row in enumerate(rows, start=1):
        row.setdefault("序号", str(index))
        rows_by_hospital[clean(row.get("医院"))].append(row)

    hospital_stats = {
        hospital: hospital_stat_template(len(hospital_rows))
        for hospital, hospital_rows in rows_by_hospital.items()
    }
    all_skipped_rows: list[dict[str, str]] = []
    generated_paths: list[Path] = []
    index_paths: list[Path] = []

    for hospital, hospital_rows in sorted(rows_by_hospital.items()):
        if hospital in skip_hospitals:
            for row in hospital_rows:
                row["_skip_reason"] = "整院跳过：该医院医生画像已生成"
                all_skipped_rows.append(row)
            hospital_stats[hospital]["skipped"] = len(hospital_rows)
            hospital_stats[hospital]["anomalies"] = sum(1 for row in hospital_rows if clean(row.get("异常提示")))
            continue

        hospital_dir = output_root / safe_filename(hospital)
        hospital_dir.mkdir(parents=True, exist_ok=True)
        existing_sources = extract_existing_sources(hospital_dir)
        existing_names = {path.name for path in hospital_dir.glob("*.md") if path.name != "_索引.md"}
        used_names = set(existing_names)
        created_rows: list[dict[str, str]] = []
        skipped_rows: list[dict[str, str]] = []
        path_by_source: dict[str, Path] = {}

        for fallback_index, row in enumerate(hospital_rows, start=1):
            reason = skip_reason_for_core_fields(row)
            if reason:
                row["_skip_reason"] = reason
                skipped_rows.append(row)
                all_skipped_rows.append(row)
                continue

            path, reason = select_profile_path(
                row=row,
                hospital_dir=hospital_dir,
                used_names=used_names,
                existing_names=existing_names,
                existing_sources=existing_sources,
                fallback_index=fallback_index,
                allow_overwrite=allow_overwrite,
            )
            if path is None:
                row["_skip_reason"] = reason
                skipped_rows.append(row)
                all_skipped_rows.append(row)
                continue

            row["_profile_seq"] = str(len(created_rows) + 1)
            path.write_text(build_profile(row, generated_at), encoding="utf-8", newline="\n")
            created_rows.append(row)
            generated_paths.append(path)
            path_by_source[clean(row.get("来源链接"))] = path

        index_path = hospital_dir / "_索引.md"
        index_path.write_text(
            build_index(hospital, created_rows, skipped_rows, path_by_source, generated_at),
            encoding="utf-8",
            newline="\n",
        )
        index_paths.append(index_path)

        hospital_stats[hospital]["generated"] = len(created_rows)
        hospital_stats[hospital]["skipped"] = len(skipped_rows)
        hospital_stats[hospital]["anomalies"] = sum(1 for row in hospital_rows if clean(row.get("异常提示")))
        hospital_stats[hospital]["index_path"] = str(index_path)

    report_text = build_report(
        rows=rows,
        source_path=source_path,
        report_path=report_path,
        output_root=output_root,
        hospital_stats=hospital_stats,
        skipped_rows=all_skipped_rows,
        generated_at=generated_at,
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(report_text, encoding="utf-8", newline="\n")

    return {
        "source": str(source_path),
        "output_root": str(output_root),
        "report": str(report_path),
        "total_rows": len(rows),
        "generated_profiles": len(generated_paths),
        "generated_indexes": len(index_paths),
        "skipped_rows": len(all_skipped_rows),
        "skip_hospitals": sorted(skip_hospitals),
        "hospital_stats": hospital_stats,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从统一总底表生成 Obsidian 医生画像 Markdown。")
    parser.add_argument("--source", help="统一总底表路径，支持 .csv 或 .xlsx；默认优先读取 CSV。")
    parser.add_argument("--output-root", default=str(PROFILE_ROOT), help="Obsidian 医生画像输出根目录。")
    parser.add_argument("--report", default=str(DEFAULT_REPORT), help="全量画像生成报告路径。")
    parser.add_argument("--supplement-report", default=str(DEFAULT_SUPPLEMENT_REPORT), help="补充生成报告路径。")
    parser.add_argument("--missing-report", default=str(DEFAULT_MISSING_REPORT), help="缺失画像补充生成报告路径。")
    parser.add_argument("--skip-hospital", action="append", default=[], help="额外整院跳过的医院，可重复传入。")
    parser.add_argument(
        "--apply-report-corrections",
        action="store_true",
        help="读取管理员手工修正后的全量报告，只刷新报告修正项对应的画像。",
    )
    parser.add_argument(
        "--generate-missing-only",
        action="store_true",
        help="只生成当前画像仓库中尚不存在的医生画像，已有画像不改动。",
    )
    parser.add_argument(
        "--generate-default-skipped-hospitals",
        action="store_true",
        help="生成默认跳过医院；当前默认跳过中山大学附属第五医院。",
    )
    parser.add_argument(
        "--admin-allow-overwrite",
        action="store_true",
        help="管理员明确授权后才可覆盖已有医生 Markdown。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_path = choose_source(args.source)
    rows = read_master_rows(source_path)
    missing_fields = validate_source_rows(rows)
    if missing_fields:
        raise SystemExit("总底表缺少核心字段：" + "、".join(missing_fields))

    if args.apply_report_corrections:
        result = apply_report_corrections(
            rows=rows,
            output_root=Path(args.output_root),
            report_path=Path(args.report),
            supplement_report_path=Path(args.supplement_report),
            allow_overwrite=args.admin_allow_overwrite,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    skip_hospitals = set(args.skip_hospital)
    if not args.generate_default_skipped_hospitals:
        skip_hospitals.update(DEFAULT_SKIP_HOSPITALS)

    if args.generate_missing_only:
        result = generate_missing_profiles(
            rows=rows,
            output_root=Path(args.output_root),
            skip_hospitals=skip_hospitals,
            report_path=Path(args.missing_report),
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    result = generate_profiles(
        rows=rows,
        source_path=source_path,
        output_root=Path(args.output_root),
        report_path=Path(args.report),
        skip_hospitals=skip_hospitals,
        allow_overwrite=args.admin_allow_overwrite,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
